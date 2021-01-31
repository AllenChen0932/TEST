# -*- coding: utf-8 -*-
# @Time    : 2019/11/24 16:59
# @Author  : CE
# @File    : LoadSummaryTable.py

import sys, os
import configparser
from PyQt5.QtWidgets import (QApplication,
                             QMainWindow,
                             QDesktopWidget,
                             QLabel,
                             qApp,
                             QAction,
                             QLineEdit,
                             QPushButton,
                             QFileDialog,
                             QGridLayout,
                             QVBoxLayout,
                             QHBoxLayout,
                             QWidget,
                             QGroupBox,
                             QCheckBox,
                             QRadioButton,
                             QMessageBox)
from PyQt5 import QtCore
from PyQt5.QtGui import QIcon, QFont
import logging
import traceback

from openpyxl.styles import Side, Border, Alignment, Font
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from tool.load_summary.result_output_v3 import excel_operation as new
from tool.load_summary.result_append import excel_operation as append

__version__ = "2.1.1"
'''
2020.4.21_v2.0.1 repair the bug from APE
2020.4.21_v2.0.2 repair the bug from Ultimate tower output
2020.6.06_v2.1.1 repair the bug from fatigue tower output
'''

class MyQLineEdit(QLineEdit):
    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if event.mimeData().text():
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event): # 放下文件后的动作
        file_path = event.mimeData().text()
        if '///' in file_path:
            self.setText(file_path[8:].replace('/','\\'))   # 删除local path多余开头
        else:
            self.setText(file_path[5:].replace('/','\\'))   # 删除network path多余开头

class LoadSumWindow(QMainWindow):

    def __init__(self, parent=None):

        super(LoadSumWindow, self).__init__(parent)
        self.mywidget = QWidget(self)

        self.setMinimumSize(700, 600)
        # self.resize(800, 500)
        self.setWindowTitle('Load Summary Table')
        self.setWindowIcon(QIcon("./Icon/sheet.ico"))
        # root = QFileInfo(__file__).absolutePath()
        # self.setWindowIcon(QIcon(root + "./icon/Text_Edit.ico"))

        self.xls_temp   = None      #excel template
        self.out_path   = None      #load summary table
        self.xls_name   = None      #the name of a new load summary table
        self.xls_path   = None      #the path for the new load summary table

        self.ultimate = False
        self.fatigue  = False
        self.heat_map = False
        self.fat_wind = False
        self.fat_case = False
        self.aep_dlc  = False
        self.plot_ldd = False
        self.plot_lrd = False
        self.pitch_bl = False
        self.main_bl  = False
        self.gearbox  = False
        self.nacelle  = False
        self.main_all = True

        self.ult_list = list()
        self.fat_list = list()
        self.lct_list = list()
        self.dlc_list = list()
        self.loopname = list()
        self.ldd_list = list()
        self.lrd_list = list()
        self.nac_list = list()
        self.pst_list = list()
        self.pbl_list = list()     # pitch bearing life
        self.get_list = list()     # gearbox equivalent torque
        self.mbl_list = list()     # main bearing life

        self.title_font = QFont("Calibri", 10)
        self.title_font.setItalic(True)
        self.title_font.setBold(True)

        self.cont_font = QFont("Calibri", 9)
        self.cont_font.setItalic(False)
        self.cont_font.setBold(False)

        self.initUI()
        self.load_setting()
        self.center()
        # self.display()

    def initUI(self):

        menubar = self.menuBar()
        fileMenu = menubar.addMenu('File')
        helpMenu = menubar.addMenu('Help')

        excelAction = QAction(QIcon('Icon/Excel.ico'), 'Template', self)
        excelAction.triggered.connect(self.generate_template)
        fileMenu.addAction(excelAction)

        saveAction = QAction(QIcon('Icon/save.ico'), 'Save', self)
        saveAction.setShortcut('Ctrl+S')
        saveAction.triggered.connect(self.save_setting)
        fileMenu.addAction(saveAction)

        clearAction = QAction(QIcon('Icon/clear.ico'), 'Clear', self)
        clearAction.setShortcut('Ctrl+R')
        clearAction.triggered.connect(self.clear_setting)
        fileMenu.addAction(clearAction)

        exitAction = QAction(QIcon('Icon/exit.ico'), 'Exit', self)
        exitAction.setShortcut('Ctrl+Q')
        exitAction.triggered.connect(qApp.quit)
        fileMenu.addAction(exitAction)

        helpAction = QAction(QIcon('Icon/doc.ico'), 'User Manual', self)
        helpAction.triggered.connect(self.user_manual)
        helpMenu.addAction(helpAction)

        self.main_layout = QVBoxLayout()  # 创建主部件的网格布局

        # ******************************* input groupbox **********************************
        self.gbox1 = QGroupBox('Input')
        self.gbox1.setFont(self.title_font)

        self.lab1 = QLabel('Excel Template')
        self.lab1.setFont(self.cont_font)
        self.lin1 = MyQLineEdit()
        self.lin1.setFont(self.cont_font)
        self.lin1.setPlaceholderText('Choose the excel template')
        self.btn1 = QPushButton("...")
        self.btn1.clicked.connect(self.load_template)

        # set layout
        self.glayout1 = QGridLayout()
        self.glayout1.addWidget(self.lab1, 0, 0, 1, 1)
        self.glayout1.addWidget(self.lin1, 0, 1, 1, 3)
        self.glayout1.addWidget(self.btn1, 0, 4, 1, 1)

        self.gbox1.setLayout(self.glayout1)

        # ******************************* output groupbox **********************************
        self.gbox5 = QGroupBox('Output')
        self.gbox5.setFont(self.title_font)

        # sub1
        self.gbox5_sub1 = QGroupBox()

        # two options
        self.radio_main = QRadioButton('Main result')
        self.radio_main.setFont(self.cont_font)
        self.radio_main.setChecked(True)
        self.radio_main.toggled.connect(self.main_option)
        self.radio_all  = QRadioButton('All result')
        self.radio_all.setFont(self.cont_font)
        self.radio_all.toggled.connect(self.all_option)
        if self.radio_all.isChecked():
            self.main_all = False

        self.lay_sub1 = QGridLayout()
        self.lay_sub1.addWidget(self.radio_main, 0, 0, 1, 2)
        self.lay_sub1.addWidget(self.radio_all,  0, 3, 1, 3)
        self.gbox5_sub1.setLayout(self.lay_sub1)

        # sub2
        self.gbox5_sub2 = QGroupBox()

        # two options
        self.radio_new = QRadioButton('New table')
        self.radio_new.setFont(self.cont_font)
        self.radio_new.setChecked(True)
        self.radio_new.toggled.connect(self.new_option)
        self.radio_app = QRadioButton('Append')
        self.radio_app.setFont(self.cont_font)
        self.radio_app.setChecked(False)
        # self.radio_app.setDisabled(True)
        self.radio_app.toggled.connect(self.append_option)

        self.lab2 = QLabel('Excel Directory')
        self.lab2.setFont(self.cont_font)
        self.lin2 = MyQLineEdit()
        self.lin2.setFont(self.cont_font)
        self.lin2.setPlaceholderText('Choose the path to save result')
        self.btn2 = QPushButton("...")
        self.btn2.clicked.connect(self.load_output)

        self.lab3 = QLabel('Excel Name')
        self.lab3.setFont(self.cont_font)
        self.lin3 = QLineEdit()
        self.lin3.setFont(self.cont_font)
        self.lin3.setPlaceholderText('Enter the name of the excel')

        self.lab4 = QLabel('Excel Path')
        self.lab4.setFont(self.cont_font)
        self.lin4 = QLineEdit()
        self.lin4.setFont(self.cont_font)
        self.lin4.setPlaceholderText('Choose excel to append result')
        self.btn3 = QPushButton("...")
        self.btn3.clicked.connect(self.choose_excel)
        self.lin4.setDisabled(True)
        self.btn3.setDisabled(True)

        # set layout
        # self.lay_sub1 = QGridLayout()
        # self.lay_sub1.addWidget(self.radio_new, 0, 0, 1, 1)
        # self.lay_sub1.addWidget(self.radio_app, 0, 2, 1, 1)

        self.lay_sub2 = QGridLayout()
        self.lay_sub2.addWidget(self.radio_new, 0, 0, 1, 1)
        self.lay_sub2.addWidget(self.radio_app, 0, 2, 1, 1)
        self.lay_sub2.addWidget(self.lab2, 1, 0, 1, 1)
        self.lay_sub2.addWidget(self.lin2, 1, 1, 1, 3)
        self.lay_sub2.addWidget(self.btn2, 1, 4, 1, 1)
        self.lay_sub2.addWidget(self.lab3, 2, 0, 1, 1)
        self.lay_sub2.addWidget(self.lin3, 2, 1, 1, 3)
        self.lay_sub2.addWidget(self.lab4, 3, 0, 1, 1)
        self.lay_sub2.addWidget(self.lin4, 3, 1, 1, 3)
        self.lay_sub2.addWidget(self.btn3, 3, 4, 1, 1)

        # self.gbox5_sub1.setLayout(self.lay_sub1)
        # self.gbox5_sub2.setLayout(self.lay_sub2)

        self.glayout5 = QVBoxLayout()
        # self.glayout5.addWidget(self.gbox5_sub1)
        # self.glayout5.addWidget(self.gbox5_sub2)
        # self.glayout5.addStretch(1)

        # self.gbox5.setLayout(self.glayout5)
        self.gbox5.setLayout(self.lay_sub2)

        # ******************************* main function groupbox **********************************
        self.gbox2 = QGroupBox('Ultimate Function')
        self.gbox2.setFont(self.title_font)

        self.cb_ultimate = QCheckBox("Ulitmate")
        self.cb_ultimate.setFont(self.cont_font)
        # self.cb_ultimate.setChecked(True)
        self.cb_fatigue  = QCheckBox('Fatigue')
        self.cb_fatigue.setFont(self.cont_font)
        # self.cb_fatigue.setChecked(True)
        self.cb_heatmap = QCheckBox('Heat Map')
        self.cb_heatmap.setFont(self.cont_font)
        # self.cb_aepdlc12.font(QFont('微软雅黑', 9))
        # self.checkbox4 = QCheckBox('AEP')

        self.checkbox8 = QLabel()
        self.checkbox9 = QLabel()
        self.checkbox10 = QLabel()
        # self.checkbox8.setDisabled(True)

        self.glayout2 = QGridLayout()
        self.glayout2.addWidget(self.cb_ultimate, 0, 0, 1, 1)
        # self.glayout2.addWidget(self.cb_fatigue,  1, 0, 1, 1)
        self.glayout2.addWidget(self.cb_heatmap,  1, 0, 1, 1)
        # self.glayout2.addWidget(self.checkbox4, 3, 0, 1, 1)
        self.glayout2.addWidget(self.checkbox8,  2, 0, 1, 1)
        # self.glayout2.addWidget(self.checkbox9,  3, 0, 1, 1)

        self.gbox2.setLayout(self.glayout2)

        # ******************************* fatigue function groupbox **********************************
        self.gbox3 = QGroupBox('Fatigue Function')
        self.gbox3.setFont(self.title_font)

        self.cb_fatwind = QCheckBox('Fatigue per wind')
        self.cb_fatwind.setFont(self.cont_font)
        self.cb_fatdlc  = QCheckBox('Fatigue per DLC')
        self.cb_fatdlc.setFont(self.cont_font)
        self.cb_fatwave = QCheckBox('Fatigue per wave')
        self.cb_fatwave.setFont(self.cont_font)
        self.cb_fatwave.setDisabled(True)


        self.glayout3 = QGridLayout()
        self.glayout3.addWidget(self.cb_fatigue, 0, 0, 1, 1)
        self.glayout3.addWidget(self.cb_fatwind, 1, 0, 1, 1)
        self.glayout3.addWidget(self.cb_fatdlc,  2, 0, 1, 1)
        # self.glayout3.addWidget(self.cb_fatwave, 2, 0, 1, 1)
        # self.glayout3.addWidget(self.checkbox10, 3, 0, 1, 1)

        self.gbox3.setLayout(self.glayout3)

        # ******************************* other function groupbox **********************************
        self.gbox4 = QGroupBox(' Other function')
        self.gbox4.setFont(self.title_font)
        self.cb_aepdlc12 = QCheckBox('Annual Energy Production')
        self.cb_aepdlc12.setFont(self.cont_font)
        # self.cb_towercl = QCheckBox('Tower Clearance')
        # self.cb_towercl.setFont(self.cont_font)
        # self.cb_towercl.setDisabled(True)
        self.cb_alarmid = QCheckBox('Alarm ID')
        self.cb_alarmid.setFont(self.cont_font)
        self.cb_alarmid.setDisabled(True)
        self.cb_plotlrd = QCheckBox('Plot Pitch Travel')
        self.cb_plotlrd.setFont(self.cont_font)
        # self.cb_plotlrd.setDisabled(True)
        self.cb_nacacce = QCheckBox('Nacelle Fore-Aft/Side-Side Acc')
        self.cb_nacacce.setFont(self.cont_font)
        self.cb_nacacce.setDisabled(False)
        self.cb_gearequ = QCheckBox('Gearbox Equivalent Torque')
        self.cb_gearequ.setFont(self.cont_font)
        self.cb_gearequ.setDisabled(False)
        self.cb_mainber = QCheckBox('Main Bearing Life(single bearing)')
        self.cb_mainber.setFont(self.cont_font)
        self.cb_mainber.setDisabled(False)
        self.cb_pitbear = QCheckBox('Pitch Bearing Life/Equivalent Torque')
        self.cb_pitbear.setFont(self.cont_font)
        self.cb_pitbear.setDisabled(False)
        self.cb_plotldd = QCheckBox('Plot MxHR LDD')
        self.cb_plotldd.setFont(self.cont_font)
        # self.cb_plotldd.setDisabled(True)
        self.cb_dynpowc = QCheckBox('Dynamic Power Curve')
        self.cb_dynpowc.setFont(self.cont_font)
        self.cb_dynpowc.setDisabled(False)
        self.glayout4 = QGridLayout()
        # self.glayout4.addWidget(self.cb_alarmid, 0, 0, 1, 1)
        self.glayout4.addWidget(self.cb_plotldd, 0, 0, 1, 1)
        self.glayout4.addWidget(self.cb_plotlrd, 1, 0, 1, 1)
        self.glayout4.addWidget(self.cb_dynpowc, 2, 0, 1, 1)
        self.glayout4.addWidget(self.cb_aepdlc12, 3, 0, 1, 1)
        # self.glayout4.addWidget(self.cb_towercl, 0, 1, 1, 1)
        self.glayout4.addWidget(self.cb_nacacce, 0, 1, 1, 1)
        self.glayout4.addWidget(self.cb_pitbear, 1, 1, 1, 1)
        self.glayout4.addWidget(self.cb_gearequ, 2, 1, 1, 1)
        self.glayout4.addWidget(self.cb_mainber, 3, 1, 1, 1)
        self.gbox4.setLayout(self.glayout4)

        # set main layout
        # check button
        self.btn_check = QPushButton('Check Input')
        self.btn_check.setFont(self.cont_font)
        self.btn_check.clicked.connect(self.check_input)

        # handle button
        self.btn_handle = QPushButton('Handle Data')
        self.btn_handle.setFont(self.cont_font)
        self.btn_handle.setDisabled(True)
        self.btn_handle.clicked.connect(self.handle)

        # reset button
        self.btn_reset = QPushButton('Reset Selection')
        self.btn_reset.setFont(self.cont_font)
        self.btn_reset.setDisabled(True)
        self.btn_reset.clicked.connect(self.reset)

        self.Hlayout = QHBoxLayout()
        self.Hlayout.addWidget(self.gbox2)
        self.Hlayout.addWidget(self.gbox3)

        self.main_layout.addWidget(self.gbox1)
        self.main_layout.addWidget(self.gbox5)
        self.main_layout.addLayout(self.Hlayout)
        self.main_layout.addWidget(self.gbox4)
        self.main_layout.addWidget(self.btn_check)
        self.main_layout.addWidget(self.btn_handle)
        self.main_layout.addWidget(self.btn_reset)
        self.main_layout.addStretch(0)

        self.mywidget.setLayout(self.main_layout)
        self.setCentralWidget(self.mywidget)

    def save_setting(self):
        config = configparser.ConfigParser()
        config.read('config.ini', encoding='utf-8')

        if not config.has_section('Load Summary'):
            config.add_section('Load Summary')

        config['Load Summary'] = {'Excel template':self.lin1.text(),
                                  'Excel directory':self.lin2.text(),
                                  'Excel name':self.lin3.text(),
                                  'Excel path':self.lin4.text()}

        with open("config.ini",'w') as f:
            config.write(f)

    def clear_setting(self):
        self.lin1.setText('')
        self.lin2.setText('')
        self.lin3.setText('')
        self.lin4.setText('')

    def load_setting(self):
        config = configparser.ConfigParser()
        config.read('config.ini', encoding='utf-8')

        if config.has_section('Load Summary'):
            self.lin1.setText(config.get('Load Summary', 'Excel template'))
            self.lin2.setText(config.get('Load Summary', 'Excel directory'))
            # self.lin2.home(True)
            self.lin3.setText(config.get('Load Summary', 'Excel name'))
            # self.lin3.home(True)
            self.lin4.setText(config.get('Load Summary', 'Excel path'))
            # self.lin4.home(True)

    def generate_template(self):

        template  = 'Load Summary Template.xlsx'
        path      = os.path.abspath('.')
        file_path = os.path.join(path, template)
        print(file_path)

        border = Border(left   =Side(border_style='thin', color='000000'),
                        right  =Side(border_style='thin', color='000000'),
                        top    =Side(border_style='thin', color='000000'),
                        bottom =Side(border_style='thin', color='000000'))

        if not os.path.isfile(file_path):

            table = Workbook()
            sheet = table['Sheet']
            sheet.title = 'Main'
            sheet['A1'] = 'Wind Turbine Name'
            sheet['B1'] = 'Category'
            sheet['C1'] = 'Path'
            # font for the first row
            sheet['A1'].font = Font(name='Microsoft Ya Hei', size=9, bold=True)
            sheet['B1'].font = Font(name='Microsoft Ya Hei', size=9, bold=True)
            sheet['C1'].font = Font(name='Microsoft Ya Hei', size=9, bold=True)
            # width for each column
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 10
            sheet.column_dimensions['C'].width = 70

            for col in sheet.iter_rows(min_row=1, max_col=3, max_row=1):
                for cell in col:
                    cell.alignment = Alignment(vertical='center', horizontal='center')

            for row in sheet.iter_rows(min_row=2, max_col=3, max_row=31):
                for cell in row:
                    cell.font = Font(name='Microsoft Ya Hei', size=9)
            #
            for row in sheet.iter_rows(min_row=2, max_col=1, max_row=31):
                for cell in row:
                    cell.alignment = Alignment(vertical='center', horizontal='center')

            # set data validation for column B
            dv = DataValidation(type='list', formula1='"Ultimate,Rainflow,Post,LCT,DLC12"')
            dv.prompt      = 'Make sure the selection is unique for each loop'
            dv.promptTitle = 'List Selection'
            sheet.add_data_validation(dv)
            dv.add('B2:B301')     # apply the validation to a range of cellls

            # border
            for i in range(1,52):
                for j in range(1,4):
                    sheet.cell(row=i,column=j).border = border

            # merge
            for i in range(10):
                sheet.merge_cells(start_row=5*i+2, start_column=1, end_row=5*i+6, end_column=1)
            table.save('Load Summary Template.xlsx')
            print('Template generated successfully!')

        # open template
        os.startfile(file_path)

    def user_manual(self):
        os.startfile(os.getcwd() + '\\' + 'User Manual_Load Summary.docx')

    def new_option(self):

        if self.radio_new.isChecked():

            # 初始化功能
            # self.lin1.setDisabled(False)
            # self.btn1.setDisabled(False)
            self.lin2.setDisabled(False)
            self.btn2.setDisabled(False)
            self.lin3.setDisabled(False)

            self.lin4.setDisabled(True)
            self.btn3.setDisabled(True)

            # restore the function for new option, if the append option has been choosed before
            self.radio_main.setDisabled(False)
            self.radio_all.setDisabled(False)
            self.btn_handle.setDisabled(True)

            self.lin2.setDisabled(False)
            self.lin3.setDisabled(False)

            self.cb_heatmap.setDisabled(False)
            self.cb_fatigue.setDisabled(False)
            self.cb_fatwind.setDisabled(False)
            self.cb_fatdlc.setDisabled(False)
            self.cb_ultimate.setDisabled(False)

            self.cb_aepdlc12.setDisabled(False)
            self.cb_plotldd.setDisabled(False)
            self.cb_plotlrd.setDisabled(False)
            self.cb_dynpowc.setDisabled(False)
            self.cb_nacacce.setDisabled(False)
            self.cb_pitbear.setDisabled(False)
            self.cb_gearequ.setDisabled(False)
            self.cb_mainber.setDisabled(False)

    def append_option(self):

        if self.radio_app.isChecked():

            # 初始化功能
            # self.lin1.setDisabled(False)
            # self.btn1.setDisabled(False)
            self.lin2.setDisabled(True)
            self.btn2.setDisabled(True)
            self.lin3.setDisabled(True)

            self.lin4.setDisabled(False)
            self.btn3.setDisabled(False)

            # if append option is choosed, then the results will be based on the excel
            self.radio_main.setDisabled(True)
            self.radio_all.setDisabled(True)
            self.btn_handle.setDisabled(True)

            self.cb_heatmap.setDisabled(True)
            self.cb_fatigue.setDisabled(False)
            self.cb_ultimate.setDisabled(False)
            self.cb_fatdlc.setDisabled(True)
            self.cb_fatwind.setDisabled(True)

            self.cb_aepdlc12.setDisabled(True)
            self.cb_plotldd.setDisabled(True)
            self.cb_plotlrd.setDisabled(True)
            self.cb_dynpowc.setDisabled(True)
            self.cb_nacacce.setDisabled(True)
            self.cb_pitbear.setDisabled(True)
            self.cb_gearequ.setDisabled(True)
            self.cb_mainber.setDisabled(True)

    def main_option(self):

        if self.radio_main.isChecked():

            self.btn_handle.setDisabled(True)
            # self.lin2.setDisabled(False)
            # self.lin3.setDisabled(False)
            # self.cb_aepdlc12.setDisabled(False)
            # self.cb_heatmap.setDisabled(False)
            # self.cb_fatigue.setDisabled(False)
            # self.cb_fatwind.setDisabled(False)
            # self.cb_fatdlc.setDisabled(False)
            # self.cb_ultimate.setDisabled(False)

    def all_option(self):

        if self.radio_all.isChecked():
            self.btn_handle.setDisabled(True)
            # self.lin2.setDisabled(False)
            # self.lin3.setDisabled(False)
            # self.cb_aepdlc12.setDisabled(False)
            # self.cb_heatmap.setDisabled(False)
            # self.cb_fatigue.setDisabled(False)
            # self.cb_fatwind.setDisabled(False)
            # self.cb_fatdlc.setDisabled(False)
            # self.cb_ultimate.setDisabled(False)

    def keyPressEvent(self, e):

        if e.key() == QtCore.Qt.Key_Escape:
            self.close()

    def load_template(self):
        excel_return = QFileDialog.getOpenFileName(self,
                                                   "Choose excel template dialog",
                                                   r".",
                                                   "excel(*.xlsx)")

        if excel_return[0]:
            self.lin1.setText(excel_return[0].replace('/', '\\'))

    def load_output(self):
        excel_dir = QFileDialog.getExistingDirectory(self,
                                                     "Choose excel result path",
                                                     r".")

        if excel_dir:
            self.lin2.setText(excel_dir.replace('/', '\\'))

    def choose_excel(self):
        excel_return = QFileDialog.getOpenFileName(self,
                                                   "Choose excel template dialog",
                                                   r".",
                                                   "excel(*.xlsx)")

        if excel_return[0]:
            self.lin4.setText(excel_return[0].replace('/', '\\'))

    def check_input(self):
        '''check .$TE to find out if the simulation is finished, or give warning'''
        print('Begin to check excel template...')
        if '\\' in self.lin1.text():
            self.xls_temp = self.lin1.text().replace('\\', '/')
        else:
            self.xls_temp = self.lin1.text()
        print('Excel template: ', self.xls_temp)

        if self.radio_new.isChecked():

            self.xls_path = self.lin2.text()
            self.xls_name = self.lin3.text()
            self.out_path = None

        if self.radio_app.isChecked():

            self.out_path = self.lin4.text()
            self.xls_path = None
            self.xls_name = None

        # 如果不关闭窗口，再次读入excel会重复追加excel中的路径
        # 初始化列表
        self.loopname.clear()
        self.ult_list.clear()
        self.fat_list.clear()
        self.lct_list.clear()
        self.dlc_list.clear()
        self.ldd_list.clear()
        self.lrd_list.clear()
        self.pbl_list.clear()
        self.mbl_list.clear()
        self.get_list.clear()
        self.nac_list.clear()

        # read excel template
        if self.xls_temp:

            try:
                workbook = load_workbook(self.lin1.text())
                sheet = workbook.get_sheet_by_name('Main')

                row_num = sheet.max_row
                row_max = row_num-(row_num-1)%5
                # print(row_num)>
                self.loopname = [sheet[i][0].value for i in range(2,row_max+1) if sheet[i][0].value != None ]
                self.pst_list = [sheet[i][2].value for i in range(2,row_max+1) if sheet[i][1].value == 'Post']
                self.dlc_list = [sheet[i][2].value for i in range(2,row_max+1) if sheet[i][1].value == 'DLC12']
                self.lct_list = [sheet[i][2].value for i in range(2,row_max+1) if sheet[i][1].value == 'LCT']
                self.ult_list = [sheet[i][2].value for i in range(2,row_max+1) if sheet[i][1].value == 'Ultimate']
                self.fat_list = [sheet[i][2].value for i in range(2,row_max+1) if sheet[i][1].value == 'Rainflow']

                print('LPN :', self.loopname)
                print('ULT :', self.ult_list)
                print('FAT :', self.fat_list)
                print('LCT :', self.lct_list)
                print('PST :', self.pst_list)
                print('L12 :', self.dlc_list)

                loop_num = len(self.loopname)
                # print(loop_num)
                if any(self.pst_list):
                    for i in range(loop_num):
                        post_path = self.pst_list[i]
                        if not os.path.isdir(post_path):
                            raise Exception('%s not a valid path!' %post_path)

                        file_list = os.listdir(post_path)
                        for file in file_list:
                            if '07_ultimate' == file.lower():
                                nac_acc = os.path.join(post_path, r'07_Ultimate\13_Nacelle_Acc')
                                if os.path.isdir(nac_acc):
                                    self.nac_list.append(nac_acc)
                                else:
                                    self.nac_list.append(None)

                            if '05_ldd' == file.lower():

                                ldd_path = os.path.join(post_path, file)
                                if os.path.isdir(ldd_path):
                                    self.ldd_list.append(ldd_path)
                                else:
                                    self.ldd_list.append(None)

                            if '06_lrd' == file.lower():
                                lrd_path = os.path.join(post_path, file)
                                if os.path.isdir(lrd_path):
                                    self.lrd_list.append(lrd_path)
                                else:
                                    self.lrd_list.append(None)

                            if '06_lrd' == file.lower():
                                br_mxy_lrd = os.path.join(post_path, r'06_lrd\br_mxy_64')
                                if os.path.isdir(br_mxy_lrd):
                                    self.pbl_list.append(br_mxy_lrd)
                                else:
                                    self.pbl_list.append(None)

                            if '06_lrd' == file.lower():
                                hs_lrd = os.path.join(post_path, r'06_lrd\hs_64')
                                if os.path.isdir(hs_lrd):
                                    self.get_list.append(hs_lrd)
                                else:
                                    self.get_list.append(None)

                            if '05_ldd' == file.lower():
                                hs_ldd = os.path.join(post_path, r'05_ldd\hs_144')
                                if os.path.isdir(hs_ldd):
                                    self.mbl_list.append(hs_ldd)
                                else:
                                    self.mbl_list.append(None)

                print('LDD :', self.ldd_list)
                print('LRD :', self.lrd_list)
                print('NAC :', self.nac_list)
                print('PBL :', self.pbl_list)
                print('GET :', self.get_list)
                print('MBL :', self.mbl_list)
                print()

                def check():

                    error_flag = None
                    input_isok = True

                    # check excel input
                    if not self.xls_temp:
                        error_flag = 0
                        input_isok = False
                        return error_flag, input_isok

                    if len(self.ult_list) != len(self.loopname) and self.cb_ultimate.isChecked():
                        error_flag = 1
                        input_isok = False
                        return error_flag, input_isok

                    if len(self.fat_list) != len(self.loopname) and self.cb_fatigue.isChecked():
                        error_flag = 2
                        input_isok = False
                        return error_flag, input_isok

                    if not all(self.loopname):
                        error_flag = 3
                        input_isok = False
                        return error_flag, input_isok

                    # check function
                    if self.cb_fatwind.isChecked() and (len(self.lct_list) != len(self.fat_list)):
                        error_flag = 4
                        input_isok = False
                        return error_flag, input_isok

                    if self.cb_fatdlc.isChecked() and (len(self.loopname) != len(self.fat_list)):
                        error_flag = 5
                        input_isok = False
                        return error_flag, input_isok

                    if ((self.cb_dynpowc.isChecked() or self.cb_aepdlc12.isChecked())
                        and (len(self.lct_list) != len(self.loopname) or len(self.dlc_list) != len(self.loopname))):
                        error_flag = 6
                        input_isok = False
                        return error_flag, input_isok

                    if self.cb_heatmap.isChecked() and (len(self.ult_list) != len(self.loopname)):
                        error_flag = 7
                        input_isok = False
                        return error_flag, input_isok

                    if not((self.radio_new and self.xls_path and self.xls_name) or (self.radio_app and self.out_path)):
                        error_flag = 8
                        input_isok = False
                        return error_flag, input_isok

                    if self.cb_plotldd.isChecked() and (len(self.ldd_list) != len(self.loopname)):
                        error_flag = 9
                        input_isok = False
                        return error_flag, input_isok

                    if self.cb_plotlrd.isChecked() and (len(self.lrd_list) != len(self.loopname)):
                        error_flag = 10
                        input_isok = False
                        return error_flag, input_isok

                    if self.cb_mainber.isChecked() and (len(self.mbl_list) != len(self.loopname)):
                        error_flag = 11
                        input_isok = False
                        return error_flag, input_isok

                    if self.cb_pitbear.isChecked() and (len(self.pbl_list) != len(self.loopname)):
                        error_flag = 12
                        input_isok = False
                        return error_flag, input_isok

                    if self.cb_gearequ.isChecked() and (len(self.get_list) != len(self.loopname)):
                        error_flag = 13
                        input_isok = False
                        return error_flag, input_isok

                    if self.cb_nacacce.isChecked() and (len(self.nac_list) != len(self.loopname)):
                        error_flag = 15
                        input_isok = False
                        return error_flag, input_isok

                    if not(self.cb_ultimate.isChecked() or self.cb_gearequ.isChecked() or self.cb_fatigue.isChecked()
                           or self.cb_fatdlc.isChecked() or self.cb_fatwind.isChecked() or self.cb_heatmap.isChecked()
                           or self.cb_plotldd.isChecked() or self.cb_plotlrd.isChecked() or self.cb_pitbear.isChecked()
                           or self.cb_mainber.isChecked() or self.cb_dynpowc.isChecked() or self.cb_aepdlc12.isChecked()
                           or self.cb_alarmid.isChecked() or self.cb_nacacce.isChecked()):

                        error_flag = 14
                        input_isok = False
                        return error_flag, input_isok

                    return error_flag, input_isok

                check_result = check()
                # print(self.cb_dynpowc.isChecked())
                # print(check_result[0], check_result[1])

                if check_result[0] == 0:
                    QMessageBox.about(self, 'Message', 'The excel template has not been choosed yet...\n\n'
                                                        'Please choose a excel template first!')
                elif check_result[0] == 1:
                    QMessageBox.about(self, 'Message', 'The input for ultimate function is not satisfied...\n\n'
                                                        'Please make sure that the ultimate path is right!')
                elif check_result[0] == 2:
                    QMessageBox.about(self, 'Message', 'The input for fatigue function is not satisfied...\n\n'
                                                        'Please make sure that the fatigue path is right!')
                elif check_result[0] == 3:
                    QMessageBox.about(self, 'Message', 'The template is empty...\n\n'
                                                        'Please define the loop first!')
                elif check_result[0] == 4:
                    QMessageBox.about(self, 'Message', 'The input for fatigue_wind function is not satisfied...\n\n'
                                                        'Please make sure that the lct and fatigue inputs are right!')
                elif check_result[0] == 5:
                    QMessageBox.about(self, 'Message', 'The input for fatigue_case function is not satisfied...\n\n'
                                                        'Please make sure that the lct and fatigue inputs are right!')
                elif check_result[0] == 6:
                    QMessageBox.about(self, 'Message', 'The input for aep of dlc12 function is not satisfied...\n\n'
                                                        'Please make sure that the lct and dlc inputs are right!')
                elif check_result[0] == 7:
                    QMessageBox.about(self, 'Message', 'The input for heat map function is not satisfied...\n\n'
                                                        'Please make sure that the ultimate path inputs are right!')
                elif check_result[0] == 8:
                    QMessageBox.about(self, 'Message', 'The output excel has not been choosed...\n\n'
                                                        'Please make sure that the ultimate path inputs are right!')
                elif check_result[0] == 9:
                    QMessageBox.about(self, 'Message', 'The input for ldd function is not satisfied...\n\n'
                                                        'Please make sure that the ldd path inputs are right!')

                elif check_result[0] == 10:
                    QMessageBox.about(self, 'Message', 'The input for main bearing life function is not satisfied...\n\n'
                                                        'Please make sure that the main bearing path inputs are right!')
                elif check_result[0] == 11:
                    QMessageBox.about(self, 'Message', 'The input for pitch bearing life function is not satisfied...\n\n'
                                                        'Please make sure that the pitch bearing path inputs are right!')
                elif check_result[0] == 12:
                    QMessageBox.about(self, 'Message', 'The input for lrd function is not satisfied...\n\n'
                                                        'Please make sure that the lrd path inputs are right!')
                elif check_result[0] == 13:
                    QMessageBox.about(self, 'Message', 'The input for gearbox equivalent torque function is not satisfied...\n\n'
                                                        'Please make sure that the path inputs are right!')

                elif check_result[0] == 14:
                    QMessageBox.about(self, 'Message', 'The function has not been chosen...\n\n'
                                                        'Please choose a function first!')
                elif check_result[0] == 15:
                    QMessageBox.about(self, 'Message', 'The input for nacelle acceleration function is not satisfied...\n\n'
                                                        'Please make sure that the path inputs are right!')

                # check is OK, then disable the pannel
                if check_result[1]:

                    QMessageBox.about(self, 'Message', 'The input for each functionality is OK...\n\n'
                                                        'You can continue to handle date!')
                    self.btn_handle.setDisabled(False)
                    self.btn_reset.setDisabled(False)
                    self.lin1.setDisabled(True)
                    self.btn1.setDisabled(True)
                    self.btn_check.setDisabled(True)
                    self.radio_main.setDisabled(True)
                    self.radio_all.setDisabled(True)

                    if self.radio_new.isChecked():

                        self.lin2.setDisabled(True)
                        self.btn2.setDisabled(True)
                        self.lin3.setDisabled(True)

                        self.radio_new.setDisabled(True)
                        self.radio_app.setDisabled(True)
                        self.radio_main.setDisabled(True)
                        self.radio_all.setDisabled(True)

                        self.cb_aepdlc12.setDisabled(True)
                        self.cb_heatmap.setDisabled(True)
                        self.cb_fatigue.setDisabled(True)
                        self.cb_fatwind.setDisabled(True)
                        self.cb_fatdlc.setDisabled(True)
                        self.cb_ultimate.setDisabled(True)
                        self.cb_plotldd.setDisabled(True)
                        self.cb_plotlrd.setDisabled(True)
                        self.cb_dynpowc.setDisabled(True)
                        self.cb_gearequ.setDisabled(True)
                        self.cb_pitbear.setDisabled(True)
                        self.cb_mainber.setDisabled(True)
                        self.cb_nacacce.setDisabled(True)

                    if self.radio_app.isChecked():

                        self.btn3.setDisabled(True)
                        self.lin4.setDisabled(True)

                        self.radio_new.setDisabled(True)
                        self.radio_app.setDisabled(True)

                        self.cb_aepdlc12.setDisabled(True)
                        self.cb_heatmap.setDisabled(True)
                        self.cb_fatigue.setDisabled(True)
                        self.cb_fatwind.setDisabled(True)
                        self.cb_fatdlc.setDisabled(True)
                        self.cb_ultimate.setDisabled(True)
                        self.cb_plotldd.setDisabled(True)
                        self.cb_plotlrd.setDisabled(True)
                        self.cb_gearequ.setDisabled(True)
                        self.cb_pitbear.setDisabled(True)
                        self.cb_mainber.setDisabled(True)

                else:

                    self.btn_handle.setDisabled(True)

            except Exception as message:

                QMessageBox.about(self, 'Message', 'Please check the template!\n%s' % message)

    def handle(self):

        if self.cb_ultimate.isChecked():
            self.ultimate = True
        else:
            self.ultimate = False

        if self.cb_fatigue.isChecked():
            self.fatigue = True
        else:
            self.fatigue = False

        if self.cb_fatwind.isChecked():
            self.fat_wind = True
        else:
            self.fat_wind = False

        if self.cb_heatmap.isChecked():
            self.heat_map = True
        else:
            self.heat_map = False

        if self.cb_fatdlc.isChecked():
            self.fat_case = True
        else:
            self.fat_case = False

        if self.cb_aepdlc12.isChecked() or self.cb_dynpowc.isChecked():
            self.aep_dlc = True
        else:
            self.aep_dlc = False

        if self.cb_plotlrd.isChecked():
            self.plot_lrd = True
        else:
            self.plot_lrd = False

        if self.cb_plotldd.isChecked():
            self.plot_ldd = True
        else:
            self.plot_ldd = False

        if self.cb_mainber.isChecked():
            self.main_bl = True
        else:
            self.main_bl = False

        if self.cb_pitbear.isChecked():
            self.pitch_bl = True
        else:
            self.pitch_bl = False

        if self.cb_gearequ.isChecked():
            self.gearbox = True
        else:
            self.gearbox = False

        if self.cb_nacacce.isChecked():
            self.nacelle = True
        else:
            self.nacelle = False

        if self.radio_new.isChecked():

            self.out_path = self.lin2.text().replace('\\', '/')
            self.xls_name = self.lin3.text()

            logging.basicConfig(
                level    = logging.INFO,  # 日志级别，只有日志级别大于等于设置级别的日志才会输出
                format   = '%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',  # 日志输出格式
                datefmt  = '[%Y-%m_%d %H:%M:%S]',  # 日期表示格式
                filename = os.path.dirname(os.path.realpath(__file__)) + "/" + 'logger.log',  # 输出定向的日志文件路径
                filemode ='a')

            try:
                new(excel_input=self.xls_temp,
                    output_path=self.out_path,
                    excel_name=self.xls_name,
                    ultimate=self.ultimate,
                    fatigue=self.fatigue,
                    heat_map=self.heat_map,
                    fat_case=self.fat_case,
                    fat_wind=self.fat_wind,
                    aep_dlc=self.aep_dlc,
                    ldd=self.plot_ldd,
                    lrd=self.plot_lrd,
                    mbl=self.main_bl,
                    pbl=self.pitch_bl,
                    get=self.gearbox,
                    nac=self.nacelle)

                QMessageBox.about(self, 'Message', 'Load summary table is done!')
            except Exception as error_message:
                QMessageBox.about(self, 'Message', '%s' %error_message)

                traceback.print_exc()
                logging.warning("exec failed, failed msg:" + traceback.format_exc())

        if self.radio_app.isChecked():

            self.xls_path = self.lin4.text()

            if '\\' in self.lin4.text():
                self.xls_path = self.lin4.text().replace('\\', '/')

            append(self.xls_temp,
                   self.xls_path,
                   self.ultimate,
                   self.fatigue,
                   self.heat_map,
                   self.fat_case,
                   self.fat_wind,
                   self.aep_dlc)

            QMessageBox.about(self, 'Message', 'Load summary table is done!')

    def reset(self):

        self.btn_handle.setDisabled(True)
        self.btn_reset.setDisabled(True)
        self.lin1.setDisabled(False)
        self.btn1.setDisabled(False)
        self.btn_check.setDisabled(False)

        if self.radio_new.isChecked():

            self.lin2.setDisabled(False)
            self.btn2.setDisabled(False)
            self.lin3.setDisabled(False)

            self.radio_new.setDisabled(False)
            # self.radio_app.setDisabled(False)
            self.radio_main.setDisabled(False)
            self.radio_all.setDisabled(False)

            self.cb_aepdlc12.setDisabled(False)
            self.cb_heatmap.setDisabled(False)
            self.cb_fatigue.setDisabled(False)
            self.cb_fatwind.setDisabled(False)
            self.cb_fatdlc.setDisabled(False)
            self.cb_ultimate.setDisabled(False)
            self.cb_plotldd.setDisabled(False)
            self.cb_plotlrd.setDisabled(False)
            self.cb_dynpowc.setDisabled(False)
            self.cb_gearequ.setDisabled(False)
            self.cb_pitbear.setDisabled(False)
            self.cb_mainber.setDisabled(False)

        if self.radio_app.isChecked():

            self.btn3.setDisabled(False)
            self.lin4.setDisabled(False)

            self.radio_new.setDisabled(False)
            self.radio_app.setDisabled(False)

            self.cb_aepdlc12.setDisabled(False)
            self.cb_heatmap.setDisabled(False)
            self.cb_fatigue.setDisabled(False)
            self.cb_fatwind.setDisabled(False)
            self.cb_fatdlc.setDisabled(False)
            self.cb_ultimate.setDisabled(False)
            self.cb_plotldd.setDisabled(False)
            self.cb_plotlrd.setDisabled(False)
            self.cb_dynpowc.setDisabled(False)
            self.cb_gearequ.setDisabled(False)
            self.cb_pitbear.setDisabled(False)
            self.cb_mainber.setDisabled(False)

        # if self.radio_new.isChecked():
        #
        #     self.lin1.setDisabled(False)
        #     self.lin2.setDisabled(False)
        #     self.lin3.setDisabled(False)
        #     self.lin4.setDisabled(False)
        #     self.cb_aepdlc12.setDisabled(False)
        #     self.cb_heatmap.setDisabled(False)
        #     self.cb_fatigue.setDisabled(False)
        #     self.cb_fatwind.setDisabled(False)
        #     self.cb_fatdlc.setDisabled(False)
        #     self.cb_ultimate.setDisabled(False)
        #
        # if self.radio_app.isChecked():
        #
        #     self.lin1.setDisabled(False)
        #     self.lin4.setDisabled(False)
        #     self.cb_aepdlc12.setDisabled(False)
        #     self.cb_heatmap.setDisabled(False)
        #     self.cb_fatigue.setDisabled(False)
        #     self.cb_fatwind.setDisabled(False)
        #     self.cb_fatdlc.setDisabled(False)
        #     self.cb_ultimate.setDisabled(False)

    def center(self):

        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = LoadSumWindow()
    # window.setWindowOpacity(0.9)
    window.show()
    sys.exit(app.exec_())