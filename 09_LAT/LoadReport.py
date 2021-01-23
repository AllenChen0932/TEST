# ！usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2019/10/26 11:14
# @Author  : CE
# @File    : joblist.py
'''
1 本程序只适用于通过additional controller parameters对话框来设置控制器参数，不适用于有多个控制器时的模型；
2 有多个控制器的模型需要在每个控制器中定义单独的参数文件；
'''

import sys, os
import configparser
import openpyxl
import pysnooper

from PyQt5.QtWidgets import (QMainWindow,
                             QApplication,
                             QDesktopWidget,
                             QLabel,
                             qApp,
                             QLineEdit,
                             QPushButton,
                             QFileDialog,
                             QGridLayout,
                             QVBoxLayout,
                             QWidget,
                             QAction,
                             QMessageBox,
                             QGroupBox)
from PyQt5.QtGui import QIcon, QFont
from PyQt5 import QtCore

from tool.load_report.writeUltimate import ReadPostMX
from tool.load_report.writeRainflow import writeRainflow
from tool.load_report.writeOccurrence import Occurrence


class MyQLineEdit(QLineEdit):
    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if event.mimeData().text():
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event): # 放下文件后的动作
        file_path = event.mimeData().text()
        if '///' in file_path:
            self.setText(file_path[8:].replace('/','\\'))   # 删除local path多余开头
        else:
            self.setText(file_path[5:].replace('/','\\'))   # 删除network path多余开头

class LoadReportWindow(QMainWindow):

    def __init__(self, parent=None):

        super(LoadReportWindow, self).__init__(parent)
        self.mywidget = QWidget(self)

        self.setMinimumSize(700, 250)
        self.setWindowTitle('Load Report')
        self.setWindowIcon(QIcon(r".\icon\excel1.ico"))
        # root = QFileInfo(__file__).absolutePath()
        # self.setWindowIcon(QIcon(root + "./icon/Text_Edit.ico"))

        self.job_list = None
        self.lis_name = None
        self.job_path = None
        self.res_path = None
        self.dll_path = None
        self.xml_path = None
        self.njl_path = None
        self.para_con = None
        self.othe_con = None

        self.title_font = QFont("Calibri", 10)
        self.title_font.setItalic(True)
        self.title_font.setBold(True)

        self.cont_font = QFont("Calibri", 9)
        self.cont_font.setItalic(False)
        self.cont_font.setBold(False)

        self.initUI()
        self.load_setting()
        self.center()

    def initUI(self):

        menubar = self.menuBar()
        fileMenu = menubar.addMenu('File')
        helpMenu = menubar.addMenu('Help')

        saveAction = QAction(QIcon('./Icon/save.ico'), 'Save', self)
        saveAction.setShortcut('Ctrl+S')
        saveAction.triggered.connect(self.save_setting)
        fileMenu.addAction(saveAction)

        clearAction = QAction(QIcon('Icon/clear.ico'), 'Clear', self)
        clearAction.setShortcut('Ctrl+R')
        clearAction.triggered.connect(self.clear_setting)
        fileMenu.addAction(clearAction)

        exitAction = QAction(QIcon('Icon/exit.ico'), 'Exit', self)
        exitAction.setShortcut('Ctrl+Q')
        exitAction.triggered.connect(qApp.quit)
        fileMenu.addAction(exitAction)

        helpAction = QAction(QIcon('Icon/doc.ico'), 'User Manual', self)
        helpAction.setShortcut('Ctrl+H')
        helpAction.triggered.connect(self.user_manual)
        helpMenu.addAction(helpAction)

        # Blade
        # root axes
        self.label1 = QLabel("Fatigue ")
        self.label1.setFont(self.cont_font)
        self.line1 = MyQLineEdit()
        self.line1.setFont(self.cont_font)
        self.line1.setPlaceholderText("Pick 08_Fatigue")
        self.btn1 = QPushButton("...")
        self.btn1.setFont(self.cont_font)
        self.btn1.clicked.connect(self.load_fatigue)

        self.label2 = QLabel("Ultimate")
        self.label2.setFont(self.cont_font)
        self.line2 = MyQLineEdit()
        self.line2.setFont(self.cont_font)
        self.line2.setPlaceholderText("Pick 07_Ultimate")
        self.btn2 = QPushButton("...")
        self.btn2.setFont(self.cont_font)
        self.btn2.clicked.connect(self.load_ultimate)

        self.label3 = QLabel("Load Case Table")
        self.label3.setFont(self.cont_font)
        self.line3 = MyQLineEdit()
        self.line3.setFont(self.cont_font)
        self.line3.setPlaceholderText("Pick load case table")
        self.btn3 = QPushButton("...")
        self.btn3.setFont(self.cont_font)
        self.btn3.clicked.connect(self.load_lct)

        # output
        self.label4 = QLabel("Output path")
        self.label4.setFont(self.cont_font)
        self.line4 = MyQLineEdit()
        self.line4.setFont(self.cont_font)
        self.line4.setPlaceholderText("Pick output path")
        self.btn4 = QPushButton("...")
        self.btn4.setFont(self.cont_font)
        self.btn4.clicked.connect(self.load_output)

        self.btn5 = QPushButton("Generate load report input excel")
        self.btn5.setFont(self.cont_font)
        self.btn5.clicked.connect(self.generate_excel)

        self.group1 = QGroupBox('Main input')
        self.group1.setFont(self.title_font)
        self.grid1 = QGridLayout()
        # 起始行，起始列，占用行，占用列
        self.grid1.addWidget(self.label1, 0, 0, 1, 1)
        self.grid1.addWidget(self.line1, 0, 1, 1, 5)
        self.grid1.addWidget(self.btn1, 0, 6, 1, 1)
        self.grid1.addWidget(self.label2, 1, 0, 1, 1)
        self.grid1.addWidget(self.line2, 1, 1, 1, 5)
        self.grid1.addWidget(self.btn2, 1, 6, 1, 1)
        self.grid1.addWidget(self.label3, 2, 0, 1, 1)
        self.grid1.addWidget(self.line3, 2, 1, 1, 5)
        self.grid1.addWidget(self.btn3, 2, 6, 1, 1)

        self.group1.setLayout(self.grid1)

        # output
        self.group2 = QGroupBox('Output path')
        self.group2.setFont(self.title_font)
        self.grid2 = QGridLayout()
        # 起始行，起始列，占用行，占用列
        self.grid2.addWidget(self.label4, 0, 0, 1, 1)
        self.grid2.addWidget(self.line4, 0, 1, 1, 5)
        self.grid2.addWidget(self.btn4, 0, 6, 1, 1)
        self.group2.setLayout(self.grid2)

        self.main_layout = QVBoxLayout()
        self.main_layout.addWidget(self.group1)
        self.main_layout.addWidget(self.group2)
        self.main_layout.addWidget(self.btn5)
        self.main_layout.addStretch(1)

        self.mywidget.setLayout(self.main_layout)
        self.setCentralWidget(self.mywidget)

    def keyPressEvent(self, e):
        if e.key() == QtCore.Qt.Key_Escape:
            self.close()

    def save_setting(self):
        config = configparser.ConfigParser()
        config.read('config.ini', encoding='utf-8')

        if not config.has_section('Load Report'):
            config.add_section('Load Report')

        config['Load Report'] = {'Fatigue path':self.line1.text(),
                                 'Ultimate path':self.line2.text(),
                                 'Loadcase table':self.line3.text(),
                                 'Output path':self.line4.text()}

        with open("config.ini",'w') as f:
            config.write(f)

    def load_setting(self):
        config = configparser.ConfigParser()
        config.read('config.ini', encoding='utf-8')

        if config.has_section('Load Report'):
            self.line1.setText(config.get('Load Report','Fatigue path'))
            self.line2.setText(config.get('Load Report','Ultimate path'))
            self.line3.setText(config.get('Load Report','Loadcase table'))
            self.line4.setText(config.get('Load Report','Output path'))

    def clear_setting(self):
        self.line1.setText('')
        self.line2.setText('')
        self.line3.setText('')
        self.line4.setText('')

    def user_manual(self):
        os.startfile(os.getcwd()+os.sep+'user manual\Load Report.docx')

    def load_fatigue(self):
        root_path = QFileDialog.getExistingDirectory(self, "Choose fatigue result path", ".")

        if root_path:
            self.line1.setText(root_path.replace('/', '\\'))

    def load_ultimate(self):
        root_path = QFileDialog.getExistingDirectory(self, "Choose ultimate result path", ".")

        if root_path:
            self.line2.setText(root_path.replace('/', '\\'))

    def load_lct(self):
        lct_path = QFileDialog.getOpenFileName(self, "Choose load case table path", ".", "excel(*.xlsx)")

        if lct_path:
            self.line3.setText(lct_path.replace('/', '\\'))

    def load_output(self):
        output_path = QFileDialog.getExistingDirectory(self, "Choose output path", ".")

        if output_path:
            self.line4.setText(output_path.replace('/', '\\'))
            self.line4.home(True)

    # @pysnooper.snoop()
    def generate_excel(self):

        fat_path = self.line1.text()
        ult_path = self.line2.text()
        lct_path = self.line3.text()
        out_path = self.line4.text()

        # main
        fat_main  = os.path.join(fat_path, '05_Main')
        hub_path  = os.path.join(ult_path, r'\06_HR_Onlydlc8\Inclsf')
        ult_main  = os.path.join(ult_path, '08_Main_Inclsf')
        # blade root axes
        root_inclsf  = os.path.join(ult_path, '01_BRS_Inclsf')
        root_exclsf  = os.path.join(ult_path, '02_BRS_Exclsf')
        root_fatigue = os.path.join(fat_path, r'01_BRS\brs1')
        # blade root axes
        user_inclsf  = os.path.join(ult_path, '03_BUS_Inclsf')
        user_exclsf  = os.path.join(ult_path, '04_BUS_Inclsf')
        user_fatigue = os.path.join(fat_path, r'02_BUS\bus1')
        # tower
        tower_inclsf = os.path.join(ult_path, '10_Tower_Inclsf')
        tower_exclsf = os.path.join(ult_path, '11_Tower_Exclsf')
        tower_dlc12  = os.path.join(ult_path, '12_Tower_dlc12')
        tower_fatigue= os.path.join(fat_path, '06_Tower')

        # check fatigue path:
        if not os.path.isdir(fat_path):
            QMessageBox.about(self, 'warnning', 'Please choose a right fatigue path!')
        elif not os.path.isdir(ult_path):
            QMessageBox.about(self, 'warnning', 'Please choose a right ultimate path!')
        elif not os.path.isfile(lct_path):
            QMessageBox.about(self, 'warnning', 'Please choose a right load case table!')
        elif not os.path.isdir(out_path):
            QMessageBox.about(self, 'warnning', 'Please choose a right output path!')
        elif not os.path.isdir(fat_path):
            QMessageBox.about(self, 'warnning', 'Please make sure main fatigue result exist!\n%s' %fat_path)
        elif not os.path.isdir(ult_path):
            QMessageBox.about(self, 'warnning', 'Please make sure main ultimate result exist!\n%s' %ult_path)
        elif not os.path.isdir(hub_path):
            QMessageBox.about(self, 'warnning', 'Please make sure hub without dlc8x result exist!\n%s' %ult_path)
        elif not root_inclsf:
            QMessageBox.about(self,'warnning','Please make sure blade root axes result exist!\n%s' %root_inclsf)
        elif not root_exclsf:
            QMessageBox.about(self,'warnning','Please make sure blade root axes result exist!\n%s' %root_exclsf)
        elif not root_fatigue:
            QMessageBox.about(self,'warnning','Please make sure blade root axes result exist!\n%s' %root_fatigue)
        elif not user_inclsf:
            QMessageBox.about(self,'warnning','Please make sure blade user axes result exist!\n%s' %user_inclsf)
        elif not user_exclsf:
            QMessageBox.about(self,'warnning','Please make sure blade user axes result exist!\n%s' %user_exclsf)
        elif not user_fatigue:
            QMessageBox.about(self,'warnning','Please make sure blade user axes result exist!\n%s' %user_fatigue)
        elif not tower_inclsf:
            QMessageBox.about(self,'warnning','Please make sure tower result exist!\n%s' %tower_inclsf)
        elif not tower_exclsf:
            QMessageBox.about(self,'warnning','Please make sure tower result exist!\n%s' %tower_exclsf)
        elif not tower_dlc12:
            QMessageBox.about(self,'warnning','Please make sure tower result exist!\n%s' %tower_dlc12)
        elif not tower_fatigue:
            QMessageBox.about(self,'warnning','Please make sure tower result exist!\n%s' %tower_fatigue)
        else:
            # write compare result
            print('begin to write excel...')
            table = openpyxl.Workbook()
            try:
                sheet_name = 'Ultimate'
                table.create_sheet(sheet_name)
                br       = os.path.join(ult_main, 'br')
                ultimate = ReadPostMX(br, sheet_name, column_start=1, row_start=1, row_space=2, height_flag=False)
                ultimate.write_result(table)

                hs       = os.path.join(hub_path, 'hs')
                ultimate = ReadPostMX(hs, sheet_name, column_start=1, row_start=22, row_space=2, height_flag=False)
                ultimate.write_result(table)

                hr       = os.path.join(hub_path, 'hr')
                ultimate = ReadPostMX(hr, sheet_name, column_start=1, row_start=43, row_space=2, height_flag=False)
                ultimate.write_result(table)


                sheet_name = 'Fatigue'
                table.create_sheet(sheet_name)
                br      = os.path.join(fat_main, 'br1')
                fatigue = writeRainflow(br, content=('DEL',),row_start=2,col_start=8,height_flag=False)
                fatigue.write2excel(table, sheet_name)

                hs      = os.path.join(fat_main, 'hs')
                fatigue = writeRainflow(hs, content=('DEL',),row_start=16,col_start=8,height_flag=False)
                fatigue.write2excel(table, sheet_name)

                hr      = os.path.join(fat_main, 'hr')
                fatigue = writeRainflow(hr, content=('DEL',),row_start=30,col_start=8,height_flag=False)
                fatigue.write2excel(table, sheet_name)

                excel1_path = os.path.join(out_path, 'Compare_results.xlsx')
                if os.path.isfile(excel1_path):
                    os.remove(excel1_path)
                table.remove(table['Sheet'])
                table.save(excel1_path)
                print('%s is done!' %excel1_path)

                # write blade
                table      = openpyxl.Workbook()
                sheet_name = 'extreme root section'
                table.create_sheet(sheet_name)

                ultimate   = ReadPostMX(root_inclsf, sheet_name, 0, height_flag=True)
                ultimate.write_result(table)
                ultimate   = ReadPostMX(root_exclsf, sheet_name, 13, height_flag=True)
                ultimate.write_result(table)

                # sheet_name = 'extreme user section'
                # table.create_sheet(sheet_name)
                # ultimate   = ReadPostMX(user_inclsf, sheet_name, 0)
                # ultimate.write_result(table)
                # ultimate   = ReadPostMX(user_exclsf, sheet_name, 13)
                # ultimate.write_result(table)

                sheet_name = 'fatigue root section'
                table.create_sheet(sheet_name)
                fatigue    = writeRainflow(root_fatigue, content=('DEL',), row_space=3, height_flag=True)
                fatigue.write2excel(table, sheet_name)

                # sheet_name = 'fatigue user section'
                # table.create_sheet(sheet_name)
                # fatigue    = writeRainflow(user_fatigue, content=('DEL',))
                # fatigue.write2excel(table, sheet_name)

                excel2_path = os.path.join(out_path, 'Blade.xlsx')
                if os.path.isfile(excel2_path):
                    os.remove(excel2_path)
                table.remove(table['Sheet'])
                table.save(excel2_path)
                print('%s is done!' %excel2_path)

                # write tower
                table      = openpyxl.Workbook()
                sheet_name = 'extreme flange section'
                table.create_sheet(sheet_name)
                if 'Mbr' not in ''.join(os.listdir(tower_inclsf)):
                    ultimate = ReadPostMX(tower_inclsf, sheet_name, 0, height_flag=True)
                else:
                    ultimate = ReadPostMX(tower_inclsf, sheet_name, 0, height_flag=False)
                ultimate.write_result(table)

                if 'Mbr' not in ''.join(os.listdir(tower_exclsf)):
                    ultimate = ReadPostMX(tower_exclsf, sheet_name, 13, height_flag=True)
                else:
                    ultimate = ReadPostMX(tower_exclsf, sheet_name, 13, height_flag=False)
                ultimate.write_result(table)

                sheet_name = 'occurrence'
                table.create_sheet(sheet_name)
                Occurrence(lct_path, table, sheet_name)

                sheet_name = 'extreme flange section DLC12'
                table.create_sheet(sheet_name)
                if 'Mbr' not in ''.join(os.listdir(tower_dlc12)):
                    ultimate = ReadPostMX(tower_dlc12, sheet_name, 0, height_flag=True)
                else:
                    ultimate = ReadPostMX(tower_dlc12, sheet_name, 0, height_flag=False)
                ultimate.write_result(table)

                sheet_name = 'fatigue flange section'
                table.create_sheet(sheet_name)

                if 'Mbr' not in ''.join(os.listdir(tower_fatigue)):
                    fatigue = writeRainflow(tower_fatigue, content=('DEL',), height_flag=True)
                    fatigue.write2excel(table, sheet_name)
                else:
                    fatigue = writeRainflow(tower_fatigue, content=('DEL',), height_flag=False)
                    fatigue.write2excel(table, sheet_name)

                excel3_path = os.path.join(out_path, 'Tower.xlsx')
                if os.path.isfile(excel3_path):
                    os.remove(excel3_path)
                table.remove(table['Sheet'])
                table.save(excel3_path)
                print('%s is done!' %excel3_path)

                if os.path.isfile(excel1_path) and os.path.isfile(excel2_path) and os.path.isfile(excel3_path):
                    QMessageBox.about(self, 'Window', 'Generating excels is done!')

            except Exception as e:
                QMessageBox.about(self, 'Warnning', 'Error occurs when generating excels!\n%s' %e)

    def center(self):

        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

if __name__ == '__main__':

    app = QApplication(sys.argv)
    window = LoadReportWindow()
    # window.setWindowOpacity(0.9)
    window.show()
    sys.exit(app.exec_())

