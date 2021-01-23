# -*- coding: utf-8 -*-
# @Time    : 2020/07/31 11:10
# @Author  : TJ
# @File    : writeOccur.py

import openpyxl
from openpyxl.utils import get_column_letter

class Occurrence(object):

    def __init__(self, project, table_object, sheet_name, time_step=None, row_start=1, col_start=1, simple=True):
        
        self.lct_table  = project
        self.table_obj  = table_object
        self.sheet_name = sheet_name
        self.time_step  = time_step
        self.row_start  = row_start
        self.col_start  = col_start
        self.simple     = simple

        # self.dlc_list  = ['DLC12', 'DLC24a', 'DLC24b', 'DLC24c', 'DLC24d', 'DLC31', 'DLC41', 'DLC64']

        self.dlc_data  = {}
        self.lc_list   = []
        self.life_time = None

        self.read_proj()
        if self.simple:
            self.write_result()
        else:
            self.write_all_info()

    def search_table(self, sheet, name_list):

        position = {}
        for i in range(1, sheet.max_column+1):
            for j in range(1, sheet.max_row+1):
                for name in name_list:
                    value = sheet.cell(row=i, column=j).value

                    if type(value) == str and str.lower(value) == str.lower(name):
                        # print(sheet.cell(row=i, column=j))
                        position[name] = [i, j]

                        if len(position) == len(name_list):
                            break
        return position

    def read_proj(self):

        xls = openpyxl.load_workbook(self.lct_table, keep_vba=True, data_only=True)

        # read setting
        sheet   = xls['Setting']
        col_max = sheet.max_row
        sheets  = [name for name in xls.sheetnames if name.startswith('DLC')]
        sheets.sort()

        stand   = '3' if 'IEC61400-3' in sheet.cell(1,2).value else '1'

        for i in range(1, col_max+1):
            if sheet.cell(i,2).value == 'T_life':
                self.life_time = float(sheet.cell(i,3).value)
                break

        var_name = ['Run Name', 'VHub', 'Yaw', 'tout', 'tSim', 'Hour/Year', 'Times/year'] \
            if stand=='1' else ['Run Name', 'VHub', 'Yaw Error', 'tout', 'tSim', 'Hour/Year', 'Times/year']

        # DLC12
        for dlc in sheets:
            
            if dlc == 'DLC12' or dlc == 'DLC16' or dlc == 'DLC64' or dlc == 'DLC65' or dlc == 'DLC72':
                sheet   = xls[dlc]
                row_max = sheet.max_row
                var_pos = self.search_table(sheet, var_name)
    
                row_start = var_pos['Run Name'][0]
                col_run   = get_column_letter(var_pos['Run Name'][1])
                col_tout  = get_column_letter(var_pos['tout'][1])
                col_tSim  = get_column_letter(var_pos['tSim'][1])
                col_hours = get_column_letter(var_pos['Hour/Year'][1])
                col_wind  = get_column_letter(var_pos['VHub'][1])
                col_yaw   = get_column_letter(var_pos['%s' %('Yaw Error' if stand=='3' else 'Yaw')][1])

                for row in range(row_start+2, row_max+1):

                    lc  = sheet[col_run+str(row)].value
                    if lc:
                        tout = float(sheet[col_tout+str(row)].value)
                        tsim = float(sheet[col_tSim+str(row)].value)
                        hour = float(sheet[col_hours+str(row)].value)
                        if 'REF' in str(hour):
                            raise Exception('Error hour definition for %s' %lc)
                        wind = float(sheet[col_wind+str(row)].value)
                        yaw  = float(sheet[col_yaw+str(row)].value)
                        self.dlc_data[lc] = [tout, tsim, hour, wind, yaw]
                        self.lc_list.append(lc)

            elif dlc == 'DLC31' or dlc == 'DLC41' or dlc.startswith('DLC24'):
                sheet   = xls[dlc]
                row_max = sheet.max_row
                var_pos = self.search_table(sheet, var_name)

                row_start = var_pos['Run Name'][0]
                col_run   = get_column_letter(var_pos['Run Name'][1])
                col_tout  = get_column_letter(var_pos['tout'][1])
                col_tSim  = get_column_letter(var_pos['tSim'][1])
                col_time  = get_column_letter(var_pos['Times/year'][1])
                col_wind  = get_column_letter(var_pos['VHub'][1])
                col_yaw   = get_column_letter(var_pos['%s' %('Yaw Error' if stand=='3' else 'Yaw')][1])
                
                for row in range(row_start+2, row_max+1):
                    lc  = sheet[col_run+str(row)].value
                    if lc:
                        tout = float(sheet[col_tout+str(row)].value)
                        tsim = float(sheet[col_tSim+str(row)].value)
                        time = float(sheet[col_time+str(row)].value)
                        if 'REF' in str(time):
                            raise Exception('Error times definition for %s' %lc)
                        wind = float(sheet[col_wind+str(row)].value)
                        yaw  = float(sheet[col_yaw+str(row)].value)
                        self.dlc_data[lc] = [tout, tsim, time, wind, yaw]
                        self.lc_list.append(lc)

    def write_result(self):

        if self.sheet_name in self.table_obj.sheetnames:
            sheet = self.table_obj[self.sheet_name]
        else:
            sheet = self.table_obj.create_sheet(self.sheet_name)
        # print(sheet)

        sheet.cell(self.row_start, self.col_start,   value='dlc')
        sheet.cell(self.row_start, self.col_start+3, value='hours in %s years' %int(self.life_time))
        sheet.cell(self.row_start, self.col_start+2, value='occurrences in %s years' %int(self.life_time))
        sheet.cell(self.row_start, self.col_start+1, value='simulation')

        sheet.cell(self.row_start+1, self.col_start,   value='(-)')
        sheet.cell(self.row_start+1, self.col_start+3, value='(hour)')
        sheet.cell(self.row_start+1, self.col_start+2, value='(-)')
        sheet.cell(self.row_start+1, self.col_start+1, value='(s)')

        for i, lc in enumerate(self.lc_list):
            sim_time = self.dlc_data[lc][1] - self.dlc_data[lc][0]

            # load case defined by hours
            if lc.startswith('12') or lc.startswith('16') or lc.startswith('24c') or lc.startswith('24d')\
                    or lc.startswith('64') or lc.startswith('65') or lc.startswith('L12') or lc.startswith('L16')\
                    or lc.startswith('L64') or lc.startswith('L65') or lc.startswith('72'):
                sheet.cell(self.row_start+i+2, self.col_start,   value = lc)
                sheet.cell(self.row_start+i+2, self.col_start+3, value = self.dlc_data[lc][2]*self.life_time)
                sheet.cell(self.row_start+i+2, self.col_start+2, value = self.dlc_data[lc][2]*self.life_time*3600/sim_time)
                sheet.cell(self.row_start+i+2, self.col_start+1, value = sim_time)

            # load case defined by times
            else:
                sheet.cell(self.row_start+i+2, self.col_start,   value = lc)
                sheet.cell(self.row_start+i+2, self.col_start+3, value = self.dlc_data[lc][2]*self.life_time*sim_time/3600)
                sheet.cell(self.row_start+i+2, self.col_start+2, value = self.dlc_data[lc][2]*self.life_time)
                sheet.cell(self.row_start+i+2, self.col_start+1, value = sim_time)

        print('occurrence is done!')

    def write_all_info(self):

        if self.sheet_name in self.table_obj.sheetnames:
            sheet = self.table_obj[self.sheet_name]
        else:
            sheet = self.table_obj.create_sheet(self.sheet_name)
        # print(sheet)

        sheet.cell(self.row_start, self.col_start,   value='dlc')
        sheet.cell(self.row_start, self.col_start+1, value='wind speed')
        sheet.cell(self.row_start, self.col_start+2, value='yaw error')
        sheet.cell(self.row_start, self.col_start+3, value='simulation')
        sheet.cell(self.row_start, self.col_start+4, value='simulation step')
        sheet.cell(self.row_start, self.col_start+5, value='No. of steps')
        sheet.cell(self.row_start, self.col_start+7, value='hours in %s years' %int(self.life_time))
        sheet.cell(self.row_start, self.col_start+6, value='occurrences in %s years' %int(self.life_time))

        sheet.cell(self.row_start+1, self.col_start,   value='(-)')
        sheet.cell(self.row_start+1, self.col_start+2, value='(m/s)')
        sheet.cell(self.row_start+1, self.col_start+3, value='(°)')
        sheet.cell(self.row_start+1, self.col_start+4, value='(s)')
        sheet.cell(self.row_start+1, self.col_start+5, value='(-)')
        sheet.cell(self.row_start+1, self.col_start+7, value='(hour)')
        sheet.cell(self.row_start+1, self.col_start+6, value='(-)')

        for i, lc in enumerate(self.lc_list):

            sim_time = self.dlc_data[lc][1] - self.dlc_data[lc][0]

            # load case defined by hours
            if lc.startswith('12') or lc.startswith('16') or lc.startswith('24c') or lc.startswith('24d')\
                    or lc.startswith('64') or lc.startswith('65') or lc.startswith('L12') or lc.startswith('L16')\
                    or lc.startswith('L64') or lc.startswith('L65') or lc.startswith('72'):
                sheet.cell(self.row_start+i+2, self.col_start,   value = lc)
                sheet.cell(self.row_start+i+2, self.col_start+1, value = self.dlc_data[lc][3])
                sheet.cell(self.row_start+i+2, self.col_start+2, value = self.dlc_data[lc][4])
                sheet.cell(self.row_start+i+2, self.col_start+7, value = self.dlc_data[lc][2]*self.life_time)
                sheet.cell(self.row_start+i+2, self.col_start+6, value = self.dlc_data[lc][2]*self.life_time*3600/sim_time)
                sheet.cell(self.row_start+i+2, self.col_start+3, value = sim_time)
                sheet.cell(self.row_start+i+2, self.col_start+4, value = self.time_step)
                sheet.cell(self.row_start+i+2, self.col_start+5, value = int(sim_time/float(self.time_step)))
                print(lc, self.dlc_data[lc][2]*self.life_time)

            # load case defined by times
            else:
                sheet.cell(self.row_start+i+2, self.col_start,   value = lc)
                sheet.cell(self.row_start+i+2, self.col_start+1, value = self.dlc_data[lc][3])
                sheet.cell(self.row_start+i+2, self.col_start+2, value = self.dlc_data[lc][4])
                sheet.cell(self.row_start+i+2, self.col_start+7, value = self.dlc_data[lc][2]*self.life_time*sim_time/3600)
                sheet.cell(self.row_start+i+2, self.col_start+6, value = self.dlc_data[lc][2]*self.life_time)
                sheet.cell(self.row_start+i+2, self.col_start+3, value = sim_time)
                sheet.cell(self.row_start+i+2, self.col_start+4, value = self.time_step)
                sheet.cell(self.row_start+i+2, self.col_start+5, value = int(sim_time/float(self.time_step)))

            if lc.startswith('1'):
                sheet.cell(self.row_start+i+2, self.col_start+8,   value = 'Power production')
            elif lc.startswith('2'):
                sheet.cell(self.row_start+i+2, self.col_start+8,   value = 'Power production with fault')
            elif lc.startswith('3'):
                sheet.cell(self.row_start+i+2, self.col_start+8,   value = 'Start up')
            elif lc.startswith('4'):
                sheet.cell(self.row_start+i+2, self.col_start+8,   value = 'Normal stop')
            elif lc.startswith('6'):
                sheet.cell(self.row_start+i+2, self.col_start+8,   value = 'Idling')
        print('occurrence is done!')
        
if __name__ == '__main__':

    project = r"\\172.20.4.132\fs02\CE\WE3600NB-167\project\3.6_167_105_IEC61400-1_ed4_model_oneclick_new_name_V1.3.xlsm"
    sheet   = 'occurrence'
    excel   = r"\\172.20.4.132\fs02\CE\WE3600NB-167\report\Occurrence.xlsx"
    table   = openpyxl.load_workbook(excel)

    result  = Occurrence(project, table, sheet, time_step=0.05, row_start=3, col_start=2, simple=True)
    table.save(excel)