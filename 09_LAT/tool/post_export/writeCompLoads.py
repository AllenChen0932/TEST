# -*- coding: utf-8 -*-
# @Time    : 2020/07/25 11:04
# @Author  : CJG
# @File    : writeCompLoads.py

import os
import shutil
import openpyxl
from tool.post_export.writeRainflow_v1_1 import writeRainflow
from tool.post_export.writeLDDLRD_v1_0 import writeLDDLRD
from tool.post_export.writeBasicStats import writeBasicStats
from tool.post_export.writeUltimate_v2 import writeUltimate
from tool.post_export.writeOccurrence import Occurrence

class writeCompLoads(object):

    def __init__(self, postpath, resultpath, templpath, selcomp, timestep, lct):

        self.postpath   = postpath
        self.resultpath = resultpath
        # self.projxls = projxls      # load case table
        self.templpath  = templpath   # path holding template excels for component loads output
        self.selcomp    = selcomp     # {component: {'fat': True/False, 'ult': True/False} determining whether output Fatigue/Ultimate loads
        self.timestep   = timestep
        self.lc_table   = lct

        self.comppost   = {'ult': {}, 'fat': {}, 'ldd': {}, 'lrd': {}, 'bstats': {}, 'extpl': None, 'towcl': None}
        self.compout    = {}    # {component: outpath}
        self.templfile  = {}    # {component: excel}

        self.parsePostPath()
        self.parseResultPath()
        self.parseTemplPath()
        self.writeComponents()

    def parsePostPath(self):
        # how about read post paths for each component from a configuration file?
        '''
        define post data path for fatigue/ultimate/bstats/ldd/lrd etc.
        :return: dict {'ult': {compload: path}, 'fat': {compload: path}, 'ldd': {compload: path}, etc...}
        '''

        self.comppost['fat']['main']         = self.parseMainPath(os.path.join(self.postpath, r'08_Fatigue\05_Main'), flag='F')
        self.comppost['fat']['brs']          = os.path.join(self.postpath, r'08_Fatigue\01_BRS\brs1')
        self.comppost['fat']['bus']          = os.path.join(self.postpath, r'08_Fatigue\02_BUS\bus1')
        self.comppost['fat']['br1_Mxy_seg']  = os.path.join(self.postpath, r'08_Fatigue\03_BR_Mxy_Seg\br_comb')
        self.comppost['fat']['brs1_Mxy_seg'] = os.path.join(self.postpath, r'08_Fatigue\04_BRS_Mxy_Seg\brs1_comb')
        self.comppost['fat']['brs2_Mxy_seg'] = os.path.join(self.postpath, r'08_Fatigue\04_BRS_Mxy_Seg\brs2_comb')
        self.comppost['fat']['brs3_Mxy_seg'] = os.path.join(self.postpath, r'08_Fatigue\04_BRS_Mxy_Seg\brs3_comb')
        self.comppost['fat']['tower']        = os.path.join(self.postpath, r'08_Fatigue\06_Tower')
        self.comppost['fat']['foundation']   = os.path.join(self.postpath, r'08_Fatigue\07_Foundation')
        self.comppost['fat']['tower_s0']     = os.path.join(self.postpath, r'08_Fatigue\06_Tower')

        self.comppost['ult']['main_inclsf']  = self.parseMainPath(os.path.join(self.postpath, r'07_Ultimate\08_Main_Inclsf'), flag='U')  # /br,brs,hs,hr,tt,tb
        self.comppost['ult']['main_exclsf']  = self.parseMainPath(os.path.join(self.postpath, r'07_Ultimate\09_Main_Exclsf'), flag='U')  # /br,brs,hs,hr,tt,tb
        self.comppost['ult']['main2_inclsf'] = os.path.join(self.postpath, r'07_Ultimate\08_Main_Inclsf')
        self.comppost['ult']['main2_exclsf'] = os.path.join(self.postpath, r'07_Ultimate\09_Main_Exclsf')
        self.comppost['ult']['brs_inclsf']   = os.path.join(self.postpath, r'07_Ultimate\01_BRS_Inclsf')
        self.comppost['ult']['brs_exclsf']   = os.path.join(self.postpath, r'07_Ultimate\02_BRS_Exclsf')
        self.comppost['ult']['bus_inclsf']   = os.path.join(self.postpath, r'07_Ultimate\03_BUS_Inclsf')
        self.comppost['ult']['bus_exclsf']   = os.path.join(self.postpath, r'07_Ultimate\04_BUS_Exclsf')
        self.comppost['ult']['hrbr_inclsf']  = os.path.join(self.postpath, r'07_Ultimate\05_HR_BR_Inclsf\hr')
        self.comppost['ult']['hrwo8_inclsf'] = os.path.join(self.postpath, r'07_Ultimate\07_HR_Wodlc8\Inclsf\hr')
        self.comppost['ult']['hrwo8_exclsf'] = os.path.join(self.postpath, r'07_Ultimate\07_HR_Wodlc8\Exclsf\hr')
        self.comppost['ult']['hswo8_inclsf'] = os.path.join(self.postpath, r'07_Ultimate\07_HR_Wodlc8\Inclsf\hs')
        self.comppost['ult']['hswo8_exclsf'] = os.path.join(self.postpath, r'07_Ultimate\07_HR_Wodlc8\Exclsf\hs')
        self.comppost['ult']['tower_inclsf'] = os.path.join(self.postpath, r'07_Ultimate\10_Tower_Inclsf')
        self.comppost['ult']['tower_exclsf'] = os.path.join(self.postpath, r'07_Ultimate\11_Tower_Exclsf')
        self.comppost['ult']['tower12']      = os.path.join(self.postpath, r'07_Ultimate\12_Tower_dlc12')
        self.comppost['ult']['nacacc']       = os.path.join(self.postpath, r'07_Ultimate\13_Nacelle_Acc\nac')
        self.comppost['ult']['hronly8_inclsf'] = os.path.join(self.postpath, r'07_Ultimate\06_HR_Onlydlc8\Inclsf\hr')
        self.comppost['ult']['hronly8_exclsf'] = os.path.join(self.postpath, r'07_Ultimate\06_HR_Onlydlc8\Exclsf\hr')

        self.comppost['lrd']['hs_64']  = os.path.join(self.postpath, r'06_LRD\hs_64')
        self.comppost['lrd']['hs_144'] = os.path.join(self.postpath, r'06_LRD\hs_144')
        self.comppost['ldd']['hs_64']  = os.path.join(self.postpath, r'05_LDD\hs_64')  # bins=64
        self.comppost['ldd']['hs_144'] = os.path.join(self.postpath, r'05_LDD\hs_144')  # bins=64
        self.comppost['ldd']['hr'] = os.path.join(self.postpath, r'05_LDD\hr_64')
        self.comppost['ldd']['yb'] = os.path.join(self.postpath, r'05_LDD\yb_64')


        self.comppost['extpl']           = os.path.join(self.postpath, r'01_Extrapolation')
        self.comppost['towcl']           = os.path.join(self.postpath, r'03_Clearance')
        self.comppost['bstats']['hs_Mx'] = os.path.join(self.postpath, r'02_Bstats\mb')

    def parseResultPath(self):
        # how about read result paths for each component from a configuration file?
        '''
        define result output path for each component
        :return: dict {component: result path}
        '''
        if not os.path.exists(self.resultpath):
            os.mkdir(self.resultpath)
        self.compout['blade']        = os.path.join(self.resultpath, r'Blade_Section')
        self.compout['bladeroot']    = os.path.join(self.resultpath, r'Blade_Section\blade_root')
        self.compout['drivetrain']   = os.path.join(self.resultpath, r'drive_train')
        self.compout['gearbox']      = os.path.join(self.resultpath, r'Gearbox')
        # self.compout['gentorque']    = os.path.join(self.resultpath, r'generator_torque')
        self.compout['hub']          = os.path.join(self.resultpath, r'Hub_Check')
        self.compout['mainshaft']    = os.path.join(self.resultpath, r'Main_Shaft&Bearing_r')
        self.compout['mainbearing']  = os.path.join(self.resultpath, r'Main_Shaft&Bearing_r')
        self.compout['nacacc']       = os.path.join(self.resultpath, r'Nacelle_Acceleration')
        self.compout['pitchbearing'] = os.path.join(self.resultpath, r'Pitch_Bearing_blade1')
        self.compout['pitchlock']    = os.path.join(self.resultpath, r'Pitch_Lock')
        self.compout['pitchsystem']  = os.path.join(self.resultpath, r'Pitch_System')
        self.compout['tower']        = os.path.join(self.resultpath, r'Tower_Section')
        self.compout['yawbearing']   = os.path.join(self.resultpath, r'Yaw_Bearing')

        # for k, path in self.compout.items():
        #     if not os.path.exists(path):
        #         os.mkdir(path)

    def parseTemplPath(self):
        '''
        define template excel for component loads output (contains only components requiring fat/ult/ldd/lrd/bstats loads)
        :return: dict {component: templfile}
        '''
        # template containing Ultimate,DEL,RFC,Markov,LDD/LRD
        self.templfile['pitchbearing'] = os.path.join(self.templpath, 'pitch_bearing_load_information.xlsx')
        self.templfile['yawbearing']   = os.path.join(self.templpath, 'yaw_bearing_load_information.xlsx')
        self.templfile['gearbox']      = os.path.join(self.templpath, 'main_bearing_load_information.xlsx')
        self.templfile['mainshaft']    = os.path.join(self.templpath, 'main_shaft_load_information.xlsx')
        self.templfile['mainbearing']  = os.path.join(self.templpath, 'main_bearing_load_information.xlsx')
        # template containing only Ultimate
        self.templfile['blade']       = os.path.join(self.templpath, 'ultimate_blank.xlsx')
        self.templfile['tower']       = os.path.join(self.templpath, 'ultimate_blank.xlsx')
        self.templfile['hub']         = os.path.join(self.templpath, 'ultimate_blank.xlsx')
        self.templfile['pitchlock']   = os.path.join(self.templpath, 'ultimate_blank.xlsx')
        self.templfile['nacacc']      = os.path.join(self.templpath, 'ultimate_blank.xlsx')
        self.templfile['pitchsystem'] = os.path.join(self.templpath, 'ultimate_blank.xlsx')
        self.templfile['drivetrain']  = os.path.join(self.templpath, 'ultimate_blank.xlsx')

    def writeComponents(self):

        selcomp  = self.selcomp
        compout  = self.compout
        comppost = self.comppost

        # writeOccur(self.projxls, self.postpath, self.resultpath, table)

        # blade
        if 'blade' in selcomp.keys():
            if selcomp['blade']['fat'] or selcomp['blade']['ult']:
                print('Begin to write component: blade')
                if not os.path.exists(compout['blade']):
                    os.mkdir(compout['blade'])

            if selcomp['blade']['fat'] and os.path.isdir(comppost['fat']['brs']):
                writeRainflow(comppost['fat']['brs'],
                              content=('DEL',)   ).write2singletxt(os.path.join(compout['blade'], 'brs1_del.txt'))
                writeRainflow(comppost['fat']['brs'],
                              content=('Markov',)).write2singletxt(os.path.join(compout['blade'], 'brs1_Markov.txt'))

            if selcomp['blade']['fat'] and os.path.isdir(comppost['fat']['bus']):
                writeRainflow(comppost['fat']['bus'],
                              content=('DEL',)   ).write2singletxt(os.path.join(compout['blade'], 'bus1_del.txt'))
                writeRainflow(comppost['fat']['bus'],
                              content=('Markov',)).write2singletxt(os.path.join(compout['blade'], 'bus1_Markov.txt'))

            if selcomp['blade']['ult']:
                table = openpyxl.load_workbook(self.templfile['blade'])
                if os.path.isdir(comppost['ult']['brs_inclsf']):
                    writeUltimate(comppost['ult']['brs_inclsf'], table, sheetname='brs_inclsf', heightflag=True)
                    writeUltimate(comppost['ult']['brs_exclsf'], table, sheetname='brs_exclsf', heightflag=True)

                if os.path.isdir(comppost['ult']['bus_inclsf']):
                    writeUltimate(comppost['ult']['bus_inclsf'], table, sheetname='bus_inclsf', heightflag=True)
                    writeUltimate(comppost['ult']['bus_exclsf'], table, sheetname='bus_exclsf', heightflag=True)
                table.save(os.path.join(compout['blade'], 'ultimate_blade.xlsx'))

        # blade root
        if 'bladeroot' in selcomp.keys():
            if selcomp['bladeroot']['fat']:
                print('Begin to write component: blade root')
                if not os.path.exists(compout['bladeroot']):
                    os.mkdir(compout['bladeroot'])

            if os.path.isdir(comppost['fat']['brs1_Mxy_seg']) and selcomp['bladeroot']['fat']:
                writeRainflow(comppost['fat']['brs1_Mxy_seg'], content=('Markov',),
                              variable=()).write2singletxt(os.path.join(compout['bladeroot'],'brs1_Mxy_seg_Markov.txt'))
            if os.path.isdir(comppost['fat']['brs2_Mxy_seg']) and selcomp['bladeroot']['fat']:
                writeRainflow(comppost['fat']['brs2_Mxy_seg'], content=('Markov',),
                              variable=()).write2singletxt(os.path.join(compout['bladeroot'],'brs2_Mxy_seg_Markov.txt'))
            if os.path.isdir(comppost['fat']['brs3_Mxy_seg']) and selcomp['bladeroot']['fat']:
                writeRainflow(comppost['fat']['brs3_Mxy_seg'],content=('Markov',),
                              variable=()).write2singletxt(os.path.join(compout['bladeroot'],'brs3_Mxy_seg_Markov.txt'))

        # hub check
        if 'hub' in selcomp.keys():
            if selcomp['hub']['fat'] or selcomp['hub']['ult']:
                print('Begin to write component: hub')
                if not os.path.exists(compout['hub']):
                    os.mkdir(compout['hub'])
                    if not os.path.exists(os.path.join(compout['hub'],'Markov')):
                        os.mkdir(os.path.join(compout['hub'],'Markov'))

            if selcomp['hub']['fat'] and os.path.isdir(comppost['fat']['main']['br1']):
                writeRainflow(comppost['fat']['main']['br1'],
                              content=('Markov',)).write2multitxt(os.path.join(compout['hub'], r'Markov\Blade_root1'))
            if selcomp['hub']['fat'] and os.path.isdir(comppost['fat']['main']['br2']):
                writeRainflow(comppost['fat']['main']['br2'],
                              content=('Markov',)).write2multitxt(os.path.join(compout['hub'], r'Markov\Blade_root2'))
            if selcomp['hub']['fat'] and os.path.isdir(comppost['fat']['main']['br3']):
                writeRainflow(comppost['fat']['main']['br3'],
                              content=('Markov',)).write2multitxt(os.path.join(compout['hub'], r'Markov\Blade_root3'))
            if selcomp['hub']['fat'] and os.path.isdir(comppost['fat']['main']['hr']):
                writeRainflow(comppost['fat']['main']['hr'],
                              content=('Markov',)).write2multitxt(os.path.join(compout['hub'], r'Markov\Hub_rotating'))
            if selcomp['hub']['ult'] and os.path.isdir(comppost['ult']['hrbr_inclsf']):
                table = openpyxl.load_workbook(self.templfile['hub'])
                writeUltimate(comppost['ult']['hrbr_inclsf'], table, rowstart=3, sheetname='extreme')
                table.save(os.path.join(compout['hub'], 'ultimate_hub.xlsx'))

        # pitch bearing
        if 'pitchbearing' in selcomp.keys():
            if selcomp['pitchbearing']['fat'] or selcomp['pitchbearing']['ult']:
                print('Begin to write component: pitch bearing')
                if not os.path.exists(compout['pitchbearing']):
                    os.mkdir(compout['pitchbearing'])

                table = openpyxl.load_workbook(self.templfile['pitchbearing'])

                if selcomp['pitchbearing']['fat'] and os.path.isdir(comppost['fat']['main']['br1']):
                    writeRainflow(comppost['fat']['main']['br1'], content=('DEL', 'RFC')).write2excel(table)
                if selcomp['pitchbearing']['fat'] and os.path.isdir(comppost['fat']['br1_Mxy_seg']):
                    writeRainflow(comppost['fat']['br1_Mxy_seg'], content=('Markov',), variable=()).write2excel(table)
                if selcomp['pitchbearing']['ult'] and os.path.isdir(comppost['ult']['hrbr_inclsf']):
                    writeUltimate(comppost['ult']['hrbr_inclsf'], table, rowstart=3, sheetname='extreme')
                Occurrence(self.lc_table, table, 'occurrence', self.timestep, row_start=3, col_start=2, simple=False)
                table.save(os.path.join(compout['pitchbearing'], 'pitch_bearing_load_info.xlsx'))
                print('pitch bearing is done!')

        # yaw bearing
        if 'yawbearing' in selcomp.keys():
            if selcomp['yawbearing']['fat'] or selcomp['yawbearing']['ult']:
                print('Begin to write component: yaw bearing')
                if not os.path.exists(compout['yawbearing']):
                    os.mkdir(compout['yawbearing'])

                table = openpyxl.load_workbook(self.templfile['yawbearing'])

                if selcomp['yawbearing']['fat'] and os.path.isdir(comppost['fat']['main']['yb']):
                    writeRainflow(comppost['fat']['main']['yb'], content=('DEL',)).write2excel(table)
                    writeLDDLRD([comppost['ldd']['yb']], table.get_sheet_by_name('LDD'))
                if selcomp['yawbearing']['ult'] and os.path.isdir(comppost['ult']['main_inclsf']['yb']):
                    writeUltimate(comppost['ult']['main_inclsf']['yb'], table, rowstart=3, sheetname='extreme')
                if selcomp['yawbearing']['ult'] and os.path.isdir(comppost['ult']['main_exclsf']['yb']):
                    writeUltimate(comppost['ult']['main_exclsf']['yb'], table, rowstart=25, sheetname='extreme')
                Occurrence(self.lc_table, table, 'occurrence', self.timestep, row_start=3, col_start=2, simple=False)
                table.save(os.path.join(compout['yawbearing'], 'yaw_bearing_load_info.xlsx'))

        # main bearing:
        if 'mainbearing' in selcomp.keys():
            if selcomp['mainbearing']['fat'] or selcomp['mainbearing']['ult']:
                print('Begin to write component: main bearing')
                if not os.path.exists(compout['mainbearing']):
                    os.mkdir(compout['mainbearing'])

                table = openpyxl.load_workbook(self.templfile['mainbearing'])

                if selcomp['mainbearing']['fat'] and os.path.isdir(comppost['fat']['main']['hs']):
                    writeRainflow(comppost['fat']['main']['hs'],
                                  content=('DEL', 'RFC')).write2excel(table, del_rowstart=5, del_colstart=2)
                if selcomp['mainbearing']['fat'] and os.path.isdir(comppost['fat']['main']['hr']):
                    writeRainflow(comppost['fat']['main']['hr'],
                                  content=('DEL',)).write2excel(table, del_rowstart=5, del_colstart=11)
                if selcomp['mainbearing']['fat'] and os.path.isdir(comppost['ldd']['hs_64']) \
                        and os.path.isdir(comppost['lrd']['hs_64']):
                    writeLDDLRD([comppost['ldd']['hs_64'],
                                 comppost['lrd']['hs_64']], table.get_sheet_by_name('LDD'))
                if selcomp['mainbearing']['fat'] and os.path.isdir(comppost['bstats']['hs_Mx']):
                    writeBasicStats(comppost['bstats']['hs_Mx'], table.get_sheet_by_name('mean'))
                if selcomp['mainbearing']['ult'] and os.path.isdir(comppost['ult']['main_inclsf']['hs']):
                    writeUltimate(comppost['ult']['main_inclsf']['hs'], table, rowstart=3, colstart=2)
                if selcomp['mainbearing']['ult'] and os.path.isdir(comppost['ult']['main_inclsf']['hr']):
                    writeUltimate(comppost['ult']['main_inclsf']['hr'], table, rowstart=3, colstart=15)
                if selcomp['mainbearing']['ult'] and os.path.isdir(comppost['ult']['hswo8_inclsf']):
                    writeUltimate(comppost['ult']['hswo8_inclsf'], table, rowstart=24, colstart=2)
                if selcomp['mainbearing']['ult'] and os.path.isdir(comppost['ult']['hrwo8_inclsf']):
                    writeUltimate(comppost['ult']['hrwo8_inclsf'], table, rowstart=24, colstart=15)
                Occurrence(self.lc_table, table, 'occurrence', row_start=3, col_start=2)
                table.save(os.path.join(compout['mainbearing'], 'main_bearing_load_info.xlsx'))

        # gearbox_64:
        if 'gearbox_64' in selcomp.keys():
            if selcomp['gearbox_64']['fat'] or selcomp['gearbox_64']['ult']:
                print('Begin to write component: gearbox')
                if not os.path.exists(compout['gearbox']):
                    os.mkdir(compout['gearbox'])

                table = openpyxl.load_workbook(self.templfile['gearbox'])
                if selcomp['gearbox_64']['fat'] and os.path.isdir(comppost['fat']['main']['hs']):
                    writeRainflow(comppost['fat']['main']['hs'],
                                  content=('DEL', 'RFC')).write2excel(table, del_rowstart=5, del_colstart=2)
                if selcomp['gearbox_64']['fat'] and os.path.isdir(comppost['fat']['main']['hr']):
                    writeRainflow(comppost['fat']['main']['hr'],
                                  content=('DEL',)).write2excel(table, del_rowstart=5, del_colstart=11)
                if selcomp['gearbox_64']['fat'] and os.path.isdir(comppost['ldd']['hs_64']):
                    writeLDDLRD([comppost['ldd']['hs_64'],
                                 comppost['lrd']['hs_64']], table.get_sheet_by_name('LDD'))
                if selcomp['gearbox_64']['fat'] and os.path.isdir(comppost['bstats']['hs_Mx']):
                    writeBasicStats(comppost['bstats']['hs_Mx'], table.get_sheet_by_name('mean'))
                if selcomp['gearbox_64']['ult'] and os.path.isdir(comppost['ult']['main_inclsf']['hs']):
                    writeUltimate(comppost['ult']['main_inclsf']['hs'], table, rowstart=3, colstart=2)
                if selcomp['gearbox_64']['ult'] and os.path.isdir(comppost['ult']['main_inclsf']['hr']):
                    writeUltimate(comppost['ult']['main_inclsf']['hr'], table, rowstart=3, colstart=15)
                if selcomp['gearbox_64']['ult'] and os.path.isdir(comppost['ult']['hswo8_inclsf']):
                    writeUltimate(comppost['ult']['hswo8_inclsf'], table, rowstart=24, colstart=2)
                if selcomp['gearbox_64']['ult'] and os.path.isdir(comppost['ult']['hrwo8_inclsf']):
                    writeUltimate(comppost['ult']['hrwo8_inclsf'], table, rowstart=24, colstart=15)
                Occurrence(self.lc_table, table, 'occurrence', row_start=3, col_start=2)
                table.save(os.path.join(compout['gearbox'], 'gearbox_load_info_64.xlsx'))

        # gearbox_144
        if 'gearbox_144' in selcomp.keys():
            if selcomp['gearbox_144']['fat'] or selcomp['gearbox_144']['ult']:
                print('Begin to write component: gearbox')
                if not os.path.exists(compout['gearbox']):
                    os.mkdir(compout['gearbox'])

                table = openpyxl.load_workbook(self.templfile['gearbox'])
                if selcomp['gearbox_144']['fat'] and os.path.isdir(comppost['fat']['main']['hs']):
                    writeRainflow(comppost['fat']['main']['hs'],
                                  content=('DEL', 'RFC')).write2excel(table, del_rowstart=5, del_colstart=2)
                if selcomp['gearbox_144']['fat'] and os.path.isdir(comppost['fat']['main']['hr']):
                    writeRainflow(comppost['fat']['main']['hr'],
                                  content=('DEL',)).write2excel(table, del_rowstart=5, del_colstart=11)
                if selcomp['gearbox_144']['fat'] and os.path.isdir(comppost['ldd']['hs_144']):
                    writeLDDLRD([comppost['ldd']['hs_144'],
                                 comppost['lrd']['hs_144']], table.get_sheet_by_name('LDD'))
                    writeBasicStats(comppost['bstats']['hs_Mx'], table.get_sheet_by_name('mean'))
                if selcomp['gearbox_144']['ult'] and os.path.isdir(comppost['ult']['main_inclsf']['hs']):
                    writeUltimate(comppost['ult']['main_inclsf']['hs'], table, rowstart=3, colstart=2)
                if selcomp['gearbox_144']['ult'] and os.path.isdir(comppost['ult']['main_inclsf']['hr']):
                    writeUltimate(comppost['ult']['main_inclsf']['hr'], table, rowstart=3, colstart=15)
                if selcomp['gearbox_144']['ult'] and os.path.isdir(comppost['ult']['hswo8_inclsf']):
                    writeUltimate(comppost['ult']['hswo8_inclsf'], table, rowstart=24, colstart=2)
                if selcomp['gearbox_144']['ult'] and os.path.isdir(comppost['ult']['hrwo8_inclsf']):
                    writeUltimate(comppost['ult']['hrwo8_inclsf'], table, rowstart=24, colstart=15)
                Occurrence(self.lc_table, table, 'occurrence', row_start=3, col_start=2)
                table.save(os.path.join(compout['gearbox'], 'gearbox_load_info_144.xlsx'))

        # main shaft
        if 'mainshaft' in selcomp.keys():
            if selcomp['mainshaft']['fat'] or selcomp['mainshaft']['ult']:
                print('Begin to write component: main shaft')
                if not os.path.exists(compout['mainshaft']):
                    os.mkdir(compout['mainshaft'])
                table = openpyxl.load_workbook(self.templfile['mainshaft'])

                if selcomp['mainshaft']['fat'] and os.path.isdir(comppost['fat']['main']['hr']):
                    writeRainflow(comppost['fat']['main']['hr'], content=('DEL', 'RFC')).write2excel(table)
                if selcomp['mainshaft']['fat'] and os.path.isdir(comppost['ldd']['hr']):
                    writeLDDLRD([comppost['ldd']['hr']], table.get_sheet_by_name('LDD'))
                if selcomp['mainshaft']['fat'] and os.path.isdir(comppost['bstats']['hs_Mx']):
                    writeBasicStats(comppost['bstats']['hs_Mx'], table.get_sheet_by_name('mean'))
                if selcomp['mainshaft']['ult'] and os.path.isdir(comppost['ult']['main_inclsf']['hr']):
                    writeUltimate(comppost['ult']['main_inclsf']['hr'], table, rowstart=3)
                if selcomp['mainshaft']['ult'] and os.path.isdir(comppost['ult']['hrwo8_inclsf']):
                    writeUltimate(comppost['ult']['hrwo8_inclsf'], table, rowstart=24)
                Occurrence(self.lc_table, table, 'occurrence', row_start=3, col_start=2)
                table.save(os.path.join(compout['mainshaft'], 'main_shaft_load_info.xlsx'))

        # pitch lock
        if 'pitchlock' in selcomp.keys():
            if selcomp['pitchlock']['ult']:
                print('Begin to write component: pitch lock')
                if not os.path.exists((compout['pitchlock'])):
                    os.mkdir(compout['pitchlock'])
                table = openpyxl.load_workbook(self.templfile['pitchlock'])
                if os.path.isdir(comppost['ult']['hronly8_inclsf']):
                    writeUltimate(comppost['ult']['hronly8_inclsf'], table, sheetname='hronly8_inclsf')
                if os.path.isdir(comppost['ult']['hronly8_exclsf']):
                    writeUltimate(comppost['ult']['hronly8_exclsf'], table, sheetname='hronly8_exclsf')
                table.save(os.path.join(compout['pitchlock'], 'ultimate_pitch_lock.xlsx'))

        # pitch system
        if 'pitchsystem' in selcomp.keys():
            if selcomp['pitchsystem']['ult']:
                print('Begin to write component: pitch system')
                if not os.path.exists((compout['pitchsystem'])):
                    os.mkdir(compout['pitchsystem'])
                table = openpyxl.load_workbook(self.templfile['pitchsystem'])
                if os.path.isdir(comppost['ult']['main2_inclsf']):
                    writeUltimate(comppost['ult']['main2_inclsf'], table, sheetname='main_inclsf')
                table.save(os.path.join(compout['pitchsystem'], 'ultimate_pitch_system.xlsx'))

        # nac acc
        if 'nacacc' in selcomp.keys():
            if selcomp['nacacc']['ult']:
                print('Begin to write component: nacelle acceleration')
                if not os.path.exists((compout['nacacc'])):
                    os.mkdir(compout['nacacc'])
                table = openpyxl.load_workbook(self.templfile['nacacc'])
                if os.path.isdir(comppost['ult']['nacacc']):
                    writeUltimate(comppost['ult']['nacacc'], table, sheetname='extreme')
                table.save(os.path.join(compout['nacacc'], 'ultimate_nacacc.xlsx'))

        # tower
        if 'tower' in selcomp.keys():
            if selcomp['tower']['fat'] or selcomp['tower']['ult']:
                print('Begin to write component: tower')
                if not os.path.exists(compout['tower']):
                    os.mkdir(compout['tower'])
                tower_bottom = os.path.join(comppost['fat']['tower_s0'], os.listdir(comppost['fat']['tower_s0'])[0])
                if selcomp['tower']['fat'] and os.path.isdir(tower_bottom):
                    writeRainflow(tower_bottom,
                                  content=('Markov',),
                                  variable=('Mxy',)).write2singletxt(os.path.join(compout['tower'],
                                                                                  'tower_0_Mxy_Markov.txt'))
                if selcomp['tower']['fat'] and os.path.isdir(comppost['fat']['tower']):
                    writeRainflow(comppost['fat']['tower'],
                                  content=('DEL',)).write2singletxt(os.path.join(compout['tower'], 'tower_del.txt'))

                table = openpyxl.load_workbook(self.templfile['tower'])
                if selcomp['tower']['ult'] and os.path.isdir(comppost['ult']['tower_inclsf']):
                    writeUltimate(comppost['ult']['tower_inclsf'], table, sheetname='tower_inclsf', heightflag=False)
                if selcomp['tower']['ult'] and os.path.isdir(comppost['ult']['tower_exclsf']):
                    writeUltimate(comppost['ult']['tower_exclsf'], table, sheetname='tower_exclsf', heightflag=False)
                if selcomp['tower']['ult'] and os.path.isdir(comppost['ult']['tower12']):
                    writeUltimate(comppost['ult']['tower12'], table, sheetname='tower_dlc12', heightflag=False)
                table.save(os.path.join(compout['tower'], 'ultimate_tower.xlsx'))

    def parseMainPath(self, MainPostPath, flag):
        main_path_dic = {}
        if flag == 'F':
            # br, brs
            main_path_dic['br1'] = MainPostPath + '\\br1'
            main_path_dic['br2'] = MainPostPath + '\\br2'
            main_path_dic['br3'] = MainPostPath + '\\br3'

            main_path_dic['brs1'] = MainPostPath + '\\brs1_0'
            main_path_dic['brs2'] = MainPostPath + '\\brs2_0'
            main_path_dic['brs3'] = MainPostPath + '\\brs3_0'
        elif flag == 'U':
            main_path_dic['br']  = MainPostPath + '\\br'
            main_path_dic['brs'] = MainPostPath + '\\brs_0.000'
        # hr, hs, yb
        main_path_dic['hr'] = MainPostPath + '\\hr'
        main_path_dic['hs'] = MainPostPath + '\\hs'
        main_path_dic['yb'] = MainPostPath + '\\yb'
        # tower top, tower bottom
        twr_subpath = []
        twr_height  = []
        for subdir in os.listdir(MainPostPath):
            if os.path.isdir(os.path.join(MainPostPath, subdir)):
                if subdir.startswith('tr_') or subdir.startswith('Mbr_'):
                    # print(subdir.split('tr_')[1])
                    twr_subpath.append(subdir)
                    twr_height.append(float(subdir.split('_')[1]))
        if twr_height:
            twr_height.sort()
        else:
            raise Exception('No tower section loads end with height!\n%s' %MainPostPath)
        # tt_idx = twr_height[-1]
        # tb_idx = twr_height[0]
        main_path_dic['tt'] = MainPostPath + '\\' + twr_subpath[-1]
        main_path_dic['tb'] = MainPostPath + '\\' + twr_subpath[0]
        return main_path_dic

if __name__ == '__main__':

    postpath   = r'\\172.20.0.4\fs03\CJG\V2B\loop6\run_vr11.9\post'
    resultpath = r'\\172.20.0.4\fs03\CJG\V2B\loop6\run_vr11.9\result_test'
    templpath = r'E:\01 tool dev\02 post tool\06_readBladedPost\output_template\20200803'

    selcomp = {'blade':        {'fat': False, 'ult': False},
               'bladeroot':    {'fat': False},
               'hub':          {'fat': False, 'ult': False},
               'pitchbearing': {'fat': False, 'ult': False},
               'yawbearing':   {'fat': False, 'ult': False},
               'mainbearing':  {'fat': False, 'ult': False},
               'mainshaft':    {'fat': False, 'ult': False},
               'tower':        {'fat': True, 'ult': True},
               'nacacc':                      {'ult': False},
               'pitchlock':                   {'ult': False},
               'pitchsystem':                 {'ult': False}}

    # 'drivetrain': {'fat': False, 'ult': False},
    # 'gearbox': {'fat': False, 'ult': False},
    writeCompLoads(postpath, resultpath, templpath, selcomp)

# # drivetrain
# if selcomp['drivetrain']['fat'] or selcomp['drivetrain']['ult']:
#     # if not os.path.exists(compout['drivetrain']):
#     #     os.mkdir(compout['drivetrain'])
#     # table = openpyxl.load_workbook(self.templfile['drivetrain'])
#     # if selcomp['drivetrain']['fat']:
#     #     writeLDDLRD([comppost['ldd']['hs'], comppost['lrd']['hs']], table.get_sheet_by_name('LDD_hs'))
#     # if selcomp['drivetrain']['ult']:
#     #     writeUltimate(comppost['ult']['main_inclsf']['hs'], table, rowstart=3, colstart=2)
#     #     writeUltimate(comppost['ult']['main_inclsf']['hr'], table, rowstart=3, colstart=15)
#     #     writeUltimate(comppost['ult']['hswo8_inclsf'], table, rowstart=24, colstart=2)
#     #     writeUltimate(comppost['ult']['hrwo8_inclsf'], table, rowstart=24, colstart=15)
#     # table.save(os.path.join(compout['drivetrain'], 'drivetrain_load_info.xlsx'))
#     pass
#
# # gearbox
# if selcomp['gearbox']['fat'] or selcomp['gearbox']['ult']:   # same to main_bearing/drivetrain
#     # if not os.path.exists(compout['gearbox']):
#     #     os.mkdir(compout['gearbox'])
#     # table = openpyxl.load_workbook(self.templfile['gearbox'])
#     # if selcomp['gearbox']['fat']:
#     #     writeRainflow(comppost['fat']['main']['hs'], content=('DEL', 'RFC')).write2excel(table)
#     #     writeRainflow(comppost['fat']['main']['hr'], content=('DEL',)).write2excel(table, del_colstart=11)
#     #     writeLDDLRD([comppost['ldd']['hs'], comppost['lrd']['hs']], table.get_sheet_by_name('LDD_hs'))
#     #     writeBasicStats(comppost['bstats']['hr_Mx'], table.get_sheet_by_name('mean'))
#     # if selcomp['gearbox']['ult']:
#     #     writeUltimate(comppost['ult']['main_inclsf']['hs'], table, rowstart=3, colstart=2)
#     #     writeUltimate(comppost['ult']['main_inclsf']['hr'], table, rowstart=3, colstart=15)
#     #     writeUltimate(comppost['ult']['hswo8_inclsf'], table, rowstart=24, colstart=2)
#     #     writeUltimate(comppost['ult']['hrwo8_inclsf'], table, rowstart=24, colstart=15)
#     # table.save(os.path.join(compout['gearbox'], 'gearbox_load_info.xlsx'))
#     pass



