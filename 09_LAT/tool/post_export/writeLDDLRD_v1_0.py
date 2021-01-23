# -*- coding: utf-8 -*-
# @Time    : 2020/07/16 15:55
# @Author  : CJG
# @File    : writeLDDLRD_v1_0.py

import os
import numpy as np
import openpyxl

class writeLDDLRD(object):

    def __init__(self, probdpath, sheetout):
        '''
        read and write probdist (LDD/LRD) result from BLADED v4.8
        :param probdpath: define LDD (and LRD) path in a list, you can include only LDD path.
        '''
        self.probdpath = probdpath
        self.sheetout  = sheetout
        self.probd = ''
        self.ldd = {'perfilelist': [], 'varlist': [], 'per_var': {}, 'var_ldd': {}}
        self.lrd = {'perfilelist': [], 'varlist': [], 'per_var': {}, 'var_lrd': {}}
        self.pi = 3.141592653589793

        # self.read_probd_result()
        self.write_probd_result()

    def read_probd_result(self):
        for path in self.probdpath:
            for file in os.listdir(path):
                if os.path.splitext(file)[1][0:2] == '.%':

                    [pername, perext] = os.path.splitext(file)
                    perext = perext[2:]

                    # read percent and dollar file USING 'read_Bladed'
                    res = read_Bladed(pername, path, perext)

                    # print(res.var_list)
                    if res.var_list[1] == 'Time at level per bin':

                        self.probd = 'ldd'
                        self.ldd['perfilelist'].append(file)
                        var = res.var_list[0]
                        self.ldd['varlist'].append(var)
                        self.ldd['per_var'][perext] = var
                        self.ldd['var_ldd'][var] = res.data[:, 0:2]

                    elif res.var_list[1] == 'Revs at level per bin':

                        self.probd = 'lrd'
                        self.lrd['perfilelist'].append(file)
                        var = res.var_list[0]
                        self.lrd['varlist'].append(var)
                        self.lrd['per_var'][perext] = var
                        self.lrd['var_lrd'][var] = res.data[:, 0:2]
                        # print(self.lrd['var_lrd']['Stationary hub Mx'])

                    else:
                        print('The given path does NOT content LDD or LRD result!')

        # print(self.per_var)
        # print(self.var_ldd['Stationary hub Mz'])
        # print(self.probd)

    def write_probd_result(self):

        # read LDD and/or LRD result
        self.read_probd_result()

        # write LDD and/or LRD result
        sheet = self.sheetout
        row_start = 5
        col_start = 2

        if len(self.probdpath) == 2:
            # write both LDD and LRD into one single excel sheet (e.g. Stationary hub LDD)
            # Format: [Load, Time at level per bin, Revs at level per bin, Rotor speed], where Load = Mx,My,Mz,Fx,Fy,Fz

            try:
                for i in range(6):
                    var_intemple = self.sheetout.cell(row=row_start-2, column=col_start+4*i).value
                    # print(var_intemple)
                    if var_intemple in self.ldd['varlist'] and var_intemple in self.lrd['varlist']:
                        var = var_intemple
                        ldddata = self.ldd['var_ldd'][var]
                        lrddata = self.lrd['var_lrd'][var]
                        # print(ldddata.shape)
                        for j in range(ldddata.shape[0]):
                            sheet.cell(row_start+j, col_start+4*i,   ldddata[j, 0]/1000)  # load per bin
                            sheet.cell(row_start+j, col_start+1+4*i, ldddata[j, 1]/3600)  # time per bin
                            sheet.cell(row_start+j, col_start+2+4*i, lrddata[j, 1]/2/self.pi)
                            # sheet.cell(row_start+j, col_start+2+4*i, lrddata[j,1]/ldddata[j,1]/2/self.pi*60)
                    else:
                        print('Variables in template does NOT match variables in given LDD/LRD result!')

            except IOError:
                print('Make sure both LDD and LRD paths are defined!')

        elif len(self.probdpath) == 1 and self.probd == 'ldd':
            # write LDD into given excel sheet (e.g. Rotationary hub LDD)
            # Format: [Load, Time at level per bin, Acc. time above level], where Load = Mx,My,Mz,Fx,Fy,Fz

            if 'Yaw bearing Mz' in self.ldd['varlist']:
                var_intemple = self.sheetout.cell(row=row_start-2, column=col_start).value  # yaw bearing Mz
                # print(var_intemple)
                if var_intemple in self.ldd['varlist']:
                    var = var_intemple
                    ldddata = self.ldd['var_ldd'][var]
                    for j in range(ldddata.shape[0]):
                        sheet.cell(row_start+j, col_start,   ldddata[j,0]/1000)  # load per bin
                        sheet.cell(row_start+j, col_start+1, ldddata[j,1]/3600)  # time per bin
                else:
                    print('Variables in template does NOT match variables in given LDD/LRD result!')

            else:
                for i in range(6):
                    var_intemple = self.sheetout.cell(row=row_start-2, column=col_start+3*i).value
                    # print(var_intemple)
                    if var_intemple in self.ldd['varlist']:
                        var = var_intemple
                        ldddata = self.ldd['var_ldd'][var]
                        for j in range(ldddata.shape[0]):
                            sheet.cell(row_start+j, col_start+3*i,   ldddata[j,0]/1000)  # load per bin
                            sheet.cell(row_start+j, col_start+1+3*i, ldddata[j,1]/3600)  # time per bin
                    else:
                        print('Variables in template does NOT match variables in given LDD/LRD result!')


class read_Bladed(object):

    def __init__(self, pj_name, pj_path, ext):
        '''
        COPIED from 'Read_Bladed.py' by CE.
        read a single bladed percent-dollar file,
        output variable list and data in narray format.
        :param name: pj name
        :param path: pj path
        :param ext: file extension ID
        '''
        self.pj_name = pj_name
        self.pj_path = pj_path
        self.ext = str(ext)
        self.format = ''
        self.type = ''
        self.dim = []
        self.var_list = []
        self.data = None

        self.read_percent()
        self.read_dollar()

    def read_percent(self):
        file = ''.join([self.pj_name, '.%', self.ext])
        path = os.sep.join([self.pj_path, file])
        if not os.path.isfile(path):
            raise Exception('%s is not a file!'%path)
        try:
            with open(path, 'r') as f:
                for line in f.readlines():
                    if 'ACCESS' in line:
                        self.format = line.strip("\n\t").split()[-1]
                    if 'RECL' in line:
                        self.type = line.strip().split()[-1]
                    if line.startswith('DIMENS'):
                        self.dim = line.strip().split()[1:]
                    if 'VARIAB' in line:
                        self.var_list = line.strip().split("'")[1::2]
                        # print(self.var_list)
                        break
            return self.var_list
        except IOError:
            print('Open "%s" failed!' % path)

    def read_dollar(self):

        # read percent to get format info
        # self.read_percent()
        dollar_path = os.path.join(self.pj_path, ''.join([self.pj_name, '.$', self.ext]))
        if not os.path.isfile(dollar_path):
            raise Exception('%s is not a file!' %dollar_path)

        # read dollar file to get stored data in original format
        if self.format == 'D':  # binary format
            try:
                with open(dollar_path, 'rb') as f:
                    if self.type == '4':
                        self.data = np.fromfile(f, np.float32)
                    elif self.type == '8':
                        self.data = np.fromfile(f, np.float64)
            except IOError:
                print('open "%s" failed' % dollar_path)

        elif self.format == 'S':  # txt format
            try:
                with open(dollar_path, 'r') as f:
                    if self.type == '4':
                        self.data = np.loadtxt(f, np.float32)
                    elif self.type == '8':
                        self.data = np.loadtxt(f, np.float64)
            except IOError:
                print('open "%s" failed' % dollar_path)

        # reshape data for output.
        if len(self.dim) == 2:
            self.data = self.data.reshape((int(self.dim[1]), int(self.dim[0])))
            # self.data_len = int(self.dim[1])
        elif len(self.dim) == 3:
            self.data = self.data.reshape((int(self.dim[2]), int(self.dim[1]), int(self.dim[0])))
            # self.data_len = int(self.dim[2])

        return self.data


if __name__ == '__main__':
    # input:
    probdpath = [r'\\172.20.4.132\fs02\CE\V3\loop06\post_1119\05_LDD\hs_144',
                 r'\\172.20.4.132\fs02\CE\V3\loop06\post_1119\06_LRD\hs_144']
    excel_temp = r"\\172.20.4.132\fs02\CE\V3\loop06\post_1119\gearbox_LDD_LRD.xlsx"
    excel_out  = r"\\172.20.4.132\fs02\CE\V3\loop06\post_1119\gearbox_LDD_LRD.xlsx"

    # probdpath = [r'\\172.20.0.4\fs02\LSY\V3\loop04_backup20200623\BackUp\post\ldd\Yaw']
    # excelouttemple = r"E:\01 tool dev\02 post tool\06_readBladedPost\test_LDD_yb - 副本.xlsx"
    # excelout = r"E:\01 tool dev\02 post tool\06_readBladedPost\test_LDD_yb.xlsx"

    # probdpath = [r'\\172.20.0.4\fs03\CJG\V2B\loop2.2\post\LDD\hr']
    # excelouttemple = r"E:\01 tool dev\02 post tool\06_readBladedPost\test_LDD_hr - 副本.xlsx"
    # excelout = r"E:\01 tool dev\02 post tool\06_readBladedPost\test_LDD_hr.xlsx"

    table = openpyxl.load_workbook(excel_temp)
    sheet = table.get_sheet_by_name('LDD_hs_loop6')

    writeLDDLRD(probdpath, sheet)

    table.save(excel_out)

    # read_Bladed('hs', probdpath, '003').read_dollar()

