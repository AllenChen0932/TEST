# -*- coding: utf-8 -*-
# @Time    : 2020/07/31 11:10
# @Author  : TJ
# @File    : writeUltimate.py
'''
if given a path with subfolders, write ultimate data in a column of tables;
if given a path containing ult data, write data to a single table.
Modification from CJG:
1. reorgonize function calling;
2. write Ultimate to 'extreme' sheet and given row&col;
3. add function to sort tower/blade sections;
4. get load & variable names from .MX file.
'''

import os
import openpyxl
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment

class writeUltimate(object):

    def __init__(self, ultpath, table, sheetname='extreme', rowstart=2, colstart=2, rowspace=1, heightflag=False):

        self.ultpath = ultpath
        self.table   = table
        self.sheetname = sheetname
        self.rowstart  = rowstart
        self.colstart  = colstart
        self.rowspace  = rowspace
        self.heightflag  = heightflag
        self.MXfile_list = []

        self.get_MXfilelist()
        self.write_result()

    def get_MXfilelist(self):
        '''
        get .$MX file (NOT case-specified $MX) list in the given post path.
        :return: $PJ-specified $MX file list.
        '''
        height_file = {}
        for root, dirs, files in os.walk(self.ultpath, False):
            for file in files:
                if '.$PJ' in file.upper():

                    pj_name = file.split('.$')[0]
                    mx_name = pj_name + '.$MX'
                    mx_file = os.path.join(root, mx_name)
                    if not os.path.isfile(mx_file):
                        raise Exception('No results in %s' %mx_file)

                    self.MXfile_list.append(mx_file)

                    if self.heightflag:
                        height = float(pj_name.split('_')[-1])
                        height_file.update({height: mx_file})

        if self.heightflag:
            heightlist = [key for key in height_file.keys()]
            heightlist.sort()
            self.MXfile_list = [height_file[h] for h in heightlist]
            # print(self.MXfile_list)

    def read_dollarMX(self, file):

        data = []  # ultimate table
        var_list = []  # variable name without force/moment component, e.g. Blade root 1, Rotationary hub
        load_list = []  # force/moment component, e.g. Fx, My ...

        with open(file, 'r') as f:

            lines = f.readlines()
            for line in lines:

                if ('MAXIMUM' in line) or ('MINIMUM' in line):

                    item = line.strip().split('\t')

                    # get variable and force/moment components list in the given .$MX file
                    item[0] = item[0].split("'")[1]  # variable name
                    loadflag = False
                    var_parse = item[0].split(" ")
                    # print(var_parse)
                    for term in var_parse:
                        if term in ['Fx','Fy','Fz','Fxy','Fyz','Fxz','Mx','My','Mz','Mxy','Myz','Mxz']:
                            var_parse.remove(term)
                            var_list.append(' '.join(var_parse))
                            load_list.append(term)
                            loadflag = True
                        if term in ['Fx,','Fy,','Fz,','Fxy,','Fyz,','Fxz,','Mx,','My,','Mz,','Mxy,','Myz,','Mxz,']:
                            var_parse[var_parse.index(term)] = ','
                            var_list.append(' '.join(var_parse))
                            load_list.append(term.split(',')[0])
                            loadflag = True
                    if not loadflag:  # variable is not force/moment, e.g. acc...
                        var_list.append(item[0])
                        load_list.append(item[0])

                    # translate loads to float numbers, keeping IS units
                    cols = len(item)
                    for j in range(3, cols-1):
                        item[j] = float(item[j])
                    # translate safe factor to 1.xx format
                    item[cols-1] = round(float(item[cols-1]), 2)

                    item = ['Max' if ix == 'MAXIMUM ' else ix for ix in item]
                    item = ['Min' if ix == 'MINIMUM ' else ix for ix in item]

                    # print(item)
                    data.append(item)
        # print(load_list[0:-1:2])
        # print(var_list)

        # check whether more than 1 variable exist in the .$MX file (e.g. Hr_Br_ultimate)
        onevar_flag = True
        for i in range(1, len(var_list)):
            if var_list[i] != var_list[i-1]:
                onevar_flag = False
                break

        # if only 1 variable exists, reduce variable name to force/moment only(Fx, Fy, Mx...)
        variable = None
        if onevar_flag:
            variable = var_list[0]  # variable name
            for i in range(len(data)):
                data[i][0] = load_list[i]  # load component

        # check whether safety factor included
        sf_flag = '_Exclsf: '
        for i in range(len(var_list)):
            if data[i][-1] != 1.00:
                sf_flag = '_Inclsf: '
                break

        # print(data)
        return data, variable, sf_flag, load_list

    def write_result(self):

        xls = self.table
        # new_sheet = self.sheetname
        # if new_sheet in xls.sheetnames:
        #     xls.remove(xls[new_sheet])
        # sheet = xls.create_sheet(new_sheet)
        if self.sheetname not in xls.sheetnames:
            sheet = xls.create_sheet(self.sheetname)
        else:
            sheet = xls.get_sheet_by_name(self.sheetname)
        if 'Sheet1' in xls.sheetnames:
            xls.remove(xls['Sheet1'])

        rowstart = self.rowstart
        colstart = self.colstart
        # print('Begin to write: ',self.ultpath)
        for f_idx, file in enumerate(self.MXfile_list):
            # print('Open file [%d of %d]: %s' % (f_idx + 1, len(self.MXfile_list), file))
            data, var, sf, load = self.read_dollarMX(file)
            nrow = len(data)
            ncol = len(data[0])

            # ---------unit translation--------
            for icol in range(int(len(load)/2)):  # len(load)=nrow-2
                if load[icol*2] in ['Fx','Fy','Fz','Fxy','Fyz','Fxz','Mx','My','Mz','Mxy','Myz','Mxz']:
                    for irow in range(nrow):
                       data[irow][icol+3] /= 1000

            # ---------Sheet <extreme>-------
            sheet.cell(rowstart, colstart, 'Ultimate'+sf+[var if var else ''][0])
            sheet.cell(rowstart+1, colstart+ncol-1, 'Safety factor')
            sheet.cell(rowstart+2, colstart+ncol-1, '-')
            sheet.cell(rowstart+2, colstart+2, 'Load case')

            # write variable name and unit in first two rows
            for i in range(ncol-4):
                sheet.cell(rowstart + 1, colstart + 3 + i, data[i*2][0])

                if load[i*2][0] == 'F':
                    sheet.cell(rowstart + 2, colstart + 3 + i, 'kN')
                elif load[i*2][0] == 'M':
                    sheet.cell(rowstart + 2, colstart + 3 + i, 'kNm')
                elif load[i*2].split(' ')[-1] == 'acceleration':
                    if load[i*2].split(' ')[-2] == ('roll' or 'yaw' or 'nod'):
                        sheet.cell(rowstart + 2, colstart + 3 + i, 'rad/s^2')
                    else:
                        sheet.cell(rowstart + 2, colstart + 3 + i, 'm/s^2')
                else:
                    sheet.cell(rowstart + 2, colstart + 3 + i, '?')

            # write ultimate data (including case name, min/max, variable)
            for i in range(nrow):
                for j in range(ncol):
                    sheet.cell(rowstart + 3 + i, colstart + j, data[i][j])

            # -------set font and border-------
            # sheet.column_dimensions[get_column_letter((colstart+ncol-1))].width = 11
            left, right, top, bottom = [Side(style='thin')] * 4
            for j in range(ncol):
                sheet.cell(rowstart + 1, colstart + j).fill = PatternFill('solid', fgColor='FFEFD5')
                sheet.cell(rowstart + 1, colstart + j).alignment = Alignment(wrap_text=True, vertical='center')
                for i in range(nrow + 2):
                    sheet.cell(rowstart + 1 + i, colstart + j).border = Border(left=left, right=right, top=top, bottom=bottom)
                    sheet.cell(rowstart + 1 + i, colstart + j).font = Font(name='Times New Roman')
            for i in range(nrow):
                for j in range(ncol - 4):
                    sheet.cell(rowstart + j * 2 + 3, colstart + j + 3).fill = PatternFill('solid', fgColor='FFEFD5')
                    sheet.cell(rowstart + j * 2 + 4, colstart + j + 3).fill = PatternFill('solid', fgColor='FFEFD5')

            rowstart += nrow + 3 + self.rowspace

        print('%s is done!' % self.ultpath)

if __name__ == '__main__':

    # ultpath = r'\\172.20.0.4\fs03\CJG\V2B\loop6\run_vr11.9\post\07_Ultimate\02_BRS_Exclsf'
    ultpath = r'\\172.20.0.4\fs03\CJG\V2B\loop6\run_vr11.9\post\07_Ultimate\05_HR_BR_Inclsf\hr'
    ultpath = r'\\172.20.0.4\fs03\CJG\V2B\loop6\run_vr11.9\post\07_Ultimate\09_Main_Exclsf'
    template = r"E:\01 tool dev\02 post tool\06_readBladedPost\output_template\20200803\ultimate_blank.xlsx"
    outpath = r'\\172.20.0.4\fs03\CJG\V2B\loop6\run_vr11.9\result_test'
    table = openpyxl.load_workbook(template)
    writeUltimate(ultpath, table, sheetname='extreme', rowstart=2, colstart=2, rowspace=1, heightflag=False)
    table.save(os.path.join(outpath, 'ult_test.xlsx'))


    # table = openpyxl.load_workbook('blade_load_info.xlsx')
    #
    # mx = ReadPostMX(r'\\172.20.0.4\fs02\LSY\V3\loop04_backup20200623\BackUp\post\07_Ultimate\01_BRS_Inclsf\BladeR_0.0000000')
    # MXfile_list = mx.get_MXfilelist()
    # mx.write_result(MXfile_list, table, 'sheet1')
    #
    # table.save('blade_load_info.xlsx')

