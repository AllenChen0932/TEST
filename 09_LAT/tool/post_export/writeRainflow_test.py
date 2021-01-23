# -*- coding: utf-8 -*-
# @Time    : 2020/07/17 15:49
# @Author  : CJG
# @File    : writeRainflow_v1_1.py
'''
v1.1: write_rainflow -> write2singletxt(), write2multitxt(), write2excel()
'''

import os
import numpy as np
try:
    from tool.post_export.readRainflow_v1_0 import readRainflow
    import tool.post_export.reorderVariable as rv
except:
    from readRainflow_v1_0 import readRainflow
    import reorderVariable as rv


class writeRainflow(object):

    def __init__(self,
                 rfpath,
                 variable=('Mx','My','Mz','Fx','Fy','Fz'),
                 content =('DEL','RFC','Markov'),
                 height_flag=False):

        self.rfpath   = rfpath
        self.content  = content       # content tuple to output ('DEL','RFC','Markov')
        self.variable = variable      # variable tuple to output ('Mx','My','Fz')
        self.hei_flag = height_flag
        self.rainflow = []            # rainflow data extracted from rfpath
        self.dirlist  = []            # sub directions (blade/tower stations)

        self.read_rainflow()
        # self.write_rainflow()

    def read_rainflow(self):
        # for now, ALL variables in the given rfpath are extracted,
        # though for example only 'MxyBR' are required to ouput to excel.
        # AND, contents ('DEL','RFC','Markov') are optional.

        rfpath = self.rfpath
        dirs   = os.listdir(rfpath)

        if os.path.isdir(os.path.join(rfpath, dirs[0])):
            # if sub folders exist in rfpath (e.g. blade stations, tower stations),
            # read all subfolder results into list 'rainflow'
            height_flag = None
            height_list = []
            height_dir  = {}

            for file in dirs:
                if '_' not in file:
                    height_flag = False
                    break
                else:
                    height = float(file.split('_')[-1])
                    height_list.append(height)
                    height_dir[height] = file
                    height_flag = True

            if height_flag:
                height_list.sort()

                for h in height_list:
                    dir  = height_dir[h]
                    path = os.path.join(rfpath, dir)
                    # print(path)

                    if os.path.isdir(path):
                        self.dirlist.append(dir)
                        res = readRainflow(path, self.content)

                        self.rainflow.append(res)
            else:
                for dir in dirs:
                    path = os.path.join(rfpath, dir)
                    # print(path)
                    if os.path.isdir(path):
                        # print(path)
                        self.dirlist.append(dir)
                        res = readRainflow(path, self.content)

                        self.rainflow.append(res)

        else:
            # rfpath has no sub folders (blade root, hub, ...)
            res  = readRainflow(self.rfpath, self.content)

            self.rainflow.append(res)

    def write2singletxt(self, txtfile):
        # ---------------------------------------- #
        # write DEL, Markov to TXTs (blade, tower)
        # additional input: txt file name with full path.
        # DEL/Markov arranged firstly by subfolder, then by variable.
        # ---------------------------------------- #
        rfpath = self.rfpath
        dirs   = os.listdir(rfpath)

        # if os.path.isdir(os.path.join(rfpath, dirs[0])) and '.$' not in dirs:

        if 'DEL' in self.content:

            with open(txtfile,'w') as f:
                f.write('')  # clear existing content

            with open(txtfile,'ab') as f:

                for i in range(len(self.rainflow)):

                    DEL = self.rainflow[i].DEL
                    varout = rv.reorderVariable(DEL['varlist'], self.variable)
                    if i == 0:
                        header = self.dirlist[i] + '\r\n'
                    else:
                        header = '\r\n' + self.dirlist[i] + '\r\n'
                    header += 'SN\t' + '\t'.join(varout) + '\r\n'
                    header += '-' + '\tkNm'*int(len(varout)/2) + '\tkN'*int(len(varout)/2) + '\r'
                    data = np.array([round(float(mvalue),1) for mvalue in DEL['m']]).reshape(-1,1)
                    for var in varout:
                        data = np.hstack((data, DEL['var_del'][var].reshape(-1,1)/1000))
                    np.savetxt(f, data, header=header, fmt='%.f'+'\t%.2f'*len(varout)+'\r', delimiter='\t', comments='')

        if 'Markov' in self.content:

            with open(txtfile,'w') as f:
                f.write('')  # clear existing content
            with open(txtfile,'ab') as f:
                for i in range(len(self.rainflow)):
                    Markov = self.rainflow[i].markov

                    # output variable are arranged according to self.variable.
                    if self.variable:
                        varout = rv.reorderVariable(Markov['varlist'], self.variable)
                    else:
                        varout = Markov['varlist']
                    for var in varout:
                        mean_col = Markov['var_mean'][var].reshape(-1, 1)/1000
                        range_row = np.array(Markov['var_range'][var] / 1000)
                        header  = 'Number of cycles [' + var + '] [.] against Cycle range [kNm]\r\n'
                        header += 'Cycle mean [kNm]\t' + '\t'.join([str(m) for m in range_row]) + '\r'
                        data = np.hstack((mean_col, Markov['var_cycles'][var]))
                        np.savetxt(f, data, header=header, fmt='%.2f' + '\t%.2f'*len(range_row) + '\r', delimiter='\t', comments='')

        if 'RFC' in self.content:
            raise Exception('For now RFC output at blade/tower stations are not supported!')

    def write2multitxt(self, outpath):
        # ---------------------------------------- #
        # write multi variables' Markov to TXT
        # additional input: out path
        # Markov out txt files named by variable name.
        # ---------------------------------------- #
        rfpath = self.rfpath
        dirs = os.listdir(rfpath)
        if not os.path.exists(outpath):
            os.mkdir(outpath)

        if not os.path.isdir(os.path.join(rfpath, dirs[0])) or ('.$' or '.in') in dirs:

            if len(self.rainflow) != 1:
                print('More than one rainflow result exist in %s, output ONLY the first one!' % self.rfpath)

            if 'Markov' in self.content:
                Markov = self.rainflow[0].markov
                varout = rv.reorderVariable(Markov['varlist'], self.variable)

                for var in varout:
                    txtfile = os.path.join(outpath, var+'_Markov.txt')
                    with open(txtfile,'wb') as f:
                        mean_col = Markov['var_mean'][var].reshape(-1, 1)/1000
                        range_row = np.array(Markov['var_range'][var] / 1000)
                        header = rfpath+'\r\n'
                        header += 'Number of cycles [' + var + '] [.] against Cycle range [kNm]\r\n'
                        header += 'Cycle mean [kNm]\t' + '\t'.join([str(m) for m in range_row]) + '\r'
                        data = np.hstack((mean_col, Markov['var_cycles'][var]))
                        np.savetxt(f, data, header=header, fmt='%.2f' + '\t%.2f'*len(range_row) + '\r', delimiter='\t', comments='')

            if ('DEL' or 'RFC') in self.content:
                pass

    def write2excel(self, table, del_rowstart=6, del_colstart=2):
        # ---------------------------------------- #
        # write DEL, Markov, RFC to EXCEL
        # additional input: opened table;
        # Markov/DEL/RFC are output to sheets of this table with specified sheetnames.
        # ---------------------------------------- #
        rfpath = self.rfpath
        dirs = os.listdir(rfpath)

        if not os.path.isdir(os.path.join(rfpath, dirs[0])) or ('.$' or '.in') in dirs:

            if len(self.rainflow) != 1:
                raise Exception('More than one rainflow result exist in %s, output ONLY the first one!' % self.rfpath)
            print('Begin to write: ',self.rfpath)
            rfres = self.rainflow[0]

            if 'DEL' in self.content:

                DEL = rfres.DEL
                sheetname = 'equivalent_fatigue'
                try:
                    if sheetname not in table.sheetnames:
                        sheet = table.create_sheet(sheetname)
                    else:
                        sheet = table.get_sheet_by_name(sheetname)
                    # row_start = 6
                    # col_start = 2
                    row_start = del_rowstart
                    col_start = del_colstart
                    data_len = len(DEL['m'])

                    for i in range(data_len):
                        mvalue = DEL['m'][i]
                        sheet.cell(row_start+i, col_start, round(float(mvalue),1))

                    # output variable are arranged according to the excel template,
                    # though in fact the variables has been defined by self.variable.
                    varout = rv.reorderVariable(DEL['varlist'], self.variable)
                    for i in range(len(varout)):
                        try:
                            var_intemple = sheet.cell(row=row_start-2, column=col_start+1+i).value
                        except AttributeError:
                            print('DEL: variable "%s" is not defined in the template!' % varout[i])
                            break
                        if var_intemple in varout:
                            var = var_intemple
                            for j in range(data_len):
                                # print(DEL['var_del'][var])
                                sheet.cell(row_start+j, col_start+1+i, DEL['var_del'][var][j,0]/1000)  # DEL
                        else:
                            print('DEL: variables in template does NOT match variables in given rainflow result!')
                except IOError:
                    print('no sheet named %s found in template table!' % sheetname)

            if 'RFC' in self.content:

                RFC = rfres.RFC
                sheetname = 'RFC'
                try:
                    if sheetname not in table.sheetnames:
                        sheet = table.create_sheet(sheetname)
                    else:
                        sheet = table.get_sheet_by_name(sheetname)

                    # row_start = 5
                    # col_start = 2
                    # data_len = RFC['var_rfc'][RFC['varlist'][0]].shape[0]
                    data_len = RFC['datalen']

                    # output variable are arranged according to the excel template,
                    # though in fact the variables has been defined by self.variable.
                    varout = rv.reorderVariable(RFC['varlist'], self.variable)
                    print(varout)
                    for i,var in enumerate(varout):
                        sheet.cell(del_rowstart+i*(data_len+2),   del_colstart, var)
                        sheet.cell(del_rowstart+i*(data_len+2)+1, del_colstart, 'Range')
                        sheet.cell(del_rowstart+i*(data_len+2)+1, del_colstart+1, 'Cycle')
                        for j in range(data_len):
                            sheet.cell(del_rowstart+i*(data_len+2)+2+j, del_colstart,   RFC['var_rfc'][var][j,0]/1000)  # load range
                            sheet.cell(del_rowstart+i*(data_len+2)+2+j, del_colstart+1, RFC['var_rfc'][var][j,1])       # cycles at per range

                except IOError:
                    print('no sheet named %s found in template table!' % sheetname)

            if 'Markov' in self.content:

                Markov = rfres.markov
                if len(self.variable) == 1:
                    sheetname = 'Markov_'+self.variable[0]
                else:
                    sheetname = 'Markov'
                sheetname = 'Markov'
                try:
                    if sheetname not in table.sheetnames:
                        sheet = table.create_sheet(sheetname)
                    else:
                        sheet = table.get_sheet_by_name(sheetname)
                    row_start = 5
                    col_start = 3
                    data_len = Markov['datalen']  # [len_mean, len_range]

                    # output variable are arranged according to the excel template,
                    # though in fact the variables has been defined by self.variable.
                    if self.variable:  #
                        varout = rv.reorderVariable(Markov['varlist'], self.variable)
                        for i in range(len(varout)):
                            row_starti = row_start + i*(data_len[0]+3)
                            try:
                                string = sheet.cell(row=row_starti-2, column=col_start-1).value
                                var_intemple = string.strip().split('[')[1].split(']')[0]
                            except AttributeError:
                                print('Markov: variable "%s" is not defined in the template!' % varout[i])
                                break
                            if var_intemple in varout:
                                var = var_intemple
                                # cycles:
                                for j in range(data_len[0]):
                                    for k in range(data_len[1]):
                                        sheet.cell(row_starti+j, col_start+k, Markov['var_cycles'][var][j,k])
                                # mean:
                                for j in range(data_len[0]):
                                    sheet.cell(row_starti+j, col_start-1, Markov['var_mean'][var][j] / 1000)
                                # range:
                                for k in range(data_len[1]):
                                    sheet.cell(row_starti-1, col_start+k, Markov['var_range'][var][k] / 1000)
                            else:
                                print('Markov: variables in template does NOT match variables in given rainflow result!')

                    else:  # ignore the template contents, write out all vars Markov data, e.g. combined BRMxy
                        # print('variable is none')
                        varout = Markov['varlist']
                        for i in range(len(varout)):
                            row_starti = row_start + i*(data_len[0]+3)
                            var = varout[i]
                            sheet.cell(row_starti-2, col_start-1, 'Number of cycles ['+var+'] [.] against Cycle range [kNm]')
                            sheet.cell(row_starti-1, col_start-1, 'Cycle mean [kNm]')
                            # cycles:
                            for j in range(data_len[0]):
                                for k in range(data_len[1]):
                                    sheet.cell(row_starti+j, col_start+k, Markov['var_cycles'][var][j,k])
                            # mean:
                            for j in range(data_len[0]):
                                sheet.cell(row_starti+j, col_start-1, Markov['var_mean'][var][j] / 1000)
                            # range:
                            for k in range(data_len[1]):
                                sheet.cell(row_starti-1, col_start+k, Markov['var_range'][var][k] / 1000)
                except IOError:
                    print('no sheet named %s found in template table!' % sheetname)

if __name__ == '__main__':

    import openpyxl
    # # rfpath = r'\\172.20.0.4\fs02\LSY\V3\loop04_backup20200623\BackUp\post\rainflow\br_25y_1E7'
    # rfpath = r'\\172.20.0.4\fs02\LSY\V3\loop04_backup20200623\BackUp\post\rainflow\main_25y_1E7\Blade_root'
    # excelouttemple = r"E:\01 tool dev\02 post tool\06_readBladedPost\out_test\test_RF_pb - 副本 - 副本.xlsx"
    excel = r"\\172.20.4.132\fs02\CE\V3\loop06\post_test\RFC\RFC.xlsx"
    table = openpyxl.load_workbook(excel)
    #
    # # writeRainflow(rfpath, table, variable=('Mx','My','Mz'), content=('DEL','RFC','Markov'))
    # writeRainflow(rfpath, table, content=('DEL','RFC','Markov'))
    #


    # rfpath = r'E:\01 tool dev\02 post tool\06_readBladedPost\out_test\V3-LSY\post\br_25y_1E7'
    # txtout = r"E:\01 tool dev\02 post tool\06_readBladedPost\out_test\br_markov.txt"
    # writeRainflow(rfpath, txt=txtout, content=('Markov',))
    # txtout = r"E:\01 tool dev\02 post tool\06_readBladedPost\out_test\br_del.txt"
    # writeRainflow(rfpath, txt=txtout, content=('DEL'))

    rfpath  = r'\\172.20.4.132\fs02\CE\V3\loop06\post_new\08_Fatigue\05_Main\hr'
    # excel   = r'\\172.20.4.132\fs02\CE\V3\loop06\post_new'
    # txt     = r"\\172.20.4.132\fs02\CE\V3\loop06\post_test\RFC_loop6_KpKi.txt"
    writeRainflow(rfpath, content=('RFC',)).write2excel(table, 1, 5)
    # writeRainflow(rfpath, content=('RFC',)).write2singletxt(txt)
    table.save(excel)

