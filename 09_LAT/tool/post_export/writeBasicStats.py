# -*- coding: utf-8 -*-
# @Time    : 2020/07/22 15:49
# @Author  : CJG
# @File    : writeBasicStatistics.py


import os
import openpyxl
try:
    from tool.post_export.Read_Bladed_v3 import read_bladed as ReadBladed
except:
    from Read_Bladed_v3 import read_bladed as ReadBladed

class writeBasicStats:

    def __init__(self,
                 bstatspath,
                 sheet,
                 varout=('Hub wind speed magnitude','Rotor speed','Stationary hub Mx'),
                 content=('mean',)):

        self.bstatspath = bstatspath
        self.sheetout   = sheet
        self.varout     = varout    # variables to output in excel
        self.content    = content  # statistic data for each variable to output
        self.bstats     = {'perfilelist': [], 'varlist': [], 'per_var': {}, 'dlclist': [], 'var_bstats': {}}

        self.pi = 3.141592653589793

        self.writeBasicStatistics()

    def readBasicStatistics(self):
        '''
        read basic statistics data according to user-requested variable and content
        :return: self.bstats
        '''

        path = self.bstatspath

        for file in os.listdir(path):

            if os.path.splitext(file)[1][0:2] == '.%':
                [pername, perext] = os.path.splitext(file)
                perext = perext[2:]

                res      = ReadBladed(pername, path, perext)
                var      = res.channel.split('[')[1].split(']')[0]  # one var in each percent file
                Contlist = [varib.strip().split('[')[0].strip().upper() for varib in res.var_list]
                # print(Contlist)
                if var in self.varout:

                    self.bstats['perfilelist'].append(file)
                    self.bstats['varlist'].append(var)
                    self.bstats['per_var'][perext] = var
                    self.bstats['dlclist'] = res.axis[0]['tick']

                    # statistic data for each variable with contents specified by self.content (e.g. mean, max, etc.)
                    ind = []
                    for cont in self.content:
                        if cont.upper() in Contlist:
                            ind.append(Contlist.index(cont.upper()))
                        else:
                            raise Exception('No bstats content "%s" found in %s' %(cont, file))

                    self.bstats['var_bstats'][var] = res.data[:, ind]
        print(self.bstats['varlist'])
        for tmp in self.bstats['varlist']:
            if tmp not in self.varout:
                raise Exception('Variable %s not found in %s!' %(tmp, path))

    def writeBasicStatistics(self):

        self.readBasicStatistics()

        sheet = self.sheetout
        row_start = 5
        col_start = 3

        # write dlc
        dlclist = self.bstats['dlclist']
        for j in range(len(dlclist)):
            sheet.cell(row_start+j, col_start, dlclist[j])

        # write mean value for chosen var (wind speed, rotor speed, MxHR)
        for i in range(1, 4):

            var_intemple = self.sheetout.cell(row=row_start-2, column=col_start+i).value
            # print(var_intemple)
            if var_intemple in self.bstats['varlist']:
                var = var_intemple

                bstatdata = self.bstats['var_bstats'][var][:,0]  # here 'mean' is assumed in first column
                if var == 'Rotor speed':
                    bstatdata = bstatdata/2/self.pi*60
                elif var == 'Stationary hub Mx':
                    bstatdata = bstatdata/1000

                for j in range(len(dlclist)):
                    sheet.cell(row_start+j, col_start+i, bstatdata[j])
            else:
                raise Exception('Variables "%s" in template does NOT match variables in given bstats result!' %var_intemple)

if __name__ == '__main__':

    # bstatspath = r'\\172.20.0.4\fs02\LSY\V3\loop04_backup20200623\BackUp\post\ldd\Yaw'
    # excelouttemple = r"E:\01 tool dev\02 post tool\06_readBladedPost\test_LDD_yb - 副本.xlsx"
    # excelout = r"E:\01 tool dev\02 post tool\06_readBladedPost\test_LDD_yb.xlsx"
    bstats_path  = r'\\172.20.0.4\fs02\CE\V3\post_test\02_Bstats\mb'
    excel_temple = r"C:\Users\10700700\Desktop\tool\py\20_Load Assistant Tool\template\main_bearing_load_information.xlsx"
    excel_path   = r"C:\Users\10700700\Desktop\tool\py\20_Load Assistant Tool\test\post export\test_Mean_hr.xlsx"

    table = openpyxl.load_workbook(excel_temple)
    sheet = table.get_sheet_by_name('mean')

    writeBasicStats(bstats_path, sheet)

    table.save(excel_path)
