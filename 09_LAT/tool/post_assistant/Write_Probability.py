# -*- coding: utf-8 -*-
# @Time    : 2020/5/3 9:10
# @Author  : CE
# @File    : Write_Probability.py

import os
from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter, column_index_from_string

class Probability(object):

    def __init__(self,
                 run_path,
                 fat_path,
                 lct_path,
                 dlc_list,
                 num_bins,
                 des_life,
                 eq_cycle,
                 fat_chan,
                 type_flag,
                 lrd_angle=None):

        self.run_path = run_path.replace('/', '\\')
        self.fat_path = fat_path.replace('/', '\\')
        self.lct_path = lct_path.replace('/', '\\')
        self.num_bins = str(num_bins)
        self.dlc_list = dlc_list
        self.cha_dict = fat_chan         #{channel: variable list}, e.g. yb:[Yaw bearing Fx, Yaw bearing Fy]
        self.des_life = des_life
        self.eq_cycle = eq_cycle
        self.type_flag  = str(type_flag)   # LDD:1; LRD:2
        self.lrd_var  = lrd_angle

        self.dlc_num = 0
        self.lc_list = []    # load case list
        self.lc_path = {}    # lc name: load case path
        self.lc_lct  = {}    # load case: [hour, type]

        # self.error_fat   = False

        self.PAI = 3.14153186157

        # self.get_dlc()
        # self.get_subgroup()
        try:
            if self.type_flag == '1':
                self.get_loadcase()
                self.read_lct()
                self.write_ldd_pj()
                self.write_ldd_in()
                # self.finish_fat = True
            elif self.type_flag == '2':
                self.get_loadcase()
                self.read_lct()
                self.write_lrd_pj()
                self.write_lrd_in()
                # self.finish_fat = True
        except:
            self.finish_fat = False
        # except IOError:
        #     self.error_flag = True

    def get_loadcase(self):
        '''
        get subgroup sorted by mean and half
        :return:
        '''
        self.dlc_list.sort()
        # print(self.dlc_list)

        for dlc in self.dlc_list:

            dlc_path = os.path.join(self.run_path, dlc)
            lc_list  = os.listdir(dlc_path)

            for lc in lc_list:

                lc_path = os.path.join(dlc_path, lc)

                if os.path.isdir(lc_path):

                    self.dlc_num += 1
                    self.lc_path[lc] = lc_path
                    self.lc_list.append(lc)

        # print(self.dlc_num)

    def write_ldd_pj(self):

        for key, value in self.cha_dict.items():

            # header
            content = '<?xml version="1.0" encoding="ISO-8859-1" ?>\n' \
                      '<BladedProject ApplicationVersion="4.8.0.41">\n' \
                      '	<DataModelVersion>0.8</DataModelVersion>\n' \
                      '	<BladedData dataFormat="project">\n' \
                      '<![CDATA[\n' \
                      'VERSION	4.8.0.34\n' \
                      'MULTIBODY	 1\n' \
                      'CALCULATION	22\n' \
                      'OPTIONS	0\n' \
                      'PROJNAME	\n' \
                      'DATE	\n' \
                      'ENGINEER	\n' \
                      'NOTES	""\n' \
                      'PASSWORD	\n' \
                      'MSTART PROBD\n' \
                      'TYPE	2\n' \
                      'XFLAG	0\n' \
                      "ATTRIBAZ	''\n" \
                      'VARIAB	""\n' \
                      'NDIMENS	0\n' \
                      'AZNAME	"Time"\n' \
                      'MIN	 0\n' \
                      'MAX	 0\n' \
                      'NBINS	0\n' \
                      'RMMEAN	N\n' \
                      "ATTRIBF	''\n" \
                      'VARIAB	""\n' \
                      'NDIMENS	0\n' \
                      'NBINDIM	1\n' \
                      'MEND\n\n'

            # variable
            content += self.write_pj_variable(key, value[0])

            # load case
            content += 'MSTART MULTIRUN\n' \
                       'NCASES	%s\n' %str(self.dlc_num)

            for lc in self.lc_list:

                occurrence = None
                if self.lc_lct[lc][1] == 'H':
                    occurrence = float(self.lc_lct[lc][0])*3600
                elif self.lc_lct[lc][1] == 'T':
                    occurrence = float(self.lc_lct[lc][0])/(8766*3600)

                content += 'DIRECTORY	%s\n' %self.lc_path[lc]
                content += 'OLDVERSION	0\n'
                content += 'RUNNAME	%s\n' %lc
                content += 'CONFIG	%s\n' %self.lc_lct[lc][1]
                content += 'OCCFREQ	 %.14f\n' %occurrence

            content += 'LIFE	 %d\n' %(int(self.des_life)*8766*3600)
            content += 'BINTYPE	 1\n'
            content += 'MEND\n\n'

            content += '0PROBD\n'
            content += '0MULTIRUN\n'
            content += '0MULTIVAR\n'

            content += '		]]>\n' \
                       '	</BladedData>\n' \
                       '</BladedProject>'

            pj_name = key + '.$PJ'
            # pj_path = os.path.join(self.fat_path, key).replace(' ', '_')
            pj_path = os.path.join(self.fat_path, '_'.join([key, self.num_bins])).replace(' ', '_')
            if not os.path.isdir(pj_path):
                os.makedirs(pj_path)

            with open(os.sep.join([pj_path, pj_name]), 'w+') as f:
                f.write(content)
            print('%s pj is done!' %key)

    def write_ldd_in(self):

        for key, value in self.cha_dict.items():

            # variable
            # pj_path = os.path.join(self.fat_path, key).replace(' ', '_')
            pj_path = os.path.join(self.fat_path, '_'.join([key, self.num_bins])).replace(' ', '_')

            # header
            content  = 'PTYPE	7\n' \
                       'SDSTAT	1\n' \
                       'DONGLOG\t0\n' \
                       'OUTSTYLE	B\n'
            content += 'PATH	%s\n' %pj_path
            content += 'RUNNAME	%s\n' %key
            content += 'MSTART PROBD\n' \
                       'TYPE	2\n' \
                       'MEND\n\n'

            # variable
            content += self.write_in_variable(key, value[0])

            # load case
            content += 'MSTART MULTIRUN\n' \
                       'NCASES	%s\n' %str(self.dlc_num)
            content += 'OUTCHOICE	1\n'

            for lc in self.lc_list:

                occurrence = None
                if self.lc_lct[lc][1] == 'H':
                    occurrence = float(self.lc_lct[lc][0])*3600
                elif self.lc_lct[lc][1] == 'T':
                    occurrence = float(self.lc_lct[lc][0])/(8766*3600)

                content += 'DIRECTORY	%s\n' %(self.lc_path[lc]+'\\')
                content += 'RUNNAME	%s\n' %lc
                content += 'CONFIG	%s\n' %self.lc_lct[lc][1]
                content += 'OCCFREQ	 %.14f\n' %occurrence

            content += 'LIFE	 %d\n' %(int(self.des_life)*8766*3600)
            content += 'MEND\n\n'

            content += 'MSTART WINDREGIME\n' \
                       'WTYPE	1\n' \
                       'AMWS	999\n' \
                       'WEIBLK	1\n' \
                       'WEIBLC	1\n' \
                       'MEND'

            if not os.path.isdir(pj_path):
                os.makedirs(pj_path)

            with open(os.sep.join([pj_path, 'dtsignal.in']), 'w+') as f:
                f.write(content)

            print('%s in is done!' %key)

    def write_lrd_pj(self):

        for key, value in self.cha_dict.items():

            # header
            content = '<?xml version="1.0" encoding="ISO-8859-1" ?>\n' \
                      '<BladedProject ApplicationVersion="4.8.0.41">\n' \
                      '	<DataModelVersion>0.8</DataModelVersion>\n' \
                      '	<BladedData dataFormat="project">\n' \
                      '<![CDATA[\n' \
                      'VERSION	4.8.0.34\n' \
                      'MULTIBODY	 1\n' \
                      'CALCULATION	22\n' \
                      'OPTIONS	0\n' \
                      'PROJNAME	\n' \
                      'DATE	\n' \
                      'ENGINEER	\n' \
                      'NOTES	""\n' \
                      'PASSWORD	\n' \
                      'MSTART PROBD\n' \
                      'TYPE	2\n' \
                      'XFLAG	-1\n' \
                      "ATTRIBAZ	%%%s\n" \
                      'VARIAB	"%s"\n' \
                      'NDIMENS	0\n' \
                      'AZNAME	"Revs"\n' \
                      'MIN	 0\n' \
                      'MAX	 0\n' \
                      'NBINS	0\n' \
                      'RMMEAN	N\n' \
                      "ATTRIBF	''\n" \
                      'VARIAB	""\n' \
                      'NDIMENS	0\n' \
                      'NBINDIM	1\n' \
                      'MEND\n\n' %(self.lrd_var[key][0],self.lrd_var[key][1])

            # variable
            content += self.write_pj_variable(key, value[0])

            # load case
            content += 'MSTART MULTIRUN\n' \
                       'NCASES	%s\n' %str(self.dlc_num)

            for lc in self.lc_list:

                occurrence = None
                if self.lc_lct[lc][1] == 'H':
                    occurrence = float(self.lc_lct[lc][0])*3600
                elif self.lc_lct[lc][1] == 'T':
                    occurrence = float(self.lc_lct[lc][0])/(8766*3600)

                content += 'DIRECTORY	%s\n' %self.lc_path[lc]
                content += 'OLDVERSION	0\n'
                content += 'RUNNAME	%s\n' %lc
                content += 'CONFIG	%s\n' %self.lc_lct[lc][1]
                content += 'OCCFREQ	 %.14f\n' %occurrence

            content += 'LIFE	 %d\n' %(int(self.des_life)*8766*3600)
            content += 'BINTYPE	 1\n'
            content += 'MEND\n\n'

            content += '0PROBD\n'
            content += '0MULTIRUN\n'
            content += '0MULTIVAR\n'

            content += '		]]>\n' \
                       '	</BladedData>\n' \
                       '</BladedProject>'

            pj_name = ''.join([key,'.$PJ'])
            pj_path = os.path.join(self.fat_path, '_'.join([key,self.num_bins])).replace(' ', '_')
            if not os.path.isdir(pj_path):
                os.makedirs(pj_path)

            with open(os.sep.join([pj_path, pj_name]), 'w+') as f:
                f.write(content)
            print('%s pj is done!' % pj_name)

    def write_lrd_in(self):

        for key, value in self.cha_dict.items():
            # print(key, value)

            # variable
            pj_path = os.path.join(self.fat_path, '_'.join([key,self.num_bins])).replace(' ', '_')

            # header
            content  = 'PTYPE	7\n' \
                       'SDSTAT	1\n' \
                       'DONGLOG\t0\n' \
                       'OUTSTYLE	B\n'
            content += 'PATH	%s\n' %pj_path
            content += 'RUNNAME	%s\n' %key
            content += 'MSTART PROBD\n' \
                       'TYPE	2\n' \
                       'ATTRIBAZ	%%%s\n' \
                       "VARIAB	'%s'\n" \
                       'AZNAME	Revs\n' \
                       'MEND\n\n' %(self.lrd_var[key][0],self.lrd_var[key][1])

            # variable
            content += self.write_in_variable(key, value[0])

            # load case
            content += 'MSTART MULTIRUN\n' \
                       'NCASES	%s\n' %str(self.dlc_num)
            content += 'OUTCHOICE	1\n'

            for lc in self.lc_list:

                occurrence = None
                if self.lc_lct[lc][1] == 'H':
                    occurrence = float(self.lc_lct[lc][0])*3600
                elif self.lc_lct[lc][1] == 'T':
                    occurrence = float(self.lc_lct[lc][0])/(8766*3600)

                content += 'DIRECTORY	%s\n' %(self.lc_path[lc]+'\\')
                content += 'RUNNAME	%s\n' %lc
                content += 'CONFIG	%s\n' %self.lc_lct[lc][1]
                content += 'OCCFREQ	 %.14f\n' %occurrence

            content += 'LIFE	 %d\n' %(int(self.des_life)*8766*3600)
            content += 'MEND\n\n'

            content += 'MSTART WINDREGIME\n' \
                       'WTYPE	1\n' \
                       'AMWS	999\n' \
                       'WEIBLK	1\n' \
                       'WEIBLC	1\n' \
                       'MEND'

            if not os.path.isdir(pj_path):
                os.makedirs(pj_path)

            with open(os.sep.join([pj_path, 'dtsignal.in']), 'w+') as f:
                f.write(content)

            print('%s in is done!' %key)

    def write_in_variable(self, channel, var_list, height=None):

        content = ''

        if channel == 'br':
            # br1, br2, br3
            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	%s\n' %len(var_list)
            for var in var_list:
                
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%22\n'
                content += "VARIAB	'%s'\n" %var
                content += "FULL_NAME	'%s'\n" %var

        elif channel == 'br_mxy':
            # br1, br2, br3
            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	%s\n' %len(var_list)
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' % self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%22\n'
                content += "VARIAB	'%s'\n" % var
                content += "FULL_NAME	'%s'\n" % var

        elif channel == 'hs':

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	%s\n' %len(var_list)
            for var in var_list:

                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%23\n'
                content += "VARIAB	'%s'\n" %var
                content += "FULL_NAME	'%s'\n" %var

        elif channel == 'hr':

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	%s\n' %len(var_list)
            for var in var_list:

                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%22\n'
                content += "VARIAB	'%s'\n" %var
                content += "FULL_NAME	'%s'\n" %var

        elif channel == 'yb':

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	%s\n' %len(var_list)
            for var in var_list:

                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%24\n'
                content += "VARIAB	'%s'\n"  %var
                content += "FULL_NAME	'%s'\n" %var

        content += 'MEND\n\n'
        # print(content)
        return content

    def write_pj_variable(self, channel, var_list, height=None):
        # print(channel[:3])
        content = ''

        if channel == 'br':
            # br1, br2, br3
            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	%s\n' %len(var_list)
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%22\n'
                content += 'VARIAB	"%s"\n'  %var
                content += 'DESCRIPTION	"%s"\n' %var
                content += 'NDIMENS	2\n'
                content += 'DIMFLAG	-1\n'

        elif channel == 'br_mxy':
            # br1, br2, br3
            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	%s\n' %len(var_list)
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%22\n'
                content += 'VARIAB	"%s"\n'  %var
                content += 'DESCRIPTION	"%s"\n' %var
                content += 'NDIMENS	2\n'
                content += 'DIMFLAG	-1\n'

        elif channel == 'hs':

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	%s\n' %len(var_list)
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%23\n'
                content += 'VARIAB	"%s"\n'  %var
                content += 'DESCRIPTION	"%s"\n' %var
                content += 'NDIMENS	2\n'
                content += 'DIMFLAG	-1\n'

        elif channel == 'hr':

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	%s\n' %len(var_list)
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%22\n'
                content += 'VARIAB	"%s"\n'  %var
                content += 'DESCRIPTION	"%s"\n' %var
                content += 'NDIMENS	2\n'
                content += 'DIMFLAG	-1\n'

        elif channel == 'yb':

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	%s\n' %len(var_list)
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%24\n'
                content += 'VARIAB	"%s"\n'  %var
                content += 'DESCRIPTION	"%s"\n' %var
                content += 'NDIMENS	2\n'
                content += 'DIMFLAG	-1\n'

        content += 'MEND\n\n'
        # print(content)
        return content

    def read_lct(self):
        # print(self.lct_path)
        table = load_workbook(self.lct_path, keep_vba=True, data_only=True)

        ws_setting = table['Setting']
        IEC_stand  = '1' if '61400-1' in ws_setting['B1'].value else ('3' if '61400-3' in ws_setting['B1'].value else None)
        # print(IEC_stand)
        if IEC_stand == '1':

            for dlc in self.dlc_list:
                lc = dlc.split('_')[0] if '_' in dlc else dlc.split()[0]
                name_col, hour_col, type_col = None, None, None

                ws = table[lc]
                col_num = ws.max_column
                row_num = ws.max_row
                row_start = None

                # for onshore, the table header are line9 and line10
                for i in range(1, col_num+1):

                    for j in range(1, row_num):

                        cell1 = ws.cell(j,   i)
                        cell2 = ws.cell(j+1, i)

                        if cell1.value == 'Run Name' and cell2.value in [None, '']:
                            name_col  = cell1.column
                            row_start = j+2
                            continue
                        if (cell1.value == 'Hour/Year' and cell2.value == 'hours') or \
                            (cell1.value == 'Times/year' and cell2.value == '-'):
                            hour_col = cell1.column
                            continue
                        if cell1.value == 'Rainflow' and cell2.value == 'type':
                            type_col = cell1.column
                            break

                    if name_col and hour_col and type_col:
                        break
                # print(name_col, hour_col, type_col)

                for i in range(row_start, row_num+1):

                    self.lc_lct[ws.cell(i, name_col).value] = [ws.cell(i, hour_col).value, ws.cell(i, type_col).value]

        elif IEC_stand == '3':

            for dlc in self.dlc_list:

                lc = dlc.split('_')[0] if '_' in dlc else dlc.split()[0]

                name_col = None
                hour_col = None
                type_col = None

                ws = table[lc]
                col_num = ws.max_column
                row_num = ws.max_row
                row_start = None

                # for onshore, the table header are line13 and line14
                for i in range(1, col_num+1):

                    for j in range(1, row_num):

                        cell1 = ws.cell(j,  i)
                        cell2 = ws.cell(j+1, i)

                        if cell1.value == 'Run Name' and cell2.value in [None, '']:
                            # print(cell1.value, cell2.value, cell1.column)
                            name_col  = cell1.column
                            row_start = j+2
                            continue
                        if (cell1.value == 'Hour/Year' and cell2.value == 'hours') or \
                           (cell1.value == 'Times/year' and cell2.value == '-'):
                            # print(cell1.value, cell2.value, cell1.column)
                            hour_col = cell1.column
                            continue
                        if cell1.value == 'Rainflow' and cell2.value == 'type':
                            # print(cell1.value, cell2.value, cell1.column)
                            type_col = cell1.column
                            break

                    if name_col and hour_col and type_col:
                        break
                # print(name_col, hour_col, type_col)

                for i in range(row_start, row_num+1):

                    if ws.cell(i, name_col).value:
                        self.lc_lct[ws.cell(i, name_col).value] = [ws.cell(i, hour_col).value, ws.cell(i, type_col).value]
                        # print(ws.cell(i, name_col).value, [ws.cell(i, hour_col).value, ws.cell(i, type_col).value])

        # print(self.lc_lct)

if __name__ == '__main__':

    run_path = r'\\172.20.0.4\fs03\CE\W6250-172-14m\run_0430_0.8'


