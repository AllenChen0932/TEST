
# -*- coding: utf-8 -*-
# @Time    : 2020/5/3 9:10
# @Author  : CE
# @File    : TC.py

import os
from openpyxl import load_workbook
import pysnooper
# from openpyxl.utils import get_column_letter, column_index_from_string

class Fatigue(object):

    def __init__(self,
                 run_path,
                 fat_path,
                 lct_path,
                 dlc_list,
                 num_bins,
                 des_life,
                 eq_cycle,
                 fat_chan):

        self.run_path = run_path.replace('/', '\\')
        self.fat_path = fat_path.replace('/', '\\')
        self.lct_path = lct_path.replace('/', '\\')
        self.num_bins = str(num_bins)
        self.dlc_list = dlc_list
        self.cha_dict = fat_chan    #{channel: variable list}, e.g. yb:[Yaw bearing Fx, Yaw bearing Fy]
        self.des_life = des_life
        self.eq_cycle = eq_cycle

        if not os.path.exists(self.fat_path):
            os.makedirs(self.fat_path)
            # print(self.fat_path)

        self.dlc_num = 0
        self.lc_list = []    # load case list
        self.lc_path = {}    # lc name: load case path
        self.lc_lct  = {}    # load case: [hour, type]

        # self.error_fat   = False

        self.PAI = 3.14153186157

        # self.get_dlc()
        # self.get_subgroup()

        self.get_loadcase()
        self.read_lct()
        self.write_pj()
        self.write_in()
        # self.finish_fat = True

        # except IOError:
        #     self.error_flag = True

    def get_loadcase(self):
        '''
        get subgroup sorted by mean and half
        :return:
        '''
        self.dlc_list.sort()

        for dlc in self.dlc_list:
            dlc_path = os.path.join(self.run_path, dlc)
            lc_list  = os.listdir(dlc_path)

            for lc in lc_list:
                lc_path = os.path.join(dlc_path, lc)
                if os.path.isdir(lc_path):

                    self.dlc_num += 1
                    self.lc_path[lc] = lc_path
                    self.lc_list.append(lc)

    def write_pj(self):

        for key, value in self.cha_dict.items():
            # variable
            if len(value) == 2 and len(value[0]) == 3:
                # write bpa, bra, baa, bua
                # bpa1, bpa2, bpa3
                # print(key, value)
                for i in range(1, 4):
                    # variable
                    for sec in value[-1]:
                        # header
                        content = '<?xml version="1.0" encoding="ISO-8859-1" ?>\n' \
                                  '<BladedProject ApplicationVersion="4.8.0.41">\n' \
                                  '	<DataModelVersion>0.8</DataModelVersion>\n' \
                                  '	<BladedData dataFormat="project">\n' \
                                  '<![CDATA[\n' \
                                  'VERSION	4.8.0.34\n' \
                                  'MULTIBODY	 1\n' \
                                  'CALCULATION	25\n' \
                                  'OPTIONS	0\n' \
                                  'PROJNAME	\n' \
                                  'DATE	\n' \
                                  'ENGINEER	\n' \
                                  'NOTES	""\n' \
                                  'PASSWORD	\n' \
                                  'MSTART RAINF\n' \
                                  'EQLOAD	-1\n' \
                                  'NSLOP	10\n' \
                                  'SLOPES	  .3333333  .25  .2  .1666667  .1428571  .125  .1111111  .1  9.090909E-02  8.333334E-02\n'

                        content += 'ELFREQ	 %.8f\n' %float(2*self.PAI*float(self.eq_cycle)/(float(self.des_life)*8766*3600))
                        content += 'LIFE	 0\n' \
                                   'TYPE	2\n' \
                                   'MIN	 0\n' \
                                   'MAX	 0\n' \
                                   'NBINS	0\n' \
                                   'MIN_RANGE	 0\n'
                        content += "ATTRIBF	%s\n" %(r"'c:\program files (x86)\dnv gl\bladed 4.8\'")
                        content += 'VARIAB	""\n' \
                                   'NDIMENS	0\n' \
                                   'MEND\n\n'

                        # variable
                        # str(i) was used to identify the different %for different blade
                        content += self.write_pj_variable(key+str(i), value[0][i-1], sec)

                        # load case
                        content += 'MSTART MULTIRUN\n' \
                                   'NCASES	%s\n' %str(self.dlc_num)

                        for lc in self.lc_list:

                            occurrence = None
                            if self.lc_lct[lc][1] == 'H':
                                occurrence = float(self.lc_lct[lc][0])*3600
                            elif self.lc_lct[lc][1] == 'T':
                                occurrence = float(self.lc_lct[lc][0])/(8766*3600)

                            content += 'DIRECTORY	%s\n' %self.lc_path[lc]
                            content += 'OLDVERSION	0\n'
                            content += 'RUNNAME	%s\n' %lc
                            content += 'CONFIG	%s\n' %self.lc_lct[lc][1]
                            content += 'OCCFREQ	 %.14f\n' %occurrence

                        content += 'LIFE	 %d\n' %(int(self.des_life)*8766*3600)
                        content += 'BINTYPE	 1\n'
                        content += 'MEND\n\n'

                        content += '0RAINF\n'
                        content += '0MULTIRUN\n'
                        content += '0MULTIVAR\n'

                        se_name = str('%.3f' %float(sec)) if '.' in sec else sec
                        pj_name = key+str(i)+'_'+se_name
                        # print(pj_name)
                        content += '		]]>\n' \
                                   '	</BladedData>\n' \
                                   '<RunConfiguration>\n' \
                                   '  <Name>%s</Name>\n' \
                                   '  <Calculation>0</Calculation>\n' \
                                   '</RunConfiguration>\n\n' \
                                   '</BladedProject>' %pj_name

                        pj_file = pj_name +'.$PJ'
                        pj_path = os.path.join(self.fat_path, pj_name).replace(' ', '_')
                        if not os.path.isdir(pj_path):
                            os.makedirs(pj_path)

                        with open(os.sep.join([pj_path, pj_file]), 'w+') as f:
                            f.write(content)
                        print('%s pj is done!' %pj_name)

            elif len(value) == 2 and len(value[0]) != 3:
                # write tr
                # print(value)
                for sec in value[-1]:
                    # header
                    content = '<?xml version="1.0" encoding="ISO-8859-1" ?>\n' \
                              '<BladedProject ApplicationVersion="4.8.0.41">\n' \
                              '	<DataModelVersion>0.8</DataModelVersion>\n' \
                              '	<BladedData dataFormat="project">\n' \
                              '<![CDATA[\n' \
                              'VERSION	4.8.0.34\n' \
                              'MULTIBODY	 1\n' \
                              'CALCULATION	25\n' \
                              'OPTIONS	0\n' \
                              'PROJNAME	\n' \
                              'DATE	\n' \
                              'ENGINEER	\n' \
                              'NOTES	""\n' \
                              'PASSWORD	\n' \
                              'MSTART RAINF\n' \
                              'EQLOAD	-1\n' \
                              'NSLOP	10\n' \
                              'SLOPES	  .3333333  .25  .2  .1666667  .1428571  .125  .1111111  .1  9.090909E-02  8.333334E-02\n'

                    content += 'ELFREQ	 %.8f\n' %float(2*self.PAI*float(self.eq_cycle)/(float(self.des_life)*8766*3600))
                    content += 'LIFE	 0\n' \
                               'TYPE	2\n' \
                               'MIN	 0\n' \
                               'MAX	 0\n' \
                               'NBINS	0\n' \
                               'MIN_RANGE	 0\n'
                    content += "ATTRIBF	%s\n" %(r"'c:\program files (x86)\dnv gl\bladed 4.8\'")
                    content += 'VARIAB	""\n' \
                               'NDIMENS	0\n' \
                               'MEND\n\n'

                    # variable
                    content += self.write_pj_variable(key, value[0], sec)

                    # load case
                    content += 'MSTART MULTIRUN\n' \
                               'NCASES	%s\n' %str(self.dlc_num)

                    for lc in self.lc_list:

                        occurrence = None
                        if self.lc_lct[lc][1] == 'H':
                            occurrence = float(self.lc_lct[lc][0])*3600
                        elif self.lc_lct[lc][1] == 'T':
                            occurrence = float(self.lc_lct[lc][0])/(8766*3600)

                        content += 'DIRECTORY	%s\n' %self.lc_path[lc]
                        content += 'OLDVERSION	0\n'
                        content += 'RUNNAME	%s\n' %lc
                        content += 'CONFIG	%s\n' %self.lc_lct[lc][1]
                        content += 'OCCFREQ	 %.14f\n' %occurrence

                    content += 'LIFE	 %d\n' %(int(self.des_life)*8766*3600)
                    content += 'BINTYPE	 1\n'
                    content += 'MEND\n\n'

                    content += '0RAINF\n'
                    content += '0MULTIRUN\n'
                    content += '0MULTIVAR\n'

                    pj_name = (key+'_'+str('%.3f' %float(sec))) if 'Mbr' not in sec else sec
                    content += '		]]>\n' \
                               '	</BladedData>\n' \
                               '<RunConfiguration>\n' \
                               '  <Name>%s</Name>\n' \
                               '  <Calculation>0</Calculation>\n' \
                               '</RunConfiguration>\n\n' \
                               '</BladedProject>' %(pj_name if 'Mbr' not in pj_name else ("'%s'" %pj_name))

                    pj_file = pj_name+'.$PJ'
                    pj_path = os.path.join(self.fat_path, pj_name).replace(' ', '_')
                    if not os.path.isdir(pj_path):
                        os.makedirs(pj_path)

                    with open(os.sep.join([pj_path, pj_file]), 'w+') as f:
                        f.write(content)
                    print('%s pj is done!' %pj_name)

            elif len(value) == 1 and len(value[0]) == 3:
                # write br
                # br1, br2, br3
                for i in range(1, 4):

                    # header
                    content = '<?xml version="1.0" encoding="ISO-8859-1" ?>\n' \
                              '<BladedProject ApplicationVersion="4.8.0.41">\n' \
                              '	<DataModelVersion>0.8</DataModelVersion>\n' \
                              '	<BladedData dataFormat="project">\n' \
                              '<![CDATA[\n' \
                              'VERSION	4.8.0.34\n' \
                              'MULTIBODY	 1\n' \
                              'CALCULATION	25\n' \
                              'OPTIONS	0\n' \
                              'PROJNAME	\n' \
                              'DATE	\n' \
                              'ENGINEER	\n' \
                              'NOTES	""\n' \
                              'PASSWORD	\n' \
                              'MSTART RAINF\n' \
                              'EQLOAD	-1\n' \
                              'NSLOP	10\n' \
                              'SLOPES	  .3333333  .25  .2  .1666667  .1428571  .125  .1111111  .1  9.090909E-02  8.333334E-02\n'

                    content += 'ELFREQ	 %.8f\n' %float(2*self.PAI*float(self.eq_cycle)/(float(self.des_life)*8766*3600))
                    content += 'LIFE	 0\n' \
                               'TYPE	2\n' \
                               'MIN	 0\n' \
                               'MAX	 0\n' \
                               'NBINS	0\n' \
                               'MIN_RANGE	 0\n'
                    content += "ATTRIBF	%s\n" %(r"'c:\program files (x86)\dnv gl\bladed 4.8\'")
                    content += 'VARIAB	""\n' \
                               'NDIMENS	0\n' \
                               'MEND\n\n'

                    # variable
                    content += self.write_pj_variable(key, value[0][i-1])

                    # load case
                    content += 'MSTART MULTIRUN\n' \
                               'NCASES	%s\n' %str(self.dlc_num)

                    for lc in self.lc_list:

                        occurrence = None
                        if self.lc_lct[lc][1] == 'H':
                            occurrence = float(self.lc_lct[lc][0])*3600
                        elif self.lc_lct[lc][1] == 'T':
                            occurrence = float(self.lc_lct[lc][0])/(8766*3600)

                        content += 'DIRECTORY	%s\n' %self.lc_path[lc]
                        content += 'OLDVERSION	0\n'
                        content += 'RUNNAME	%s\n' %lc
                        content += 'CONFIG	%s\n' %self.lc_lct[lc][1]
                        content += 'OCCFREQ	 %.14f\n' %occurrence

                    content += 'LIFE	 %d\n' %(int(self.des_life)*8766*3600)
                    content += 'BINTYPE	 1\n'
                    content += 'MEND\n\n'

                    content += '0RAINF\n'
                    content += '0MULTIRUN\n'
                    content += '0MULTIVAR\n'

                    pj_name = key+str(i)
                    content += '		]]>\n' \
                               '	</BladedData>\n' \
                               '<RunConfiguration>\n' \
                               '  <Name>%s</Name>\n' \
                               '  <Calculation>0</Calculation>\n' \
                               '</RunConfiguration>\n\n' \
                               '</BladedProject>' %pj_name

                    pj_file = pj_name+'.$PJ'
                    pj_path = os.path.join(self.fat_path, pj_name).replace(' ', '_')
                    if not os.path.isdir(pj_path):
                        os.makedirs(pj_path)

                    with open(os.sep.join([pj_path, pj_file]), 'w+') as f:
                        f.write(content)
                    print('%s pj is done!' %pj_name)

            elif len(value) == 1 and len(value[0]) != 3:
                # write hs, hr, yb
                # header
                content = '<?xml version="1.0" encoding="ISO-8859-1" ?>\n' \
                          '<BladedProject ApplicationVersion="4.8.0.41">\n' \
                          '	<DataModelVersion>0.8</DataModelVersion>\n' \
                          '	<BladedData dataFormat="project">\n' \
                          '<![CDATA[\n' \
                          'VERSION	4.8.0.34\n' \
                          'MULTIBODY	 1\n' \
                          'CALCULATION	25\n' \
                          'OPTIONS	0\n' \
                          'PROJNAME	\n' \
                          'DATE	\n' \
                          'ENGINEER	\n' \
                          'NOTES	""\n' \
                          'PASSWORD	\n' \
                          'MSTART RAINF\n' \
                          'EQLOAD	-1\n' \
                          'NSLOP	10\n' \
                          'SLOPES	  .3333333  .25  .2  .1666667  .1428571  .125  .1111111  .1  9.090909E-02  8.333334E-02\n'

                content += 'ELFREQ	 %.8f\n' %float(2*self.PAI*float(self.eq_cycle)/(float(self.des_life)*8766*3600))
                content += 'LIFE	 0\n' \
                           'TYPE	2\n' \
                           'MIN	 0\n' \
                           'MAX	 0\n' \
                           'NBINS	0\n' \
                           'MIN_RANGE	 0\n'
                content += "ATTRIBF	%s\n" %(r"'c:\program files (x86)\dnv gl\bladed 4.8\'")
                content += 'VARIAB	""\n' \
                           'NDIMENS	0\n' \
                           'MEND\n\n'

                # variable
                content += self.write_pj_variable(key, value[0])

                # load case
                content += 'MSTART MULTIRUN\n' \
                           'NCASES	%s\n' %str(self.dlc_num)
                # print(self.lc_list)

                for lc in self.lc_list:
                    occurrence = None
                    if self.lc_lct[lc][1] == 'H':
                        occurrence = float(self.lc_lct[lc][0])*3600
                    elif self.lc_lct[lc][1] == 'T':
                        occurrence = float(self.lc_lct[lc][0])/(8766*3600)
                    # print(occurrence)

                    content += 'DIRECTORY	%s\n' %self.lc_path[lc]
                    content += 'OLDVERSION	0\n'
                    content += 'RUNNAME	%s\n' %lc
                    content += 'CONFIG	%s\n' %self.lc_lct[lc][1]
                    content += 'OCCFREQ	 %.14f\n' %occurrence

                content += 'LIFE	 %d\n' %(int(self.des_life)*8766*3600)
                content += 'BINTYPE	 1\n'
                content += 'MEND\n\n'

                content += '0RAINF\n'
                content += '0MULTIRUN\n'
                content += '0MULTIVAR\n'

                content += '		]]>\n' \
                           '	</BladedData>\n' \
                           '<RunConfiguration>\n' \
                           '  <Name>%s</Name>\n' \
                           '  <Calculation>0</Calculation>\n' \
                           '</RunConfiguration>\n\n' \
                           '</BladedProject>' %key

                pj_file = key+'.$PJ'
                pj_path = os.path.join(self.fat_path, key).replace(' ', '_')
                if not os.path.isdir(pj_path):
                    os.makedirs(pj_path)

                with open(os.sep.join([pj_path, pj_file]), 'w+') as f:
                    f.write(content)
                print('%s pj is done!' %key)

    def write_in(self):

        for key, value in self.cha_dict.items():
            # print(key, value)

            # variable
            if len(value) == 2 and len(value[0]) == 3:
                # write bpa, bra, baa, bua
                # bpa1, bpa2, bpa3
                for i in range(1, 4):

                    # variable
                    for sec in value[-1]:

                        se_name = str('%.3f' %float(sec)) if '.' in sec else sec
                        pj_name = key+str(i)+'_'+se_name
                        pj_path = os.path.join(self.fat_path, pj_name).replace(' ', '_')

                        # header
                        content  = 'PTYPE	10\n' \
                                   'SDSTAT	1\n' \
                                   'OUTSTYLE	B\n'
                        content += 'PATH	%s\n' %pj_path
                        content += 'RUNNAME	%s\n' %pj_name
                        content += 'MSTART RAINF\n' \
                                   'NSLOP	10\n' \
                                   'SLOPES	  .3333333  .25  .2  .1666667  .1428571  .125  .1111111  .1  9.090909E-02  8.333334E-02\n'

                        content += 'ELFREQ	 %.8f\n' %float(2*self.PAI*float(self.eq_cycle)/(float(self.des_life)*8766*3600))
                        content += 'LIFE	 0\n' \
                                   'TYPE	2\n' \
                                   'MEND\n\n'

                        # variable
                        # str(i) was used to identify the different %for different blade
                        content += self.write_in_variable(key+str(i), value[0][i-1], sec)

                        # load case
                        content += 'MSTART MULTIRUN\n' \
                                   'NCASES	%s\n' %str(self.dlc_num)
                        content += 'OUTCHOICE	1\n'

                        for lc in self.lc_list:

                            occurrence = None
                            if self.lc_lct[lc][1] == 'H':
                                occurrence = float(self.lc_lct[lc][0])*3600
                            elif self.lc_lct[lc][1] == 'T':
                                occurrence = float(self.lc_lct[lc][0])/(8766*3600)

                            content += 'DIRECTORY	%s\n' %(self.lc_path[lc]+'\\')
                            content += 'RUNNAME	%s\n' %lc
                            content += 'CONFIG	%s\n' %self.lc_lct[lc][1]
                            content += 'OCCFREQ	 %.14f\n' %occurrence

                        content += 'LIFE	 %d\n' %(int(self.des_life)*8766*3600)
                        content += 'MEND\n\n'

                        content += 'MSTART WINDREGIME\n' \
                                   'WTYPE	1\n' \
                                   'AMWS	999\n' \
                                   'WEIBLK	1\n' \
                                   'WEIBLC	1\n' \
                                   'MEND'

                        if not os.path.isdir(pj_path):
                            os.makedirs(pj_path)

                        with open(os.sep.join([pj_path, 'dtsignal.in']), 'w+') as f:
                            f.write(content)
                        print('%s in is done!' %pj_name)

            elif len(value) == 2 and len(value[0]) != 3:
                # write tr
                # print(value)
                for sec in value[-1]:

                    pj_name = (key+'_'+str('%.3f' %float(sec))) if 'Mbr' not in sec else sec
                    pj_path = os.path.join(self.fat_path, pj_name).replace(' ', '_')

                    # header
                    content  = 'PTYPE	10\n' \
                               'SDSTAT	1\n' \
                               'OUTSTYLE	B\n'
                    content += 'PATH	%s\n' %pj_path
                    content += 'RUNNAME	%s\n' %(pj_name if 'Mbr' not in pj_name else ("'%s'" %pj_name))
                    content += 'MSTART RAINF\n' \
                               'NSLOP	10\n' \
                               'SLOPES	  .3333333  .25  .2  .1666667  .1428571  .125  .1111111  .1  9.090909E-02  8.333334E-02\n'

                    content += 'ELFREQ	 %.8f\n' %float(2*self.PAI*float(self.eq_cycle)/(float(self.des_life)*8766*3600))
                    content += 'LIFE	 0\n' \
                               'TYPE	2\n' \
                               'MEND\n\n'

                    # variable
                    content += self.write_in_variable(key, value[0], sec)

                    # load case
                    content += 'MSTART MULTIRUN\n' \
                               'NCASES	%s\n' %str(self.dlc_num)
                    content += 'OUTCHOICE	1\n'

                    for lc in self.lc_list:

                        occurrence = None
                        if self.lc_lct[lc][1] == 'H':
                            occurrence = float(self.lc_lct[lc][0])*3600
                        elif self.lc_lct[lc][1] == 'T':
                            occurrence = float(self.lc_lct[lc][0])/(8766*3600)

                        content += 'DIRECTORY	%s\n' %(self.lc_path[lc]+'\\')
                        content += 'RUNNAME	%s\n' %lc
                        content += 'CONFIG	%s\n' %self.lc_lct[lc][1]
                        content += 'OCCFREQ	 %.14f\n' %occurrence

                    content += 'LIFE	 %d\n' %(int(self.des_life)*8766*3600)
                    content += 'MEND\n\n'

                    content += 'MSTART WINDREGIME\n' \
                               'WTYPE	1\n' \
                               'AMWS	999\n' \
                               'WEIBLK	1\n' \
                               'WEIBLC	1\n' \
                               'MEND'

                    if not os.path.isdir(pj_path):
                        os.makedirs(pj_path)

                    with open(os.sep.join([pj_path, 'dtsignal.in']), 'w+') as f:
                        f.write(content)

                    print('%s in is done!' %pj_name)

            elif len(value) == 1 and len(value[0]) == 3:
                # write br
                # br1, br2, br3
                for i in range(1, 4):

                    pj_name = key+str(i)
                    pj_path = os.path.join(self.fat_path, pj_name).replace(' ', '_')

                    # header
                    content = 'PTYPE	10\n' \
                              'SDSTAT	1\n' \
                              'OUTSTYLE	B\n'
                    content += 'PATH	%s\n' %pj_path
                    content += 'RUNNAME	%s\n' %pj_name
                    content += 'MSTART RAINF\n' \
                               'NSLOP	10\n' \
                               'SLOPES	  .3333333  .25  .2  .1666667  .1428571  .125  .1111111  .1  9.090909E-02  8.333334E-02\n'

                    content += 'ELFREQ	 %.8f\n' %float(
                        2 * self.PAI * float(self.eq_cycle) / (float(self.des_life) * 8766 * 3600))
                    content += 'LIFE	 0\n' \
                               'TYPE	2\n' \
                               'MEND\n\n'

                    # variable
                    # str(i) was used to identify the different %for different blade
                    content += self.write_in_variable(key, value[0][i-1])

                    # load case
                    content += 'MSTART MULTIRUN\n' \
                               'NCASES	%s\n' %str(self.dlc_num)
                    content += 'OUTCHOICE	1\n'

                    for lc in self.lc_list:

                        occurrence = None
                        if self.lc_lct[lc][1] == 'H':
                            occurrence = float(self.lc_lct[lc][0]) * 3600
                        elif self.lc_lct[lc][1] == 'T':
                            occurrence = float(self.lc_lct[lc][0]) / (8766 * 3600)

                        content += 'DIRECTORY	%s\n' %(self.lc_path[lc]+'\\')
                        content += 'RUNNAME	%s\n' %lc
                        content += 'CONFIG	%s\n' %self.lc_lct[lc][1]
                        content += 'OCCFREQ	 %.14f\n' %occurrence

                    content += 'LIFE	 %d\n' %(int(self.des_life) * 8766 * 3600)
                    content += 'MEND\n\n'

                    content += 'MSTART WINDREGIME\n' \
                               'WTYPE	1\n' \
                               'AMWS	999\n' \
                               'WEIBLK	1\n' \
                               'WEIBLC	1\n' \
                               'MEND'

                    if not os.path.isdir(pj_path):
                        os.makedirs(pj_path)

                    with open(os.sep.join([pj_path, 'dtsignal.in']), 'w+') as f:
                        f.write(content)

                    print('%s in is done!' %pj_name)

            elif len(value) == 1 and len(value[0]) != 3:
                # write hs, hr, yb
                # header

                pj_path = os.path.join(self.fat_path, key).replace(' ', '_')

                # header
                content  = 'PTYPE	10\n' \
                           'SDSTAT	1\n' \
                           'OUTSTYLE	B\n'
                content += 'PATH	%s\n' %pj_path
                content += 'RUNNAME	%s\n' %key
                content += 'MSTART RAINF\n' \
                           'NSLOP	10\n' \
                           'SLOPES	  .3333333  .25  .2  .1666667  .1428571  .125  .1111111  .1  9.090909E-02  8.333334E-02\n'

                content += 'ELFREQ	 %.8f\n' %float(2*self.PAI*float(self.eq_cycle)/(float(self.des_life)*8766*3600))
                content += 'LIFE	 0\n' \
                           'TYPE	2\n' \
                           'MEND\n\n'

                # variable
                content += self.write_in_variable(key, value[0])

                # load case
                content += 'MSTART MULTIRUN\n' \
                           'NCASES	%s\n' %str(self.dlc_num)
                content += 'OUTCHOICE	1\n'

                for lc in self.lc_list:

                    occurrence = None
                    if self.lc_lct[lc][1] == 'H':
                        occurrence = float(self.lc_lct[lc][0])*3600
                    elif self.lc_lct[lc][1] == 'T':
                        occurrence = float(self.lc_lct[lc][0])/(8766*3600)

                    content += 'DIRECTORY	%s\n' %(self.lc_path[lc]+'\\')
                    content += 'RUNNAME	%s\n' %lc
                    content += 'CONFIG	%s\n' %self.lc_lct[lc][1]
                    content += 'OCCFREQ	 %.14f\n' %occurrence

                content += 'LIFE	 %d\n' %(int(self.des_life)*8766*3600)
                content += 'MEND\n\n'

                content += 'MSTART WINDREGIME\n' \
                           'WTYPE	1\n' \
                           'AMWS	999\n' \
                           'WEIBLK	1\n' \
                           'WEIBLC	1\n' \
                           'MEND'

                if not os.path.isdir(pj_path):
                    os.makedirs(pj_path)

                with open(os.sep.join([pj_path, 'dtsignal.in']), 'w+') as f:
                    f.write(content)

                print('%s in is done!' %key)

    def write_in_variable(self, channel, var_list, height=None):

        content = ''

        if channel == 'br':
            # br1, br2, br3
            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:
                
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%22\n'
                content += "VARIAB	'%s'\n" %var
                content += "FULL_NAME	'%s'\n" %var

        elif channel == 'hr':

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:

                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%22\n'
                content += "VARIAB	'%s'\n" %var
                content += "FULL_NAME	'%s'\n" %var

        elif channel == 'hs':

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:

                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%23\n'
                content += "VARIAB	'%s'\n"  %var
                content += "FULL_NAME	'%s'\n" %var

        elif channel == 'yb':

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:

                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%24\n'
                content +=  "VARIAB	'%s'\n"  %var
                content += "FULL_NAME	'%s'\n" %var

        elif channel == 'tr' and 'Mbr' in height:

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:

                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%25\n'
                content += "VARIAB	'%s'\n" %var
                content += "AXITICK2	'%s'\n" %height
                content += "FULL_NAME	'%s, Location=%s'\n" %(var, height+'.')

        elif channel == 'tr' and 'Mbr' not in height:

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%25\n'
                content += "VARIAB	'%s'\n" %var
                content += 'DIM2	 %s\n' %height
                content += "FULL_NAME	'%s, Tower station height= %sm'\n" %(var, float(height))

        elif channel.startswith('bps'):

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%'+'%s\n' %(['15','16','17'][int(channel[-1])-1])
                content += "VARIAB	'%s'\n" %var
                content += 'DIM2	 %s\n' %height
                content += "FULL_NAME	'%s, Distance along blade= %sm'\n" %(var, height)

        elif channel.startswith('brs') and ('comb'not in channel):
            # print(channel)
            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%'+'%s\n' %(['41','42','43'][int(channel[-1])-1])
                content += "VARIAB	'%s'\n" %var
                content += 'DIM2	 %s\n' %height
                content += "FULL_NAME	'%s, Distance along blade= %sm'\n" %(var, height)

        elif channel.startswith('bas'):

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%'+'%s\n' %(['59','60','61'][int(channel[-1])-1])
                content += "VARIAB	'%s'\n" %var
                content += 'DIM2	 %s\n' %height
                content += "FULL_NAME	'%s, Distance along blade= %sm'\n" %(var, height)

        elif channel.startswith('bus'):

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%'+'%s\n' %(['62', '63', '64'][int(channel[-1])-1])
                content += "VARIAB	'%s'\n" %var
                content += 'DIM2	 %s\n' %height
                content += "FULL_NAME	'%s, Distance along blade= %sm'\n" %(var, height)

        elif channel.startswith('brs') and 'comb' in channel:
            chan_ext = {'brs1':'800', 'brs2':'810','brs3':'820'}
            # brs_comb
            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	%s\n' %len(var_list)
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' % self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%%%s\n' %chan_ext[channel[:4]]
                content += "VARIAB	'%s'\n" % var
                content += "FULL_NAME	'%s'\n" % var

        elif channel == 'br_comb':
            # brs_comb
            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	%s\n' %len(var_list)
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' % self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%700\n'
                content += "VARIAB	'%s'\n" % var
                content += "FULL_NAME	'%s'\n" % var

        content += 'MEND\n\n'
        # print(channel)
        return content

    def write_pj_variable(self, channel, var_list, height=None):
        # print(channel, var_list)
        content = ''

        if channel == 'br':
            # br1, br2, br3
            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%22\n'
                content +=  'VARIAB	"%s"\n'  %var
                content += 'DESCRIPTION	"%s"\n' %var
                content += 'NDIMENS	2\n'
                content += 'DIMFLAG	-1\n'

        elif channel == 'hr':

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%22\n'
                content +=  'VARIAB	"%s"\n'  %var
                content += 'DESCRIPTION	"%s"\n' %var
                content += 'NDIMENS	2\n'
                content += 'DIMFLAG	-1\n'

        elif channel == 'hs':

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%23\n'
                content +=  'VARIAB	"%s"\n'  %var
                content += 'DESCRIPTION	"%s"\n' %var
                content += 'NDIMENS	2\n'
                content += 'DIMFLAG	-1\n'

        elif channel == 'yb':

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%24\n'
                content +=  'VARIAB	"%s"\n'  %var
                content += 'DESCRIPTION	"%s"\n' %var
                content += 'NDIMENS	2\n'
                content += 'DIMFLAG	-1\n'

        elif channel == 'tr' and 'Mbr' in height:

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%25\n'
                content +=  'VARIAB	"%s"\n'  %var
                content += 'DESCRIPTION	"%s, Location=%s"\n' %(var, height+'.')
                content += 'NDIMENS	3\n'
                content += 'DIMFLAG 0\n'
                content += "AXITICK2	'%s'\n" %height

        elif channel == 'tr' and 'Mbr' not in height:

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%25\n'
                content +=  'VARIAB	"%s"\n'  %var
                content += 'DESCRIPTION	"%s, Tower station height= %sm"\n' %(var, float(height))
                content += 'NDIMENS	3\n'
                content += 'DIMFLAG -1\n'
                content += 'DIM2	 %s\n' %height

        elif channel.startswith('bps'):

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%'+'%s\n' %(['15', '16', '17'][int(channel[-1])-1])
                content +=  'VARIAB	"%s"\n'  %var
                content += 'DESCRIPTION	"%s, Distance along blade= %sm"\n' %(var, height)
                content += 'NDIMENS	3\n'
                content += 'DIMFLAG	-1\n'
                content += 'DIM2	 %s\n' %height

        elif channel.startswith('brs') and ('comb' not in channel):

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%'+'%s\n' %(['41', '42', '43'][int(channel[-1])-1])
                content +=  'VARIAB	"%s"\n'  %var
                content += 'DESCRIPTION	"%s, Distance along blade= %sm"\n' %(var, height)
                content += 'NDIMENS	3\n'
                content += 'DIMFLAG	-1\n'
                content += 'DIM2	 %s\n' %height

        elif channel.startswith('bas'):

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%'+'%s\n' %(['59', '60', '61'][int(channel[-1])-1])
                content +=  'VARIAB	"%s"\n'  %var
                content += 'DESCRIPTION	"%s, Distance along blade= %sm"\n' %(var, height)
                content += 'NDIMENS	3\n'
                content += 'DIMFLAG	-1\n'
                content += 'DIM2	 %s\n' %height

        elif channel.startswith('bus'):

            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	6\n'
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%'+'%s\n' %(['62', '63', '64'][int(channel[-1])-1])
                content +=  'VARIAB	"%s"\n'  %var
                content += 'DESCRIPTION	"%s, Distance along blade= %sm"\n' %(var, height)
                content += 'NDIMENS	3\n'
                content += 'DIMFLAG	-1\n'
                content += 'DIM2	 %s\n' %height

        elif channel.startswith('brs') and 'comb' in channel:
            chan_ext = {'brs1':'800', 'brs2':'810','brs3':'820'}
            # brs_comb
            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	%s\n' %len(var_list)
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%%%s\n' %chan_ext[channel[:4]]
                content +=  'VARIAB	"%s"\n'  %var
                content += 'DESCRIPTION	"%s"\n' %var
                content += 'NDIMENS	2\n'
                content += 'DIMFLAG	-1\n'

        elif channel == 'br_comb':
            # print(channel)
            # brs_comb
            content += 'MSTART MULTIVAR\n'
            content += 'NVARS	%s\n' %len(var_list)
            for var in var_list:
                content += 'MIN	 0\n'
                content += 'MAX	 0\n'
                content += 'NBINS	%s\n' %self.num_bins
                content += 'MIN_RANGE	 0\n'
                content += 'ATTRIBF	%700\n'
                content +=  'VARIAB	"%s"\n'  %var
                content += 'DESCRIPTION	"%s"\n' %var
                content += 'NDIMENS	2\n'
                content += 'DIMFLAG	-1\n'

        content += 'MEND\n\n'
        # print(content)
        return content

    # @pysnooper.snoop()
    def read_lct(self):
        # print(self.lct_path)
        table = load_workbook(self.lct_path, keep_vba=True, data_only=True)

        ws_setting = table['Setting']
        IEC_stand  = '1' if '61400-1' in ws_setting['B1'].value else ('3' if '61400-3' in ws_setting['B1'].value else None)
        # print(IEC_stand)
        if IEC_stand == '1':

            for dlc in self.dlc_list:
                # print(dlc)
                lc = dlc.split('_')[0] if '_' in dlc else dlc.split()[0]
                name_col, hour_col, type_col = None, None, None

                if lc in table.sheetnames:
                    ws = table[lc]
                else:
                    raise Exception('%s not in table!' %dlc)
                col_num = ws.max_column
                row_num = ws.max_row
                row_start = None

                # for onshore, the table header are line9 and line10
                for j in range(1, row_num):
                    for i in range(1, col_num+1):

                        cell1 = ws.cell(j,   i)
                        cell2 = ws.cell(j+1, i)

                        if cell1.value == 'Run Name' and cell2.value in [None, '']:
                            name_col  = cell1.column
                            row_start = j+2
                            continue
                        if (cell1.value == 'Hour/Year' and cell2.value == 'hours') or \
                           (cell1.value == 'Times/year' and cell2.value == '-'):
                            hour_col = cell1.column
                            continue
                        if cell1.value == 'Rainflow' and cell2.value == 'type':
                            type_col = cell1.column
                            break

                    if name_col and hour_col and type_col:
                        break

                if not name_col:
                    raise Exception('%s has no Run Name colunm!' %dlc)
                if not hour_col:
                    raise Exception('%s has no Hour/Year or Times/year colunm!'%dlc)
                if not type_col:
                    raise Exception('%s has no Rainflow type colunm!'%dlc)

                for i in range(row_start, row_num+1):
                    # print(ws.cell(i, name_col).value)
                    if ws.cell(i, name_col).value:
                        if not isinstance(ws.cell(i, hour_col).value, (float,int)):
                            raise Exception('Times/Hour definition error for %s row number %s' %(dlc, i))

                        self.lc_lct[ws.cell(i, name_col).value] = [ws.cell(i, hour_col).value, ws.cell(i, type_col).value]

        elif IEC_stand == '3':

            for dlc in self.dlc_list:
                print(dlc)
                lc = dlc.split('_')[0] if '_' in dlc else dlc.split()[0]

                name_col = None
                hour_col = None
                type_col = None

                if lc in table.sheetnames:
                    ws = table[lc]
                else:
                    raise Exception('%s not in table!' %dlc)

                col_num = ws.max_column
                row_num = ws.max_row
                row_start = None

                # for onshore, the table header are line13 and line14
                for j in range(1, row_num):
                    for i in range(1, col_num+1):

                        cell1 = ws.cell(j,  i)
                        cell2 = ws.cell(j+1, i)

                        if cell1.value == 'Run Name' and cell2.value in [None, '']:
                            # print(cell1.value, cell2.value, cell1.column)
                            name_col  = cell1.column
                            row_start = j+2
                            continue
                        if (cell1.value == 'Hour/Year' and cell2.value == 'hours') or \
                           (cell1.value == 'Times/year' and cell2.value == '-'):
                            # print(cell1.value, cell2.value, cell1.column)
                            hour_col = cell1.column
                            continue
                        if cell1.value == 'Rainflow' and cell2.value == 'type':
                            # print(cell1.value, cell2.value, cell1.column)
                            type_col = cell1.column
                            break

                    if name_col and hour_col and type_col:
                        break

                if not name_col:
                    raise Exception('%s has no Run Name colunm!' %dlc)
                if not hour_col:
                    raise Exception('%s has no Hour/Year or Times/year colunm!'%dlc)
                if not type_col:
                    raise Exception('%s has no Rainflow type colunm!'%dlc)

                # print(name_col, hour_col, type_col)

                for i in range(row_start, row_num+1):
                    # print(ws.cell(i, name_col).value)
                    if ws.cell(i, name_col).value:
                        print(ws.cell(i, name_col).value, ws.cell(i, hour_col).value, type(ws.cell(i, hour_col).value))
                        if not isinstance(ws.cell(i, hour_col).value, (float,int)):
                            raise Exception('Times/Hour definition error for %s row number %s' %(dlc, i))

                        self.lc_lct[ws.cell(i, name_col).value] = [ws.cell(i, hour_col).value, ws.cell(i, type_col).value]
                        # print(ws.cell(i, name_col).value, [ws.cell(i, hour_col).value, ws.cell(i, type_col).value])

if __name__ == '__main__':

    run_path = r'\\172.20.0.4\fs03\CE\W6250-172-14m\run_0430_0.8'

    # ultimate(run_path)

