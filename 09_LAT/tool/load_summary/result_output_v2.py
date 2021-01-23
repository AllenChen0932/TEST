# -*- coding: utf-8 -*-
# @Time    : 2019/11/28 12:54
# @Author  : CE
# @File    : excel_output.py

'''
2019.12.4 版本012在011的基础上增加了fatigue的部分
'''

import os
import time
import openpyxl
import configparser
from openpyxl.utils import get_column_letter

from tool.load_summary.pitch_bearing_life import handle as pitch_bearing_life
from tool.load_summary.main_bearing_life import handle as main_bearing_life
from tool.load_summary.gearbox_equivalent_torque import handle as gearbox_eq_torque
from tool.load_summary.dynamic_power_curve import output_PCs as powercurve

from tool.load_summary.bladed_ultimate import BladedUltimate
from tool.load_summary.bladed_rainflow import BladedRainflow
from tool.load_summary.excel_format import SetFormat
from tool.load_summary.bladed_ldd import BladedLDD
from tool.load_summary.bladed_lrd import BladedLRD
from tool.load_summary.read_LCT import read_LCT_dlc12

__version__ = "2.0.2"
'''
2020.4.21_v2.0.1 repair the bug from APE
2020.4.21_v2.0.2 repair the bug from Ultimate tower output
'''

class excel_operation:

    def __init__(self,
                 excel_input,
                 output_path,
                 excel_name,
                 ultimate=False,
                 fatigue=False,
                 heat_map=False,
				 fat_case=False,
                 fat_wind=False,
                 aep_dlc=False,
                 ldd=False,
                 lrd=False,
                 mbl=False,
                 pbl=False,
                 get=False,
                 main_all=True,
                 new_app=True):

        self.template = excel_input
        self.out_path = output_path
        self.exl_name = excel_name
        self.ultimate = ultimate
        self.fatigue  = fatigue
        self.heat_map = heat_map
        self.fat_wind = fat_wind
        self.fat_case = fat_case
        self.aep_dlc  = aep_dlc
        self.ldd      = ldd
        self.lrd      = lrd
        self.mbl      = mbl
        self.pbl      = pbl
        self.get      = get
        self.main_all = main_all
        self.new_app  = new_app            #True: new; False: append
        self.var_m    = {}

        # input excel:
        self.loopname = []      # loopname list
        self.pst_list = []      # post path list
        self.ult_list = []      # ultimate path list
        self.fat_list = []      # fatigue path list
        self.ldd_list = []      # ldd path list
        self.lrd_list = []      # lrd path list
        self.dlc_list = []      # dlc12 path list
        self.lct_list = []      # load case table list
        self.pbl_list = []      # pitch bearing life
        self.get_list = []      # gearbox equivalent torque
        self.mbl_list = []      # main bearing life

        self.loop_dlc = {}      # {loopname: dlc}

        # for ultimate and heat map
        self.loop_channels  = {}     #{loopname: channel(which is in order)}
        self.loop_vari_dlc  = {}     #{loopname: {var: {dlc:value}}}, maximum value
        self.loop_var_ulti  = {}     #{loopname: {var: {dlc:value}}}, ultimate value
        self.loop_chan_var  = {}     #{loopname: {channel: variable}}
        self.loop_dlc_path  = {}     #{loopname: {dlc: path}}
        self.loop_dlcs_dlc  = {}     #{loopname: {dlcs: DLC}}
        self.loop_br_mx     = {}

        # for fatigue
        self.lp_channels = {}    # {loopname: [channel list]}
        self.lp_var_fat  = {}    # {loopname: {var: {m: {dlc: value}}}}
        self.lp_chan_var = {}    # {loopname: {channel: [variable list]}}
        self.lp_dlc_path = {}    # {loopname: {dlc: path}}

        # for wind fatigue
        self.lp_case_V = {}
        self.lp_V_case = {}
        self.lp_v_list = {}

        # for ldd
        self.loop_chan_pj = {}     # {loopname: {channel: project name}}
        self.loop_prj_ext = {}     # {loopname: {pj name: extension list}
        self.loop_prj_var = {}     # {loopname: {pj name: variable list}}
        self.lp_ch_pj_var = {}     # {loopname: {channel: {pj name: {variable: data}}}}

        # for lrd
        self.loop_chan_pj_lrd = {}     # {loopname: {channel: project name}}
        self.loop_prj_ext_lrd = {}     # {loopname: {pj name: extension list}
        self.loop_prj_var_lrd = {}     # {loopname: {pj name: variable list}}
        self.lp_ch_pj_var_lrd = {}     # {loopname: {channel: {pj name: {variable: data}}}}

        # get current data
        self.date = time.strftime("%Y%m%d", time.localtime())

        self.read_excel()
        self.write_result()

    # get variables
    def get_vars(self, path_list):

        dollar_list = []
        var_list    = []

        for path in path_list:

            for root, dir, names in os.walk(path):
                # print(root, dir, names)

                for name in names:

                    if '.%' in name:

                        file_dir = root + os.sep + name

                        dollar_list.append(file_dir)
        # print(dollar_list)
        for dollar in dollar_list:

            with open(dollar) as f:

                for line in f.readlines():

                    if 'VARIAB' in line:

                        variab = line.split("'")[1]

                        if ',' in variab and 'Blade' in variab:
                            temp = variab.split(',')

                            height = temp[1].split('=')[1].strip()[0].split('m')[0]

                            variab = temp[0] + ',' + '%.2f' % float(height)

                            var_list.append(variab)

                        if ',' in variab and 'Tower' in variab:

                            temp = variab.split(',')

                            if 'height' in temp[1]:

                                height = temp[1].split('=')[1].strip()[0].split('m')[0]

                                variab = temp[0] + ',' + '%.2f' % float(height)

                                var_list.append(variab)

                            elif 'Location' in temp[1]:

                                height = temp[1].split('=')[1]

                                variab = temp[0] + ',' + height

                                var_list.append(variab)

        return var_list

    # get path from template
    def read_excel(self):

        if not os.path.isfile(self.template):
            raise Exception('%s not valid!' %self.template)

        workbook = openpyxl.load_workbook(self.template)

        sheet = workbook.get_sheet_by_name('Main')

        row_num = sheet.max_row
        row_max = row_num-(row_num-1)%5
        # print(row_max)

        print('Begin to read excel...')
        self.loopname = [sheet[i][0].value for i in range(2,row_max+1) if sheet[i][0].value]
        self.pst_list = [sheet[i][2].value for i in range(2,row_max+1) if sheet[i][1].value == 'Post']
        self.dlc_list = [sheet[i][2].value for i in range(2,row_max+1) if sheet[i][1].value == 'DLC12']
        self.lct_list = [sheet[i][2].value for i in range(2,row_max+1) if sheet[i][1].value == 'LCT']
        self.ult_list = [sheet[i][2].value for i in range(2,row_max+1) if sheet[i][1].value == 'Ultimate']
        self.fat_list = [sheet[i][2].value for i in range(2,row_max+1) if sheet[i][1].value == 'Rainflow']

        print('Loop name    :', self.loopname)
        print('Post path    :', self.pst_list)
        print('DLC12 path   :', self.dlc_list)
        print('LCT path     :', self.lct_list)

        loop_num = len(self.loopname)

        if self.pst_list:
            for i in range(loop_num):
                post_path = self.pst_list[i]
                file_list = os.listdir(post_path)

                for file in file_list:
                    if '05_ldd' == file.lower():
                        self.ldd_list.append(os.path.join(post_path, file))

                    if '06_lrd' == file.lower():
                        self.lrd_list.append(os.path.join(post_path, file))

                    if '06_lrd' == file.lower():
                        br_mxy_lrd = os.path.join(os.path.join(post_path, r'06_lrd\br_mxy_64'))
                        if os.path.isdir(br_mxy_lrd):
                            self.pbl_list.append(br_mxy_lrd)

                    if '06_lrd' == file.lower():
                        hs_lrd = os.path.join(os.path.join(post_path, r'06_lrd\hs_64'))
                        if os.path.isdir(hs_lrd):
                            self.get_list.append(hs_lrd)

                    if '05_ldd' == file.lower():
                        hs_ldd = os.path.join(os.path.join(post_path, r'05_ldd\hs_144'))
                        if os.path.isdir(hs_ldd):
                            self.mbl_list.append(hs_ldd)

        print('Ultimate path:', self.ult_list)
        print('Fatigue path :', self.fat_list)
        print('LDD path     :', self.ldd_list)
        print('LRD path     :', self.lrd_list)
        print('pitch bearing:', self.pbl_list)
        print('gearbox path :', self.get_list)
        print('main bearing :', self.mbl_list)
        print()

    def get_result(self):

        print('Begin to get result...')
        lp_num = len(self.loopname)

        for i in range(lp_num):

            # ==============================================
            # get ultimate:
            if self.ultimate or self.heat_map:

                print('Begin to read "%s"' %self.ult_list[i])
                ult_path = self.ult_list[i]

                if os.path.isdir(ult_path):
                    result   = BladedUltimate(ult_path, self.main_all)

                    self.loop_channels[self.loopname[i]]  = result.main_cha
                    self.loop_var_ulti[self.loopname[i]]  = result.ult_val
                    self.loop_vari_dlc[self.loopname[i]]  = result.var_ult
                    self.loop_chan_var[self.loopname[i]]  = result.chan_var
                    self.loop_dlc_path[self.loopname[i]]  = result.dlc_path
                    self.loop_dlcs_dlc[self.loopname[i]]  = result.dlcs_DLC
                    self.loop_br_mx[self.loopname[i]]     = result.br_mx if result.chan_br else None
                    print('Read is done!')
                else:
                    raise Exception('%s not valid!' %ult_path)

            # ==============================================
            # get fatigue:
            if self.fatigue or self.fat_wind or self.fat_case:

                fat_path = self.fat_list[i]
                print('Begin to read "%s"' %self.fat_list[i])

                if os.path.isdir(fat_path):
                    result2 = BladedRainflow(fat_path, self.main_all)

                    self.lp_channels[self.loopname[i]] = result2.main_chan
                    self.lp_var_fat[self.loopname[i]]  = result2.fatigue_val
                    self.lp_chan_var[self.loopname[i]] = result2.chan_var
                    self.lp_dlc_path[self.loopname[i]] = result2.dlc_path
                    print('Read is done!')

                else:
                    raise  Exception('%s not valid!' %fat_path)

            # ==============================================
            # get Vhub and hours for dlc12 cases: (only used in wind fatigue)
            if self.fat_wind:

                lct_path = self.lct_list[i]
                print('Begin to read "%s"' % self.lct_list[i])

                if os.path.isfile(lct_path):
                    if self.lct_list[:i+1].count(lct_path) == 1:
                        result3 = read_LCT_dlc12(lct_path)
                        self.lp_case_V[self.loopname[i]] = result3[0]
                        self.lp_V_case[self.loopname[i]] = result3[1]
                        self.lp_v_list[self.loopname[i]] = result3[2]
                    # for loop with the same load case table
                    else:
                        lp_name  = self.loopname[self.lct_list[:i+1].index(lct_path)]
                        self.lp_case_V[self.loopname[i]] = self.lp_case_V[lp_name]
                        self.lp_V_case[self.loopname[i]] = self.lp_V_case[lp_name]
                        self.lp_v_list[self.loopname[i]] = self.lp_v_list[lp_name]

                    print('Read is done!')

                else:
                    raise  Exception('%s not valid!' %lct_path)

            # ==============================================
            # get ldd
            if self.ldd:

                ldd_path = self.ldd_list[i]
                print('Begin to read "%s"' % self.ldd_list[i])

                if os.path.isdir(ldd_path):
                    result4 = BladedLDD(ldd_path)
                    self.loop_chan_pj[self.loopname[i]] = result4.chan_pj
                    self.loop_prj_ext[self.loopname[i]] = result4.prj_ext
                    self.loop_prj_var[self.loopname[i]] = result4.prj_var
                    self.lp_ch_pj_var[self.loopname[i]] = result4.ch_pj_var_data
                    print('Read is done!')

                else:
                    raise Exception('%s not valid!' % ldd_path)

            # ==============================================
            # get lrd
            if self.lrd:
                lrd_path = self.lrd_list[i]
                print('Begin to read "%s"' % self.lrd_list[i])

                if os.path.isdir(lrd_path):
                    result5 = BladedLRD(lrd_path)
                    self.loop_chan_pj_lrd[self.loopname[i]] = result5.chan_pj
                    self.loop_prj_ext_lrd[self.loopname[i]] = result5.prj_ext
                    self.loop_prj_var_lrd[self.loopname[i]] = result5.prj_var
                    self.lp_ch_pj_var_lrd[self.loopname[i]] = result5.ch_pj_var_data
                    print('Read is done!')

                else:
                    raise Exception('%s not valid!' % lrd_path)

        print('Get result is done! \n')

    def get_heatmap(self, dlc_DLC, dlc_val):

        DLC = []
        DLC_val  = {}

        for key, val in dlc_val.items():

            if dlc_DLC[key] not in DLC:

                DLC.append(dlc_DLC[key])
                DLC_val[dlc_DLC[key]] = val[0]

            elif dlc_DLC[key] in DLC:

                if val[0] > DLC_val[dlc_DLC[key]]:

                    DLC_val[dlc_DLC[key]] = val[0]

                else:

                    continue

            DLC.sort()

        return DLC_val, DLC

    def get_m_slope(self):

        try:
            config = configparser.ConfigParser()
            config.read('config_m.dat', encoding='utf-8')

            if config.has_section('Slope M'):
                options = config.options("Slope M")
                print(options)
                self.var_m['Blade principal axes'] = config.get('Slope M','blade principal axes') \
                    if 'blade principal axes' in options else '10.0'
                self.var_m['Blade aerodynamic axes'] = config.get('Slope M','blade aerodynamic axes') \
                    if 'blade aerodynamic axes' in options else '10.0'
                self.var_m['Blade root axes'] = config.get('Slope M','blade root axes') \
                    if 'blade root axes' in options else '10.0'
                self.var_m['Blade user axes'] = config.get('Slope M','blade user axes') \
                    if 'blade user axes'.lower() in options else '10.0'
                self.var_m['Blade root'] = config.get('Slope M','blade root') \
                    if 'blade root' in options else '4.0'
                self.var_m['Rotating hub'] = config.get('Slope M','rotating hub') \
                    if 'rotating hub' in options else '4.0'
                self.var_m['Stationary hub'] = config.get('Slope M','stationary hub') \
                    if 'stationary hub' in options else '4.0'
                self.var_m['Yaw bearing'] = config.get('Slope M','yaw bearing') \
                    if 'yaw bearing' in options else '4.0'
                self.var_m['Tower'] = config.get('Slope M','tower') \
                    if 'tower' in options else '4.0'
                print(self.var_m.items())

            if config.has_section('Gearbox Equivalent Torque'):
                self.get_m = config.get('Gearbox Equivalent Torque', 'm')

        except Exception:
            raise Exception('Error occurs in reading config_m.dat!')

    def write_result(self):

        # **************************** New excel table ****************************************
        table       = openpyxl.Workbook()
        sheet_names = table.sheetnames
        # print(sheet_names)
        # **************************** New excel table ****************************************

        self.get_result()
        lp_num = len(self.loopname)

        # ********************** Ultimate and Heat map ****************************************
        if self.ultimate or self.heat_map:
            ult_chan_list = []
            for loop, channels in self.loop_channels.items():
                for chan in channels:
                    if 'Tower' not in chan and chan:
                        ult_chan_list.append(chan)

            ult_chan_list = list(set(ult_chan_list))
            ult_chan_list.sort()

        # ***************************** Ultimate **********************************************
        if self.new_app and self.ultimate:

            print('Begin to write ultimate result...')
            if 'Sheet' in sheet_names:
                sheet_ultimate       = table['Sheet']
                sheet_ultimate.title = 'Ultimate'

            # write ultimate
            for i in range(lp_num):
                print('Begin to write: ', self.loopname[i])
                loop_name = self.loopname[i]

                # 写表头，前4行
                sheet_ultimate.cell(row=1, column=i*6+1, value=i+1)
                sheet_ultimate.cell(row=2, column=i*6+1, value=self.loopname[i])
                hyper1 = '=HYPERLINK("{}","{}")' .format(self.ult_list[i], self.ult_list[i])
                sheet_ultimate.cell(row=3, column=i*6+1, value=hyper1)
                sheet_ultimate.cell(row=4, column=i*6+1, value='Variable')
                sheet_ultimate.cell(row=4, column=i*6+2, value='DLC')
                sheet_ultimate.cell(row=4, column=i*6+3, value='Value\n(kN/kNm)')
                sheet_ultimate.cell(row=4, column=i*6+4, value='Path')
                sheet_ultimate.cell(row=4, column=i*6+5, value= 1)

                # 写变量和数据
                tow_ind  = len(ult_chan_list)                       # start row for tower channel
                chan_num = len(self.loop_channels[loop_name])

                tower_num = 0
                for j in range(chan_num):
                    cha = self.loop_channels[loop_name][j]

                    if 'Tower' in cha:
                        tower_num += 1

                    for k in range(8):

                        var = self.loop_chan_var[loop_name][cha][k]
                        dlc = [key for key in self.loop_var_ulti[loop_name][var].keys()]
                        val = [val for val in self.loop_var_ulti[loop_name][var].values()]
                        # pat = self.loop_dlc_path[loop_name][dlc[0]]
                        # different statistics method: absolute, maximum/minimum
                        dlc_name = '_'.join(dlc[0].split('_')[:-1]) if 'max' in dlc[0].lower() or 'min' in dlc[0] else dlc[0]
                        pat      = self.loop_dlc_path[loop_name][dlc_name]
                        # print(pat)

                        char = get_column_letter(6*i+5)
                        rat  = '=INDIRECT(ADDRESS(ROW(),COLUMN()-2))/INDIRECT(ADDRESS(ROW(),(%s$4*6-3)))' % char

                        # get row index for each variables
                        br_index = ult_chan_list.index('Blade root') if 'Blade root' in ult_chan_list else -1
                        if 'Tower' not in cha:
                            if cha == 'Blade root':
                                if k < 2:
                                    row_index = 5+ult_chan_list.index(cha)*11+2*k
                                else:
                                    row_index = 5+ult_chan_list.index(cha)*11+k+2
                            else:
                                chan_index = ult_chan_list.index(cha)
                                row_index = 5+(ult_chan_list.index(cha)-1)*9+k+11 if chan_index>br_index and br_index!=-1 \
                                    else 5+ult_chan_list.index(cha)*9+k+11
                        else:
                            row_index = 5+(tow_ind-1)*9+k+11 if tow_ind>br_index and br_index!=-1 else 5+tow_ind*9+k
                            tow_ind  += int((k+1)/8)

                        if tower_num == 1:
                            var = var.split(',')[0].split()[-1]+'TT'
                        elif tower_num == 2:
                            var = var.split(',')[0].split()[-1]+'TB'

                        if cha == 'Blade root' and k == 0:
                            var_list = list(self.loop_br_mx[loop_name].keys())
                            var_list.sort()
                            var = var_list[0]
                            # print(self.loop_br_mx[loop_name][var])
                            sheet_ultimate.cell(row=row_index, column=i*6+1, value=var)
                            sheet_ultimate.cell(row=row_index, column=i*6+2, value=self.loop_br_mx[loop_name][var][0])
                            sheet_ultimate.cell(row=row_index, column=i*6+3, value=self.loop_br_mx[loop_name][var][1]/1000)
                            pat = self.loop_dlc_path[loop_name][self.loop_br_mx[loop_name][var][0]]
                            hyper2 = '=HYPERLINK("{}","{}")'.format(pat, pat)
                            sheet_ultimate.cell(row=row_index+1, column=i*6+4, value=hyper2)
                            sheet_ultimate.cell(row=row_index+1, column=i*6+5, value=rat)
                            var = var_list[1]
                            sheet_ultimate.cell(row=row_index+1, column=i*6+1, value=var)
                            sheet_ultimate.cell(row=row_index+1, column=i*6+2, value=self.loop_br_mx[loop_name][var][0])
                            sheet_ultimate.cell(row=row_index+1, column=i*6+3, value=self.loop_br_mx[loop_name][var][1]/1000)
                            pat = self.loop_dlc_path[loop_name][self.loop_br_mx[loop_name][var][0]]
                            hyper2 = '=HYPERLINK("{}","{}")'.format(pat, pat)
                            sheet_ultimate.cell(row=row_index, column=i*6+4, value=hyper2)
                            sheet_ultimate.cell(row=row_index, column=i*6+5, value=rat)
                        elif cha == 'Blade root' and k == 1:
                            var_list = list(self.loop_br_mx[loop_name].keys())
                            var_list.sort()
                            var = var_list[2]
                            sheet_ultimate.cell(row=row_index, column=i*6+1, value=var)
                            sheet_ultimate.cell(row=row_index, column=i*6+2, value=self.loop_br_mx[loop_name][var][0])
                            sheet_ultimate.cell(row=row_index, column=i*6+3, value=self.loop_br_mx[loop_name][var][1]/1000)
                            pat = self.loop_dlc_path[loop_name][self.loop_br_mx[loop_name][var][0]]
                            hyper2 = '=HYPERLINK("{}","{}")'.format(pat, pat)
                            sheet_ultimate.cell(row=row_index, column=i*6+4, value=hyper2)
                            sheet_ultimate.cell(row=row_index, column=i*6+5, value=rat)
                            var = var_list[3]
                            sheet_ultimate.cell(row=row_index+1, column=i*6+1, value=var)
                            sheet_ultimate.cell(row=row_index+1, column=i*6+2, value=self.loop_br_mx[loop_name][var][0])
                            sheet_ultimate.cell(row=row_index+1, column=i*6+3, value=self.loop_br_mx[loop_name][var][1]/1000)
                            pat = self.loop_dlc_path[loop_name][self.loop_br_mx[loop_name][var][0]]
                            hyper2 = '=HYPERLINK("{}","{}")'.format(pat, pat)
                            sheet_ultimate.cell(row=row_index+1, column=i*6+4, value=hyper2)
                            sheet_ultimate.cell(row=row_index+1, column=i*6+5, value=rat)
                        else:
                            sheet_ultimate.cell(row=row_index, column=i*6+1, value=var)
                            sheet_ultimate.cell(row=row_index, column=i*6+2, value=dlc[0])
                            # 此处读取的array数据带格式说明，实际上是一个包含数值和类型两个元素的元组，因此需要加上[0]来取其值
                            sheet_ultimate.cell(row=row_index, column=i*6+3, value=val[0][0]/1000)
                            hyper2 = '=HYPERLINK("{}","{}")'.format(pat, pat)
                            sheet_ultimate.cell(row=row_index, column=i*6+4, value=hyper2)
                            sheet_ultimate.cell(row=row_index, column=i*6+5, value=rat)

                print('loop %s is done!' %(i+1))

            SetFormat().set_ultimate(sheet_ultimate)
            print('Ultimate loads is done! \n')

        # ***************************** Heat map **********************************************
        if self.new_app and self.heat_map:

            print('Begin to read heat map...')
            sheet_heatmap = table.create_sheet('Heat Map')

            # column values for the variables
            col_var = 1

            for i in range(lp_num):
                print('Begin to write: ', self.loopname[i])

                # 写表头
                loop_name = self.loopname[i]
                # the column for loop name and variables are the same
                sheet_heatmap.cell(row=1, column=col_var, value=loop_name)

                # 写变量和数据
                tow_ind  = len(ult_chan_list)                       # start row for tower channel
                chan_num = len(self.loop_channels[loop_name])

                tower_num = 0
                for j in range(chan_num):
                    ch_name = self.loop_channels[loop_name][j]

                    if 'Tower' in ch_name:
                        tower_num += 1

                    # 记录工况
                    dlc_list = []

                    for k in range(8):

                        d_dlc = self.loop_dlcs_dlc[loop_name]              #dict of dlcs and DLC for each loop
                        v_nam = self.loop_chan_var[loop_name][ch_name][k]  #variable name
                        d_val = self.loop_vari_dlc[loop_name][v_nam]       #dict of dlc and value for each variable
                        loop_result = self.get_heatmap(d_dlc, d_val)       #{dlcs: DLC}, {DLC: value}

                        # get row index for each variables
                        if 'Tower' not in ch_name:
                            row_index = 2+ult_chan_list.index(ch_name)*10+k+2
                        else:
                            row_index = 2+tow_ind*10+k+2
                            tow_ind   += int((k+1)/8)

                        if k == 0:
                            # 记录第一个变量的工况名称，后面的工况按照该变量工况的顺序进行写
                            dlc_list = loop_result[1]
                            # write title according to the index of the first variable
                            for _ in range(len(dlc_list)):
                                # 写工况名称
                                sheet_heatmap.cell(row=row_index-1, column=col_var+1+_, value=dlc_list[_])

                        # 写变量名称
                        if tower_num == 1:
                            v_nam = v_nam.split(',')[0].split()[-1]+'TT'
                        elif tower_num == 2:
                            v_nam = v_nam.split(',')[0].split()[-1]+'TB'

                        sheet_heatmap.cell(row=row_index, column=col_var, value=v_nam)

                        # write result
                        # maximum value for each variable
                        max_val = max(loop_result[0].values())
                        dlc_num = len(dlc_list)

                        for __ in range(dlc_num):

                            dlc_val = float(loop_result[0][dlc_list[__]])
                            sheet_heatmap.cell(row=row_index, column=col_var+1+__, value=dlc_val/max_val)

                col_var = sheet_heatmap.max_column + 2
                print('loop %s is done!' %(i+1))

            SetFormat().set_heat_map(sheet_heatmap)
            print('Heat map is done! \n')

        # ***************************** Fatigue ***********************************************
        if self.fatigue or self.fat_case or self.fat_wind:
            fat_chan_list = []

            for loop, channels in self.lp_channels.items():

                for chan in channels:

                    if 'Tower' not in chan:
                        fat_chan_list.append(chan)

            fat_chan_list = list(set(fat_chan_list))
            fat_chan_list.sort()
            # print(fat_chan_list)
            self.get_m_slope()

        if self.new_app and self.fatigue:
            print('Begin to write fatigue...')

            fatigue = table.create_sheet('Fatigue')

            for i in range(lp_num):

                print('Begin to write: ', self.loopname[i])

                # 表头：
                fatigue.cell(row=1, column=i*5+1, value=i+1)
                fatigue.cell(row=2, column=i*5+1, value=self.loopname[i])
                hyper3 = '=HYPERLINK("{}", "{}")'.format(self.fat_list[i], self.fat_list[i])
                fatigue.cell(row=3, column=i*5+1, value=hyper3)
                fatigue.cell(row=4, column=i*5+1, value='Variable')
                fatigue.cell(row=4, column=i*5+2, value='m')
                fatigue.cell(row=4, column=i*5+3, value='Value\n(KN/KNm)')
                fatigue.cell(row=4, column=i*5+4, value=1)

                # 数据区：
                lp_name  = self.loopname[i]
                channels = self.lp_channels[lp_name]  # list in order
                # print(channels)

                tow_ind = 0

                tower_num = 0

                for j in range(len(channels)):

                    ch_name = channels[j]
                    print(ch_name)

                    if 'Tower' in ch_name:
                        tower_num += 1

                    if ch_name:

                        chan_var = self.lp_chan_var[lp_name][channels[j]]  # list, len=6
                        # print(ch_name, chan_var)

                        if len(chan_var) != 6:

                            print('Warning: not 6 variables in ' + lp_name + os.sep + channels[j])

                        for k in range(6):     #for k in range(len(chan_var)):
                            # print(k)

                            # get row index for each variables
                            if 'Tower' not in ch_name:
                                row_index = 4+fat_chan_list.index(ch_name)*7+k+1
                            else:
                                row_index = 4+(len(fat_chan_list)+tow_ind)*7+k+1
                                tow_ind  += int((k+1)/6)

                            var = chan_var[k]
                            print(ch_name)
                            # ch  = ''.join([_ for _ in ch_name if _.isalpha()])
                            if 'Tower' not in ch_name:
                                m = '4.0' if ch_name not in self.var_m.keys() else self.var_m[ch_name]
                            else:
                                m = '4.0' if 'Tower' not in self.var_m.keys() else self.var_m['Tower']

                            # equivalent fatigue load
                            fat_eq = float(self.lp_var_fat[lp_name][var][m]['Total'])/1000
                            char = get_column_letter(5*i+4)
                            rat  = '=INDIRECT(ADDRESS(ROW(),COLUMN()-1))/INDIRECT(ADDRESS(ROW(),(%s$4*5-2)))' %char

                            if tower_num == 1:
                                var = var.split(',')[0].split()[-1]+'TT'
                            elif tower_num == 2:
                                var = var.split(',')[0].split()[-1]+'TB'

                            fatigue.cell(row=row_index, column=i*5+1, value=var)
                            fatigue.cell(row=row_index, column=i*5+2, value='%.1f' % float(m))
                            fatigue.cell(row=row_index, column=i*5+3, value=fat_eq)
                            fatigue.cell(row=row_index, column=i*5+4, value=rat)

                print('loop %s is done!' %(i+1))

            SetFormat().set_fatigue(fatigue)

            print('Fatigue loads is done! \n')

        # ***************************** Fatigue_wind ******************************************
        if self.new_app and self.fat_wind:

            fat_wind = table.create_sheet('Fatigue_Wind')
            print('Begin to write fatigue_wind...')
            # loops
            wind_list = []
            ws_max    = 0
            for lp, ws in self.lp_v_list.items():
                if len(ws) > ws_max:
                    ws_max    = len(ws)
                    wind_list = ws

            for i in range(lp_num):
                print('Begin to write: ', self.loopname[i])

                lp_name  = self.loopname[i]
                channels = self.lp_channels[lp_name]  # list in order
                V_case   = self.lp_V_case[lp_name]
                # ws_loop  = list(V_case.keys())        # wind speed in each loop
                # ws_loop.sort()

                ws_list = self.lp_v_list[lp_name]
                # ws_num  = len(ws_list)

                tow_ind = 0

                # 表头1：loop信息
                fat_wind.cell(1, i*14+1, i+1)
                fat_wind.cell(2, i*14+1, lp_name)
                hyper3 = '=HYPERLINK("{}", "{}")'.format(self.fat_list[i], self.fat_list[i])
                fat_wind.cell(3, i*14+1, hyper3)

                tower_num = 0
                # channels
                for j in range(len(channels)):
                    chan_name = channels[j]

                    if 'Tower' in chan_name:
                        tower_num += 1

                    if chan_name:

                        chan_var = self.lp_chan_var[lp_name][chan_name]  # var list, len=6
                        # wind speed
                        for k, ws in enumerate(ws_list):
                            # variables
                            for l in range(6):  # for k in range(len(chan_var))

                                var = chan_var[l]

                                if 'Tower' not in chan_name:
                                    chan_index = fat_chan_list.index(chan_name)
                                    row_start  = 4+(2+ws_max+1)*chan_index
                                else:
                                    row_start = 4+(2+ws_max+1)*(len(fat_chan_list)+tow_ind)
                                    if (k == (len(ws_list)-1)) and (l == 5):
                                        tow_ind += 1

                                if 'Tower' not in chan_name:
                                    m     = '4.0' if chan_name not in self.var_m.keys() else self.var_m[chan_name]
                                    m_var = 'm=4.0' if chan_name not in self.var_m.keys() else 'm=%s' %self.var_m[chan_name]
                                else:
                                    m     = '4.0' if 'Tower' not in self.var_m.keys() else self.var_m['Tower']
                                    m_var = 'm=4.0' if 'Tower' not in self.var_m.keys() else 'm=%s' %self.var_m['Tower']

                                # m and index
                                fat_wind.cell(row_start+2, i*14+1, 'm/s')
                                fat_wind.cell(row_start+2, i*14+2*(l+1), m_var)

                                if l == 0:
                                    # Mx
                                    fat_wind.cell(row_start+2, i*14+2+1, 1)
                                else:
                                    # Mx m值列
                                    col_value = '=%s%s' %(get_column_letter(i*14+2+1), row_start+2)
                                    fat_wind.cell(row_start+2, i*14+2*(l+1)+1, col_value)

                                # 数据区：风速、等效疲劳载荷
                                fat_tmp = 0

                                for case in V_case[ws]['case']:
                                    fat_tmp += self.lp_var_fat[lp_name][var][m][case]**float(m)

                                fatigue = fat_tmp**(1/float(m))/1000
                                # column letter for the ratio
                                char = get_column_letter(i*14+2*(l+1)+1)
                                rat = '=INDIRECT(ADDRESS(ROW(),COLUMN()-1))/' \
                                      'INDIRECT(ADDRESS(ROW(),((%s$%s-1)*14+2*(%i+1))))' %(char, row_start+2, l)

                                if tower_num == 1:
                                    var = var.split(',')[0].split()[-1]+'TT'
                                elif tower_num == 2:
                                    var = var.split(',')[0].split()[-1]+'TB'

                                # 表头2：变量名
                                fat_wind.cell(row_start+1, i*14+1, value='Wind Speed')
                                fat_wind.cell(row_start+1, i*14+2*(l+1), value=var)
                                # write wind
                                row_ws = row_start+2+wind_list.index(ws)+1
                                fat_wind.cell(row=row_ws, column=i*14+1,         value=ws)
                                fat_wind.cell(row=row_ws, column=i*14+2*(l+1),   value=fatigue)
                                fat_wind.cell(row=row_ws, column=i*14+2*(l+1)+1, value=rat)

                print('loop %s is done!' %(i+1))

            SetFormat().set_fatigue_wind(fat_wind)
            print('Wind fatigue is done!\n')
        # ***************************** Fatigue_wind ******************************************

        # ***************************** Fatigue_case ******************************************
        # case fatigue: (all fatigue cases)
        if self.new_app and self.fat_case:

            fat_case = table.create_sheet('Fatigue_Case')

            # col_start = 1
            # row_start = fat_case.max_row  # 1
            row_start = 1
            # print(len(fat_chan_list))

            print('Begin to write fatigue_case...')
            # loop cycle
            for i in range(lp_num):
                # print(i)

                print('Begin to write: ', self.loopname[i])

                lp_name  = self.loopname[i]
                channels = self.lp_channels[lp_name]  # list in order
                # print(channels)
                # col_len = len(self.lp_var_fat[lp_name]) + 3

                # 工况名、路径链接
                dlc_list = [k for k in self.lp_dlc_path[lp_name].keys()]
                dlc_list.remove('Total')
                dlc_list.sort()
                dlc_path = self.lp_dlc_path[lp_name]

                # 添加表头：
                # 当loop变量个数不一样，i*col_len+1 的方式不适用，表头写入移到数据循环后
                # 当取3叶片最大时，len(self.lp_var_fat[lp_name])确定列数的方式也不适用
                fat_case.cell(row=row_start, column=1, value=i+1)
                fat_case.cell(row=row_start+1, column=1, value=self.loopname[i])
                hyper3 = '=HYPERLINK("{}", "{}")'.format(self.fat_list[i], self.fat_list[i])
                fat_case.cell(row=row_start+2, column=1, value=hyper3)

                # 添加变量名
                # fat_case.cell(row=4, column=col_start, value='m')
                fat_case.cell(row=row_start+4, column=1, value='DLC-Wind Speed')
                fat_case.cell(row=row_start+5, column=1, value='RunName')
                fat_case.cell(row=row_start+6, column=1, value='Path')
                # dlc cycle
                for p in range(len(dlc_list)):
                    hyper4 = '=HYPERLINK("{}", "{}")'.format(dlc_path[dlc_list[p]], dlc_path[dlc_list[p]])
                    fat_case.cell(row=row_start+5, column=2+p+1, value=dlc_list[p])
                    fat_case.cell(row=row_start+6, column=2+p+1, value=hyper4)

                tower_num = 0
                # channel cycle
                for j in range(len(channels)):
                    chan_name = channels[j]

                    if chan_name:
                        chan_var = self.lp_chan_var[lp_name][chan_name]  # list, len=6
                        # print(len(chan_var))

                        if len(chan_var) != 6:
                            print('Warning: not 6 variables in ' + lp_name + os.sep + chan_name)

                        chan_num = len(fat_chan_list) + 2

                        if 'Tower' not in chan_name:
                            row_start = (chan_num*13+7)*i+7+13*fat_chan_list.index(chan_name)
                        else:
                            row_start  = (chan_num*13+7)*i+7+13*(len(fat_chan_list)+tower_num)
                            tower_num += 1
                            # print(row_start)

                        # variable cycle
                        for k in range(6):
                            var = chan_var[k]
                            # print(var)

                            if 'Tower' not in chan_name:
                                m     = '4.0' if chan_name not in self.var_m.keys() else self.var_m[chan_name]
                                m_var = 'm=4.0' if chan_name not in self.var_m.keys() else 'm=%s' %self.var_m[chan_name]
                            else:
                                m     = '4.0' if 'Tower' not in self.var_m.keys() else self.var_m['Tower']
                                m_var = 'm=4.0' if 'Tower' not in self.var_m.keys() else 'm=%s' %self.var_m['Tower']

                            # Mx的index
                            if k == 0:
                                fat_case.cell(row=row_start+2*k+1, column=2, value=1)
                            else:
                                fat_case.cell(row=row_start+2*k+1, column=2, value='=B$%s' %(row_start+1))

                            # 各工况疲劳载荷：
                            for p in range(len(dlc_list)):
                                fat_eq = self.lp_var_fat[lp_name][var][m][dlc_list[p]]/1000
                                fat_case.cell(row=row_start+2*k+2, column=2+p+1, value=fat_eq)

                                rat = '=INDIRECT(ADDRESS(ROW()+1,COLUMN()))/' \
                                      'INDIRECT(ADDRESS((%s+($B%s-%s)*(7+13*%s)),COLUMN()))' \
                                      %(row_start+2*k+2,row_start+2*k+1, i+1, chan_num)
                                fat_case.cell(row=row_start+2*k+1, column=2+p+1, value=rat)

                            # 添加表头变量名信息：
                            if tower_num == 1:
                                var = var.split(',')[0].split()[-1]+'TT'
                            elif tower_num == 2:
                                var = var.split(',')[0].split()[-1]+'TB'

                            fat_case.cell(row=row_start+2*k+1, column=1, value=var)
                            fat_case.cell(row=row_start+2*k+2, column=2, value=m_var)

                # 记录当前loop所占列数：
                row_start = fat_case.max_row+2
                # print(row_start)
                print('loop %s is done!' % (i+1))

            # 设置格式
            SetFormat().set_fatigue_case(fat_case)
            print('Case fatigue is done!\n')
        # ***************************** Fatigue_case ******************************************

        # ***************************** AEP for DLC12 *****************************************
        if self.aep_dlc:

            sheet_aep = table.create_sheet('AEP_DLC12')

            powercurve(sheet_aep, self.loopname, self.dlc_list, self.lct_list)

            SetFormat().set_aep(sheet_aep)
        # ***************************** AEP for DLC12 *****************************************

        # ********************************* LDD ***********************************************
        if self.ldd:

            print('Begin to create ldd...')
            sheet_name = table.sheetnames

            if 'LDD' in sheet_name:
                sheet = table.get_sheet_by_name('LDD')
                table.remove(sheet)
            sheet_ldd = table.create_sheet('LDD')

            col_start = 1

            # write result
            for i in range(lp_num):

                lp_name = self.loopname[i]
                print('Begin to write %s' %lp_name)
                
                # channel: project file
                chan_pj = self.loop_chan_pj[lp_name]
                # print('chan_pj:',chan_pj)
                chan_list = list(chan_pj.keys())
                chan_list.sort()
                # print(chan_list)

                chan_row_max = 0
                chan_col_max = 0

                row_start = 1
                # path
                sheet_ldd.cell(row=row_start,   column=col_start, value=lp_name)
                sheet_ldd.cell(row=row_start+1, column=col_start, value=self.ldd_list[i])
                
                for chan in chan_list:
                    
                    pj_list = chan_pj[chan]
                    # pj_list.sort()
                    pj_num  = len(pj_list)
                    
                    # record the variable number for each channel
                    chan_var_num = 0

                    for j in range(pj_num):

                        pj_name  = pj_list[j]

                        var_list = list(self.loop_prj_var[lp_name][pj_name])
                        var_list.sort()
                        var_len  = len(var_list)

                        # record the variable of the previous project
                        if j > 0:
                            chan_var_num += len(self.loop_prj_var[lp_name][pj_list[j-1]])

                        # record the max column for each loop
                        if pj_num*var_len > chan_col_max:
                            chan_col_max = pj_num*var_len*2

                        for k in range(var_len):

                            var      = var_list[k]
                            data     = self.lp_ch_pj_var[lp_name][chan][pj_name][var]
                            data_len = len(data[:,0])
                            if data_len > chan_row_max:
                                chan_row_max = data_len

                            # header
                            if pj_num == 1:
                                sheet_ldd.cell(row=row_start+3, column=col_start+2*chan_var_num+2*k,   value=var)
                                sheet_ldd.cell(row=row_start+3, column=col_start+2*chan_var_num+2*k+1, value='ProbDensity')
                            else:
                                sheet_ldd.cell(row=row_start+3, column=col_start+2*chan_var_num+2*k+1, value=str(pj_name))
                                sheet_ldd.cell(row=row_start+4, column=col_start+2*chan_var_num+2*k,   value=var)
                                sheet_ldd.cell(row=row_start+4, column=col_start+2*chan_var_num+2*k+1, value='ProbDensity')

                            # data
                            for l in range(data_len):
                                if pj_num == 1:
                                    sheet_ldd.cell(row=row_start+3+l+1, column=col_start+2*chan_var_num+2*k,   value=data[l, 0])
                                    sheet_ldd.cell(row=row_start+3+l+1, column=col_start+2*chan_var_num+2*k+1, value=data[l, 4])
                                else:
                                    sheet_ldd.cell(row=row_start+4+l+1, column=col_start+2*chan_var_num+2*k,   value=data[l, 0])
                                    sheet_ldd.cell(row=row_start+4+l+1, column=col_start+2*chan_var_num+2*k+1, value=data[l, 4])

                    row_start += chan_row_max+2

                col_start += chan_col_max+1
                # print(row_start)
                print('loop %s is done!' % (i+1))

            # set format
            SetFormat().set_ldd(sheet_ldd)
            print('ldd is done!\n')
        # ********************************* LDD ***********************************************

        # ********************************* LRD ***********************************************
        if self.lrd:

            print('Begin to write lrd...')
            sheet_name = table.sheetnames

            if 'LRD' in sheet_name:
                sheet = table.get_sheet_by_name('LRD')
                table.remove(sheet)
            sheet_lrd = table.create_sheet('LRD')

            col_start = 1

            # write result
            for i in range(lp_num):

                lp_name = self.loopname[i]
                print('Begin to write %s',lp_name)
                # channel: project file
                chan_pj = self.loop_chan_pj_lrd[lp_name]
                # print('chan_pj:',chan_pj)
                chan_list = list(chan_pj.keys())
                chan_list.sort()
                # print(chan_list)

                chan_row_max = 0
                chan_col_max = 0

                row_start = 1
                # path
                sheet_lrd.cell(row=row_start,   column=col_start, value=lp_name)
                sheet_lrd.cell(row=row_start+1, column=col_start, value=self.lrd_list[i])

                for chan in chan_list:

                    pj_list = chan_pj[chan]
                    # pj_list.sort()
                    pj_num  = len(pj_list)

                    # record the variable number for each channel
                    chan_var_num = 0

                    for j in range(pj_num):

                        pj_name  = pj_list[j]

                        var_list = list(self.loop_prj_var_lrd[lp_name][pj_name])
                        var_list.sort()
                        var_len  = len(var_list)

                        # record the variable of the previous project
                        if j > 0:
                            chan_var_num += len(self.loop_prj_var_lrd[lp_name][pj_list[j-1]])

                        # record the max column for each loop
                        if pj_num*var_len > chan_col_max:
                            chan_col_max = pj_num*var_len*2

                        for k in range(var_len):

                            var      = var_list[k]
                            data     = self.lp_ch_pj_var_lrd[lp_name][chan][pj_name][var]
                            data_len = len(data[:,0])
                            if data_len > chan_row_max:
                                chan_row_max = data_len

                            # header
                            if pj_num == 1:
                                sheet_lrd.cell(row=row_start+3, column=col_start+2*chan_var_num+2*k, value=var)
                                sheet_lrd.cell(row=row_start+3, column=col_start+2*chan_var_num+2*k, value='Revs')
                            else:
                                sheet_lrd.cell(row=row_start+3, column=col_start+2*chan_var_num+2*k+1, value=str(pj_name))
                                sheet_lrd.cell(row=row_start+4, column=col_start+2*chan_var_num+2*k,   value=var)
                                sheet_lrd.cell(row=row_start+4, column=col_start+2*chan_var_num+2*k+1, value='Revs')

                            # data
                            for l in range(data_len):
                                if pj_num == 1:
                                    sheet_lrd.cell(row=row_start+3+l+1, column=col_start+2*chan_var_num+2*k,   value=data[l, 0])
                                    sheet_lrd.cell(row=row_start+3+l+1, column=col_start+2*chan_var_num+2*k+1, value=data[l, 1])
                                else:
                                    sheet_lrd.cell(row=row_start+4+l+1, column=col_start+2*chan_var_num+2*k,   value=data[l, 0])
                                    sheet_lrd.cell(row=row_start+4+l+1, column=col_start+2*chan_var_num+2*k+1, value=data[l, 1])

                    row_start += chan_row_max+2

                col_start += chan_col_max+1
                # print(row_start)
                print('loop %s is done!' % (i+1))

            # set format
            SetFormat().set_lrd(sheet_lrd)
            print('lrd is done!\n')
        # ********************************* LRD ***********************************************

        # ********************************* Load Summary **************************************
        lp_num = len(self.loopname)
        print('Begin to write summary...')
        if self.mbl:
            for i in range(lp_num):
                print('Begin to write main bearing life %s' %self.loopname[i])
                main_bearing_life(self.mbl_list[i], table, self.loopname[i], i)
                print('loop %s is done!' %(i+1))

        if self.pbl:
            for i in range(lp_num):
                print('Begin to write pitch bearing life %s' % self.loopname[i])
                pitch_bearing_life(self.pbl_list[i], table, self.loopname[i], i)
                print('loop %s is done!' % (i+1))

        if self.get:
            for i in range(lp_num):
                print('Begin to write gearbox equivalent torque %s' % self.loopname[i])
                gearbox_eq_torque(self.mbl_list[i], table, self.loopname[i], i)
                print('loop %s is done!' % (i+1))

        if 'Load Summary' in table.sheetnames:
            sheet_ls = table.get_sheet_by_name('Load Summary')
            SetFormat().set_load_summary(sheet_ls)

        print('load summary is done!')
        # ********************************* Load Summary **************************************

        # **************************** Save excel *********************************************
        table_path = self.out_path + os.sep + self.exl_name + '_' + self.date + ".xlsx"

        # 以下命令似乎不起作用
        # row = sheet_ultimate.row_dimensions[4]
        # row.Font = Font(name='Microsoft Ya Hei', size=9)

        try:
            table.save(table_path)
        except Exception:
            table_path = self.out_path + os.sep + self.exl_name + '_' + self.date + "_1" + ".xlsx"
            table.save(table_path)

        print('Result is written to: ' + table_path)
        print('Load summary table is done!')
        # **************************** Save excel *********************************************

if __name__ == '__main__':

    input = r"E:\00_Tools_Dev\02_LoadSummaryTable\Load Summary Template.xlsx"
    path  = r'E:\00_Tools_Dev\02_LoadSummaryTable\Test_Result'
    excel = 'Test'

    excel_operation(excel_input=input,
                    output_path=path,
                    excel_name=excel,
                    ultimate=False,
                    fatigue=False,
                    heat_map=False,
                    fat_case=False,
                    fat_wind=False,
                    ldd=True,
                    lrd=True,
                    aep_dlc=False,
                    main_all=True,
                    new_app=True)