# -*- coding: utf-8 -*-
# @Time    : 2019/11/30 14:26
# @Author  : CE
# @File    : excel_format.py
'''
2019.12.18: 极限载荷根据实际row和column来进行设置
'''

from openpyxl.styles import Side, Border, Alignment, Font, colors, Color, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormatObject, ColorScale, Rule, CellIsRule
import numpy as np
# from openpyxl.formatting import Rule

class SetFormat:

    def __init__(self, loop_num=10):

        # self.ultimate = ultimate
        # self.heat_map = heat_map
        self.loop_num = loop_num     # 默认设置10个loop

    # 定义边框
    def set_boader(self, sheet, area, each=True):
        '''
        定义边框
        :param sheet: 表格名称
        :param area: 区域（area[0]:起始行；area[1]:终止行；area[2]:起始列；area[3]:终止列）
        :return:
        '''

        boader1 = Side(style='medium', color=colors.BLACK)

        cell_start1 = str(get_column_letter(area[2])) + str(area[0])
        cell_end1   = str(get_column_letter(area[3])) + str(area[0])
        cell_start2 = str(get_column_letter(area[2])) + str(area[1])
        cell_end2   = str(get_column_letter(area[3])) + str(area[1])

        if area[0] == area[1] and each:
            # 该判断使每个单元格都加边框；如果去掉该判断，那么所有行添加一个边框

            for row in sheet[cell_start1: cell_end1]:
                for cell in row:
                    cell.border = Border(left=boader1, right=boader1,top=boader1,bottom=boader1)

        else:

            # set upper boader
            for row in sheet[cell_start1: cell_end1]:
                # sheet返回元组
                for cell in row:
                    # cell.border.bottom.style指默认的格式
                    cell.border = Border(top   =boader1,
                                         bottom=Side(style=cell.border.bottom.style, color=colors.BLACK),
                                         left  =Side(style=cell.border.left.style,   color=colors.BLACK),
                                         right =Side(style=cell.border.right.style,  color=colors.BLACK))


            # set bottom boader
            for row in sheet[cell_start2: cell_end2]:
                for cell in row:
                    cell.border = Border(top   =Side(style=cell.border.top.style,   color=colors.BLACK),
                                         bottom=boader1,
                                         left  =Side(style=cell.border.left.style,  color=colors.BLACK),
                                         right =Side(style=cell.border.right.style, color=colors.BLACK))

            # set left boader
            for col in sheet[cell_start1: cell_start2]:
                for cell in col:
                    cell.border = Border(top   =Side(style=cell.border.top.style,    color=colors.BLACK),
                                         bottom=Side(style=cell.border.bottom.style, color=colors.BLACK),
                                         left  =boader1,
                                         right =Side(style=cell.border.right.style,  color=colors.BLACK))

            # set right boader
            for col in sheet[cell_end1: cell_end2]:
                for cell in col:
                    cell.border = Border(top   =Side(style=cell.border.top.style,    color=colors.BLACK),
                                         bottom=Side(style=cell.border.bottom.style, color=colors.BLACK),
                                         left  =Side(style=cell.border.left.style,   color=colors.BLACK),
                                         right =boader1)

    # ------------------------------------------------------------------------------
    # 定义ultimate表格的格式
    def set_ultimate(self, ultimate):

        # wb = openpyxl.Workbook()
        # ultimate = wb.create_sheet('test')

        font1 = Font(name='Microsoft Ya Hei', size=11, bold=True)
        font2 = Font(name='Microsoft Ya Hei', size=9)
        font3 = Font(name='Microsoft Ya Hei', size=9, underline='single', color=colors.BLUE, bold=True)
        font4 = Font(name='Microsoft Ya Hei', size=9, bold=True)
        font5 = Font(name='Microsoft Ya Hei', size=9, underline='single', color=colors.BLUE)

        align1 = Alignment(horizontal='center', vertical='center')
        align2 = Alignment(horizontal='left', vertical='center')
        align3 = Alignment(horizontal='center', vertical='center', wrapText=True)

        redfill    = PatternFill(start_color='F5A9BC', end_color='F5A9BC', fill_type='solid')
        greenfill  = PatternFill(start_color='CEF6E3', end_color='CEF6E3', fill_type='solid')
        yellowfill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')  # FFF68F

        # 实际的loop个数以及channel的个数
        num_loop = int(np.ceil(ultimate.max_column/6))
        num_chan = int(np.ceil((ultimate.max_row-4)/8))
        # print(num_loop, num_chan)

        # 设置前3行的格式
        for i in range(num_loop):
            ultimate.merge_cells(start_row=1, start_column=i*6+1, end_row=1, end_column=i*6+5)
            ultimate.cell(row=1, column=i*6+1).font      = font1
            ultimate.cell(row=1, column=i*6+1).alignment = align1

            ultimate.merge_cells(start_row=2, start_column=i*6+1, end_row=2, end_column=i*6+5)
            ultimate.cell(row=2, column=i*6+1).font      = font1
            ultimate.cell(row=2, column=i*6+1).alignment = align2

            ultimate.merge_cells(start_row=3, start_column=i*6+1, end_row=3, end_column=i*6+5)
            ultimate.cell(row=3, column=i*6+1).font      = font3
            ultimate.cell(row=3, column=i*6+1).alignment = align2

        # 设置第4行的格式
        # ultimate.row_dimensions[4].font = font4     #该行似乎不起作用
        for i in range(1, num_loop*6+1):
            ultimate.cell(row=4, column=i).font      = font4
            ultimate.cell(row=4, column=i).alignment = align3

        ultimate.row_dimensions[4].height = 25

        # row_num = 150
        row_num = ultimate.max_row
        col_num = ultimate.max_column

        # 设置所有数据的格式
        for row in ultimate.iter_rows(min_row=5, max_col=col_num, max_row=row_num):
            for cell in row:
                cell.font      = font2
                cell.alignment = align2

        for i in range(num_loop):
            for j in range(5, row_num+1):
                # 设置所有path的格式，下划线+蓝色字体
                ultimate.cell(row=j, column=i*6+4).font = font5
                # 设置ratio的格式，保存两位小数
                ultimate.cell(row=j, column=i*6+5).number_format = '0.00'
                ultimate.cell(row=j, column=i*6+3).number_format = '0.00'
                # 设置value和ratio列居中
                ultimate.cell(row=j, column=i*6+1).alignment = Alignment(horizontal='right')
                ultimate.cell(row=j, column=i*6+2).alignment = Alignment(horizontal='right')
                ultimate.cell(row=j, column=i*6+3).alignment = Alignment(horizontal='right')
                ultimate.cell(row=j, column=i*6+5).alignment = Alignment(horizontal='center')

        # 设置每列的宽度
        for i in range(num_loop):
            ultimate.column_dimensions[get_column_letter(i*6+1)].width = 10
            ultimate.column_dimensions[get_column_letter(i*6+2)].width = 20
            ultimate.column_dimensions[get_column_letter(i*6+3)].width = 10
            ultimate.column_dimensions[get_column_letter(i*6+4)].width = 5
            ultimate.column_dimensions[get_column_letter(i*6+5)].width = 5
            ultimate.column_dimensions[get_column_letter(i*6+6)].width = 3

            # header format
            self.set_boader(ultimate, [1, 1, i*6+1, i*6+5])
            self.set_boader(ultimate, [2, 2, i*6+1, i*6+5])
            self.set_boader(ultimate, [3, 3, i*6+1, i*6+5])
            self.set_boader(ultimate, [4, 4, i*6+1, i*6+5])

        row_index = 5
        while row_index<row_num:
            br_flag = [True if ultimate.cell(row_index, i*6+1).value=='MxBR_Max' else False for i in range(num_loop)]
            if br_flag[0]:
                # hide
                ultimate.row_dimensions.group(row_index+6, row_index+9, hidden=True)
                # border
                for i in range(num_loop):
                    if ultimate.cell(row_index, column=i*6+1).value:
                        self.set_boader(ultimate, [row_index, row_index+5, i*6+1, i*6+5])
                        self.set_boader(ultimate, [row_index+6, row_index+9, i*6+1, i*6+5])
                    else:
                        continue
                    # ratio format
                    if ultimate.cell(row_index, column=i*6+5).value:
                        cell_start = get_column_letter(5+i*6)+str(row_index)
                        # 默认为8个变量
                        cell_end = get_column_letter(5+i*6)+str(row_index+9)

                        cells = cell_start+':'+cell_end
                        # print(cells)
                        ultimate.conditional_formatting.add(cells, CellIsRule(operator='lessThan', formula=['1.0'],
                                                                              stopIfTrue=True, fill=greenfill))
                        ultimate.conditional_formatting.add(cells,
                                                            CellIsRule(operator='between', formula=['1.0', '1.03'],
                                                                       stopIfTrue=True, fill=yellowfill))
                        ultimate.conditional_formatting.add(cells,
                                                            CellIsRule(operator='greaterThan', formula=['1.03'],
                                                                       stopIfTrue=True, fill=redfill))

                row_index += 11
            else:
                # hide
                ultimate.row_dimensions.group(row_index+4, row_index+7, hidden=True)

                for i in range(num_loop):
                    # border
                    if ultimate.cell(row_index, column=i*6+1).value:
                        self.set_boader(ultimate, [row_index, row_index+3, i*6+1, i*6+5])
                        self.set_boader(ultimate, [row_index+4, row_index+7, i*6+1, i*6+5])
                    else:
                        continue
                    # ratio format
                    if ultimate.cell(row_index, column=i*6+5).value:
                        cell_start = get_column_letter(5+i*6)+str(row_index)
                        # 默认为8个变量
                        cell_end = get_column_letter(5+i*6)+str(row_index+7)

                        cells = cell_start+':'+cell_end
                        # print(cells)
                        ultimate.conditional_formatting.add(cells, CellIsRule(operator='lessThan', formula=['1.0'],
                                                                              stopIfTrue=True, fill=greenfill))
                        ultimate.conditional_formatting.add(cells,
                                                            CellIsRule(operator='between', formula=['1.0', '1.03'],
                                                                       stopIfTrue=True, fill=yellowfill))
                        ultimate.conditional_formatting.add(cells,
                                                            CellIsRule(operator='greaterThan', formula=['1.03'],
                                                                       stopIfTrue=True, fill=redfill))

                row_index += 9

    # ------------------------------------------------------------------------------
    # 定义heatmap表格的格式
    def set_heat_map(self, heat_map):

        font1 = Font(name='Microsoft Ya Hei', size=11, bold=True)
        font2 = Font(name='Microsoft Ya Hei', size=9, bold=True)
        font3 = Font(name='Microsoft Ya Hei', size=9)

        align1 = Alignment(horizontal='center', vertical='center')
        align2 = Alignment(horizontal='left', vertical='center')

        # 定义色阶
        first = FormatObject(type='max')
        last  = FormatObject(type='min')
        mid   = FormatObject(type='percentile', val=50)

        color_list = [Color('63BE7B'), Color('FFEB84'), Color('F8696B')]
        cs         = ColorScale(cfvo=[first, mid, last], color=color_list)
        rule       = Rule(type='colorScale', colorScale=cs)

        col_max = heat_map.max_column
        # print(col_max)

        # 记录每个loop的第一列和最后一列
        loop_pos = [1]

        for i in range(2, col_max+1):

            cell1 = heat_map.cell(row=4, column=i)
            cell2 = heat_map.cell(row=4, column=i+1)

            # 记录每个loop的结束列
            if cell1.value and (not cell2.value):
                loop_pos.append(i)

            # 记录每个loop的开始列
            if (not cell1.value) and cell2:
                loop_pos.append(i+1)

        # print(loop_pos)

        loop_sta = loop_pos[0::2]
        loop_end = loop_pos[1::2]
        loop_num = len(loop_sta)
        # print(len(loop_sta), len(loop_end))

        # 设置第一行的字体
        for i in loop_sta:

            heat_map.cell(row=1, column=i).font = font1
            heat_map.cell(row=1, column=i).alignment = align1

        # 合并每个loop的第一行
        for i in range(loop_num):

            col_sta = loop_sta[i]
            col_end = loop_end[i]
            # print(col_sta, col_end)

            heat_map.merge_cells(start_row=1, start_column=col_sta, end_row=1, end_column=col_end)

        # 设置列宽
        for i in range(loop_num):

            # set up the width of variable column
            heat_map.column_dimensions[get_column_letter(loop_sta[i])].width = 7

            # set up the width of dlc column
            for j in range(loop_sta[i]+1, loop_end[i]+1):
                # print(get_column_letter(j))

                heat_map.column_dimensions[get_column_letter(j)].width = 7

            # set up the width of blank column
            heat_map.column_dimensions[get_column_letter(loop_end[i]+1)].width = 3

        row_num = heat_map.max_row

        # 设置变量名的格式
        for i in range(3, row_num+1):

            for j in range(loop_num):

                cell = heat_map.cell(row=i, column=loop_sta[j])

                if cell:

                    cell.font      = font2
                    cell.alignment = align2

        # 设置工况名称的格式
        for i in range(3, row_num+1):

            if (i-3)%10 == 0:

                for j in range(1, loop_end[-1]+1):

                    heat_map.cell(row=i, column=j).font      = font2
                    heat_map.cell(row=i, column=j).alignment = align2

        # 设置数据的格式
        for i in range(3, row_num+1):

            if (i-3)%10 != 0:

                for j in range(loop_num):

                    for k in range(loop_sta[j], loop_end[j]+1):

                        cell = heat_map.cell(row=i, column=k)

                        if cell:

                            cell.font          = font3
                            cell.alignment     = align1
                            cell.number_format = '0.00'

        # 第一行添加边框
        for i in range(loop_num):

            self.set_boader(heat_map, [1, 1, loop_sta[i], loop_end[i]])

        for i in range(3, row_num):

            # 工况行添加边框
            if (i-3)%10 == 0:

                for j in range(loop_num):
                    # to check the first value is empty or not
                    if heat_map.cell(row=i+1, column=loop_sta[j]+1).value:

                        # dlc row
                        self.set_boader(heat_map, [i,i,loop_sta[j]+1,loop_end[j]])
                        # variable column
                        self.set_boader(heat_map, [i+1,i+8,loop_sta[j],loop_sta[j]])
                        # data area
                        self.set_boader(heat_map, [i+1,i+8,loop_sta[j]+1,loop_end[j]])

                        for k in range(1, 9):

                            cell_start = get_column_letter(loop_sta[j]+1) + str(i+k)
                            cell_end   = get_column_letter(loop_end[j]) + str(i+k)

                            heat_map.conditional_formatting.add(cell_start+':'+cell_end, rule)

    # ------------------------------------------------------------------------------
    # 定义fatigue表格的格式
    def set_fatigue(self, sheet_fatigue):

        font1 = Font(name='Microsoft Ya Hei', size=11, bold=True)
        font2 = Font(name='Microsoft Ya Hei', size=9)
        font3 = Font(name='Microsoft Ya Hei', size=9, underline='single', color=colors.BLUE, bold=True)
        font4 = Font(name='Microsoft Ya Hei', size=9, bold=True)

        align1 = Alignment(horizontal='center', vertical='center')
        align2 = Alignment(horizontal='left',   vertical='center')
        align3 = Alignment(horizontal='center', vertical='center', wrapText=True)
        align4 = Alignment(horizontal='right',  vertical='center')

        redfill    = PatternFill(start_color='F5A9BC', end_color='F5A9BC', fill_type='solid')
        greenfill  = PatternFill(start_color='CEF6E3', end_color='CEF6E3', fill_type='solid')
        yellowfill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')  # FFF68F

        sheet_fatigue.row_dimensions[4].height = 25

        # print((sheet_fatigue.max_column + 1) / 5)
        lp_num = int(np.ceil(sheet_fatigue.max_column/5))
        # for i in range(self.loop_num):
        for i in range(lp_num):

            # 设置前3行的格式
            sheet_fatigue.merge_cells(start_row=1, start_column=i*5+1, end_row=1, end_column=i*5+4)
            sheet_fatigue.cell(row=1, column=i*5+1).font      = font1
            sheet_fatigue.cell(row=1, column=i*5+1).alignment = align1

            sheet_fatigue.merge_cells(start_row=2, start_column=i*5+1, end_row=2, end_column=i*5+4)
            sheet_fatigue.cell(row=2, column=i*5+1).font      = font1
            sheet_fatigue.cell(row=2, column=i*5+1).alignment = align2
            
            sheet_fatigue.merge_cells(start_row=3, start_column=i*5+1, end_row=3, end_column=i*5+4)
            sheet_fatigue.cell(row=3, column=i*5+1).font      = font3
            sheet_fatigue.cell(row=3, column=i*5+1).alignment = align2
            
            # 设置第4行的格式
            for j in range(i*5+1, i*5+5):
                sheet_fatigue.cell(row=4, column=j).font      = font4
                sheet_fatigue.cell(row=4, column=j).alignment = align3

            # 设置列宽
            sheet_fatigue.column_dimensions[get_column_letter(i*5+1)].width = 9
            sheet_fatigue.column_dimensions[get_column_letter(i*5+2)].width = 5
            sheet_fatigue.column_dimensions[get_column_letter(i*5+3)].width = 9
            sheet_fatigue.column_dimensions[get_column_letter(i*5+4)].width = 5
            sheet_fatigue.column_dimensions[get_column_letter(i*5+5)].width = 3

            # 记录每个loop最大行数、channel个数
            for jj in range(5, sheet_fatigue.max_row+1):

                if sheet_fatigue.cell(row=jj, column=i*5+4).value:

                    col_max_row = jj

            num_chan = int(np.ceil((col_max_row-4)/7))  # 默认每个channel下存有6个变量：Mx~Fz

            # 设置第5行起的数据格式
            for row in sheet_fatigue.iter_rows(min_col=i*5+1, min_row=5, max_col=5*i+4, max_row=col_max_row):

                for cell in row:

                    cell.font      = font2
                    cell.alignment = align1

            for j in range(5, col_max_row+1):

                # data
                if sheet_fatigue.cell(row=j, column=i*5+4).value:
                    # set number format for value
                    sheet_fatigue.cell(row=j, column=i*5+3).number_format = '0.00'
                    # set ratio format, '0.00'
                    sheet_fatigue.cell(row=j, column=i*5+4).number_format = '0.00'
                    # set alignment for m/value/ratio
                    # sheet_fatigue.cell(row=j, column=i*5+2).alignment = Alignment(horizontal='center')
                    sheet_fatigue.cell(row=j, column=i*5+3).alignment = Alignment(horizontal='right')
                    # sheet_fatigue.cell(row=j, column=i*5+4).alignment = Alignment(horizontal='center')

            # 添加表头边框
            if sheet_fatigue.cell(row=1, column=i*5+1).value:

                self.set_boader(sheet_fatigue, [1, 1, i*5+1, i*5+4])
                self.set_boader(sheet_fatigue, [2, 2, i*5+1, i*5+4])
                self.set_boader(sheet_fatigue, [3, 3, i*5+1, i*5+4])
                self.set_boader(sheet_fatigue, [4, 4, i*5+1, i*5+4])

            # 添加数据区边框
            for j in range(num_chan):

                if sheet_fatigue.cell(row=j*7+5, column=i*5+1).value:
                    self.set_boader(sheet_fatigue, [j*7+5,   j*7+5+2, i*5+1, i*5+4])
                    self.set_boader(sheet_fatigue, [j*7+5+3, j*7+5+5, i*5+1, i*5+4])
                else:
                    continue

            # 添加ratio颜色
            for i in range(1, lp_num):

                for j in range(num_chan):

                    # 判断是否有值
                    if sheet_fatigue.cell(row=j*7+5, column=i*5+4).value:

                        cell_start = get_column_letter(4+i*5) + str(j*7+5)
                        # 默认为6个变量
                        cell_end   = get_column_letter(4+i*5) + str(j*7+5+5)

                        cells      = cell_start + ':' + cell_end
                        # print(cells)
                        sheet_fatigue.conditional_formatting.add(cells, CellIsRule(operator='lessThan', formula=['1.0'],
                                                                                   stopIfTrue=True, fill=greenfill))
                        sheet_fatigue.conditional_formatting.add(cells, CellIsRule(operator='between', formula=['1.0', '1.03'],
                                                                                   stopIfTrue=True, fill=yellowfill))
                        sheet_fatigue.conditional_formatting.add(cells, CellIsRule(operator='greaterThan', formula=['1.03'],
                                                                                   stopIfTrue=True, fill=redfill))

        # 隐藏Fx, Fy, Fz
        num_chan_max = int(np.ceil((sheet_fatigue.max_row-4)/7))   # 默认每个channel下存有6个变量：Mx~Fz

        for i in range(int(num_chan_max)):

            sheet_fatigue.row_dimensions.group(4+i*7+4, 4+7*(i+1)-1, hidden=True)

    # ------------------------------------------------------------------------------
    # 定义wind fatigue表格的格式
    def set_fatigue_wind(self, fat_wind):

        font1 = Font(name='Microsoft Ya Hei', size=11, bold=True)
        font2 = Font(name='Microsoft Ya Hei', size=9)
        font3 = Font(name='Microsoft Ya Hei', size=10, underline='single', color=colors.BLUE, bold=True)
        font4 = Font(name='Microsoft Ya Hei', size=9, bold=True)

        align1 = Alignment(horizontal='center', vertical='center')
        align2 = Alignment(horizontal='left', vertical='center')
        align3 = Alignment(horizontal='center', vertical='center', wrapText=True)
        align4 = Alignment(horizontal='right', vertical='center')

        redfill    = PatternFill(start_color='F5A9BC', end_color='F5A9BC', fill_type='solid')
        greenfill  = PatternFill(start_color='CEF6E3', end_color='CEF6E3', fill_type='solid')
        yellowfill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')  # FFF68F
        grayfill   = PatternFill('solid', fgColor='808080')

        lp_num = int(np.ceil((fat_wind.max_column)/14))
        # ch_num = int((fat_wind-4+1)/26)                       #有可能风速只有24m/s
        row_max = fat_wind.max_row
        ch_num  = 0
        ch_row  = []
        for i in range(5, row_max):
            if fat_wind.cell(i, 1).value == 'Wind Speed':
                ch_row.append(i)
                ch_num += 1

        wind_num = ch_row[1]-ch_row[0] if len(ch_row)>1 else row_max-ch_row[0]

        for i in range(lp_num):
            # 设置前3行的格式
            fat_wind.merge_cells(start_row=1, start_column=i*14+1, end_row=1, end_column=i*14+13)
            fat_wind.cell(row=1, column=i*14+1).font = font1
            fat_wind.cell(row=1, column=i*14+1).alignment = align1

            fat_wind.merge_cells(start_row=2, start_column=i*14+1, end_row=2, end_column=i*14+13)
            fat_wind.cell(row=2, column=i*14+1).font = font1
            fat_wind.cell(row=2, column=i*14+1).alignment = align1

            fat_wind.merge_cells(start_row=3, start_column=i*14+1, end_row=3, end_column=i*14+13)
            fat_wind.cell(row=3, column=i*14+1).font = font3
            fat_wind.cell(row=3, column=i*14+1).alignment = align1

            # 设置格式
            for j in range(i*14+1, i*14+14):
                # 设置数据区格式
                for k in range(ch_num):
                    # 设置变量名及单位的格式
                    row = 4+k*wind_num
                    fat_wind.cell(row=row+1, column=j).font = font4
                    fat_wind.cell(row=row+1, column=j).alignment = align3
                    fat_wind.cell(row=row+2, column=j).font = font4
                    fat_wind.cell(row=row+2, column=j).alignment = align3

                    # 设置数据行的格式：
                    row = 4+k*wind_num+2
                    for l in range(1, wind_num-2):

                        cell = fat_wind.cell(row=row+l, column=j)
                        cell.font = font2
                        cell.alignment = align1

                        if j%2 == 1 and j != i*14+1:
                            cell.number_format = '.00'
                            cell.alignment = align1
                        elif j%2 == 0:
                            cell.number_format = '.00'
                            cell.alignment = align4
                        elif j != i*14+1:
                            cell.alignment = align1

            # 设置列宽
            for j in range(i*14+1, i*14+14):
                # 每个loop的第一列
                if j == i*14+1:
                    fat_wind.column_dimensions[get_column_letter(j)].width = 6
                # ratio列
                elif j%2 == 1 and j>1:
                    fat_wind.column_dimensions[get_column_letter(j)].width = 5
                # m值列
                elif j%2 == 0:
                    fat_wind.column_dimensions[get_column_letter(j)].width = 9

            # 空白列宽度
            fat_wind.column_dimensions[get_column_letter(i*14+14)].width = 3

            # 合并单元格
            for j in range(ch_num):

                row = 4+wind_num*j+1
                fat_wind.merge_cells(start_row=row-1, start_column=i*14+1, end_row=row-1, end_column=i*14+13)
                fat_wind.merge_cells(start_row=row, start_column=i*14+2, end_row=row, end_column=i*14+3)
                fat_wind.merge_cells(start_row=row, start_column=i*14+4, end_row=row, end_column=i*14+5)
                fat_wind.merge_cells(start_row=row, start_column=i*14+6, end_row=row, end_column=i*14+7)
                fat_wind.merge_cells(start_row=row, start_column=i*14+8, end_row=row, end_column=i*14+9)
                fat_wind.merge_cells(start_row=row, start_column=i*14+10, end_row=row, end_column=i*14+11)
                fat_wind.merge_cells(start_row=row, start_column=i*14+12, end_row=row, end_column=i*14+13)

            # 设置边框c
            self.set_boader(fat_wind, [1, 1, i*14+1, i*14+13])
            self.set_boader(fat_wind, [2, 2, i*14+1, i*14+13])
            self.set_boader(fat_wind, [3, 3, i*14+1, i*14+13])

            # 数据
            for j in range(ch_num):
                # 设置表头
                row = 4+j*wind_num
                if fat_wind.cell(row+1, i*14+1).value:
                    self.set_boader(fat_wind, [row+1, row+1, i*14+1, i*14+13])
                    self.set_boader(fat_wind, [row+2, row+2, i*14+1, i*14+13])

                    for k in range(1, 14):
                        self.set_boader(fat_wind, [row+3, row+wind_num-1, i*14+k, i*14+k])

            # 设置变量名的行高
            for j in range(ch_num):
                row = 4+j*wind_num+1
                fat_wind.row_dimensions[row].height = 35

            # 隐藏空的风速结果
            for j in range(ch_num):
                for k in range(4+wind_num*j+1, 4+wind_num*(j+1)):

                    if not fat_wind.cell(k, 1).value:
                        fat_wind.row_dimensions.group(k, k, hidden=True)

            # 添加ratio颜色
            for j in range(ch_num):
                row = 4+j*wind_num+3
                
                for k in range(1, 7):
                    cell1 = fat_wind.cell(row=row, column=i*14+2)
                    cell2 = fat_wind.cell(row=row+1, column=i*14+2)

                    # 判断是否有值
                    if cell1.value or cell2.value:

                        cell_start = get_column_letter(14*i+2*k+1) + str(row)
                        cell_end   = get_column_letter(14*i+2*k+1) + str(row+wind_num-3-1)
                        cells      = cell_start + ':' + cell_end
                        # print(cells)
                        fat_wind.conditional_formatting.add(cells, CellIsRule(operator='lessThan',
                                                                              formula=['1.0'],
                                                                              stopIfTrue=True, fill=greenfill))
                        fat_wind.conditional_formatting.add(cells, CellIsRule(operator='between',
                                                                              formula=['1.0', '1.03'],
                                                                              stopIfTrue=True, fill=yellowfill))
                        fat_wind.conditional_formatting.add(cells, CellIsRule(operator='greaterThan',
                                                                              formula=['1.03'],
                                                                              stopIfTrue=True, fill=redfill))

            # 填充index颜色
            for j in range(ch_num):
                row = 4+wind_num*j+2

                for k in range(2, 7):
                    col = 14*i+2*k+1

                    if fat_wind.cell(row, col).value:
                        fat_wind.cell(row, col).fill = grayfill

    # ------------------------------------------------------------------------------
    # 定义case fatigue表格的格式
    def set_fatigue_case(self, fat_case):

        font1 = Font(name='Microsoft Ya Hei', size=11, bold=True)
        font2 = Font(name='Microsoft Ya Hei', size=9)
        font3 = Font(name='Microsoft Ya Hei', size=9, underline='single', color=colors.BLUE, bold=True)
        font4 = Font(name='Microsoft Ya Hei', size=9, bold=True)
        font5 = Font(name='Microsoft Ya Hei', size=9, underline='single', color=colors.BLUE)

        align1 = Alignment(horizontal='center', vertical='center')
        align2 = Alignment(horizontal='left',   vertical='center')
        align3 = Alignment(horizontal='center', vertical='center', wrapText=True)
        align4 = Alignment(horizontal='right',  vertical='center')
        align5 = Alignment(horizontal='center', vertical='center', text_rotation=90)

        redfill    = PatternFill(start_color='F5A9BC', end_color='F5A9BC', fill_type='solid')
        greenfill  = PatternFill(start_color='CEF6E3', end_color='CEF6E3', fill_type='solid')
        yellowfill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')  # FFF68F
        grayfill   = PatternFill('solid', fgColor='808080')

        row_max = fat_case.max_row
        col_max = fat_case.max_column

        loop_list = list(range(1,31))
        loop_row  = {}

        for i in range(1, row_max+1):
            cell = fat_case.cell(i, 1)

            if cell.value in loop_list:
                loop_row[cell.value] = i

        lp_num = len(loop_row)
        # print(loop_row)

        # 记录第一个loop的channel数
        if len(loop_row) == 1:
            chan_num = int(np.ceil((row_max-7)/13))
        else:
            chan_num = int(np.ceil(loop_row[2]-loop_row[1]-7)/13)
        # print(chan_num)

        # for i in range(self.loop_num):
        for i in range(lp_num):

            # 记录各个loop所占列数
            row_start = loop_row[i+1]

            # 设置前4行的格式
            fat_case.merge_cells(start_row=row_start, start_column=1, end_row=row_start, end_column=col_max)
            fat_case.cell(row=row_start, column=1).font      = font1
            fat_case.cell(row=row_start, column=1).alignment = align2

            fat_case.merge_cells(start_row=row_start+1, start_column=1, end_row=row_start+1, end_column=col_max)
            fat_case.cell(row=row_start+1, column=1).font      = font1
            fat_case.cell(row=row_start+1, column=1).alignment = align2

            fat_case.merge_cells(start_row=row_start+2, start_column=1, end_row=row_start+2, end_column=col_max)
            fat_case.cell(row=row_start+2, column=1).font      = font3
            fat_case.cell(row=row_start+2, column=1).alignment = align2

            fat_case.merge_cells(start_row=row_start+3, start_column=1, end_row=row_start+3, end_column=col_max)

            # 'DLC-Wind Speed'
            fat_case.merge_cells(start_row=row_start+4, start_column=1, end_row=row_start+4, end_column=2)
            fat_case.cell(row=row_start+4, column=1).font      = font4
            fat_case.cell(row=row_start+4, column=1).alignment = align1

            # 'RunName'
            fat_case.merge_cells(start_row=row_start+5, start_column=1, end_row=row_start+5, end_column=2)
            fat_case.cell(row=row_start+5, column=1).font      = font4
            fat_case.cell(row=row_start+5, column=1).alignment = align1

            # 'Path'
            fat_case.merge_cells(start_row=row_start+6, start_column=1, end_row=row_start+6, end_column=2)
            fat_case.cell(row=row_start+6, column=1).font      = font4
            fat_case.cell(row=row_start+6, column=1).alignment = align1

            # RunName-row 6
            for row in fat_case.iter_rows(min_col=3, min_row=row_start+5, max_col=col_max, max_row=row_start+5):

                for cell in row:

                    cell.font      = font4
                    cell.alignment = align5

            # 设置7行的格式
            for row in fat_case.iter_rows(min_col=3, min_row=row_start+6, max_col=col_max, max_row=row_start+6):

                for cell in row:

                    cell.font      = font5
                    cell.alignment = align2

            # 设置8-行的格式
            for i in range(chan_num):

                for j in range(6):
                    row  = row_start+6+13*i+2*j+1
                    cell = fat_case.cell(row, 1)

                    if cell.value:

                        # variable name
                        fat_case.cell(row, 1).font = font4
                        fat_case.cell(row, 1).alignment = align2

                        # index
                        fat_case.cell(row, 2).font = font4
                        fat_case.cell(row, 2).alignment = align1

                        # m值
                        fat_case.cell(row+1, 2).font = font4
                        fat_case.cell(row+1, 2).alignment = align1

                    for col in range(3, col_max+1):

                        cell1 = fat_case.cell(row, col)
                        cell1.font = font2
                        cell1.alignment = align2
                        cell1.number_format = '.00'

                        cell2 = fat_case.cell(row+1, col)
                        cell2.font = font2
                        cell2.alignment = align4
                        cell2.number_format = '.00'

            # 添加边框：前5行
            self.set_boader(fat_case, [row_start,   row_start,   1, col_max])
            self.set_boader(fat_case, [row_start+1, row_start+1, 1, col_max])
            self.set_boader(fat_case, [row_start+2, row_start+2, 1, col_max])
            self.set_boader(fat_case, [row_start+4, row_start+4, 1, col_max])
            self.set_boader(fat_case, [row_start+5, row_start+5, 1, col_max])
            self.set_boader(fat_case, [row_start+6, row_start+6, 1, 2])
            self.set_boader(fat_case, [row_start+6, row_start+6, 3, col_max], False)

            for i in range(chan_num):

                for j in range(6):
                    row = row_start+6+i*13+j*2+1

                    if fat_case.cell(row, 1).value:

                        # 合并变量名的两个单元格
                        fat_case.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=1)

                        # variable
                        self.set_boader(fat_case, [row,   row+1, 1, 1])
                        # m值
                        self.set_boader(fat_case, [row,   row,   2, 2])
                        # index
                        self.set_boader(fat_case, [row+1, row+1, 2, 2])

                        if j > 0:

                            fat_case.cell(row, 2).fill = grayfill

                row = row_start+6+i*13+1
                if fat_case.cell(row,3).value:
                    self.set_boader(fat_case, [row, row+11, 3, col_max])

                # 添加ratio颜色
                for k in range(6):
                    row  = row_start+6+i*13+2*k+1
                    cell = fat_case.cell(row=row, column=2)

                    # 判断是否有值
                    if cell.value:
                        cell_start = get_column_letter(3) + str(row)
                        # 默认为6个变量
                        cell_end   = get_column_letter(col_max) + str(row)
                        cells      = cell_start + ':' + cell_end
                        # print(cells)
                        fat_case.conditional_formatting.add(cells, CellIsRule(operator='lessThan',
                                                                              formula=['1.0'],
                                                                              stopIfTrue=True, fill=greenfill))
                        fat_case.conditional_formatting.add(cells, CellIsRule(operator='between',
                                                                              formula=['1.0', '1.03'],
                                                                              stopIfTrue=True, fill=yellowfill))
                        fat_case.conditional_formatting.add(cells, CellIsRule(operator='greaterThan',
                                                                              formula=['1.03'],
                                                                              stopIfTrue=True, fill=redfill))

            # 设置行高
            fat_case.row_dimensions[row_start+5].height = 45

            # 设置列宽
            fat_case.column_dimensions['A'].width = 10
            for col in range(2, col_max+1):
                fat_case.column_dimensions[get_column_letter(col)].width = 5

            # 隐藏m值行及Fx/Fy/Fz
            for i in range(chan_num):

                row = row_start+6+13*i
                if fat_case.cell(row+1, 1).value:

                    fat_case.row_dimensions.group(row+2, row+2, hidden=True)
                    fat_case.row_dimensions.group(row+4, row+4, hidden=True)
                    fat_case.row_dimensions.group(row+6, row+12, hidden=True)
                else:
                    fat_case.row_dimensions.group(row+1, row+13, hidden=True)

    # ------------------------------------------------------------------------------
    # set aep format
    def set_aep(self, aep_dlc):

        font1 = Font(name='Microsoft Ya Hei', size=11, bold=True)
        font2 = Font(name='Microsoft Ya Hei', size=9)
        font3 = Font(name='Microsoft Ya Hei', size=9, underline='single', color=colors.BLUE, bold=True)
        font4 = Font(name='Microsoft Ya Hei', size=9, bold=True)
        # font5 = Font(name='Microsoft Ya Hei', size=9, underline='single', color=colors.BLUE)

        align1 = Alignment(horizontal='center', vertical='center')
        align2 = Alignment(horizontal='left', vertical='center')
        align3 = Alignment(horizontal='right', vertical='center')

        row_max = aep_dlc.max_row
        col_max = aep_dlc.max_column

        loop_num = int((col_max+1)/6)
        # print(row_max, col_max, loop_num)

        for i in range(loop_num):
            # 设置表头
            aep_dlc.merge_cells(start_row=1, start_column=i*6+1, end_row=1, end_column=i*6+5)
            aep_dlc.cell(row=1, column=i*6+1).font      = font1
            aep_dlc.cell(row=1, column=i*6+1).alignment = align1

            aep_dlc.merge_cells(start_row=2, start_column=i*6+1, end_row=2, end_column=i*6+5)
            aep_dlc.cell(row=2, column=i*6+1).font      = font4
            aep_dlc.cell(row=2, column=i*6+1).alignment = align2

            aep_dlc.merge_cells(start_row=3, start_column=i*6+1, end_row=3, end_column=i*6+5)
            aep_dlc.cell(row=3, column=i*6+1).font      = font3
            aep_dlc.cell(row=3, column=i*6+1).alignment = align2

        row_rname = [row for row in range(7, row_max) if 'RunName' == aep_dlc.cell(row=row, column=1).value]

        # header
        for row in [5, 6, row_rname[0], row_rname[0]+1]:
            for j in range(loop_num):
                for k in range(6):
                    aep_dlc.cell(row=row, column=6*j+k+1).font      = font4
                    aep_dlc.cell(row=row, column=6*j+k+1).alignment = align1

        # result for each wind speed
        row_sum = []
        for row in range(7, row_rname[0]):
            for loop in range(loop_num):
                for col in range(6):
                    cell = aep_dlc.cell(row=row, column=6*loop+col+1)

                    if cell.value:
                        cell.font      = font2
                        cell.alignment = align1

                    if col in range(2,5):
                        cell.number_format = '0.00'
                        cell.alignment     = align3

                    if 'AEP sum' == cell.value:
                        cell.font      = font4
                        cell.alignment = align1

                        row_sum.append(row)
        # print(row_sum)
        # result for each load case
        for row in range(row_rname[0]+2, row_max+1):
            for loop in range(loop_num):
                for col in range(6):
                    cell = aep_dlc.cell(row=row, column=6*loop+col+1)

                    if cell.value:
                        cell.font      = font2
                        cell.alignment = align1

                    if col in range(2, 5):
                        cell.number_format = '0.00'
                        cell.alignment = align3

        for i in range(loop_num):
            # 设置前三行
            self.set_boader(aep_dlc, [1, 1, i*6+1, i*6+5])
            self.set_boader(aep_dlc, [2, 2, i*6+1, i*6+5])
            self.set_boader(aep_dlc, [3, 3, i*6+1, i*6+5])

            # wind-power header
            self.set_boader(aep_dlc, [5, 5, i*6+1, i*6+5])
            self.set_boader(aep_dlc, [6, 6, i*6+1, i*6+5])

            # wind-power aera
            self.set_boader(aep_dlc, [7, row_sum[i]-1, i*6+1, i*6+1])
            self.set_boader(aep_dlc, [7, row_sum[i]-1, i*6+2, i*6+2])
            self.set_boader(aep_dlc, [7, row_sum[i]-1, i*6+3, i*6+3])
            self.set_boader(aep_dlc, [7, row_sum[i]-1, i*6+4, i*6+4])
            self.set_boader(aep_dlc, [7, row_sum[i]-1, i*6+5, i*6+5])

            # AEP
            self.set_boader(aep_dlc, [row_sum[i], row_sum[i], i*6+4, i*6+5])

            # dlc-power header
            self.set_boader(aep_dlc, [row_rname[0],   row_rname[0],   i*6+1, i*6+5])
            self.set_boader(aep_dlc, [row_rname[0]+1, row_rname[0]+1, i*6+1, i*6+5])

            self.set_boader(aep_dlc, [row_rname[0]+2, row_max, 6*i+1, 6*i+5])

    # ------------------------------------------------------------------------------
    # set ldd format
    def set_ldd(self, ldd):

        font1 = Font(name='Microsoft Ya Hei', size=11, bold=True)
        font2 = Font(name='Microsoft Ya Hei', size=9)
        font3 = Font(name='Microsoft Ya Hei', size=9, underline='single', color=colors.BLUE, bold=True)
        font4 = Font(name='Microsoft Ya Hei', size=9, bold=True)
        # font5 = Font(name='Microsoft Ya Hei', size=9, underline='single', color=colors.BLUE)

        align1 = Alignment(horizontal='center', vertical='center')
        align2 = Alignment(horizontal='left',   vertical='center')
        align3 = Alignment(horizontal='right',  vertical='center')
        align4 = Alignment(horizontal='center', vertical='center', wrapText=True)

        row_max = ldd.max_row
        col_max = ldd.max_column
        # print(row_max, col_max)

        loop_pos = [i for i in range(1,col_max+1) if ldd.cell(row=1, column=i).value]
        loop_pos.append(col_max+2)
        print(loop_pos)
        loop_num = len(loop_pos)-1

        for i in range(loop_num):

            chan_pos = [j for j in range(1,row_max+1) if not ldd.cell(row=j, column=loop_pos[i]).value]
            chan_pos.append(row_max+1)
            chan_num = len(chan_pos)
            # print(chan_num)

            # set font for the first two rows
            ldd.cell(row=1, column=loop_pos[i]).font = font1
            ldd.cell(row=2, column=loop_pos[i]).font = font3
            # merge cell
            ldd.merge_cells(start_row=1, start_column=loop_pos[i], end_row=1, end_column=loop_pos[i+1]-2)
            ldd.merge_cells(start_row=2, start_column=loop_pos[i], end_row=2, end_column=loop_pos[i+1]-2)
            self.set_boader(ldd, [1, 1, loop_pos[i], loop_pos[i+1]-2])
            self.set_boader(ldd, [2, 2, loop_pos[i], loop_pos[i+1]-2])

            for j in range(chan_num-1):
                # column
                for k in range(loop_pos[i],loop_pos[i+1]):
                    # data
                    for l in range(chan_pos[j]+1, chan_pos[j+1]):
                        ldd.cell(row=l, column=k).font = font2
                    # variable header
                    for l in range(chan_pos[j]+1, chan_pos[j]+2):

                        if ldd.cell(row=chan_pos[j]+1, column=loop_pos[i]):
                            ldd.cell(row=chan_pos[j]+1, column=k).font      = font4
                            ldd.cell(row=chan_pos[j]+1, column=k).alignment = align4
                            # ldd.cell(row=chan_pos[j]+1, column=k).alignment = Alignment(wrapText=True)

                        else:
                            ldd.cell(row=chan_pos[j]+1, column=k).font      = font4
                            ldd.cell(row=chan_pos[j]+1, column=k).alignment = align4
                            # ldd.cell(row=chan_pos[j]+1, column=k).alignment = Alignment(wrapText=True)
                            ldd.cell(row=chan_pos[j]+2, column=k).font      = font4
                            ldd.cell(row=chan_pos[j]+2, column=k).alignment = align4
                            # ldd.cell(row=chan_pos[j]+2, column=k).alignment = Alignment(wrapText=True)

            for j in range(chan_num-1):

                if ldd.cell(row=chan_pos[j]+1, column=loop_pos[i]):
                    ldd.row_dimensions[chan_pos[j]+1].height = 25
                    self.set_boader(ldd, [chan_pos[j]+1, chan_pos[j]+1, loop_pos[i], loop_pos[i+1]-2])
                    self.set_boader(ldd, [chan_pos[j]+2, chan_pos[j+1]-1, loop_pos[i], loop_pos[i+1]-2])
                    # print(loop_pos[i],loop_pos[i+1]-1)
                else:
                    ldd.row_dimensions[chan_pos[j]+1].height = 25
                    ldd.row_dimensions[chan_pos[j]+2].height = 25
                    self.set_boader(ldd, [chan_pos[j]+1, chan_pos[j]+2, loop_pos[i], loop_pos[i+1]-2])
                    self.set_boader(ldd, [chan_pos[j]+3, chan_pos[j+1]-1, loop_pos[i], loop_pos[i+1]-2])

            # 设置每列的宽度
            ldd.column_dimensions[get_column_letter(i*3+1)].width = 10
            ldd.column_dimensions[get_column_letter(i*3+2)].width = 12
    # ------------------------------------------------------------------------------
    # set lrd format
    def set_lrd(self, lrd):

        font1 = Font(name='Microsoft Ya Hei', size=11, bold=True)
        font2 = Font(name='Microsoft Ya Hei', size=9)
        font3 = Font(name='Microsoft Ya Hei', size=9, underline='single', color=colors.BLUE, bold=True)
        font4 = Font(name='Microsoft Ya Hei', size=9, bold=True)
        # font5 = Font(name='Microsoft Ya Hei', size=9, underline='single', color=colors.BLUE)

        align1 = Alignment(horizontal='center', vertical='center')
        align2 = Alignment(horizontal='left',   vertical='center')
        align3 = Alignment(horizontal='right',  vertical='center')
        align4 = Alignment(horizontal='center', vertical='center', wrapText=True)

        row_max = lrd.max_row
        col_max = lrd.max_column
        # print(row_max, col_max)

        loop_pos = [i for i in range(1,col_max+1) if lrd.cell(row=1, column=i).value]
        loop_pos.append(col_max+2)
        loop_num = len(loop_pos)-1

        for i in range(loop_num):

            chan_pos = [j for j in range(1,row_max+1) if not lrd.cell(row=j, column=loop_pos[i]).value]
            chan_pos.append(row_max+1)
            chan_num = len(chan_pos)
            # print(chan_num)

            # set font for the first two rows
            lrd.cell(row=1, column=loop_pos[i]).font = font1
            lrd.cell(row=2, column=loop_pos[i]).font = font3
            # merge cell
            lrd.merge_cells(start_row=1, start_column=loop_pos[i], end_row=1, end_column=loop_pos[i+1]-2)
            lrd.merge_cells(start_row=2, start_column=loop_pos[i], end_row=2, end_column=loop_pos[i+1]-2)
            self.set_boader(lrd, [1, 1, loop_pos[i], loop_pos[i+1]-2])
            self.set_boader(lrd, [2, 2, loop_pos[i], loop_pos[i+1]-2])

            for j in range(chan_num-1):
                # column
                for k in range(loop_pos[i],loop_pos[i+1]):
                    # data
                    for l in range(chan_pos[j]+1, chan_pos[j+1]):
                        lrd.cell(row=l, column=k).font = font2
                    # variable header
                    for l in range(chan_pos[j]+1, chan_pos[j]+2):

                        if lrd.cell(row=chan_pos[j]+1, column=loop_pos[i]):
                            lrd.cell(row=chan_pos[j]+1, column=k).font      = font4
                            lrd.cell(row=chan_pos[j]+1, column=k).alignment = align4
                        else:
                            lrd.cell(row=chan_pos[j]+1, column=k).font      = font4
                            lrd.cell(row=chan_pos[j]+1, column=k).alignment = align4
                            lrd.cell(row=chan_pos[j]+2, column=k).font      = font4
                            lrd.cell(row=chan_pos[j]+2, column=k).alignment = align4

            for j in range(chan_num-1):

                if lrd.cell(row=chan_pos[j]+1, column=loop_pos[i]):
                    lrd.row_dimensions[chan_pos[j]+1].height = 25
                    self.set_boader(lrd, [chan_pos[j]+1,chan_pos[j]+1,  loop_pos[i],loop_pos[i+1]-2])
                    self.set_boader(lrd, [chan_pos[j]+2,chan_pos[j+1]-1,loop_pos[i],loop_pos[i+1]-2])
                    # print(loop_pos[i],loop_pos[i+1]-1)
                else:
                    lrd.row_dimensions[chan_pos[j]+1].height = 30
                    lrd.row_dimensions[chan_pos[j]+2].height = 25
                    self.set_boader(lrd, [chan_pos[j]+1,chan_pos[j]+2,  loop_pos[i],loop_pos[i+1]-2])
                    self.set_boader(lrd, [chan_pos[j]+3,chan_pos[j+1]-1,loop_pos[i],loop_pos[i+1]-2])

            lrd.column_dimensions[get_column_letter(i*3+1)].width = 10
            lrd.column_dimensions[get_column_letter(i*3+2)].width = 15
    # ------------------------------------------------------------------------------
    # set load summary format
    def set_load_summary(self, load_summary):

        font1 = Font(name='Microsoft Ya Hei', size=11, bold=True)
        font2 = Font(name='Microsoft Ya Hei', size=9)
        # font3 = Font(name='Microsoft Ya Hei', size=9, underline='single', color=colors.BLUE, bold=True)
        font4 = Font(name='Microsoft Ya Hei', size=9, bold=True)
        font5 = Font(name='Microsoft Ya Hei', size=9, underline='single', color=colors.BLUE)

        align1 = Alignment(horizontal='center', vertical='center')
        align2 = Alignment(horizontal='left',   vertical='center')

        # row_max = load_summary.max_row
        col_max = load_summary.max_column

        loop_num = int((col_max+1)/3)

        for i in range(loop_num):
            # set header format
            load_summary.cell(row=1, column=3*i+1).font      = font1
            load_summary.cell(row=1, column=3*i+1).alignment = align1
            load_summary.merge_cells(start_row=1, start_column=3*i+1, end_row=1, end_column=3*i+2)
            self.set_boader(load_summary,[1,1,3*i+1,3*i+2])

            load_summary.column_dimensions[get_column_letter(3*i+1)].width = 15
            load_summary.column_dimensions[get_column_letter(3*i+2)].width = 30
            load_summary.column_dimensions[get_column_letter(3*i+3)].width = 3

            row_start = 1
            for j in range(4):

                load_summary.cell(row=row_start+1, column=3*i+1).font      = font4
                load_summary.cell(row=row_start+1, column=3*i+1).alignment = align2
                load_summary.merge_cells(start_row=row_start+1, start_column=3*i+1, end_row=row_start+1, end_column=3*i+2)
                self.set_boader(load_summary, [row_start+1, row_start+1, 3*i+1, 3*i+2])
                load_summary.cell(row=row_start+2, column=3*i+1).font      = font2
                load_summary.cell(row=row_start+2, column=3*i+1).alignment = align2
                load_summary.cell(row=row_start+3, column=3*i+1).font      = font2
                load_summary.cell(row=row_start+3, column=3*i+1).alignment = align2

                load_summary.cell(row=row_start+2, column=3*i+2).font      = font5
                load_summary.cell(row=row_start+2, column=3*i+2).alignment = align2
                load_summary.cell(row=row_start+3, column=3*i+2).font      = font2
                load_summary.cell(row=row_start+3, column=3*i+2).alignment = align2

                if j==1 or j==3:
                    load_summary.cell(row=row_start+4, column=3*i+2).font = font2
                    load_summary.cell(row=row_start+4, column=3*i+2).alignment = align2

                if j == 1 or j == 3:
                    self.set_boader(load_summary, [row_start+1, row_start+4, 3*i+1, 3*i+2])
                    row_start += 5
                else:
                    self.set_boader(load_summary, [row_start+1, row_start+3, 3*i+1, 3*i+2])
                    row_start += 4


if __name__ == '__main__':

    import openpyxl
    path = r"E:\05 TASK\02_Tools\01_Load Summary\test.xlsx"

    table = openpyxl.load_workbook(path)

    sheet1 = table.get_sheet_by_name('Ultimate')
    # sheet2 = table.get_sheet_by_name('Heat Map')

    SetFormat().set_ultimate(sheet1)
    # SetFormat().set_heat_map(sheet2)
    print('Setting format is done!')

    table.save(path)
