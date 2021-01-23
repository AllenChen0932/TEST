# -*- coding: utf-8 -*-
# @Time    : 2019/11/28 12:54
# @Author  : CE
# @File    : excel_output.py

'''
2019.12.04 版本012在011的基础上增加了fatigue的部分
2019.12.29 根据excel中表格的内容来添加新的结果
'''

import time
import os

from numpy import ceil
from openpyxl import load_workbook
from openpyxl.styles import Font, colors
from openpyxl.utils import get_column_letter

from tool.load_summary.bladed_ultimate import BladedUltimate
from tool.load_summary.bladed_rainflow import BladedRainflow
from tool.load_summary.excel_format import SetFormat
from tool.load_summary.read_LCT import read_LCT_dlc12
from tool.load_summary.dynamic_power_curve import output_PCs as powercurve
# import pysnooper

class excel_operation:

    def __init__(self,
                 excel_input,
                 excel_output,
                 ultimate=False,
                 fatigue=False,
                 heat_map=False,
				 fat_case=False,
                 fat_wind=False,
                 aep_dlc=False):

        self.template = excel_input
        self.xls_path = excel_output
        self.ultimate = ultimate
        self.fatigue  = fatigue
        self.heat_map = heat_map
        self.fat_wind = fat_wind
        self.fat_case = fat_case
        self.aep_dlc  = aep_dlc

        # read all results in default
        # write results corresponding to the variables in the excel
        # it works for fatigue results
        self.main_all = True

        # input excel:
        self.loopname = []      #loopname list
        self.ult_list = []      #ultimate路径list
        self.fat_list = []      #fatigue路径list
        self.dlc_list = []      #dlc12路径list
        self.lct_list = []      #load case table list

        # 记录已有结果中的变量
        self.ult_vars = {}      #var:index
        self.fat_vars = {}
        self.hea_vars = {}
        self.fws_vars = {}
        self.fca_vars = {}

        self.loop_dlc = {}      #{loopname: dlc}

        # for ultimate and heat map
        self.loop_channels  = {}     #{loopname: channel(in order)}
        self.loop_vari_dlc  = {}     #{loopname: {var: {dlc:value}}}，最大值
        self.loop_var_ulti  = {}     #{loopname: {var: {dlc:value}}}，极值
        self.loop_chan_var  = {}     #{loopname: {channel: variable}}
        self.loop_dlc_path  = {}     #{loopname: {dlc: path}}
        self.loop_dlcs_dlc  = {}     #{loopname: {dlcs: DLC}}

        # for fatigue
        self.lp_channels = {}    # {loopname: [channel list]}
        self.lp_var_fat  = {}    # {loopname: {var: {m: {dlc: value}}}}
        self.lp_chan_var = {}    # {loopname: {channel: [variable list]}}
        self.lp_dlc_path = {}    # {loopname: {dlc: path}}

        # for wind fatigue
        self.lp_case_V = {}
        self.lp_V_case = {}

        # new excel path
        # get current year/month/day
        self.data = time.strftime("%Y%m%d", time.localtime())

        # new path
        if '\\' in self.xls_path:
            old_path = self.xls_path.replace('\\', '/')
        else:
            old_path = self.xls_path

        dir_list  = old_path.split('/')
        name_list = [dir_list[-1].split('.')[0], self.data+'.xlsx']
        new_name  = '_'.join(name_list)
        name_list = [os.sep.join(dir_list[:-1]), new_name]
        self.out_path = os.sep.join(name_list)
        print(self.out_path)

        self.read_excel()
        self.get_excel_result()
        self.append_result()

    # 获取所有结果中的变量
    def get_vars(self, path_list):

        dollar_list = []
        var_list    = []

        for path in path_list:

            for root, dir, names in os.walk(path):
                # print(root, dir, names)

                for name in names:

                    if '.%' in name:

                        file_dir = root + os.sep + name

                        dollar_list.append(file_dir)
        # print(dollar_list)
        for dollar in dollar_list:

            with open(dollar) as f:

                for line in f.readlines():

                    if 'VARIAB' in line:

                        variab = line.split("'")[1]

                        if ',' in variab and 'Blade' in variab:
                            temp = variab.split(',')

                            height = temp[1].split('=')[1].strip()[0].split('m')[0]

                            variab = temp[0] + ',' + '%.2f' % float(height)

                            var_list.append(variab)

                        if ',' in variab and 'Tower' in variab:

                            temp = variab.split(',')

                            if 'height' in temp[1]:

                                height = temp[1].split('=')[1].strip()[0].split('m')[0]

                                variab = temp[0] + ',' + '%.2f' % float(height)

                                var_list.append(variab)

                            elif 'Location' in temp[1]:

                                height = temp[1].split('=')[1]

                                variab = temp[0] + ',' + height

                                var_list.append(variab)

        return var_list

    # 读取模板中的路径
    def read_excel(self):

        workbook = load_workbook(self.template)

        sheet = workbook['Main']

        row_num = sheet.max_row
        row_max = row_num - (row_num-1)%4
        # print(row_num)

        for i in range(2, row_max+1):

            if sheet[i][0].value:

                # print(sheet[i][0].value)

                self.loopname.append(sheet[i][0].value)

        # print(self.loopname)

        for i in range(2, row_max+1):

            if sheet[i][1].value == 'Ultimate':
                self.ult_list.append(sheet[i][2].value)

            if sheet[i][1].value == 'Fatigue':
                self.fat_list.append(sheet[i][2].value)

            if sheet[i][1].value == 'DLC12':
                self.dlc_list.append(sheet[i][2].value)

            if sheet[i][1].value == 'LCT':
                self.lct_list.append(sheet[i][2].value)

        print('Begin to read excel...')
        print('Ultimate path:')
        print(self.ult_list)
        print('Fatigue path:')
        print(self.fat_list)
        print('DLC12 path:')
        print(self.dlc_list)
        print('Load case table:')
        print(self.lct_list)
        print()

    def get_result(self):

        print('Begin to get result...')

        for i in range(0, len(self.loopname)):

            ult_path = self.ult_list[i]

            fat_path = self.fat_list[i]

            lct_path = self.lct_list[i]

            if self.ultimate or self.heat_map:

                print('Begin to read "%s"' %self.ult_list[i])

                # get ultimate:
                # for ultimate results, we just need to read all results, and write corresponding to the excel
                result = BladedUltimate(ult_path, False)

                self.loop_channels[self.loopname[i]]  = result.channel
                self.loop_var_ulti[self.loopname[i]]  = result.ult_val
                self.loop_vari_dlc[self.loopname[i]]  = result.var_ult
                self.loop_chan_var[self.loopname[i]]  = result.chan_var
                self.loop_dlc_path[self.loopname[i]]  = result.dlc_path
                self.loop_dlcs_dlc[self.loopname[i]]  = result.dlcs_DLC

                print('Read is done!')

            # ==============================================
            if self.fatigue or self.fat_wind or self.fat_case:

                print('Begin to read "%s"' %self.fat_list[i])

                # get fatigue:
                # for fatigue results, we need to determine the result type because it combine all blade results for blade
                result2 = BladedRainflow(fat_path, self.main_all)

                self.lp_channels[self.loopname[i]] = result2.channel
                self.lp_var_fat[self.loopname[i]]  = result2.fatigue_val
                self.lp_chan_var[self.loopname[i]] = result2.chan_var
                self.lp_dlc_path[self.loopname[i]] = result2.dlc_path

                print('Read is done!')

            # ==============================================
            # get Vhub and hours for dlc12 cases: (only used in wind fatigue)
            if self.fat_wind:

                result3 = read_LCT_dlc12(lct_path)

                self.lp_case_V[self.loopname[i]] = result3[0]
                self.lp_V_case[self.loopname[i]] = result3[1]

        print('Get result is done! \n')

    def get_heatmap(self, dlc_DLC, dlc_val):

        DLC = []
        DLC_val  = {}

        for key, val in dlc_val.items():

            if dlc_DLC[key] not in DLC:

                DLC.append(dlc_DLC[key])

                DLC_val[dlc_DLC[key]] = val[0]

            elif dlc_DLC[key] in DLC:

                if val[0] > DLC_val[dlc_DLC[key]]:

                    DLC_val[dlc_DLC[key]] = val[0]

                else:

                    continue

            DLC.sort()

        return DLC_val, DLC

    def get_excel_result(self):
        '''
        1 Read information from the excel which contains the result
        2 Find the result type from fatigue results, all or main
        3 Get the variable list from the excel
        :return:
        '''
        table = load_workbook(self.xls_path)

        sheet_names = table.sheetnames

        # for ultimate result, main or all is useless

        if 'Ultimate' in sheet_names:

            sheet = table['Ultimate']

            row_max = sheet.max_row
            col_max = sheet.max_column

            chan_num = int(ceil((row_max-4)/9))
            loop_num = int(ceil(col_max/6))

            # there is not need to judge main or all types for ultimate loads
            # what we need to do is to write as the variables listed in the excel no matter it is main or all type.
            if chan_num != 6:

                self.main_all = False

            for chan in range(chan_num):

                for i in range(1, 9):

                    row = chan*9+4+i

                    for loop in range(loop_num):

                        variable = sheet.cell(row=row, column=6*loop+1).value

                        if variable:

                            self.ult_vars[variable] = row

        if 'Fatigue' in sheet_names:

            sheet = table['Fatigue']

            row_max = sheet.max_row
            col_max = sheet.max_column
            # print(col_max)

            chan_num = int(ceil((row_max-4)/7))
            loop_num = int(ceil(col_max/5))

            # check the type of the result in the excel
            
            # if chan_num != 6:
            #     self.main_all = False
            
            # 根据前两个通道所有的变量名称是否多余一个来判断main或all
            for loop in range(loop_num):

                # check the first two channels
                for chan in range(2):

                    row = chan*7+4
                    col = loop*5+1

                    if sheet.cell(row=row+1, column=col).value:

                        var1 = sheet.cell(row=row+1, column=col).value.split()[2]
                        var2 = sheet.cell(row=row+2, column=col).value.split()[2]
                        var3 = sheet.cell(row=row+3, column=col).value.split()[2]
                        var4 = sheet.cell(row=row+4, column=col).value.split()[2]
                        var5 = sheet.cell(row=row+5, column=col).value.split()[2]
                        var6 = sheet.cell(row=row+6, column=col).value.split()[2]

                        var_list = [var1, var2, var3, var4, var5, var6]

                        # main result
                        if len(set(var_list)) > 1:

                            self.main_all = True

                            break

            for chan in range(chan_num):

                for i in range(1, 7):

                    row = chan*7+4+i

                    for loop in range(loop_num):

                        variable = sheet.cell(row=row, column=5*loop+1).value

                        if variable:

                            self.fat_vars[variable] = row

        if 'Heat Map' in sheet_names:

            sheet = table['Heat Map']

            row_max = sheet.max_row
            col_max = sheet.max_column

            chan_num = int(ceil((row_max-2)/10))
            loop_num = 0
            vars_col = {}

            for i in range(1, col_max):

                if sheet.cell(row=1, column=i).value:

                    loop_num += 1

                    vars_col[loop_num] = i

            for chan in range(chan_num):

                for i in range(2, 10):

                    row = chan*10+2+i

                    for col in vars_col.values():

                        variable = sheet.cell(row=row, column=col).value

                        if variable:

                            self.hea_vars[variable] = row

        if 'Fatigue_Wind' in sheet_names:

            sheet = table['Fatigue_Wind']

            row_max = sheet.max_row
            col_max = sheet.max_column
            # print(col_max)

            chan_num = int(ceil((row_max-4)/26))
            loop_num = int(ceil(col_max/5))

            # check the type of the result in the excel
            if chan_num != 6:
                
                self.main_all = False
                
            # 读取变量
            if not self.main_all:

                for i in range(chan_num):
    
                    for j in range(loop_num):
    
                        for k in range(1, 7):
                            
                            row = 4+i*26+1
                            col = 14*j+2*k
    
                            variable = sheet.cell(row=row, column=col).value
    
                            if variable:

                                self.fws_vars[variable] = row

        if 'Fatigue_Case' in sheet_names:

            sheet = table['Fatigue_Case']

            row_max = sheet.max_row
            col_max = sheet.max_column
            print(row_max)

            loop_index = list(range(1,31))

            # 记录每个loop的起始行
            loop_row = {}

            for i in range(1, row_max+1):

                cell = sheet.cell(row=i, column=1)

                if cell.value in loop_index:

                    loop_row[cell.value] = i

            # 记录第一个loop的channel数
            if len(loop_row) == 1:
                chan_num = int(ceil((row_max-7)/13))
            else:
                chan_num = int(ceil(loop_row[2]-loop_row[1]-7)/13)

            # 记录变量
            if chan_num != 6:
                for i in range(chan_num):

                    var_list = []
                    for k in range(6):

                        for j in range(len(loop_row)):

                            row_start = j*(7+13*chan_num)
                            cell1 = sheet.cell(row_start+7+13*i+2*k+1, 1)
                            var = cell1.value

                            if var:

                                self.fca_vars[var] = 7+i*13+2*k+1

        # print(self.ult_vars)
        # print(self.fat_vars)
        # print(self.hea_vars)
        # print(self.fws_vars)
        print(self.fca_vars)

        table.close()

    # @pysnooper.snoop()
    def append_result(self):

        # get ultimate and fatigue loads
        self.get_result()

        # **************************** load excel *********************************************
        table = load_workbook(self.xls_path)

        sheet_names = table.sheetnames

        # ********************** Ultimate and Heat map ****************************************
        # ultimate
        if self.ultimate and 'Ultimate' in sheet_names:

            print('Begin to write ultimate result...')

            sheet_ultimate = table['Ultimate']

            loop_num_start = int(ceil(sheet_ultimate.max_column/6))

            col_start = sheet_ultimate.max_column

            # write ultimate
            for i in range(len(self.loopname)):

                print('Begin to write: ', self.loopname[i])

                loop_name = self.loopname[i]

                # 写表头，前4行
                sheet_ultimate.cell(row=1, column=col_start+i*6+1, value=loop_num_start+i+1)
                sheet_ultimate.cell(row=2, column=col_start+i*6+1, value=self.loopname[i])
                hyper1 = '=HYPERLINK("{}","{}")' .format(self.ult_list[i], self.ult_list[i])
                sheet_ultimate.cell(row=3, column=col_start+i*6+1, value=hyper1)
                sheet_ultimate.cell(row=4, column=col_start+i*6+1, value='Variable')
                sheet_ultimate.cell(row=4, column=col_start+i*6+2, value='DLC')
                sheet_ultimate.cell(row=4, column=col_start+i*6+3, value='Value\n(kN/kNm)')
                sheet_ultimate.cell(row=4, column=col_start+i*6+4, value='Path')
                sheet_ultimate.cell(row=4, column=col_start+i*6+5, value= 1)

                chan_num = len(self.loop_channels[loop_name])

                for j in range(chan_num):

                    cha = self.loop_channels[loop_name][j]

                    for k in range(8):

                        var = self.loop_chan_var[loop_name][cha][k]
                        dlc = [key for key in self.loop_var_ulti[loop_name][var].keys()]
                        val = [val for val in self.loop_var_ulti[loop_name][var].values()]
                        pat = self.loop_dlc_path[loop_name][dlc[0]]
                        char = get_column_letter(6*i+5)
                        rat = '=INDIRECT(ADDRESS(ROW(),COLUMN()-2))/INDIRECT(ADDRESS(ROW(),(%s$4*6-3)))' % char

                        if var in self.ult_vars.keys():

                            row_index = self.ult_vars[var]

                            sheet_ultimate.cell(row=row_index, column=col_start+i*6+1, value=var)
                            sheet_ultimate.cell(row=row_index, column=col_start+i*6+2, value=dlc[0])
                            # 此处读取的array数据带格式说明，实际上是一个包含数值和类型两个元素的元组，因此需要加上[0]来取其值
                            sheet_ultimate.cell(row=row_index, column=col_start+i*6+3, value=val[0][0]/1000)
                            hyper2 = '=HYPERLINK("{}","{}")'.format(pat, pat)
                            sheet_ultimate.cell(row=row_index, column=col_start+i*6+4, value=hyper2)
                            sheet_ultimate.cell(row=row_index, column=col_start+i*6+5, value=rat)

                print('loop %s is done!' %(i+1))

            SetFormat().set_ultimate(sheet_ultimate)

            print('Ultimate loads is done! \n')
            # table.save(self.out_path)

        # -------------------------------------------------------------------------------------
        # heat map
        if self.heat_map and 'Heat Map' in sheet_names:

            print('Begin to read heat map...')

            sheet_heatmap = table['Heat Map']

            # 最大列数
            col_start = sheet_heatmap.max_column

            for i in range(len(self.loopname)):

                print('Begin to write: ', self.loopname[i])

                loop_name = self.loopname[i]

                # 表头
                sheet_heatmap.cell(row=1, column=col_start+2, value=loop_name)

                # channels
                chan_num = len(self.loop_channels[loop_name])

                for j in range(chan_num):

                    ch_name = self.loop_channels[loop_name][j]

                    dlc_list = []

                    # variables
                    for k in range(8):

                        d_dlc = self.loop_dlcs_dlc[loop_name]                #dict of dlcs and DLC for each loop
                        v_nam = self.loop_chan_var[loop_name][ch_name][k]    #variable name
                        d_val = self.loop_vari_dlc[loop_name][v_nam]         #dict of dlc and value for each variable
                        loop_result = self.get_heatmap(d_dlc, d_val)         #{dlcs: DLC}, {DLC: value}

                        if k == 0:

                            # 记录第一个变量的工况名称，后面的工况按照该变量工况的顺序进行写
                            dlc_list = loop_result[1]

                        if v_nam in self.hea_vars.keys():

                            row_index = self.hea_vars[v_nam]

                            # 写变量名称
                            sheet_heatmap.cell(row=row_index, column=col_start+2, value=v_nam)

                            if k == 0:

                                for _ in range(len(dlc_list)):
                                    # 写工况名称
                                    sheet_heatmap.cell(row=row_index-1, column=col_start+2+1+_, value=dlc_list[_])

                            max_val = max(loop_result[0].values())

                            for __ in range(len(dlc_list)):

                                dlc_val = float(loop_result[0][dlc_list[__]])

                                sheet_heatmap.cell(row=row_index, column=col_start+2+1+__, value=dlc_val/max_val)

                print('loop %s is done!' %(i+1))

                col_start = sheet_heatmap.max_column
                # print(col_start)

            SetFormat().set_heat_map(sheet_heatmap)
            print('Heat map is done! \n')
        # ********************** Ultimate and Heat map ****************************************

        # ***************************** Fatigue ***********************************************
        if self.fatigue and 'Fatigue' in sheet_names:

            print('Begin to write fatigue...')
            # print(self.main_all)

            sheet_fatigue = table['Fatigue']

            col_start = sheet_fatigue.max_column + 1

            loop_num = int(ceil(sheet_fatigue.max_column/5))

            for i in range(len(self.loopname)):

                print('Begin to write: ', self.loopname[i])

                # 表头：
                sheet_fatigue.cell(row=1, column=col_start+i*5+1, value=loop_num+i+1)
                sheet_fatigue.cell(row=2, column=col_start+i*5+1, value=self.loopname[i])
                hyper3 = '=HYPERLINK("{}", "{}")'.format(self.fat_list[i], self.fat_list[i])
                sheet_fatigue.cell(row=3, column=col_start+i*5+1, value=hyper3)
                sheet_fatigue.cell(row=4, column=col_start+i*5+1, value='Variable')
                sheet_fatigue.cell(row=4, column=col_start+i*5+2, value='m')
                sheet_fatigue.cell(row=4, column=col_start+i*5+3, value='Value\n(KN/KNm)')
                sheet_fatigue.cell(row=4, column=col_start+i*5+4, value=1)

                # 数据区：
                lp_name = self.loopname[i]
                channels = self.lp_channels[lp_name]  # list in order

                for j in range(len(channels)):

                    # 去掉占位符
                    if channels[j]:

                        chan_var = self.lp_chan_var[lp_name][channels[j]]  # list, len=6
                        # print(len(chan_var))

                        for k in range(6):     #for k in range(len(chan_var)):

                            var = chan_var[k]

                            if self.main_all:

                                row_index = 4+7*j+1+k

                            else:

                                if var in self.fat_vars.keys():

                                    row_index = self.fat_vars[var]

                            if 'Blade' in var and 'axes' in var:
                                m = '10.0'
                            else:
                                m = '4.0'

                            fatigue = self.lp_var_fat[lp_name][var][m]['Total']/1000

                            char = get_column_letter(5*i+4)
                            rat = '=INDIRECT(ADDRESS(ROW(),COLUMN()-1))/INDIRECT(ADDRESS(ROW(),(%s$4*5-2)))' % char

                            sheet_fatigue.cell(row=row_index, column=col_start+i*5+1, value=var)
                            sheet_fatigue.cell(row=row_index, column=col_start+i*5+2, value='%.1f' % float(m))
                            sheet_fatigue.cell(row=row_index, column=col_start+i*5+3, value=fatigue)
                            sheet_fatigue.cell(row=row_index, column=col_start+i*5+4, value=rat)

                print('loop %s is done!' %(i+1))

            SetFormat().set_fatigue(sheet_fatigue)

            print('Fatigue loads is done! \n')
        # ***************************** Fatigue ***********************************************

        # ***************************** Fatigue_wind*******************************************
        if self.fat_wind and 'Fatigue_Wind' in sheet_names:

            fat_wind = table['Fatigue_Wind']

            ws_list = list(range(3,26))
            # print(ws_list)
            ws_num  = len(ws_list)

            col_start = fat_wind.max_column+1

            loop_num = int(ceil(fat_wind.max_column/14))

            print('Begin to write fatigue_wind...')

            for i in range(len(self.loopname)):

                print('Begin to write: ', self.loopname[i])

                lp_name  = self.loopname[i]
                channels = self.lp_channels[lp_name]  # list in order
                V_case   = self.lp_V_case[lp_name]
                ws_loop  = list(V_case.keys())        # wind speed in each loop
                # print(ws_num)
                ws_loop.sort()
                # print(ws_num)

                # 表头1：loop信息
                fat_wind.cell(1, col_start+i*14+1, loop_num+i+1)
                fat_wind.cell(2, col_start+i*14+1, lp_name)
                hyper3 = '=HYPERLINK("{}", "{}")'.format(self.fat_list[i], self.fat_list[i])
                fat_wind.cell(3, col_start+i*14+1, hyper3)

                # 数据区：
                for j in range(len(channels)):

                    chan_name = channels[j]

                    if chan_name:

                        chan_var = self.lp_chan_var[lp_name][chan_name]  # var list, len=6

                        for k in range(len(ws_loop)):

                            Vhub = ws_loop[k]

                            if Vhub in ws_loop:

                                for l in range(6):  # for k in range(len(chan_var))

                                    var = chan_var[l]

                                    if self.main_all:
                                        # row index for each channel
                                        row_index = 4+(2+ws_num+1)*j+1

                                    else:

                                        if var in self.fws_vars.keys():

                                            row_index = self.fws_vars[var]
                                    # print(self.main_all, row_index)

                                    # 表头2：变量名
                                    fat_wind.cell(row_index, col_start+i*14+1, value='Wind Speed')
                                    fat_wind.cell(row_index, col_start+i*14+2*(l+1), value=var)

                                    if 'Blade' in var and 'axes' in var:
                                        m = '10.0'
                                        m_var = 'm=10'
                                    else:
                                        m = '4.0'
                                        m_var = 'm=4'

                                    # m and index
                                    fat_wind.cell(row_index+1, col_start+i*14+1,         'm/s')
                                    fat_wind.cell(row_index+1, col_start+i*14+2*(l+1),   m_var)

                                    if l == 0:
                                        # Mx
                                        fat_wind.cell(row_index+1, col_start+i*14+2+1, 1)
                                    else:
                                        # Mx m值列
                                        col_value = '=%s%s' %(get_column_letter(col_start+i*14+2+1), row_index+1)
                                        fat_wind.cell(row_index+1, col_start+i*14+2*(l+1)+1, col_value)

                                    # 数据区：风速、等效疲劳载荷
                                    fat_tmp = 0

                                    for case in V_case[Vhub]['case']:

                                        fat_tmp += self.lp_var_fat[lp_name][var][m][case] ** float(m)

                                    fatigue = fat_tmp ** (1 / float(m)) / 1000
                                    # column letter for the ratio
                                    char = get_column_letter(col_start+i*14+2*(l+1)+1)
                                    rat = '=INDIRECT(ADDRESS(ROW(),COLUMN()-1))/' \
                                          'INDIRECT(ADDRESS(ROW(),((%s$%s-1)*14+2*(%i+1))))' %(char, row_index+1, l)

                                    # write wind
                                    row_ws = row_index+1+ws_list.index(Vhub)+1

                                    fat_wind.cell(row=row_ws, column=col_start+i*14+1,         value=Vhub)
                                    fat_wind.cell(row=row_ws, column=col_start+i*14+2*(l+1),   value=fatigue)
                                    fat_wind.cell(row=row_ws, column=col_start+i*14+2*(l+1)+1, value=rat)

                print('loop %s is done!' %(i+1))

            SetFormat().set_fatigue_wind(fat_wind)
            print('Wind fatigue is done!\n')
        # ***************************** Fatigue_wind*******************************************

        # ***************************** AEP for DLC12 *****************************************
        if self.aep_dlc and 'AEP_DLC12' in sheet_names:

            sheet_aep = table['AEP_DLC12']
            # print(sheet_aep)

            powercurve(sheet_aep, self.loopname, self.dlc_list, self.lct_list, False)

            SetFormat().set_aep(sheet_aep)

        # ***************************** AEP for DLC12 *****************************************

        # ***************************** Fatigue_case ******************************************
        # case fatigue: (all fatigue cases)
        if self.fat_case and 'Fatigue_Case' in sheet_names:

            fat_case = table['Fatigue_Case']

            row_max = fat_case.max_row
            col_max = fat_case.max_column

            loop_index = list(range(1,31))

            loop_row = {}

            for i in range(1, row_max+1):

                cell = fat_case.cell(row=i, column=1)

                if cell.value in loop_index:

                    loop_row[cell.value] = i

            loop_num = len(loop_row)
            print(loop_row)

            # col_start = 1
            row_start = fat_case.max_row+2

            # print(len(fat_chan_list))

            print('Begin to write fatigue_case...')

            for i in range(len(self.loopname)):

                print('Begin to write: ', self.loopname[i])

                lp_name  = self.loopname[i]
                channels = self.lp_channels[lp_name]  # list in order
                # col_len = len(self.lp_var_fat[lp_name]) + 3

                # 工况名、路径链接
                dlc_list = [k for k in self.lp_dlc_path[lp_name].keys()]
                dlc_list.remove('Total')
                dlc_list.sort()
                dlc_path = self.lp_dlc_path[lp_name]

                # 添加表头：
                # 当loop变量个数不一样，i*col_len+1 的方式不适用，表头写入移到数据循环后
                # 当取3叶片最大时，len(self.lp_var_fat[lp_name])确定列数的方式也不适用
                fat_case.cell(row=row_start, column=1, value=loop_num+i+1)
                fat_case.cell(row=row_start+1, column=1, value=self.loopname[i])
                hyper3 = '=HYPERLINK("{}", "{}")'.format(self.fat_list[i], self.fat_list[i])
                fat_case.cell(row=row_start+2, column=1, value=hyper3)

                # 添加变量名
                # fat_case.cell(row=4, column=col_start, value='m')
                fat_case.cell(row=row_start+4, column=1, value='DLC-Wind Speed')
                fat_case.cell(row=row_start+5, column=1, value='RunName')
                fat_case.cell(row=row_start+6, column=1, value='Path')

                for p in range(len(dlc_list)):
                    hyper4 = '=HYPERLINK("{}", "{}")'.format(dlc_path[dlc_list[p]], dlc_path[dlc_list[p]])
                    fat_case.cell(row=row_start+5, column=2+p+1, value=dlc_list[p])
                    fat_case.cell(row=row_start+6, column=2+p+1, value=hyper4)

                # 数据区
                for j in range(len(channels)):

                    chan_name = channels[j]
                    # print(chan_name)

                    if chan_name:

                        chan_var = self.lp_chan_var[lp_name][chan_name]  # list, len=6
                        # print(len(chan_var))

                        for k in range(6):

                            var = chan_var[k]
                            if self.main_all:

                                chan_num = len(channels)     #6
                                row_start = (chan_num*13+7)*(i+loop_num)+7+13*j+2*k+1

                            else:

                                # 记录第一个loop的channel数
                                if len(loop_row) == 1:
                                    chan_num = int(ceil((row_max-7)/13))
                                else:
                                    chan_num = int(ceil(loop_row[2]-loop_row[1]-7)/13)
                                # print(chan_num)

                                if var in self.fca_vars.keys():

                                    row_start = (chan_num*13+7)*(i+loop_num)+self.fca_vars[var]
                                    print(row_start)

                            if 'Blade' in var and 'axes' in var:
                                m = '10.0'
                                m_var = 'm=10'
                            else:
                                m = '4.0'
                                m_var = 'm=4'

                            # 添加表头变量名信息：
                            fat_case.cell(row=row_start, column=1, value=var)
                            fat_case.cell(row=row_start+1, column=2, value=m_var)
                            # print(row_start+2*k+1, lp_name, var)

                            # Mx的index
                            if k == 0:
                                fat_case.cell(row=row_start, column=2, value=1)
                                fat_case.cell(row=row_start+2, column=2, value='=B$%s' %(row_start))
                                fat_case.cell(row=row_start+4, column=2, value='=B$%s' % (row_start))
                                fat_case.cell(row=row_start+6, column=2, value='=B$%s' % (row_start))
                                fat_case.cell(row=row_start+8, column=2, value='=B$%s' % (row_start))
                                fat_case.cell(row=row_start+10, column=2, value='=B$%s' % (row_start))

                            # 各工况疲劳载荷：
                            for p in range(len(dlc_list)):

                                fat_eq = self.lp_var_fat[lp_name][var][m][dlc_list[p]]/1000
                                fat_case.cell(row=row_start+1, column=2+p+1, value=fat_eq)

                                rat = '=INDIRECT(ADDRESS(ROW()+1,COLUMN()))/' \
                                      'INDIRECT(ADDRESS((%s+($B%s-%s)*(7+13*%s)),COLUMN()))' \
                                      %(row_start+1, row_start, loop_num+i+1, chan_num)
                                fat_case.cell(row=row_start, column=2+p+1, value=rat)

                # 记录当前loop所占列数：
                row_start = fat_case.max_row+2
                # print(row_start)
                print('loop %s is done!' % (i+1))

            # 设置格式
            SetFormat().set_fatigue_case(fat_case)
            print('Case fatigue is done!\n')
        # ***************************** Fatigue_case ******************************************

        # **************************** Save excel *********************************************
        table.save(self.out_path)
        # table.save(table_path)
        # **************************** Save excel *********************************************


if __name__ == '__main__':

    input  = r"E:\05 TASK\02_Tools\01_Load Summary\LoadSummaryTemplate_test.xlsx"
    output = r"E:\05 TASK\02_Tools\01_Load Summary\test_20200109.xlsx"


    # def __init__(self,
    #              excel_input,
    #              excel_output,
    #              ultimate=False,
    #              fatigue=False,
    #              heat_map=False,
    #              fat_case=False,
    #              fat_wind=False,
    #              aep_dlc=False):

    excel_operation(input, output, True, True, True, True, True, False)