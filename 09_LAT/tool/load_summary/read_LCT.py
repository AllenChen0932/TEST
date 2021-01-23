# -*- coding: utf-8 -*-
# @Time    : 2019/12/06 12:06
# @Author  : CJG
# @File    : read_LCT.py
"""
查找部分改为找到后即跳出循环，节省时间；
改为不区分大小写的查找；
"""
import openpyxl

__version__ = "1.0.1"
'''
2020.4.15_v1.0.1 modify the logic for getting load case in LCT
'''

def read_LCT_dlc12(LCT_file):
    """
    读取工况表，取工况名、风速、时间：
    {case: {'Vhub':value, 'Hour':value}}
    {Vhub: {'case':list, 'hour':list}}
    """
    case_V = {}
    V_case = {}

    table = openpyxl.load_workbook(LCT_file, keep_vba=True, data_only=True)
    sheet = table.get_sheet_by_name('DLC12')

    name_list = ['Run Name', 'VHub', 'Hour/Year']
    unit_list = [None, 'm/s', 'hours']      # openpyxl：空单元格为None或NoneType
    pos       = search_table(sheet, name_list, unit_list)
    # print(pos)

    row_start = pos[name_list[0]][0]+2
    row_end   = sheet.max_row

    Vhub_1    = sheet.cell(row=row_start, column=pos['VHub'][1]).value

    case_list = []
    hour_list = []
    v_list    = []

    for i in range(row_start, row_end+1):
        if sheet.cell(row=i, column=pos['Run Name'][1]).value:

            case_name = sheet.cell(row=i, column=pos['Run Name'][1]).value
            Vhub      = sheet.cell(row=i, column=pos['VHub'][1]).value
            Hour      = sheet.cell(row=i, column=pos['Hour/Year'][1]).value
            # print(case_name, Vhub, Hour)

            if case_name:
                if Vhub not in v_list:
                    v_list.append(Vhub)
                case_V[case_name] = {'Vhub': Vhub, 'Hour': Hour}

                if Vhub == Vhub_1:
                    case_list.append(case_name)
                    hour_list.append(Hour)
                    V_case[Vhub] = {'case': case_list, 'hour': hour_list}

                else:
                    Vhub_1 = Vhub
                    case_list = []
                    hour_list = []
                    case_list.append(case_name)
                    hour_list.append(Hour)
                    V_case[Vhub] = {'case': case_list, 'hour': hour_list}
        else:
            # if the blank occurs, then stop the loop
            break
    table.close()

    return case_V, V_case, v_list

def search_table(sheet, name_list, unit_list):
    # 返回指定变量所在的行号、列号
    pos = {}

    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            for index, name in enumerate(name_list):
                # -----------不区分大小写地搜索-------------
                name_value = sheet.cell(row=i,   column=j).value
                unit_value = sheet.cell(row=i+1, column=j).value

                if unit_value == unit_list[index] and name_value == name:
                    # print(sheet.cell(row=i, column=j))
                    pos[name] = [i, j]

                    if len(pos) == len(name_list):
                        break
    return pos

if __name__ == '__main__':

    LCT_file = r"\\172.20.4.132\fs02\CE\V3\loop06\performance\power_curve\V3_loop06_v7.5_AllWindSpeedTi=10%.xlsm"
    read_LCT_dlc12(LCT_file)
