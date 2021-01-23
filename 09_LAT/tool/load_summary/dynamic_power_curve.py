#！usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2019/8/19 8:28
# @Author  : CE
# @File    : dynamic_powercurve_v1.1.py
"""
读取工况表改为自动寻找，移除人为给定的参数
AEP=sum(每个时序平均功率*对应时间)
统计power时工况名-速度关系也改为读工况表
优化函数流程（区分读取时间、计算AEP）
修复风速非整数的bug
输出每个run的功率散点
新建class优化对多个loop结果的处理
read_percent,read_dollar修复
"""

import os
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.chart import (ScatterChart, Reference, Series)
from tool.load_summary.read_LCT import read_LCT_dlc12

__version__ = "2.0.1"
'''
2020.4.21_v2.0.1 repair the bug for calculating AEP for each load case
'''

class stastic:
    def __init__(self, array):
        self.array = array

        self.max = None
        self.min = None
        self.mean = None
        self.std = None

        self.get_max()
        self.get_min()
        self.get_mean()
        self.get_std()

    def get_mean(self):
        self.mean = np.mean(self.array)

    def get_max(self):
        self.max = np.max(self.array)

    def get_min(self):
        self.min = np.min(self.array)

    def get_std(self):
        self.std = np.std(self.array)

class handle_pc:

    def __init__(self, dlc12_path, proj_excel):

        self.path      = dlc12_path       #dlc path which contains 12
        self.proj_excel= proj_excel       #load case table
        self.power_case= {}
        self.windspeed = []
        self.loadcase  = {}
        self.access    = None
        self.recl      = None
        self.dim       = None
        self.pow_ind   = None
        self.data      = None
        self.case_V = {}
        self.V_case = {}

        self.read_loadcasetable()
        self.read_percent()
        self.handle()

    def read_loadcasetable(self):

        # 读工况表得到每个工况对应风速及时间：
        # case_V: {case: {'Vhub':value, 'Hour':value}}
        # V_case: {Vhub: {'case':list, 'hour':list}}
        self.case_V, self.V_case, self.windspeed = read_LCT_dlc12(self.proj_excel)

        # 取风速列表：[vhub list]
        # windspeed = [k for k in self.V_case.keys()]
        # windspeed.sort()
        # self.windspeed = windspeed

        # 取工况名列表（以风速索引）：{vhub:[runname list]}
        for v in self.windspeed:
            self.loadcase[v] = self.V_case[v]['case']

    # @pysnooper.snoop()
    def read_percent(self):

        # 读取任意一个.%06文件，确定数据格式：
        lc   = self.loadcase[self.windspeed[0]][0]
        path = self.path + os.sep + lc + os.sep + lc + '.%06'
        path = path.replace('/', '\\')

        if os.path.isfile(path):

            with open(path, 'r') as f:
                for line in f.readlines():

                    if 'ACCESS' in line:
                        self.access = line.split()[1]

                    if 'RECL' in line:
                        self.recl = line.split()[1]

                    if line.startswith('DIMENS'):
                        self.dim = line.split()[1:]
                        self.data_len = int(self.dim[1])

                    if 'VARIAB' in line:
                        var = line.split('\'')[1::2]
                        self.pow_ind = var.index('Electrical power')

            # print('read percent is done!')

        else:

            raise Exception('%s not exist!' %path)

    # @pysnooper.snoop()
    def read_dollar(self, loadcase):

        path = self.path + os.sep + loadcase + os.sep + loadcase + '.$06'
        path = path.replace('/', '\\')
        # print(path)
        data = None

        if os.path.isfile(path):

            if self.access == 'D':

                with open(path, 'rb') as f:
                    if self.recl == '4':
                        data = np.fromfile(f, np.float32)
                    elif self.recl == '8':
                        data = np.fromfile(f, np.float64)

            elif self.access == 'S':

                with open(path, 'r') as f:

                    data = np.loadtxt(f)

            # .$06文件中存有4*data_len的数据，第2列为电功率：
            data = data.reshape(int(self.dim[1]), int(self.dim[0]))
            data = data[:, self.pow_ind]
            # print(np.shape(data))

            return data

        else:

            raise Exception('%s not exist!' % path)

    def handle(self):

        for ws in self.windspeed:

            # empty()的参数是list
            # temp = np.empty([int(self.data_len), len(self.loadcase[ws])])   #self.data_len是str类型，通过int()将其变成整数，此处很容易出错
            power_case = []

            for lc in self.loadcase[ws]:

                data = self.read_dollar(lc)

                power_case.append(stastic(data[:]).mean)

            self.power_case[ws] = power_case
        print('Read is done!')

class output_PCs:

    def __init__(self, sheet, loop_list, dlc12_list, lct_list, new_app=True):

        self.sheet    = sheet
        self.loopname = loop_list
        self.DLC12s   = dlc12_list
        self.lct_list = lct_list
        self.new_app  = new_app

        self.lp_power_case = {}
        self.lp_windspeed  = {}
        self.lp_loadcase   = {}
        self.lp_case_V     = {}
        self.lp_V_case     = {}

        self.write_result()

    # @pysnooper.snoop()
    def get_result(self):

        for i in range(len(self.loopname)):
            print('Begin to read DLC12: %s' %self.DLC12s[i])
            if not self.DLC12s[i]:
                continue

            result = handle_pc(self.DLC12s[i], self.lct_list[i])
            # print('read power data is done!')
            self.lp_power_case[self.loopname[i]] = result.power_case
            self.lp_windspeed[self.loopname[i]]  = result.windspeed
            self.lp_loadcase[self.loopname[i]]   = result.loadcase
            self.lp_case_V[self.loopname[i]]     = result.case_V
            self.lp_V_case[self.loopname[i]]     = result.V_case
        print()

    def write_result(self):

        self.get_result()

        if self.new_app:
            print('Begin to write AEP...')

            ws_num   = (len(__) for __ in self.lp_windspeed.values())
            vnum_max = max(list(ws_num))

            # 写入数据：
            for i in range(len(self.loopname)):
                print('Begin to write: {}'.format(self.loopname[i]))

                if not self.DLC12s[i]:
                    continue

                lp_name = self.loopname[i]
                ws_num   = len(self.lp_windspeed[lp_name])

                # ---------------分风速的平均功率----------------
                # 写表头
                self.sheet.cell(1, 1+6*i, i+1)
                self.sheet.cell(2, 1+6*i, self.loopname[i])
                self.sheet.cell(3, 1+6*i, self.DLC12s[i])
                self.sheet.cell(5, 1+6*i, 'Index')
                self.sheet.cell(6, 1+6*i, '-')
                self.sheet.cell(5, 2+6*i, 'Wind Speed')
                self.sheet.cell(6, 2+6*i, 'm/s')
                self.sheet.cell(5, 3+6*i, 'Power')
                self.sheet.cell(6, 3+6*i, 'Kw')
                self.sheet.cell(5, 4+6*i, 'Time/year')
                self.sheet.cell(6, 4+6*i, 'hour')
                self.sheet.cell(5, 5+6*i, 'AEP')
                self.sheet.cell(6, 5+6*i, 'Kwh')

                # ---------------分工况的平均功率----------------
                # 写表头
                row_start = 6+vnum_max+2+1          # row number for the 'RunName'
                self.sheet.cell(row_start,   1+6*i, 'RunName')
                self.sheet.cell(row_start+1, 1+6*i, '-')
                self.sheet.cell(row_start,   2+6*i, 'WindSpeed')
                self.sheet.cell(row_start+1, 2+6*i, 'm/s')
                self.sheet.cell(row_start,   3+6*i, 'Power')
                self.sheet.cell(row_start+1, 3+6*i, 'Kw')
                self.sheet.cell(row_start,   4+6*i, 'Time/Year')
                self.sheet.cell(row_start+1, 4+6*i, 'hour')
                self.sheet.cell(row_start,   5+6*i, 'AEP')
                self.sheet.cell(row_start+1, 5+6*i, 'Kwh')

                # 记录各列索引：
                col_vhub  = get_column_letter(2+6*i)
                col_power = get_column_letter(3+6*i)
                col_time  = get_column_letter(4+6*i)
                col_aep   = get_column_letter(5+6*i)

                self.sheet.column_dimensions[col_aep].width = 12

                # 写入各工况的数据
                row_start += 2            # row number for the load case name
                for vhub in self.lp_windspeed[lp_name]:
                    lc_list = self.lp_loadcase[lp_name][vhub]

                    for j in range(len(lc_list)):

                        lc = lc_list[j]
                        self.sheet.cell(row_start+j, 1+6*i, lc)
                        self.sheet.cell(row_start+j, 2+6*i, vhub)
                        self.sheet.cell(row_start+j, 3+6*i, self.lp_power_case[lp_name][vhub][j]/1000)
                        self.sheet.cell(row_start+j, 4+6*i, self.lp_V_case[lp_name][vhub]['hour'][j])
                        char = '='+col_power+str(row_start+j)+'*'+col_time+str(row_start+j)
                        self.sheet.cell(row_start+j, 5+6*i, char)

                    row_start += len(lc_list)

                # 记录数据区域：
                # row number for the average area
                row_s = 5+vnum_max+4
                row_e = row_start-1

                # 记录各列索引
                col_vhub_2 = get_column_letter(2+6*i)
                col_aep_2  = get_column_letter(5+6*i)

                # row number for windspeed start and end
                row_ws1 = 7
                row_ws2 = row_ws1+ws_num-1

                # 写入数据
                for j in range(ws_num):

                    self.sheet.cell(row_ws1+j, 1+6*i, j+1)
                    self.sheet.cell(row_ws1+j, 2+6*i, self.lp_windspeed[lp_name][j])

                    char1 = '=AVERAGEIFS('+col_power+str(row_s)+':'+col_power+str(row_e)+','\
                           +col_vhub+str(row_s)+':'+col_vhub+str(row_e)+','+'"="&'+col_vhub_2+str(row_ws1+j)+')'
                    self.sheet.cell(row_ws1+j, 3+6*i, char1)

                    char2 = '=SUMIFS('+col_time+str(row_s)+':'+col_time+str(row_e)+','\
                           +col_vhub+str(row_s)+':'+col_vhub+str(row_e)+','+'"="&'+col_vhub_2+str(row_ws1+j)+')'
                    self.sheet.cell(row_ws1+j, 4+6*i, char2)

                    char3 = '=SUMIFS('+col_aep+str(row_s)+':'+col_aep+str(row_e)+','\
                           +col_vhub+str(row_s)+':'+col_vhub+str(row_e)+','+'"="&'+col_vhub_2+str(row_ws1+j)+')'
                    self.sheet.cell(row_ws1+j, 5+6*i, char3)

                char4 = '=SUM('+col_aep_2+str(row_ws1)+':'+col_aep_2+str(row_ws2)+')'
                self.sheet.cell(row_ws2+1, 4+6*i, 'AEP sum')
                self.sheet.cell(row_ws2+1, 5+6*i, char4)

                print('loop {} is done!'.format(i+1))

        else:
            print('Begin to write AEP...')

            ws_num   = (len(__) for __ in self.lp_windspeed.values())
            vnum_max = max(list(ws_num))

            col_max = self.sheet.max_column
            row_max = self.sheet.max_row
            loop_num = int((col_max+1)/6)
            # print(col_max, row_max, loop_num)

            row_rname = [row for row in range(7, row_max+1) if 'RunName' == self.sheet.cell(row=row, column=1).value]
            # print(row_rname)

            # 判断风速的个数，如果多余原有数据，那么增加行：
            if row_rname[0] < 6+vnum_max+3:
                self.sheet.insert_rows(row_rname[0], 6+vnum_max+3)

            # 写入数据：
            for i in range(len(self.loopname)):
                print('Begin to write: {}'.format(self.loopname[i]))

                lp_name = self.loopname[i]
                ws_num  = len(self.lp_windspeed[lp_name])

                row_start = 6+vnum_max+2+1          # row number for the 'RunName'
                col_start = col_max

                # ---------------分风速的平均功率----------------
                # 写表头
                self.sheet.cell(1, 1+6*i+col_start, i+loop_num+1)
                self.sheet.cell(2, 1+6*i+col_start, self.loopname[i])
                self.sheet.cell(3, 1+6*i+col_start, self.DLC12s[i])
                self.sheet.cell(5, 1+6*i+col_start, 'Index')
                self.sheet.cell(6, 1+6*i+col_start, '-')
                self.sheet.cell(5, 2+6*i+col_start, 'Wind Speed')
                self.sheet.cell(6, 2+6*i+col_start, 'm/s')
                self.sheet.cell(5, 3+6*i+col_start, 'Power')
                self.sheet.cell(6, 3+6*i+col_start, 'Kw')
                self.sheet.cell(5, 4+6*i+col_start, 'Time/year')
                self.sheet.cell(6, 4+6*i+col_start, 'hour')
                self.sheet.cell(5, 5+6*i+col_start, 'AEP')
                self.sheet.cell(6, 5+6*i+col_start, 'Kwh')

                # ---------------分工况的平均功率----------------
                # 写表头
                self.sheet.cell(row_start,   1+6*i+col_start, 'RunName')
                self.sheet.cell(row_start+1, 1+6*i+col_start, '-')
                self.sheet.cell(row_start,   2+6*i+col_start, 'WindSpeed')
                self.sheet.cell(row_start+1, 2+6*i+col_start, 'm/s')
                self.sheet.cell(row_start,   3+6*i+col_start, 'Power')
                self.sheet.cell(row_start+1, 3+6*i+col_start, 'Kw')
                self.sheet.cell(row_start,   4+6*i+col_start, 'Time/Year')
                self.sheet.cell(row_start+1, 4+6*i+col_start, 'hour')
                self.sheet.cell(row_start,   5+6*i+col_start, 'AEP')
                self.sheet.cell(row_start+1, 5+6*i+col_start, 'Kwh')

                # 记录各列索引：
                col_vhub  = get_column_letter(2+6*i+col_start)
                col_power = get_column_letter(3+6*i+col_start)
                col_time  = get_column_letter(4+6*i+col_start)
                col_aep   = get_column_letter(5+6*i+col_start)

                # 写入数据
                row_start += 2            # row number for the load case name
                for vhub in self.lp_windspeed[lp_name]:
                    lc_list = self.lp_loadcase[lp_name][vhub]

                    for j in range(len(lc_list)):

                        lc = lc_list[j]
                        self.sheet.cell(row_start+j, 1+6*i+col_start, lc)
                        self.sheet.cell(row_start+j, 2+6*i+col_start, vhub)
                        self.sheet.cell(row_start+j, 3+6*i+col_start, self.lp_power_case[lp_name][vhub][j]/1000)
                        self.sheet.cell(row_start+j, 4+6*i+col_start, self.lp_V_case[lp_name][vhub]['hour'][j])
                        char = '='+col_power+str(row_start+j)+'*'+col_time+str(row_start+4+j)
                        self.sheet.cell(row_start+j, 5+6*i+col_start, char)

                    row_start += len(lc_list)

                # 记录数据区域：
                # row number for the average area
                row_s = 5+vnum_max+4
                row_e = row_start-1

                # 记录各列索引
                col_vhub_2 = get_column_letter(2+6*i+col_start)
                col_aep_2  = get_column_letter(5+6*i+col_start)

                # row number for windspeed start and end
                row_ws1 = 7
                row_ws2 = row_ws1+ws_num-1

                # 写入数据
                for j in range(ws_num):

                    self.sheet.cell(row_ws1+j, 1+6*i+col_start, j+1)

                    self.sheet.cell(row_ws1+j, 2+6*i+col_start, self.lp_windspeed[lp_name][j])

                    char1 = '=AVERAGEIFS('+col_power+str(row_s)+':'+col_power+str(row_e)+','\
                           +col_vhub+str(row_s)+':'+col_vhub+str(row_e)+','+'"="&'+col_vhub_2+str(row_ws1+j)+')'
                    self.sheet.cell(row_ws1+j, 3+6*i+col_start, char1)

                    char2 = '=SUMIFS('+col_time+str(row_s)+':'+col_time+str(row_e)+','\
                           +col_vhub+str(row_s)+':'+col_vhub+str(row_e)+','+'"="&'+col_vhub_2+str(row_ws1+j)+')'
                    self.sheet.cell(row_ws1+j, 4+6*i+col_start, char2)

                    char3 = '=SUMIFS('+col_aep+str(row_s)+':'+col_aep+str(row_e)+','\
                           +col_vhub+str(row_s)+':'+col_vhub+str(row_e)+','+'"="&'+col_vhub_2+str(row_ws1+j)+')'
                    self.sheet.cell(row_ws1+j, 5+6*i+col_start, char3)

                char4 = '=SUM('+col_aep_2+str(row_ws1)+':'+col_aep_2+str(row_ws2)+')'
                self.sheet.cell(row_ws2+1, 4+6*i+col_start, 'AEP sum')
                self.sheet.cell(row_ws2+1, 5+6*i+col_start, char4)

                print('loop {} is done!'.format(i + 1))

        print('Begin to create dynamic power curve...')

        # 作图
        line_chart = ScatterChart()
        line_chart.title = 'Dynamic Power Curve'  # 图的标题
        line_chart.style = 10  # 线条的style
        # line_chart.scatterStyle = 'marker'
        line_chart.y_axis.title = 'Power(kw)'  # y坐标的标题
        line_chart.x_axis.title = "Wind Speed(m/s)"  # x坐标的标题

        col_max  = self.sheet.max_column
        row_max  = self.sheet.max_row
        loop_num = int((col_max+1)/6)

        # row number for 'RunName'
        row_rname = [row for row in range(7, row_max) if 'RunName' in str(self.sheet.cell(row=row, column=1).value)]

        for i in range(loop_num):
            ws_num  = [row for row in range(7, row_max) if not self.sheet.cell(row, 1+6*i).value]
            # print(ws_num)
            lp_name = self.sheet.cell(2, 6*i+1).value

            yvalues = Reference(self.sheet, min_col=6*i+3, min_row=7, max_row=ws_num[0]-1)  # 图像的数据 起始行、起始列、终止行、终止列
            xvalues = Reference(self.sheet, min_col=6*i+2, min_row=7, max_row=ws_num[0]-1)
            series  = Series(yvalues, xvalues, title=lp_name)
            line_chart.series.append(series)

        line_chart.width = 30
        line_chart.height = 15

        cell_chart = 'A'+str(row_rname[0])
        # print(cell_chart)
        self.sheet.add_chart(line_chart, cell_chart)  # 将图表添加到 sheet中
        # self.sheet.add_chart(line_chart)

        print('Dynamic power curve is done!\n')

if __name__ == '__main__':

    from openpyxl import load_workbook

    DLC12_pathes = [r'E:\05 TASK\02_Tools\01_Load Summary\DLC12']

    res_path = r"E:\05 TASK\02_Tools\01_Load Summary\DLC12\DLC12.xlsx"
    table=load_workbook(res_path)
    sheet=table.create_sheet('AEP')

    proj_excels   = [r"E:\05 TASK\02_Tools\01_Load Summary\DLC12\W2500-135-90.xlsm"]

    output_PCs(sheet, ['test'], DLC12_pathes, proj_excels)
    table.save(res_path)

