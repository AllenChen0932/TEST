# -*- coding: utf-8 -*-
# @Time    : 2020/07/31 11:10
# @Author  : TJ
# @File    : writeOccur.py

# import os
import openpyxl
# import pandas as pd
# from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class Occurrence(object):

    def __init__(self, project, table_object, sheet_name):
        
        self.lct_table  = project
        self.table_objt = table_object
        self.sheet_name = sheet_name

        # self.dlc_list = ['DLC12', 'DLC24', 'DLC31', 'DLC41', 'DLC64', 'DLC72']

        self.dlc_data  = {}
        self.lc_list   = []
        self.life_time = None

        self.read_proj()
        self.write_result()

    def search_table(self, sheet, name_list):

        position = {}
        for i in range(1, sheet.max_column+1):
            for j in range(1, sheet.max_row+1):
                for name in name_list:
                    value = sheet.cell(row=i, column=j).value

                    if type(value) == str and str.lower(value) == str.lower(name):
                        # print(sheet.cell(row=i, column=j))
                        position[name] = [i, j]

                        if len(position) == len(name_list):
                            break
        return position

    def read_proj(self):

        xls = openpyxl.load_workbook(self.lct_table, keep_vba=True, data_only=True)

        # read setting
        sheet   = xls['Setting']
        col_max = sheet.max_row
        sheets  = [name for name in xls.sheetnames if name.startswith('DLC')]
        sheets.sort()

        for i in range(1, col_max+1):
            if sheet.cell(i,2).value == 'T_life':
                self.life_time = float(sheet.cell(i,3).value)
                break

        # DLC12
        for dlc in sheets:
            
            if dlc == 'DLC12' or dlc == 'DLC16' or dlc == 'DLC64' or dlc == 'DLC65' or dlc == 'DLC72':
                sheet   = xls[dlc]
                row_max = sheet.max_row
                var_pos = self.search_table(sheet, ['Run Name', 'tout', 'tSim', 'Hour/Year', 'Times/year'])
    
                row_start = var_pos['Run Name'][0]
                col_run   = get_column_letter(var_pos['Run Name'][1])
                col_tout  = get_column_letter(var_pos['tout'][1])
                col_tSim  = get_column_letter(var_pos['tSim'][1])
                col_hours = get_column_letter(var_pos['Hour/Year'][1])
                for row in range(row_start+2, row_max+1):

                    lc  = sheet[col_run+str(row)].value
                    if lc:
                        tout = float(sheet[col_tout+str(row)].value)
                        tsim = float(sheet[col_tSim+str(row)].value)
                        hour = float(sheet[col_hours+str(row)].value)
                        self.dlc_data[lc] = [tout, tsim, hour]
                        self.lc_list.append(lc)
            
            elif dlc == 'DLC31' or dlc == 'DLC41' or dlc.startswith('DLC24'):
                sheet   = xls[dlc]
                row_max = sheet.max_row
                var_pos = self.search_table(sheet, ['Run Name', 'tout', 'tSim', 'Times/year'])

                row_start = var_pos['Run Name'][0]
                col_run   = get_column_letter(var_pos['Run Name'][1])
                col_tout  = get_column_letter(var_pos['tout'][1])
                col_tSim  = get_column_letter(var_pos['tSim'][1])
                col_time  = get_column_letter(var_pos['Times/year'][1])
                
                for row in range(row_start+2, row_max+1):
                    lc  = sheet[col_run+str(row)].value
                    if lc:
                        tout = float(sheet[col_tout+str(row)].value)
                        tsim = float(sheet[col_tSim+str(row)].value)
                        time = float(sheet[col_time+str(row)].value)
                        self.dlc_data[lc] = [tout, tsim, time]
                        self.lc_list.append(lc)
        # print(self.lc_list)
        # print(self.dlc_data)

    def write_result(self):

        if self.sheet_name in self.table_objt.sheetnames:
            sheet = self.table_objt[self.sheet_name]
        else:
            sheet = self.table_objt.create_sheet(self.sheet_name)
        print(sheet)

        sheet.cell(1, 1, value='Load Case')
        sheet.cell(1, 2, value='Hours in %s years' %int(self.life_time))
        sheet.cell(1, 3, value='Times in %s years' %int(self.life_time))
        sheet.cell(1, 4, value='Simulation Time')

        for i, lc in enumerate(self.lc_list):

            sim_time = self.dlc_data[lc][1] - self.dlc_data[lc][0]

            # load case defined by hours
            if lc.startswith('12') or lc.startswith('16') or lc.startswith('24c') or lc.startswith('24d')\
                    or lc.startswith('64') or lc.startswith('65') or lc.startswith('L12') or lc.startswith('L16')\
                    or lc.startswith('L64') or lc.startswith('L65'):
                sheet.cell(i+2, 1, value = lc)
                sheet.cell(i+2, 2, value = self.dlc_data[lc][2]*self.life_time)
                sheet.cell(i+2, 3, value = self.dlc_data[lc][2]*self.life_time*3600/sim_time)
                sheet.cell(i+2, 4, value = sim_time)

            # load case defined by times
            else:
                sheet.cell(i+2, 1, value = lc)
                sheet.cell(i+2, 2, value = self.dlc_data[lc][2]*self.life_time*sim_time/3600)
                sheet.cell(i+2, 3, value = self.dlc_data[lc][2]*self.life_time)
                sheet.cell(i+2, 4, value = sim_time)

        # table.save(self.table_path)

if __name__ == '__main__':

    project = r"\\172.20.0.4\fs02\CE\V3\loop05\project\V3_loop05_v8.0_20200622_c2.0.xlsm"
    sheet   = 'occurrence'
    table   = r"C:\Users\10700700\Desktop\tool\py\20_Load Assistant Tool\test\report\Tower.xlsx"
    table   = openpyxl.load_workbook(table)

    result  = Occurrence(project, table, sheet)