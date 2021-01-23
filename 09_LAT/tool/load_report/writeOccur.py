# -*- coding: utf-8 -*-
# @Time    : 2020/07/31 11:10
# @Author  : TJ
# @File    : writeOccur.py

import os
import openpyxl
import pandas as pd
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment

class Occur(object):

    def __init__(self, 
                 project, 
                 inpath, 
                 outpath, 
                 DLC12=True, 
                 DLC24=True, 
                 DLC31=True, 
                 DLC41=True, 
                 DLC64=True):
        
        self.project = project
        self.inpath  = inpath
        self.outpath = outpath
        self.name_list = ['Run Name', 'VHub', 'Yaw', 'tout', 'tSim']
        self.DLC12 = DLC12
        self.DLC24 = DLC24
        self.DLC31 = DLC31
        self.DLC41 = DLC41
        self.DLC64 = DLC64
        self.sheet_list = []
        self.time_col_pos = []

    def read_proj(self):

        xls = openpyxl.load_workbook(self.project, keep_vba=True, data_only=True)
        xls_sheet = [sh.title for sh in xls]

        dict_info = {}

        if self.DLC12:
            self.sheet_list.append('DLC12')
            self.time_col_pos.extend(['W', 'X'])
        if self.DLC24:
            if 'DLC24' in xls_sheet:
                self.sheet_list.append('DLC24')
                self.time_col_pos.extend(['Z', 'AA'])
            else:
                self.sheet_list.extend(['DLC24a', 'DLC24b', 'DLC24c', 'DLC24d'])
                self.time_col_pos.extend(['Z', 'AA', 'Z', 'AA', 'Y', 'Z', 'Z', 'AA'])
        if self.DLC31:
            self.sheet_list.append('DLC31')
            self.time_col_pos.extend(['N', 'O'])
        if self.DLC41:
            self.sheet_list.append('DLC41')
            self.time_col_pos.extend(['O', 'P'])
        if self.DLC64:
            self.sheet_list.append('DLC64')
            self.time_col_pos.extend(['W', 'X'])

        # print(self.time_col_pos)

        for j in range(len(self.name_list)):
            proj_case    = []
            proj_status  = []
            proj_occur   = []
            proj_hour    = []
            self.numcase = []

            for s_id, s in enumerate(self.sheet_list):

                sheet     = xls[s]
                position  = self.search_table(sheet, self.name_list)

                row_start = position[self.name_list[0]][0] + 2
                row_end   = sheet.max_row
                num       = 0

                for i in range(row_start, row_end + 1):
                    case   = sheet.cell(row=i, column=position[self.name_list[j]][1]).value
                    occur  = sheet['{}{}'.format(self.time_col_pos[s_id*2], i)].value
                    hour   = sheet['{}{}'.format(self.time_col_pos[s_id*2+1], i)].value
                    status = sheet['E4'].value

                    proj_case.append(case)
                    proj_occur.append(occur)
                    proj_hour.append(hour)
                    proj_status.append(status)

                    num += 1

                self.numcase.append(num)

            proj_col  = {self.name_list[j]: proj_case}
            dict_info = dict(**dict_info, **proj_col)

        xls.close()

        self.df_case = pd.DataFrame(dict_info)

        self.df_case['tSim'] = self.df_case['tSim']-self.df_case['tout']
        self.df_case.drop(columns=['tout'], inplace=True)

        proj_col2 = {'occurences in 25 years': proj_occur, 'hours in 25 years': proj_hour}
        df_2      = pd.DataFrame(proj_col2)

        proj_col3 = {'status': proj_status}
        df_3      = pd.DataFrame(proj_col3)

        self.df_occu = pd.merge(df_2, df_3, left_index=True, right_index=True, how='outer')

    def search_table(self, sheet, name_list):

        position = {}
        for i in range(1, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                for name in name_list:
                    value = sheet.cell(row=i, column=j).value

                    if type(value) == str and str.lower(value) == str.lower(name):
                        # print(sheet.cell(row=i, column=j))
                        position[name] = [i, j]

                        if len(position) == len(name_list):
                            break
        return position

    def read_pj_step(self, df_case, df_occu, numcase):

        path     = os.path.dirname(self.inpath)
        simstep  = []
        timestep = None

        for i, dlc in enumerate(self.sheet_list):

            dlc_i = sum(numcase[0:i])
            pj_name = df_case.loc[dlc_i, [self.name_list[0]]].values
            pj_name = '.'.join(pj_name)       #array to str

            dlc_path = os.sep.join([path, 'run\\' + dlc])
            pj_path = os.sep.join([dlc_path, pj_name + '\\' + pj_name + '.$PJ'])

            with open(pj_path, 'r') as f:
                lines = f.readlines()

                for line in lines:
                    if 'OPSTP' in line:
                        timestep = float(line.strip().split('\t')[1])

            for j in range(dlc_i, dlc_i + numcase[i]):
                simstep.append(timestep)

        dict_simstep = {'simulation step': simstep}
        df_simstep = pd.DataFrame(dict_simstep)

        df_simstep['No. of steps'] = df_case['tSim'] / df_simstep['simulation step']

        df = pd.merge(df_case, df_simstep, left_index=True, right_index=True, how='outer')
        df = pd.merge(df, df_occu, left_index=True, right_index=True, how='outer')

        df = df.rename(columns={'Run Name': 'dlc', 'VHub': 'wind speed', 'Yaw': 'yaw error', 'tSim': 'simulation'})

        # print(df)
        return(df)

    def write_result(self, df, xls):

        df.index += 2
        df.loc[0] = ['run case', 'wind speed', 'yaw error', 'sim time', 'sim step', 'No. of steps',
                     'occurences in 25y', 'hours in 25y', 'status']
        df.loc[1] = ['(-)', '(m/s)', '(°)', '(s)', '(s)', '(-)', '(-)', '(h)', '(-)']
        df.sort_index(inplace=True)

        newsheet   = 'occurrence'
        sheet_name = xls.sheetnames

        if newsheet in sheet_name:
            sheet = xls[newsheet]
            xls.remove(sheet)
        sheet = xls.create_sheet(newsheet, 0)

        for i in range(0, df.shape[0]):
            for j in range(0, df.shape[1]):
                sheet.cell(row=i + 1, column=j + 1, value=df.iloc[i][j])

        left, right, top, bottom = [Side(style='thin')]*4
        for irow, row in enumerate(sheet.rows, start=1):
            for cell in row:
                cell.font = Font(name='Times New Roman', size=10)
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

                if irow == 1 or irow == 2:
                    cell.font = Font(name='Times New Roman', size=10, bold=True, italic=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill('solid', fgColor='FFEFD5')

        for col in ['F', 'G', 'H', 'I']:
            sheet.column_dimensions[col].width = 15

    def write_result_2(self, df, xls):

        df = df.loc[:, ['dlc', 'simulation', 'occurences in 25 years', 'hours in 25 years']]
        df.index += 2
        df.loc[0] = ['run case', 'sim time', 'occurences in 25y', 'hours in 25y']
        df.loc[1] = ['(-)', '(s)', '(-)', '(h)']
        df.sort_index(inplace=True)

        newsheet = 'occurrence'
        sheet_name = xls.sheetnames

        if newsheet in sheet_name:
            sheet = xls[newsheet]
            xls.remove(sheet)
        sheet = xls.create_sheet(newsheet, 0)

        for i in range(0, df.shape[0]):
            for j in range(0, df.shape[1]):
                sheet.cell(row=i+1, column=j+1, value=df.iloc[i][j])

        left, right, top, bottom = [Side(style='thin')]*4
        for irow, row in enumerate(sheet.rows, start=1):
            for cell in row:
                cell.font = Font(name='Times New Roman', size=10)
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

                if irow == 1 or irow == 2:
                    cell.font = Font(name='Times New Roman', size=10, bold=True, italic=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill('solid', fgColor='FFEFD5')

        for col in ['C', 'D']:
            sheet.column_dimensions[col].width = 15

if __name__ == '__main__':

    project = r"\\172.20.0.4\fs02\CE\V3\loop05\project\V3_loop05_v8.0_20200622_c2.0.xlsm"
    pj_path = r'\\172.20.0.4\fs02\CE\V3\loop05\run_0615\DLC11\11_aa-01'
    output  = r'C:\Users\10700700\Desktop\tool\py\20_Load Assistant Tool\tool\load_report'
    result  =Occur(project, pj_path, output)