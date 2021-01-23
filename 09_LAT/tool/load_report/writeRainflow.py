# -*- coding: utf-8 -*-
# @Time    : 2020/07/17 15:49
# @Author  : CJG
# @File    : writeRainflow_v1_1.py
'''
v1.1: write_rainflow -> write2singletxt(), write2multitxt(), write2excel()
'''

import os
import numpy as np

try:
    from tool.load_report.readRainflow import readRainflow
    import tool.load_report.reorderVariable as rv
except:
    from readRainflow import readRainflow
    import reorderVariable as rv

class writeRainflow(object):

    def __init__(self,
                 rf_path,
                 variable=('Mx','My','Mz','Fx','Fy','Fz'),
                 content=('DEL','RFC','Markov'),
                 row_start=0,
                 col_start=0,
                 row_space=2,
                 height_flag=False):
        
        self.rf_path  = rf_path
        self.content  = content    # content tuple to output ('DEL','RFC','Markov')
        self.variable = variable   # variable tuple to output ('Mx','My','Fz')
        self.row_start= row_start
        self.col_start= col_start
        self.row_space= row_space
        self.hei_flag = height_flag
        self.rainflow = []         # rainflow data extracted from rfpath
        self.dir_list = []         # sub directions (blade/tower stations)

        self.read_rainflow()
        # self.write_rainflow()

    def read_rainflow(self):
        # for now, ALL variables in the given rfpath are extracted,
        # though for example only 'MxyBR' are required to ouput to excel.
        # AND, contents ('DEL','RFC','Markov') are optional.

        height_file = {}
        height_path = {}

        if not os.path.isdir(self.rf_path):
            raise Exception('%s is not a path!' %self.rf_path)

            # for results with heght
        if self.hei_flag:
            files = [file for file in os.listdir(self.rf_path) if os.path.isdir(os.path.join(self.rf_path, file))]
            if not files:
                raise Exception('%s is empty!' %self.rf_path)

            for file in files:
                # for 01_BRS and tower with height
                if os.path.isdir(os.path.join(self.rf_path, file)) and ('br2' not in file) and ('br3' not in file):
                    # if sub folders exist in self.rf_path (e.g. blade stations, tower stations),
                    # read all subfolder results into list 'rainflow'
                    path = os.path.join(self.rf_path, file)

                    if os.path.isdir(path):
                        # print(path)
                        height_file.update({float(file.split('_')[-1]): file})
                        height_path.update({float(file.split('_')[-1]): path})

            height = list(height_file.keys())
            height.sort()
            # print(height)

            for h in height:

                self.dir_list.append(height_file[h])
                res = readRainflow(height_path[h], self.content)
                if not res:
                    raise Exception('No results in %s' %height_path[h])
                self.rainflow.append(res)
        # for results with out height
        else:
            files = [file for file in os.listdir(self.rf_path) if os.path.isdir(os.path.join(self.rf_path, file))]
            files.sort()
            print(files)

            # for path contains only one results
            if not files:
                self.dir_list.append(self.rf_path.split(os.sep)[-1])
                # print(self.rf_path.split(os.sep)[-1])
                res = readRainflow(self.rf_path, self.content)
                if not res:
                    raise Exception('No results in %s' %self.rf_path)
                self.rainflow.append(res)
            # for path contains many subfiles
            else:
                for file in files:
                    file_path = os.path.join(self.rf_path, file)
                    self.dir_list.append(file)
                    res = readRainflow(file_path, self.content)
                    if not res:
                        raise Exception('No results in %s'%self.rf_path)
                    self.rainflow.append(res)
        # print(self.dir_list)

    def write2singletxt(self, txtfile):
        # ---------------------------------------- #
        # write DEL, Markov to TXTs (blade, tower)
        # additional input: txt file name with full path.
        # ---------------------------------------- #
        rfpath = self.rf_path
        dirs   = os.listdir(rfpath)

        if os.path.isdir(os.path.join(rfpath, dirs[0])) and '.$' not in dirs:

            if 'DEL' in self.content:

                with open(txtfile,'w') as f:
                    f.write('')  # clear existing content
                with open(txtfile,'ab') as f:
                    for i in range(len(self.rainflow)):
                        DEL = self.rainflow[i].DEL
                        varout = rv.reorderVariable(DEL['varlist'], self.variable)
                        if i == 0:
                            header = self.dirlist[i] + '\r\n'
                        else:
                            header = '\r\n' + self.dirlist[i] + '\r\n'
                        header += 'SN\t' + '\t'.join(varout) + '\r\n'
                        header += '-' + '\tkNm'*int(len(varout)/2) + '\tkN'*int(len(varout)/2) + '\r'
                        data = np.array([round(float(mvalue),1) for mvalue in DEL['m']]).reshape(-1,1)
                        for var in varout:
                            data = np.hstack((data, DEL['var_del'][var].reshape(-1,1)/1000))
                        np.savetxt(f, data, header=header, fmt='%.f'+'\t%.2f'*len(varout)+'\r', delimiter='\t', comments='')

            if 'Markov' in self.content:

                with open(txtfile,'w') as f:
                    f.write('')  # clear existing content
                with open(txtfile,'ab') as f:
                    for i in range(len(self.rainflow)):
                        Markov = self.rainflow[i].markov

                        # output variable are arranged according to self.variable.
                        varout = rv.reorderVariable(Markov['varlist'], self.variable)
                        for var in varout:
                            mean_col = Markov['var_mean'][var].reshape(-1, 1)/1000
                            range_row = np.array(Markov['var_range'][var] / 1000)
                            header  = 'Number of cycles [' + var + '] [.] against Cycle range [kNm]\r\n'
                            header += 'Cycle mean [kNm]\t' + '\t'.join([str(m) for m in range_row]) + '\r'
                            data = np.hstack((mean_col, Markov['var_cycles'][var]))
                            np.savetxt(f, data, header=header, fmt='%.2f' + '\t%.2f'*len(range_row) + '\r', delimiter='\t', comments='')

            if 'RFC' in self.content:
                print('For now RFC output at blade/tower stations are not supported!')

    def write2multitxt(self, outpath):
        # ---------------------------------------- #
        # write multi variables' Markov to TXT
        # additional input: out path
        # Markov out txt files named by variable name.
        # ---------------------------------------- #
        rfpath = self.rf_path
        dirs = os.listdir(rfpath)
        if not os.path.exists(outpath):
            os.mkdir(outpath)

        if not os.path.isdir(os.path.join(rfpath, dirs[0])) or ('.$' or '.in') in dirs:

            if len(self.rainflow) != 1:
                print('More than one rainflow result exist in %s, output ONLY the first one!' % self.rf_path)

            if 'Markov' in self.content:
                Markov = self.rainflow[0].markov
                varout = rv.reorderVariable(Markov['varlist'], self.variable)

                for var in varout:
                    txtfile = os.path.join(outpath, var+'_Markov.txt')
                    with open(txtfile,'wb') as f:
                        mean_col = Markov['var_mean'][var].reshape(-1, 1)/1000
                        range_row = np.array(Markov['var_range'][var] / 1000)
                        header = rfpath+'\r\n'
                        header += 'Number of cycles [' + var + '] [.] against Cycle range [kNm]\r\n'
                        header += 'Cycle mean [kNm]\t' + '\t'.join([str(m) for m in range_row]) + '\r'
                        data = np.hstack((mean_col, Markov['var_cycles'][var]))
                        np.savetxt(f, data, header=header, fmt='%.2f' + '\t%.2f'*len(range_row) + '\r', delimiter='\t', comments='')

            if ('DEL' or 'RFC') in self.content:
                pass

    def write2excel(self, table, sheetname):
        # ---------------------------------------- #
        # write DEL, Markov, RFC to EXCEL
        # additional input: opened table;
        # Markov/DEL/RFC are output to sheets of this table with specified sheetnames.
        # ---------------------------------------- #

        for index, dir in enumerate(self.dir_list):
            print('begin to write fatigue： %s' %dir)

            rf_res = self.rainflow[index]
            # print(rf_res)

            if 'DEL' in self.content:

                DEL = rf_res.DEL

                try:
                    if sheetname in table.sheetnames:
                        sheet = table[sheetname]
                    else:
                        sheet = table.create_sheet(sheetname)
                    # print(sheet)
                    # row_start = 6
                    # col_start = 2
                    row_start = self.row_start+index*(12+int(self.row_space))
                    col_start = self.col_start+1
                    data_len  = len(DEL['m'])

                    # file name
                    sheet.cell(row_start+1, col_start, dir)

                    # mvariable name
                    sheet.cell(row_start+2, col_start, 'SN Slop[-]')

                    for index,var in enumerate(self.variable):
                        if 'M' in var:
                            sheet.cell(row_start+2, col_start+1+index, var+'[kNm]')
                        elif 'F' in var:
                            sheet.cell(row_start+2, col_start+1+index, var+'[kN]')

                    row_start += 3
                    for i in range(data_len):
                        mvalue = DEL['m'][i]
                        sheet.cell(row_start+i, col_start, round(float(mvalue),1))

                    # output variable are arranged according to the excel template,
                    # though in fact the variables has been defined by self.variable.
                    # print(DEL['varlist'])
                    varout = rv.reorderVariable(DEL['varlist'], self.variable)

                    for i, var in enumerate(varout):
                        for j in range(data_len):
                            # print(DEL['var_del'][var])
                            sheet.cell(row_start+j, col_start+1+i, DEL['var_del'][var][j, 0]/1000)  # DEL

                except IOError:
                    print('no sheet named %s found in template table!' % sheetname)

if __name__ == '__main__':

    # # rfpath = r'\\172.20.0.4\fs02\LSY\V3\loop04_backup20200623\BackUp\post\rainflow\br_25y_1E7'
    # rfpath = r'\\172.20.0.4\fs02\LSY\V3\loop04_backup20200623\BackUp\post\rainflow\main_25y_1E7\Blade_root'
    # excelouttemple = r"E:\01 tool dev\02 post tool\06_readBladedPost\out_test\test_RF_pb - 副本 - 副本.xlsx"
    # excelout = r"E:\01 tool dev\02 post tool\06_readBladedPost\out_test\test_RF_pb.xlsx"


    import openpyxl

    excel   = r"C:\Users\10700700\Desktop\tool\py\20_Load Assistant Tool\test\report\Compare_results.xlsx"
    table = openpyxl.load_workbook(excel)
    # print(table)
    #
    # # writeRainflow(rfpath, table, variable=('Mx','My','Mz'), content=('DEL','RFC','Markov'))
    # writeRainflow(rfpath, table, content=('DEL','RFC','Markov'))
    #
    # table.save(excelout)

    # rfpath = r'E:\01 tool dev\02 post tool\06_readBladedPost\out_test\V3-LSY\post\br_25y_1E7'
    # txtout = r"E:\01 tool dev\02 post tool\06_readBladedPost\out_test\br_markov.txt"
    # writeRainflow(rfpath, txt=txtout, content=('Markov',))
    # txtout = r"E:\01 tool dev\02 post tool\06_readBladedPost\out_test\br_del.txt"
    # writeRainflow(rfpath, txt=txtout, content=('DEL'))

    rf_path  = r'\\172.20.0.4\fs02\CE\V3\post_test\08_Fatigue\05_Main\br1'

    writeRainflow(rf_path, content=('DEL',)).write2excel(table, 'Fatigue')
    table.save(excel)


