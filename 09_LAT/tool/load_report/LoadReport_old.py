# ！usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2019/10/26 11:14
# @Author  : CE
# @File    : joblist.py
'''
1 本程序只适用于通过additional controller parameters对话框来设置控制器参数，不适用于有多个控制器时的模型；
2 有多个控制器的模型需要在每个控制器中定义单独的参数文件；
'''

import sys, os
import configparser
import openpyxl

from PyQt5.QtWidgets import (QMainWindow,
                             QApplication,
                             QDesktopWidget,
                             QLabel,
                             qApp,
                             QLineEdit,
                             QPushButton,
                             QFileDialog,
                             QGridLayout,
                             QVBoxLayout,
                             QWidget,
                             QAction,
                             QMessageBox,
                             QGroupBox)
from PyQt5.QtGui import QIcon, QFont
from PyQt5 import QtCore

from tool.load_report.writeUltimate import ReadPostMX
from tool.load_report.writeRainflow import writeRainflow
from tool.load_report.writeOccurrence import Occurrence

class LoadReportWindow(QMainWindow):

    def __init__(self, parent=None):

        super(LoadReportWindow, self).__init__(parent)
        self.mywidget = QWidget(self)

        self.setMinimumSize(800, 650)
        self.setWindowTitle('Load Report')
        self.setWindowIcon(QIcon(r".\icon\excel1.ico"))
        # root = QFileInfo(__file__).absolutePath()
        # self.setWindowIcon(QIcon(root + "./icon/Text_Edit.ico"))

        self.job_list = None
        self.lis_name = None
        self.job_path = None
        self.res_path = None
        self.dll_path = None
        self.xml_path = None
        self.njl_path = None
        self.para_con = None
        self.othe_con = None

        self.initUI()
        self.load_setting()
        self.center()

    def initUI(self):

        menubar = self.menuBar()
        fileMenu = menubar.addMenu('File')
        helpMenu = menubar.addMenu('Help')

        saveAction = QAction(QIcon('./Icon/save.ico'), 'Save', self)
        saveAction.setShortcut('Ctrl+S')
        saveAction.triggered.connect(self.save_setting)
        fileMenu.addAction(saveAction)

        clearAction = QAction(QIcon('Icon/clear.ico'), 'Clear', self)
        clearAction.setShortcut('Ctrl+R')
        clearAction.triggered.connect(self.clear_setting)
        fileMenu.addAction(clearAction)

        exitAction = QAction(QIcon('Icon/exit.ico'), 'Exit', self)
        exitAction.setShortcut('Ctrl+Q')
        exitAction.triggered.connect(qApp.quit)
        fileMenu.addAction(exitAction)

        helpAction = QAction(QIcon('Icon/doc.ico'), 'User Manual', self)
        helpAction.setShortcut('Ctrl+H')
        helpAction.triggered.connect(self.user_manual)
        helpMenu.addAction(helpAction)

        # Blade
        # root axes
        self.label1 = QLabel("Root axes(inclsf)")
        self.label1.setFont(QFont("Calibri", 10))
        self.line1 = QLineEdit()
        self.line1.setFont(QFont("Calibri", 10))
        self.line1.setPlaceholderText("Pick ultimate root axes(inclsf) path")
        self.btn1 = QPushButton("...")
        self.btn1.setFont(QFont("Calibri", 10))
        self.btn1.clicked.connect(self.load_root_inclsf)

        self.label2 = QLabel("Root axes(exclsf)")
        self.label2.setFont(QFont("Calibri", 10))
        self.line2 = QLineEdit()
        self.line2.setFont(QFont("Calibri", 10))
        self.line2.setPlaceholderText("Pick ultimate root axes(exclsf) path")
        self.btn2 = QPushButton("...")
        self.btn2.setFont(QFont("Calibri", 10))
        self.btn2.clicked.connect(self.load_root_exclsf)

        self.label3 = QLabel("Root axes(fatigue)")
        self.label3.setFont(QFont("Calibri", 10))
        self.line3 = QLineEdit()
        self.line3.setFont(QFont("Calibri", 10))
        self.line3.setPlaceholderText("Pick root axes fatigue path")
        self.btn3 = QPushButton("...")
        self.btn3.setFont(QFont("Calibri", 10))
        self.btn3.clicked.connect(self.load_root_fatigue)

        # user axes
        self.label4 = QLabel("User axes(inclsf)")
        self.label4.setFont(QFont("Calibri", 10))
        self.line4 = QLineEdit()
        self.line4.setFont(QFont("Calibri", 10))
        self.line4.setPlaceholderText("Pick ultimate user axes(inclsf) path")
        self.btn4 = QPushButton("...")
        self.btn4.setFont(QFont("Calibri", 10))
        self.btn4.clicked.connect(self.load_user_inclsf)

        self.label5 = QLabel("User axes(exclsf)")
        self.label5.setFont(QFont("Calibri", 10))
        self.line5 = QLineEdit()
        self.line5.setFont(QFont("Calibri", 10))
        self.line5.setPlaceholderText("Pick ultimate user axes(exclsf) path")
        self.btn5 = QPushButton("...")
        self.btn5.setFont(QFont("Calibri", 10))
        self.btn5.clicked.connect(self.load_user_exclsf)

        self.label6 = QLabel("User axes(fatigue)")
        self.label6.setFont(QFont("Calibri", 10))
        self.line6 = QLineEdit()
        self.line6.setFont(QFont("Calibri", 10))
        self.line6.setPlaceholderText("Pick user axes fatigue path")
        self.btn6 = QPushButton("...")
        self.btn6.setFont(QFont("Calibri", 10))
        self.btn6.clicked.connect(self.load_user_fatigue)

        # tower
        self.label7 = QLabel("Tower flange(inclsf)")
        self.label7.setFont(QFont("Calibri", 10))
        self.line7 = QLineEdit()
        self.line7.setFont(QFont("Calibri", 10))
        self.line7.setPlaceholderText("Pick tower flange(inclsf) path")
        self.btn7 = QPushButton("...")
        self.btn7.setFont(QFont("Calibri", 10))
        self.btn7.clicked.connect(self.load_tower_inclsf)

        self.label8 = QLabel("Tower flange(exclsf)")
        self.label8.setFont(QFont("Calibri", 10))
        self.line8 = QLineEdit()
        self.line8.setFont(QFont("Calibri", 10))
        self.line8.setPlaceholderText("Pick tower flange(exclsf) path")
        self.btn8 = QPushButton("...")
        self.btn8.setFont(QFont("Calibri", 10))
        self.btn8.clicked.connect(self.load_tower_exclsf)

        self.label9 = QLabel("Tower flange(dlc12)")
        self.label9.setFont(QFont("Calibri", 10))
        self.line9 = QLineEdit()
        self.line9.setFont(QFont("Calibri", 10))
        self.line9.setPlaceholderText("Pick tower flange(dlc12) path")
        self.btn9 = QPushButton("...")
        self.btn9.setFont(QFont("Calibri", 10))
        self.btn9.clicked.connect(self.load_tower_dlc12)

        self.label10 = QLabel("Tower flange(fatigue)")
        self.label10.setFont(QFont("Calibri", 10))
        self.line10 = QLineEdit()
        self.line10.setFont(QFont("Calibri", 10))
        self.line10.setPlaceholderText("Pick tower flange(fatigue) path")
        self.btn10 = QPushButton("...")
        self.btn10.setFont(QFont("Calibri", 10))
        self.btn10.clicked.connect(self.load_tower_fatigue)

        self.label11 = QLabel("Load Case Table")
        self.label11.setFont(QFont("Calibri", 10))
        self.line11 = QLineEdit()
        self.line11.setFont(QFont("Calibri", 10))
        self.line11.setPlaceholderText("Pick load case table")
        self.btn11 = QPushButton("...")
        self.btn11.setFont(QFont("Calibri", 10))
        self.btn11.clicked.connect(self.load_lct)

        # main
        self.label13 = QLabel("Main ultimate path")
        self.label13.setFont(QFont("Calibri", 10))
        self.line13 = QLineEdit()
        self.line13.setFont(QFont("Calibri", 10))
        self.line13.setPlaceholderText("Pick main ultimate path")
        self.btn13 = QPushButton("...")
        self.btn13.setFont(QFont("Calibri", 10))
        self.btn13.clicked.connect(self.load_ultimate_main)

        self.label14 = QLabel("Main fatigue path")
        self.label14.setFont(QFont("Calibri", 10))
        self.line14 = QLineEdit()
        self.line14.setFont(QFont("Calibri", 10))
        self.line14.setPlaceholderText("Pick main fatigue path")
        self.btn14 = QPushButton("...")
        self.btn14.setFont(QFont("Calibri", 10))
        self.btn14.clicked.connect(self.load_fatigue_main)

        # output
        self.label12 = QLabel("Output path")
        self.label12.setFont(QFont("Calibri", 10))
        self.line12 = QLineEdit()
        self.line12.setFont(QFont("Calibri", 10))
        self.line12.setPlaceholderText("Pick output path")
        self.btn12 = QPushButton("...")
        self.btn12.setFont(QFont("Calibri", 10))
        self.btn12.clicked.connect(self.load_output_path)

        self.btn15 = QPushButton("Generate load report input excel")
        self.btn15.setFont(QFont("Calibri", 10))
        self.btn15.clicked.connect(self.generate_excel)

        self.group1 = QGroupBox('Blade input')
        self.group1.setFont(QFont("Calibri", 8))
        self.grid1 = QGridLayout()
        # 起始行，起始列，占用行，占用列
        self.grid1.addWidget(self.label1, 0, 0, 1, 1)
        self.grid1.addWidget(self.line1, 0, 1, 1, 5)
        self.grid1.addWidget(self.btn1, 0, 6, 1, 1)
        self.grid1.addWidget(self.label2, 1, 0, 1, 1)
        self.grid1.addWidget(self.line2, 1, 1, 1, 5)
        self.grid1.addWidget(self.btn2, 1, 6, 1, 1)
        self.grid1.addWidget(self.label3, 2, 0, 1, 1)
        self.grid1.addWidget(self.line3, 2, 1, 1, 5)
        self.grid1.addWidget(self.btn3, 2, 6, 1, 1)
        self.grid1.addWidget(self.label4, 3, 0, 1, 1)
        self.grid1.addWidget(self.line4, 3, 1, 1, 5)
        self.grid1.addWidget(self.btn4, 3, 6, 1, 1)
        self.grid1.addWidget(self.label5, 4, 0, 1, 1)
        self.grid1.addWidget(self.line5, 4, 1, 1, 5)
        self.grid1.addWidget(self.btn5, 4, 6, 1, 1)
        self.grid1.addWidget(self.label6, 5, 0, 1, 1)
        self.grid1.addWidget(self.line6, 5, 1, 1, 5)
        self.grid1.addWidget(self.btn6, 5, 6, 1, 1)
        self.group1.setLayout(self.grid1)

        self.group2 = QGroupBox('Tower Input')
        self.group2.setFont(QFont("Calibri", 8))
        self.grid2 = QGridLayout()
        # 起始行，起始列，占用行，占用列
        self.grid2.addWidget(self.label7, 0, 0, 1, 1)
        self.grid2.addWidget(self.line7, 0, 1, 1, 5)
        self.grid2.addWidget(self.btn7, 0, 6, 1, 1)
        self.grid2.addWidget(self.label8, 1, 0, 1, 1)
        self.grid2.addWidget(self.line8, 1, 1, 1, 5)
        self.grid2.addWidget(self.btn8, 1, 6, 1, 1)
        self.grid2.addWidget(self.label9, 2, 0, 1, 1)
        self.grid2.addWidget(self.line9, 2, 1, 1, 5)
        self.grid2.addWidget(self.btn9, 2, 6, 1, 1)
        self.grid2.addWidget(self.label10, 3, 0, 1, 1)
        self.grid2.addWidget(self.line10, 3, 1, 1, 5)
        self.grid2.addWidget(self.btn10, 3, 6, 1, 1)
        self.group2.setLayout(self.grid2)

        # load case table
        self.group3 = QGroupBox('Loadcase Table')
        self.group3.setFont(QFont("Calibri", 8))
        self.grid3 = QGridLayout()
        # 起始行，起始列，占用行，占用列
        self.grid3.addWidget(self.label11, 0, 0, 1, 1)
        self.grid3.addWidget(self.line11, 0, 1, 1, 5)
        self.grid3.addWidget(self.btn11, 0, 6, 1, 1)
        self.group3.setLayout(self.grid3)

        # output
        self.group4 = QGroupBox('Output path')
        self.group4.setFont(QFont("Calibri", 8))
        self.grid4 = QGridLayout()
        # 起始行，起始列，占用行，占用列
        self.grid4.addWidget(self.label12, 0, 0, 1, 1)
        self.grid4.addWidget(self.line12, 0, 1, 1, 5)
        self.grid4.addWidget(self.btn12, 0, 6, 1, 1)
        self.group4.setLayout(self.grid4)

        # main
        self.group5 = QGroupBox('Main input')
        self.group5.setFont(QFont("Calibri", 8))
        self.grid5 = QGridLayout()
        # 起始行，起始列，占用行，占用列
        self.grid5.addWidget(self.label13, 0, 0, 1, 1)
        self.grid5.addWidget(self.line13, 0, 1, 1, 5)
        self.grid5.addWidget(self.btn13, 0, 6, 1, 1)
        self.grid5.addWidget(self.label14, 1, 0, 1, 1)
        self.grid5.addWidget(self.line14, 1, 1, 1, 5)
        self.grid5.addWidget(self.btn14, 1, 6, 1, 1)
        self.group5.setLayout(self.grid5)

        self.main_layout = QVBoxLayout()
        self.main_layout.addWidget(self.group5)
        self.main_layout.addWidget(self.group1)
        self.main_layout.addWidget(self.group2)
        self.main_layout.addWidget(self.group3)
        self.main_layout.addWidget(self.group4)
        self.main_layout.addWidget(self.btn15)
        self.main_layout.addStretch(1)

        self.mywidget.setLayout(self.main_layout)
        self.setCentralWidget(self.mywidget)

    def keyPressEvent(self, e):
        if e.key() == QtCore.Qt.Key_Escape:
            self.close()

    def save_setting(self):
        config = configparser.ConfigParser()
        config.read('config.ini', encoding='utf-8')

        if not config.has_section('Load Report'):
            config.add_section('Load Report')

        config['Load Report'] = {'Root axes_inclsf':self.line1.text(),
                                 'Root axes_exclsf':self.line2.text(),
                                 'Root axes_fatigue':self.line3.text(),
                                 'User axes_inclsf':self.line4.text(),
                                 'User axes_exclsf':self.line5.text(),
                                 'User axes_fatigue':self.line6.text(),
                                 'Tower_inclsf': self.line7.text(),
                                 'Tower_exclsf': self.line8.text(),
                                 'Tower_dlc12': self.line9.text(),
                                 'Tower_flange': self.line10.text(),
                                 'Loadcase table': self.line11.text(),
                                 'Output path': self.line12.text(),
                                 'Ultimate path':self.line13.text(),
                                 'Fatigue path':self.line14.text()}

        with open("config.ini",'w') as f:
            config.write(f)

    def load_setting(self):
        config = configparser.ConfigParser()
        config.read('config.ini', encoding='utf-8')

        if config.has_section('Load Report'):
            self.line1.setText(config.get('Load Report','Root axes_inclsf'))
            self.line2.setText(config.get('Load Report','Root axes_exclsf'))
            self.line3.setText(config.get('Load Report','Root axes_fatigue'))
            self.line4.setText(config.get('Load Report', 'User axes_inclsf'))
            self.line5.setText(config.get('Load Report', 'User axes_exclsf'))
            self.line6.setText(config.get('Load Report', 'User axes_fatigue'))
            self.line7.setText(config.get('Load Report', 'Tower_inclsf'))
            self.line8.setText(config.get('Load Report', 'Tower_exclsf'))
            self.line9.setText(config.get('Load Report', 'Tower_dlc12'))
            self.line10.setText(config.get('Load Report', 'Tower_flange'))
            self.line11.setText(config.get('Load Report', 'Loadcase table'))
            self.line12.setText(config.get('Load Report', 'Output path'))
            self.line13.setText(config.get('Load Report', 'Ultimate path'))
            self.line14.setText(config.get('Load Report', 'Fatigue path'))

    def clear_setting(self):
        self.line1.setText('')
        self.line2.setText('')
        self.line3.setText('')
        self.line4.setText('')
        self.line5.setText('')
        self.line6.setText('')
        self.line7.setText('')
        self.line8.setText('')
        self.line9.setText('')
        self.line10.setText('')
        self.line11.setText('')
        self.line12.setText('')
        self.line13.setText('')
        self.line14.setText('')

    def user_manual(self):
        os.startfile(os.getcwd()+os.sep+'user manual\Load Report.docx')

    def load_ultimate_main(self):
        main_path = QFileDialog.getExistingDirectory(self, "Choose result path", ".")

        if main_path:
            self.line13.setText(main_path.replace('/', '\\'))

    def load_fatigue_main(self):
        main_path = QFileDialog.getExistingDirectory(self, "Choose result path", ".")

        if main_path:
            self.line14.setText(main_path.replace('/', '\\'))

    def load_root_inclsf(self):
        root_path = QFileDialog.getExistingDirectory(self, "Choose result path", ".")

        if root_path:
            self.line1.setText(root_path.replace('/', '\\'))

    def load_root_exclsf(self):
        root_path = QFileDialog.getExistingDirectory(self, "Choose result path", ".")

        if root_path:
            self.line2.setText(root_path.replace('/', '\\'))

    def load_root_fatigue(self):
        root_path = QFileDialog.getExistingDirectory(self, "Choose result path", ".")

        if root_path:
            self.line3.setText(root_path.replace('/', '\\'))

    def load_user_inclsf(self):
        user_path = QFileDialog.getExistingDirectory(self, "Choose result path", ".")

        if user_path:
            self.line4.setText(user_path.replace('/', '\\'))

    def load_user_exclsf(self):
        user_path = QFileDialog.getExistingDirectory(self, "Choose result path", ".")

        if user_path:
            self.line5.setText(user_path.replace('/', '\\'))

    def load_user_fatigue(self):
        user_path = QFileDialog.getExistingDirectory(self, "Choose result path", ".")

        if user_path:
            self.line6.setText(user_path.replace('/', '\\'))

    def load_tower_inclsf(self):
        tower_path = QFileDialog.getExistingDirectory(self, "Choose result path", ".")

        if tower_path:
            self.line7.setText(tower_path.replace('/', '\\'))

    def load_tower_exclsf(self):
        tower_path = QFileDialog.getExistingDirectory(self, "Choose result path", ".")

        if tower_path:
            self.line8.setText(tower_path.replace('/', '\\'))

    def load_tower_dlc12(self):
        tower_path = QFileDialog.getExistingDirectory(self, "Choose result path", ".")

        if tower_path:
            self.line9.setText(tower_path.replace('/', '\\'))

    def load_tower_fatigue(self):
        tower_path = QFileDialog.getExistingDirectory(self, "Choose result path", ".")

        if tower_path:
            self.line10.setText(tower_path.replace('/', '\\'))

    def load_lct(self):
        lct_path = QFileDialog.getOpenFileName(self, "Choose load case table path", ".", "dll(*.dll)")

        if lct_path:
            self.line11.setText(lct_path.replace('/', '\\'))

    def load_output_path(self):
        output_path = QFileDialog.getExistingDirectory(self, "Choose output path", ".")

        if output_path:
            self.line12.setText(output_path.replace('/', '\\'))
            self.line12.home(True)

    def generate_excel(self):

        root_inclsf  = os.path.exists(self.line1.text())
        root_exclsf  = os.path.exists(self.line2.text())
        root_fatigue = os.path.exists(self.line3.text())
        user_inclsf  = os.path.exists(self.line4.text())
        user_exclsf  = os.path.exists(self.line5.text())
        user_fatigue = os.path.exists(self.line6.text())
        tower_inclsf = os.path.exists(self.line7.text())
        tower_exclsf = os.path.exists(self.line8.text())
        tower_dlc12  = os.path.exists(self.line9.text())
        tower_fatigue= os.path.exists(self.line10.text())
        lct_path     = os.path.exists(self.line11.text())
        output_path  = os.path.exists(self.line12.text())
        ultimate     = os.path.exists(self.line13.text())
        fatigue      = os.path.exists(self.line14.text())

        if all((root_inclsf, root_exclsf, root_fatigue, user_exclsf, user_inclsf, user_fatigue,
                tower_inclsf, tower_exclsf, tower_fatigue,tower_dlc12,
                ultimate, fatigue, lct_path,output_path)):

            # write compare result
            table      = openpyxl.Workbook()
            sheet_name = 'Ultimate'
            ultimate   = ReadPostMX(ultimate, sheet_name, column_start=1, row_start=1, row_space=2, height_flag=False)
            ultimate.write_result(table)

            sheet_name = 'Fatigue'
            table.create_sheet(sheet_name)
            fatigue    = writeRainflow(fatigue, content=('DEL',),row_start=2,col_start=8,height_flag=False)
            fatigue.write2excel(table, sheet_name)

            excel_path = os.path.join(output_path, 'Compare_results.xlsx')
            table.save(excel_path)

            # write blade
            table      = openpyxl.Workbook()
            sheet_name = 'extreme root section'
            table.create_sheet(sheet_name)
            ultimate   = ReadPostMX(root_inclsf, sheet_name, 0)
            ultimate.write_result(table)
            ultimate   = ReadPostMX(root_exclsf, sheet_name, 13)
            ultimate.write_result(table)

            sheet_name = 'extreme user section'
            table.create_sheet(sheet_name)
            ultimate   = ReadPostMX(user_inclsf, sheet_name, 0)
            ultimate.write_result(table)
            ultimate   = ReadPostMX(user_exclsf, sheet_name, 13)
            ultimate.write_result(table)

            sheet_name = 'fatigue root section'
            table.create_sheet(sheet_name)
            fatigue    = writeRainflow(root_fatigue, content=('DEL',))
            fatigue.write2excel(table, sheet_name)

            sheet_name = 'fatigue user section'
            table.create_sheet(sheet_name)
            fatigue    = writeRainflow(user_fatigue, content=('DEL',))
            fatigue.write2excel(table, sheet_name)

            excel_path = os.path.join(output_path, 'Blade.xlsx')
            table.save(excel_path)

            # write tower
            table      = openpyxl.Workbook()
            sheet_name = 'extreme flange section'
            table.create_sheet(sheet_name)
            ultimate   = ReadPostMX(tower_inclsf, sheet_name, 0)
            ultimate.write_result(table)

            sheet_name = 'occurrence'
            Occurrence(self.lin11.text(), table, sheet_name)

            ultimate   = ReadPostMX(tower_exclsf, sheet_name, 13)
            ultimate.write_result(table)

            sheet_name = 'extreme flange section DLC12'
            table.create_sheet(sheet_name)
            ultimate   = ReadPostMX(tower_dlc12, sheet_name, 0)
            ultimate.write_result(table)

            sheet_name = 'fatigue flange section'
            table.create_sheet(sheet_name)

            fatigue    = writeRainflow(tower_fatigue, content=('DEL',))
            fatigue.write2excel(table, sheet_name)

            excel_path = os.path.join(output_path, 'Tower.xlsx')
            table.save(excel_path)

        elif not ultimate:
            QMessageBox(self,'warnning','Please choose main ultimate result first!')
        elif not fatigue:
            QMessageBox(self,'warnning','Please choose main fatigue result first!')
        elif not root_inclsf:
            QMessageBox(self,'warnning','Please choose root axes(inclsf) result first!')
        elif not root_exclsf:
            QMessageBox(self,'warnning','Please choose root axes(exclsf) result first!')
        elif not root_fatigue:
            QMessageBox(self,'warnning','Please choose root axes fatigue result first!')
        elif not user_inclsf:
            QMessageBox(self,'warnning','Please choose user axes(inclsf) result first!')
        elif not user_exclsf:
            QMessageBox(self,'warnning','Please choose user axes(exclsf) result first!')
        elif not user_fatigue:
            QMessageBox(self,'warnning','Please choose user axes(fatigue) result first!')
        elif not tower_inclsf:
            QMessageBox(self,'warnning','Please choose tower flange(inclsf) result first!')
        elif not tower_exclsf:
            QMessageBox(self,'warnning','Please choose tower flange(exclsf) result first!')
        elif not tower_dlc12:
            QMessageBox(self,'warnning','Please choose tower flange(dlc12) result first!')
        elif not tower_fatigue:
            QMessageBox(self,'warnning','Please choose tower flange(fatigue) result first!')
        elif not lct_path:
            QMessageBox(self,'warnning','Please choose load case table first!')
        elif not output_path:
            QMessageBox(self,'warnning','Please choose output path first!')

    def center(self):

        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

if __name__ == '__main__':

    app = QApplication(sys.argv)
    window = LoadReportWindow()
    # window.setWindowOpacity(0.9)
    window.show()
    sys.exit(app.exec_())

