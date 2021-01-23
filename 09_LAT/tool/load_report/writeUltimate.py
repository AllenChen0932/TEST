# -*- coding: utf-8 -*-
# @Time    : 2020/07/31 11:10
# @Author  : TJ
# @File    : writeUltimate.py

import os
import openpyxl
from openpyxl.styles import Border, Side, Font, PatternFill

class ReadPostMX(object):

    def __init__(self, ultimate_path, sheet_name, column_start=0, row_start=0, row_space=1, height_flag=False):

        self.post_path    = ultimate_path
        # name = self.post_path.split('\\')[-1].split('_')
        # self.name = name[-2]+name[-1]
        self.sheet_name   = sheet_name
        self.column_start = column_start
        self.row_start    = row_start
        self.row_space    = row_space
        self.height_flag  = height_flag

        self.file_list    = []

        self.get_channel()

    def read_dollarMX(self, file):

        data = []
        with open(file, 'r') as f:

            lines = f.readlines()
            for line in lines:

                if ('MAXIMUM' in line) or ('MINIMUM' in line):
                    for item in [line.strip().split('\t')]:
                        name = item[0].split("'")[1].split(" ")

                        if 'Tower' in name:
                            item[0] = name[1].split(',')[0]
                        elif 'axes),' in name:      # blade br, ur
                            item[0] = name[1]
                        elif ('1' in name) or ('2' in name) or ('3' in name):
                            item[0] = name[1]+name[2]+name[3]
                        else:
                            item[0] = name[2]

                        icol = len(item)
                        for j in range(icol - 4):
                            item[j+3] = float(item[j+3])/1000
                        item[icol-1] = round(float(item[icol -1]), 2)

                        item = ['Max' if ix == 'MAXIMUM ' else ix for ix in item]
                        item = ['Min' if ix == 'MINIMUM ' else ix for ix in item]

                        data.append(item)
        # print(data)
        return(data)

    def get_channel(self):

        height_file = {}

        if self.height_flag:
            for root, dirs, files in os.walk(self.post_path, False):
                pj_num = 0

                for file in files:
                    if '.$PJ' in file.upper():
                        pj_num += 1
                        pj_name = file.split('.$')[0]
                        mx_name = ''.join([pj_name, '.$MX'])
                        mx_file = os.path.join(root, mx_name)
                        if not os.path.isfile(mx_file):
                            raise Exception('No results in %s' % mx_file)

                        height  = float(pj_name.split('_')[-1])
                        # print(height, mx_file)
                        height_file[height] = mx_file

            height = list(height_file.keys())
            height.sort()

            self.file_list = [height_file[h] for h in height]

        else:
            for root, dirs, files in os.walk(self.post_path, False):
                pj_num = 0

                for file in files:
                    if '.$PJ' in file.upper():
                        pj_num += 1
                        pj_name = file.split('.$')[0]
                        mx_name = ''.join([pj_name, '.$MX'])
                        mx_file = os.path.join(root, mx_name)
                        if not os.path.isfile(mx_file):
                            raise Exception('No results in %s' % mx_file)

                        self.file_list.append(mx_file)

    def write_result(self, xls):
        sheet = xls[self.sheet_name]

        irow = self.row_start
        for f_idx, file in enumerate(self.file_list):

            print('Open file [%d of %d]: %s' % (f_idx+1, len(self.file_list), file))
            data = self.read_dollarMX(file)
            name = file.split('\\')[-2]

            nrow     = len(data)
            ncol     = len(data[0])
            nrowfile = nrow+3

            # ---------Sheet <extreme>-------
            # sheet header, start from row 1
            sheet.cell(irow+1, self.column_start+1, name)
            sheet.cell(irow+2, self.column_start+12, 'Safety factor')
            sheet.cell(irow+3, self.column_start+3, 'Load case')
            sheet.cell(irow+3, self.column_start+4, 'kNm')
            sheet.cell(irow+3, self.column_start+5, 'kNm')
            sheet.cell(irow+3, self.column_start+6, 'kNm')
            sheet.cell(irow+3, self.column_start+7, 'kNm')
            sheet.cell(irow+3, self.column_start+8, 'kN')
            sheet.cell(irow+3, self.column_start+9, 'kN')
            sheet.cell(irow+3, self.column_start+10, 'kN')
            sheet.cell(irow+3, self.column_start+11, 'kN')
            sheet.cell(irow+3, self.column_start+12, '-')
            # print(irow+2)

            for i in range(nrow):
                # write variable column
                sheet.cell(irow+2, self.column_start+int(i/2)+4, data[i][0])
                for j in range(ncol):
                    sheet.cell(irow+4+i, self.column_start+j+1, data[i][j])

            # -------set font and border-------
            left, right, top, bottom = [Side(style='thin')]*4
            for i in range(nrow+2):
                for j in range(ncol):
                    sheet.cell(irow+2+i, self.column_start+j+1).border = Border(left=left, right=right, top=top, bottom=bottom)
                    sheet.cell(irow+2+i, self.column_start+j+1).font = Font(name='Times New Roman')
                    sheet.cell(irow+2, self.column_start+j+1).fill = PatternFill('solid', fgColor='FFEFD5')
            for i in range(nrow):
                for j in range(ncol - 4):
                    sheet.cell(irow+j*2+4, self.column_start+j+4).fill = PatternFill('solid', fgColor='FFEFD5')
                    sheet.cell(irow+j*2+5, self.column_start+j+4).fill = PatternFill('solid', fgColor='FFEFD5')
            irow = irow+nrowfile+self.row_space

if __name__ == '__main__':

    excel_path = r"F:\T001_Python\42_LoadReport\writeCompLoads20200807\Tower.xlsx"

    table = openpyxl.load_workbook(excel_path)

    ultimate = r'\\172.20.0.4\fs02\CE\V3\loop05\post_0615\ultimate_tower_dlc12'
    sheet    = 'extreme flange section DLC12'

    mx = ReadPostMX(ultimate, sheet, column_start=0)
    mx.write_result(table)

    table.save(excel_path)

